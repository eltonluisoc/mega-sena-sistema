#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SISTEMA DE GESTÃO DE BOLÕES PRO v5.1
Correções v5.1 (Bolão Selecionado virou aba própria):
 - "Início" agora tem 3 sub-abas: Visão Geral | Bolão Selecionado |
   Pendências por Bolão. Antes o resumo+detalhe do bolão escolhido vinha
   logo abaixo da Visão Geral, na mesma aba com rolagem — ficava uma
   rolagem longa e incômoda. Cada bloco agora tem sua própria aba, mais
   curta e sem precisar rolar tanto.
 - Cartões de seleção de bolão ganharam um badge de status (✅ em dia /
   ⚠ N atrasado(s)) calculado a partir da mesma contagem usada em
   "Participantes Atrasados" — dá pra ver o status de cada bolão sem
   precisar clicar nele.
 - KPIs "Participantes Atrasados" e "Pendente Depósito" (Visão Geral)
   ganharam um prefixo ✅/⚠ pra sinalizar de cara se está tudo em dia.
Correções v5.0 (reestruturação de "Início > Visão Geral"):
 - CORRIGIDO: coluna "Devidas" em Participantes Atrasados mostrava um
   número cumulativo (parcelas esperadas desde o início do bolão), não
   quanto realmente falta — alguém que pagou 7 de 8 parcelas esperadas
   aparecia com "Devidas: 8", parecendo que faltavam 8 e não 1. Virou
   "Faltam", mostrando a diferença real (consistente com a coluna Saldo,
   que já estava certa).
 - Tela reordenada: Visão Geral (todos os bolões) primeiro, com o que é
   essencial no dia a dia — Bolões Ativos, Arrecadado Geral, Pendente de
   Depósito Geral, Participantes Atrasados, e os totais do organizador.
   Novo bloco "Últimos Pagamentos (Geral)".
 - "Ganhos por Loteria" e "Registrar Lançamento" viraram botões que
   abrem janelas próprias — não competem mais por espaço na tela
   principal com informação essencial.
 - Novo seletor de bolão: cartões clicáveis (um por bolão ativo) no
   lugar de depender só do combo pequeno do cabeçalho.
 - Abaixo do seletor: resumo do bolão escolhido primeiro, detalhe depois
   (mesmo conteúdo de antes, só reordenado).
Correções v4.3 (revisão de UX em "Início > Visão Geral"):
 - Árvores "Situação dos Participantes"/"Últimos Pagamentos" reduzidas
   de 20 pra 10 linhas — herdaram altura de tela cheia de quando o
   Dashboard era uma aba sozinha, e agora empurravam a metade
   "Administração" pra bem longe do topo
 - Bloco "Financeiro" simplificado: removidos "Arrecadado", "Depositado"
   e "A Receber" — já apareciam nos KPIs do topo, duplicados
 - Divisória entre "este bolão" e "todos os bolões" virou uma faixa
   colorida de largura total com legenda, no lugar de uma linha de 3px
 - Cores dos KPIs da Administração trocadas — 4 das 5 repetiam cores já
   usadas nos KPIs do Dashboard acima com significados diferentes
 - Formulário "Registrar Lançamento" movido pra logo após os KPIs da
   Administração, antes das tabelas de leitura (era a única ação da
   tela inteira e ficava espremida no meio do scroll)
Correções v4.2:
 - CORRIGIDO: rolagem do mouse não funcionava na tela "Início > Visão
   Geral" — o evento só disparava com o cursor sobre a área vazia do
   canvas, quase impossível na prática já que a tela é preenchida por
   widgets. Agora liga/desliga a rolagem conforme o mouse entra/sai da
   área (cobre estar em cima de qualquer widget filho também)
Correções v4.1:
 - REMOVIDO: aba "✏ Editar" (Financeiro) — redundante com "📋
   Histórico" (busca por nome + duplo-clique já cobria tudo). Histórico
   ganhou um botão "🗑 Excluir Selecionado" pra fechar a única diferença
   real que havia (excluir um pagamento).
 - REMOVIDO: sub-aba "🔄 Sincronizar Participantes" (Site/Publicar) —
   código morto, nada escreve na fila que ela lia desde que o site
   mudou de fluxo. A limpeza de "Pagamentos Órfãos" (que ficava dentro
   dela) continua existindo, agora como botão em "📋 Histórico".
 - MELHORADO: aba "💼 Reserva / Caixa" renomeada pra "💼 Caixa por
   Loteria" — tinha nome parecido demais com "💰 Reservas Pessoais"
   apesar de serem conceitos diferentes (fundo do organizador vs.
   saldo de cada pessoa)
Correções v4.0:
 - MELHORADO: "Início" tinha 3 níveis de abas (Início > Administração >
   Visão Geral/Pendências), exigindo 2 cliques pra chegar em qualquer
   uma. Agora tem só 2 abas diretas: "🏠 Visão Geral" (Dashboard do
   bolão selecionado + resumo da Administração, uma tela só com
   rolagem) e "📅 Pendências por Bolão" (separada, como já era)
Correções v3.9:
 - CORRIGIDO (grave): "Total Esperado" no Dashboard somava o valor
   esperado de TODOS os participantes cadastrados, inclusive o ADM
   isento (que não paga) — o campo só é zerado automaticamente se o ADM
   for cadastrado DEPOIS do bolão já estar marcado como isento; se virou
   isento depois, o valor antigo ficava salvo e inflava o total. Ex.:
   bolão com 47 participantes onde o ADM não paga mostrava o total como
   se fossem 47 pagantes, não 46.
 - CORRIGIDO (grave): mesmo problema no Relatório — a linha do ADM
   isento somava o valor esperado dele tanto no "esperado" quanto no
   "arrecadado", inflando os dois totais igualmente.
 - NOVO: aviso ao tentar registrar um pagamento pro ADM isento (não
   bloqueia, mas confirma — esse tipo de lançamento aparecia escondido
   no "Total Recebido" da aba Depósitos mesmo o ADM não pagando)
 - MELHORADO: bloco "Ganhos por Loteria" (aba Administração) ficou
   menor — só tinha 3-4 loterias mas ocupava o mesmo tanto de espaço
   que o Histórico
Correções v3.8:
 - NOVO: barra de status discreta no rodapé, mostrando o que o sistema
   está fazendo ao abrir (conectando, verificando reservas do site...)
 - CORRIGIDO: quando TODOS os itens da fila de reservas do site davam
   erro ao importar, o app não avisava nada (só um print() invisível no
   .exe empacotado) — agora sempre mostra o que aconteceu, com o erro
   de cada item, tanto na barra de status quanto num aviso
Correções v3.7:
 - NOVO: sincronização de reservas passa a ser nos dois sentidos. Antes só
   ia desktop→site; agora dá pra registrar um depósito/saque de reserva no
   admin do site (inclusive pra gente nova) e, na próxima vez que o app
   desktop abrir, ele importa esses lançamentos pro banco local sozinho e
   avisa quantos importou.
Correções v3.6 (revisão de qualidade multiagente):
 - CORRIGIDO: editar um pagamento com valor inválido zerava o valor
   silenciosamente, sem avisar (2 pontos corrigidos)
 - CORRIGIDO: auto-atualização do Dashboard/Administração ao trocar de
   aba estava morta (um bind() sobrescrevia o outro, e a condição que
   sobrou nunca era verdadeira)
 - CORRIGIDO: registrar pagamento não atualizava o Dashboard na hora
 - CORRIGIDO: detecção "é o administrador?" divergia entre o Relatório
   (só checava a flag) e Dashboard/Cards/publicação (checava flag OU
   nome) — Relatório agora usa o mesmo critério
 - CORRIGIDO: publicação manual de bolão não tinha a trava de "sem
   valor de cota" que a sincronização automática já tinha
 - CORRIGIDO: excluir um bolão não limpava saques_emergenciais/taxa_adm
   (ficava lixo órfão no banco)
 - LIMPEZA: removidas funções mortas, chaves duplicadas no mapeamento
   de ID do Firebase, e uma reimplementação de fmt_brl()
Correções v3.5:
 - MELHORADO: a senha do Firebase agora é pedida logo na abertura do
   sistema, não mais no meio do fechamento — ao fechar, a sincronização
   já roda direto, sem popup de senha no caminho
Correções v3.4:
 - MELHORADO: Enter no campo "Buscar membro" já dispara a busca
 - MELHORADO: duplo-clique num resultado da busca já importa (sem precisar
   clicar em "Importar Selecionado" depois)
 - MELHORADO: popup de registrar pagamento aceita Enter pra confirmar
   (os campos já vêm preenchidos com o padrão) e o aviso final virou
   inline em vez de um popup bloqueante
Correções v3.3:
 - MELHORADO: importar membro de bolão anterior já recalcula o valor
   esperado (cotas x valor da cota), em vez de deixar "0,00" pra ajustar
   na mão
 - MELHORADO: confirmação de importação agora é um aviso na tela em vez
   de um popup bloqueante
Correções v3.2:
 - CORRIGIDO: fechar o app podia sumir com erros de sincronização sem avisar
   (agora exige fechamento manual quando algo falha)
 - CORRIGIDO: bolão excluído no site podia ser recriado no próximo fechamento
   do app (agora reflete a exclusão localmente em vez de recriar)
 - CORRIGIDO: remover bolão do site não usava o ID fixo salvo (falhava depois
   de renomear o bolão)
 - CORRIGIDO: envio ao Firebase sobrescrevia o documento inteiro em vez de
   atualizar só os campos enviados (updateMask)
Correções v3.1:
 - CORRIGIDO: combos de participantes não carregavam ao iniciar
 - NOVO: aba Administração com KPIs, ganhos por loteria e histórico
 - Versionamento interno do banco de dados
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import os, shutil
from datetime import datetime, date
import re

# ─────────────────────────────────────────────
DB_FILE    = "boloes.db"
BACKUP_DIR = "backups"

CORES = {
    "header_bg":    "#1a2a3a",
    "header_fg":    "#ffffff",
    "btn_verde":    "#27ae60",
    "btn_azul":     "#2196F3",
    "btn_laranja":  "#e67e22",
    "btn_vermelho": "#e74c3c",
    "btn_roxo":     "#8e44ad",
    "btn_cinza":    "#7f8c8d",
    "btn_teal":     "#16a085",
    "btn_dourado":  "#f39c12",
    "bg_frame":     "#f0f2f5",
    "bg_section":   "#ffffff",
    "fg_label":     "#2c3e50",
    "fg_title":     "#1a2a3a",
}

MESES    = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
LOTERIAS = ["Mega-Sena","Lotofácil","Quina","Lotomania","Loteca","Timemania",
            "Dia de Sorte","Super Sete","Dupla Sena","Federal","Outros"]

# ─────────────────────────────────────────────
def to_float(s):
    """
    Converte string para float, tratando formato brasileiro (vírgula decimal)
    e formato internacional (ponto decimal).
    Regras:
      "1.234,56" → 1234.56  (ponto = milhar, vírgula = decimal)
      "1,234.56" → 1234.56  (vírgula = milhar, ponto = decimal)
      "33,00"    → 33.0     (vírgula decimal, sem milhar)
      "33.00"    → 33.0     (ponto decimal, sem milhar)
      "33.0"     → 33.0     (ponto decimal — NÃO remove o ponto!)
      "1234"     → 1234.0
    """
    try:
        s = str(s).strip().replace(" ", "").replace("R$", "")
        if not s: return 0.0
        # Caso 1: tem vírgula E ponto → brasileiro: ponto=milhar, vírgula=decimal
        if "," in s and "." in s:
            ultimo_virgula = s.rfind(",")
            ultimo_ponto   = s.rfind(".")
            if ultimo_virgula > ultimo_ponto:
                # "1.234,56" → remove pontos, troca vírgula por ponto
                return float(s.replace(".", "").replace(",", "."))
            else:
                # "1,234.56" → remove vírgulas
                return float(s.replace(",", ""))
        # Caso 2: só vírgula → decimal brasileiro "33,00" ou "1.234" sem decimal
        elif "," in s:
            return float(s.replace(",", "."))
        # Caso 3: só ponto → pode ser decimal "33.0" ou milhar "1.234"
        elif "." in s:
            partes = s.split(".")
            # Se a parte após o ponto tem 1 ou 2 dígitos → é decimal
            if len(partes) == 2 and len(partes[1]) <= 2:
                return float(s)
            # Se tem mais de 3 dígitos após o ponto → milhar (ex: "1.234")
            else:
                return float(s.replace(".", ""))
        # Caso 4: número inteiro puro
        else:
            return float(s)
    except:
        return 0.0

def fmt_brl(v):
    try:
        return f"R$ {float(v):,.2f}".replace(",","X").replace(".",",").replace("X",".")
    except:
        return "R$ 0,00"

def btn(parent, text, color, cmd, width=22, **kw):
    b = tk.Button(parent, text=text, bg=color, fg="white",
                  font=("Arial",9,"bold"), relief="flat",
                  activebackground=color, cursor="hand2",
                  padx=10, pady=6, command=cmd, width=width, **kw)
    return b

def section(parent, title="", pady=8):
    return tk.LabelFrame(parent, text=f"  {title}  " if title else "",
                         bg=CORES["bg_section"], fg=CORES["fg_title"],
                         font=("Arial",10,"bold"), bd=1, relief="groove",
                         padx=12, pady=pady)

def entry(parent, width=40, **kw):
    return tk.Entry(parent, width=width, relief="solid", bd=1,
                    font=("Arial",9), **kw)

def make_tree(parent, cols, height=15):
    style = ttk.Style()
    style.theme_use("clam")
    style.configure("T.Treeview", background="white", foreground="#2c3e50",
                    rowheight=24, fieldbackground="white", font=("Arial",9))
    style.configure("T.Treeview.Heading", background="#1a2a3a",
                    foreground="white", font=("Arial",9,"bold"))
    style.map("T.Treeview", background=[("selected","#2196F3")])
    frame = tk.Frame(parent, bg=CORES["bg_section"])
    tree  = ttk.Treeview(frame, columns=list(cols.keys()),
                         show="headings", height=height, style="T.Treeview")
    for col, w in cols.items():
        tree.heading(col, text=col)
        tree.column(col, width=w, anchor="center")
    vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=vsb.set)
    tree.pack(side="left", fill="both", expand=True)
    vsb.pack(side="right", fill="y")
    return frame, tree

# ─────────────────────────────────────────────
#  BANCO DE DADOS
# ─────────────────────────────────────────────
class Database:
    def __init__(self):
        self.conn = sqlite3.connect(DB_FILE)
        self.conn.row_factory = sqlite3.Row
        self._create_tables()

    def _create_tables(self):
        c = self.conn.cursor()
        c.executescript("""
        CREATE TABLE IF NOT EXISTS boloes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            loteria TEXT DEFAULT 'Mega-Sena',
            data_inicio TEXT,
            num_participantes INTEGER DEFAULT 0,
            valor_total REAL DEFAULT 0,
            valor_parcela REAL DEFAULT 0,
            descricao TEXT,
            status TEXT DEFAULT 'ATIVO',
            adm_nome TEXT DEFAULT '',
            adm_paga INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS pessoas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            telefone TEXT UNIQUE,
            chave_pix TEXT,
            observacoes TEXT
        );
        CREATE TABLE IF NOT EXISTS participantes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bolao_id INTEGER,
            pessoa_id INTEGER,
            nome TEXT NOT NULL,
            telefone TEXT,
            chave_pix TEXT,
            valor_esperado REAL DEFAULT 0,
            observacoes TEXT,
            ativo INTEGER DEFAULT 1,
            is_adm INTEGER DEFAULT 0,
            FOREIGN KEY(bolao_id) REFERENCES boloes(id),
            FOREIGN KEY(pessoa_id) REFERENCES pessoas(id)
        );
        CREATE TABLE IF NOT EXISTS pagamentos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            participante_id INTEGER,
            bolao_id INTEGER,
            mes_referencia TEXT,
            valor REAL DEFAULT 0,
            data_pagamento TEXT,
            depositado INTEGER DEFAULT 0,
            data_deposito TEXT,
            observacoes TEXT,
            FOREIGN KEY(participante_id) REFERENCES participantes(id),
            FOREIGN KEY(bolao_id) REFERENCES boloes(id)
        );
        CREATE TABLE IF NOT EXISTS premiacoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bolao_id INTEGER,
            loteria TEXT DEFAULT 'Mega-Sena',
            concurso TEXT,
            data_sorteio TEXT,
            valor_premio REAL DEFAULT 0,
            descricao TEXT,
            data_registro TEXT,
            FOREIGN KEY(bolao_id) REFERENCES boloes(id)
        );
        CREATE TABLE IF NOT EXISTS reserva_caixa (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bolao_id INTEGER,
            loteria TEXT,
            tipo TEXT,
            valor REAL DEFAULT 0,
            descricao TEXT,
            data_movimento TEXT,
            FOREIGN KEY(bolao_id) REFERENCES boloes(id)
        );
        CREATE TABLE IF NOT EXISTS reservas_pessoas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            telefone TEXT,
            chave_pix TEXT,
            observacoes TEXT,
            ativo INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS reservas_movimentos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pessoa_id INTEGER NOT NULL,
            tipo TEXT NOT NULL,
            valor REAL NOT NULL,
            data_mov TEXT,
            loteria TEXT,
            concurso TEXT,
            descricao TEXT,
            FOREIGN KEY(pessoa_id) REFERENCES reservas_pessoas(id)
        );
        CREATE TABLE IF NOT EXISTS saques_emergenciais (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bolao_id INTEGER,
            valor REAL DEFAULT 0,
            data_saque TEXT,
            motivo TEXT,
            reposto INTEGER DEFAULT 0,
            data_reposicao TEXT,
            obs TEXT,
            FOREIGN KEY(bolao_id) REFERENCES boloes(id)
        );
        CREATE TABLE IF NOT EXISTS taxa_adm (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bolao_id INTEGER,
            loteria TEXT,
            concurso TEXT,
            valor_ganho REAL DEFAULT 0,
            valor_sacado REAL DEFAULT 0,
            descricao TEXT,
            data_registro TEXT,
            tipo TEXT DEFAULT 'GANHO',
            FOREIGN KEY(bolao_id) REFERENCES boloes(id)
        );
        """)
        # migrações seguras
        migs = [
            "ALTER TABLE boloes ADD COLUMN loteria TEXT DEFAULT 'Mega-Sena'",
            "ALTER TABLE boloes ADD COLUMN adm_nome TEXT DEFAULT ''",
            "ALTER TABLE boloes ADD COLUMN adm_paga INTEGER DEFAULT 0",
            "ALTER TABLE participantes ADD COLUMN is_adm INTEGER DEFAULT 0",
            "ALTER TABLE premiacoes ADD COLUMN loteria TEXT DEFAULT 'Mega-Sena'",
            "ALTER TABLE participantes ADD COLUMN pessoa_id INTEGER",
            "ALTER TABLE boloes ADD COLUMN encerrado INTEGER DEFAULT 0",
            # ID fixo do documento no Firebase: gerado uma unica vez na
            # primeira publicacao e reaproveitado sempre depois, mesmo que
            # o nome do bolao mude. Sem isso, renomear um bolao publicado
            # criava um documento novo no Firebase em vez de atualizar o
            # existente (o antigo ficava orfao, com todo o historico).
            "ALTER TABLE boloes ADD COLUMN firebase_doc_id TEXT",
        ]
        for m in migs:
            try: c.execute(m)
            except: pass
        self.conn.commit()
        # Migração: popular tabela pessoas a partir de participantes existentes
        self._migrar_pessoas(c)
        self.conn.commit()

    def _migrar_pessoas(self, c):
        """Cria registros em 'pessoas' para participantes que ainda não têm pessoa_id."""
        # Busca participantes sem pessoa_id e com telefone
        rows = c.execute(
            "SELECT id, nome, telefone, chave_pix FROM participantes "
            "WHERE pessoa_id IS NULL AND telefone IS NOT NULL AND telefone != ''").fetchall()
        for row in rows:
            tel = str(row[1] if len(row) > 1 else row["telefone"] if hasattr(row,"keys") else row[2]).strip()
            # Tenta pelo índice (sqlite3.Row)
            try:
                pid_part = row[0]; nome = row[1]; tel = row[2]; pix = row[3]
            except: continue
            if not tel: continue
            # Verifica se já existe pessoa com esse telefone
            ex = c.execute("SELECT id FROM pessoas WHERE telefone=?", (tel,)).fetchone()
            if ex:
                pessoa_id = ex[0]
            else:
                c.execute("INSERT INTO pessoas (nome,telefone,chave_pix) VALUES (?,?,?)",
                          (nome, tel, pix))
                pessoa_id = c.lastrowid
            c.execute("UPDATE participantes SET pessoa_id=? WHERE id=?",
                      (pessoa_id, pid_part))

    def execute(self, sql, params=()):
        c = self.conn.cursor()
        c.execute(sql, params)
        self.conn.commit()
        return c

    def fetchall(self, sql, params=()):
        c = self.conn.cursor()
        c.execute(sql, params)
        return c.fetchall()

    def fetchone(self, sql, params=()):
        c = self.conn.cursor()
        c.execute(sql, params)
        return c.fetchone()

    def close(self):
        self.conn.close()

# ─────────────────────────────────────────────
class BackupManager:
    def __init__(self):
        os.makedirs(BACKUP_DIR, exist_ok=True)

    def fazer_backup(self, auto=False):
        prefix = "auto" if auto else "manual"
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = os.path.join(BACKUP_DIR, f"{prefix}_backup_{ts}.db")
        shutil.copy2(DB_FILE, dest)
        if auto: self._limpar_autos()
        return dest

    def _limpar_autos(self, manter=10):
        autos = sorted(f for f in os.listdir(BACKUP_DIR) if f.startswith("auto_"))
        while len(autos) > manter:
            os.remove(os.path.join(BACKUP_DIR, autos.pop(0)))

    def listar(self):
        if not os.path.exists(BACKUP_DIR): return []
        out = []
        for f in sorted(os.listdir(BACKUP_DIR), reverse=True):
            if not f.endswith(".db"): continue
            p    = os.path.join(BACKUP_DIR, f)
            size = os.path.getsize(p)/1024
            mt   = datetime.fromtimestamp(os.path.getmtime(p)).strftime("%d/%m/%Y %H:%M:%S")
            tipo = "Manual" if f.startswith("manual") else "Automático"
            out.append({"nome":f,"path":p,"tamanho":f"{size:.2f} KB","data":mt,"tipo":tipo})
        return out

    def restaurar(self, path):
        shutil.copy2(path, DB_FILE)

# ═══════════════════════════════════════════════════════════════
#  FIREBASE — Publicar bolão no site
# ═══════════════════════════════════════════════════════════════
FIREBASE_BASE = (
    "https://firestore.googleapis.com/v1/projects/mega-sena-sistema"
    "/databases/(default)/documents/participantes"
)
FIREBASE_RESERVAS = (
    "https://firestore.googleapis.com/v1/projects/mega-sena-sistema"
    "/databases/(default)/documents/reservas_participantes"
)

def _firestore_patch_url(doc_url, campos, exigir_existente=False):
    """Monta a URL de PATCH do Firestore com updateMask (escreve só os
    campos listados, em vez de sobrescrever o documento inteiro) e,
    opcionalmente, a precondição de que o documento já precisa existir.

    exigir_existente=True faz o Firestore recusar a escrita (em vez de
    criar um documento novo) quando o doc já foi excluído do outro lado
    — usado para bolões que têm firebase_doc_id salvo (já publicados
    antes), pra não "ressuscitar" um bolão que o admin apagou no site.
    """
    import urllib.parse as _up
    params = [("updateMask.fieldPaths", c) for c in campos]
    if exigir_existente:
        params.append(("currentDocument.exists", "true"))
    return doc_url + "?" + _up.urlencode(params)

# ─────────────────────────────────────────────
#  FIREBASE — Autenticação
#  As regras do Firestore passaram a exigir login para escrever
#  (cartoes, participantes, reservas_participantes etc.). Sem isso,
#  toda escrita cai em "HTTP 403: Forbidden".
#  A senha NÃO fica salva em lugar nenhum: o programa pergunta numa
#  caixinha na primeira sincronização de cada sessão e guarda só na
#  memória enquanto o programa estiver aberto.
# ─────────────────────────────────────────────
FIREBASE_API_KEY     = "AIzaSyC5qrS22TILW6GYcg-HAgQa44J-QEgNG3Q"
FIREBASE_ADMIN_EMAIL = "eltonluisoc@gmail.com"
_firebase_token_cache = {"id_token": None, "expira_em": 0, "senha": None}

def _firebase_pedir_senha():
    """Abre uma caixinha pedindo a senha do admin (mascarada com *).
    As sincronizações rodam em threads separadas (pra não travar a
    tela), e o Tkinter só pode mostrar janelas a partir da thread
    principal — por isso, se for chamada de outra thread, agenda a
    caixinha na thread principal via root.after() e espera a resposta."""
    import tkinter as _tk
    from tkinter import simpledialog as _sd
    import threading

    root = _tk._default_root
    if root is None:
        # Fallback raro: nenhuma janela Tk ainda existe
        root = _tk.Tk()
        root.withdraw()

    def _mostrar():
        return _sd.askstring(
            "Login Firebase",
            "Digite a senha do admin (" + FIREBASE_ADMIN_EMAIL + ") para sincronizar:",
            show="*", parent=root
        )

    if threading.current_thread() is threading.main_thread():
        return _mostrar()

    resultado = {}
    evento = threading.Event()

    def _mostrar_e_avisar():
        try:
            resultado["senha"] = _mostrar()
        finally:
            evento.set()

    root.after(0, _mostrar_e_avisar)
    evento.wait()
    return resultado.get("senha")

def _firebase_login():
    """Faz login no Firebase Auth e devolve o ID token, reaproveitando
    enquanto for válido (o token dura 1h). Pede a senha por caixinha na
    primeira vez; se a senha estiver errada, pede de novo na próxima
    tentativa."""
    import time, json, urllib.request, urllib.error
    agora = time.time()
    if _firebase_token_cache["id_token"] and agora < _firebase_token_cache["expira_em"]:
        return _firebase_token_cache["id_token"]

    senha = _firebase_token_cache.get("senha")
    if not senha:
        senha = _firebase_pedir_senha()
        if not senha:
            raise RuntimeError("Login cancelado: senha não informada.")
        _firebase_token_cache["senha"] = senha

    url = ("https://identitytoolkit.googleapis.com/v1/accounts:"
           "signInWithPassword?key=" + FIREBASE_API_KEY)
    corpo = json.dumps({
        "email": FIREBASE_ADMIN_EMAIL,
        "password": senha,
        "returnSecureToken": True,
    }).encode("utf-8")
    req = urllib.request.Request(url, data=corpo, method="POST",
          headers={"Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            dados = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        _firebase_token_cache["senha"] = None  # senha errada — pede de novo na próxima
        corpo_erro = e.read().decode("utf-8") if e.fp else ""
        raise RuntimeError(
            "Falha no login do Firebase: HTTP %s — %s" % (e.code, corpo_erro[:200])
        )

    _firebase_token_cache["id_token"] = dados["idToken"]
    _firebase_token_cache["expira_em"] = agora + int(dados.get("expiresIn", "3600")) - 60
    return _firebase_token_cache["id_token"]

def _firebase_headers():
    """Cabeçalhos padrão (com login) para escritas no Firestore."""
    return {
        "Content-Type": "application/json",
        "Accept": "application/json",
        "Authorization": "Bearer " + _firebase_login(),
    }

def _make_part_id(nome, telefone):
    import unicodedata, re as _re
    def norm(s):
        s = unicodedata.normalize("NFD", s)
        s = "".join(c for c in s if unicodedata.category(c) != "Mn")
        return _re.sub(r"[^a-zA-Z0-9]", "", s)
    partes = [norm(p) for p in nome.strip().split() if p]
    id_nome = (partes[0]+"_"+partes[-1]) if len(partes)>1 else (partes[0] if partes else "x")
    tel = _re.sub(r"\D","",telefone or "")
    return (id_nome+"_"+tel) if tel else id_nome

def _tipo_fb(tipo_db):
    t = (tipo_db or "").upper()
    if t in ("CREDITO","CRÉDITO","ENTRADA","DEPOSITO","DEPÓSITO"): return "deposito"
    if t in ("DEBITO","DÉBITO","USO"): return "uso"
    return "saque"

# Mapeamento explícito: nome do bolão → ID fixo no Firebase
# Edite aqui se precisar adicionar novos bolões
FIREBASE_BOLAO_IDS = {
    "mega da virada 2026":              "bolao_mega_da_virada_2026",
    "mega da virada amigos 2026":       "bolao_mega_da_virada_amigos_2026",
    "lotofacil da independencia":       "bolao_lotofacil_da_independencia",
    "quina de sao joao 2026":           "bolao_quina_de_sao_joao_2026",
    "quina de sao joao 2026 ii":        "bolao_quina_de_sao_joao_2026_ii",
    "mega-sena 30 anos":                "bolao_mega_sena_30_anos",
    "mega sena 30 anos":                "bolao_mega_sena_30_anos",
}
# As chaves são sempre comparadas sem acento (ver _norm() em
# _bolao_doc_id() abaixo), então uma variante acentuada aqui nunca seria
# encontrada — não precisa duplicar "independencia"/"independência".



def _bolao_doc_id(titulo):
    """Retorna o ID fixo do Firebase para o bolão.
    Primeiro tenta o mapeamento explícito; se não achar, gera pelo slug."""
    import unicodedata, re as _re
    def _norm(s):
        s = unicodedata.normalize("NFD", s.lower())
        return "".join(c for c in s if unicodedata.category(c) != "Mn").strip()
    chave = _norm(titulo)
    if chave in FIREBASE_BOLAO_IDS:
        return FIREBASE_BOLAO_IDS[chave]
    # Fallback: slug sem artigos curtos
    s = unicodedata.normalize("NFD", titulo)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = _re.sub(r"[^a-zA-Z0-9]+", "_", s).strip("_").lower()
    return "bolao_" + s


def enviar_reservas_para_site(db_file):
    import sqlite3 as _sq, urllib.request, urllib.error, json
    from datetime import datetime as _dt
    conn = _sq.connect(db_file); conn.row_factory = _sq.Row
    headers = _firebase_headers()
    ts = _dt.utcnow().strftime("%Y-%m-%dT%H:%M:%S.000Z")
    pessoas = conn.execute("SELECT * FROM reservas_pessoas WHERE ativo=1 ORDER BY nome").fetchall()
    enviados = 0; erros = []
    for ps in pessoas:
        ps   = dict(ps)
        pid  = ps["id"]; nome = ps["nome"] or ""; tel = ps["telefone"] or ""
        doc_id = _make_part_id(nome, tel)
        ent = conn.execute(
            "SELECT COALESCE(SUM(valor),0) as t FROM reservas_movimentos "
            "WHERE pessoa_id=? AND UPPER(tipo) IN "
            "('CRÉDITO','CREDITO','ENTRADA','DEPOSITO','DEPÓSITO')",(pid,)).fetchone()
        sai = conn.execute(
            "SELECT COALESCE(SUM(valor),0) as t FROM reservas_movimentos "
            "WHERE pessoa_id=? AND UPPER(tipo) IN "
            "('DÉBITO','DEBITO','SAIDA','SAQUE','USO')",(pid,)).fetchone()
        saldo = float(ent["t"] if ent else 0) - float(sai["t"] if sai else 0)
        movs = conn.execute(
            "SELECT * FROM reservas_movimentos WHERE pessoa_id=? ORDER BY id ASC",(pid,)).fetchall()
        hist = []; sa = 0.0
        for m in movs:
            v = float(m["valor"] or 0); tp = _tipo_fb(m["tipo"]); ant = sa
            sa = sa-v if tp in ("saque","uso") else sa+v
            dm = m["data_mov"] or ts
            try:
                from datetime import datetime as _dt2
                if len(dm)==10 and "/" in dm:
                    dm = _dt2.strptime(dm,"%d/%m/%Y").strftime("%Y-%m-%dT00:00:00.000Z")
            except: pass
            hist.append({"mapValue":{"fields":{
                "data":{"stringValue":dm},"tipo":{"stringValue":tp},
                "valor":{"doubleValue":v},"saldoAnterior":{"doubleValue":round(ant,2)},
                "saldoNovo":{"doubleValue":round(sa,2)},
                "descricao":{"stringValue":str(dict(m).get("descricao") or dict(m).get("loteria") or tp)},
            }}})
        doc = {"fields":{
            "participanteId":{"stringValue":doc_id},"nome":{"stringValue":nome},
            "telefone":{"stringValue":tel},"saldoReserva":{"doubleValue":round(saldo,2)},
            "dataAtualizacao":{"stringValue":ts},"admin":{"booleanValue":True},
            "historico":{"arrayValue":{"values":hist}},
        }}
        corpo = json.dumps(doc,ensure_ascii=False).encode("utf-8")
        url = _firestore_patch_url(FIREBASE_RESERVAS+"/"+doc_id,
              ["participanteId","nome","telefone","saldoReserva","dataAtualizacao","admin","historico"])
        try:
            req = urllib.request.Request(url,data=corpo,method="PATCH",headers=headers)
            with urllib.request.urlopen(req,timeout=15) as resp:
                if resp.status==200: enviados+=1
                else: erros.append(nome+": status "+str(resp.status))
        except urllib.error.HTTPError as e: erros.append(nome+": HTTP "+str(e.code))
        except Exception as ex: erros.append(nome+": "+str(ex)[:40])
    conn.close()
    return enviados, erros

def _firebase_listar_docs():
    """Lista todos os documentos da coleção participantes."""
    import urllib.request, urllib.error, json
    try:
        req = urllib.request.Request(FIREBASE_BASE, method="GET",
              headers={"Accept":"application/json"})
        with urllib.request.urlopen(req, timeout=10) as r:
            dados = json.loads(r.read().decode("utf-8"))
            return dados.get("documents", [])
    except:
        return []

def _firebase_buscar_doc_por_titulo(titulo):
    """Busca o name do documento que tem o título informado. Retorna None se não achar."""
    docs = _firebase_listar_docs()
    for doc in docs:
        fields = doc.get("fields", {})
        t = fields.get("titulo", {}).get("stringValue", "")
        if t.strip().lower() == titulo.strip().lower():
            return doc.get("name", "")  # ex: "projects/.../documents/participantes/bolao_abc"
    return None

def enviar_bolao_para_site(titulo, loteria, valor_cota,
                           data_limite, participantes, doc_id_fixo=None):
    """
    Envia dados do bolão para o Firebase Firestore.

    doc_id_fixo: ID do documento já salvo localmente (ver coluna
    firebase_doc_id na tabela boloes). Quando informado, é sempre esse
    ID que é usado — mesmo que o título tenha mudado desde a última
    publicação. Sem isso, renomear um bolão gerava um ID novo (calculado
    a partir do título) e criava um documento duplicado no Firebase em
    vez de atualizar o existente.
    """
    import urllib.request, urllib.error, json
    from datetime import datetime as _dt

    # Monta o documento Firestore
    import re as _re_fb
    valores_part = []
    for p in participantes:
        tel_digits = _re_fb.sub(r"\D", "", str(p.get("telefone") or ""))
        valores_part.append({
            "mapValue": {
                "fields": {
                    "nome":            {"stringValue": str(p.get("nome",""))},
                    "telefone":        {"stringValue": tel_digits},
                    "valorPago":       {"doubleValue": float(p.get("valorPago",0))},
                    "situacao":        {"stringValue": str(p.get("situacao","em_andamento"))},
                    "quantidadeCotas": {"integerValue": str(int(p.get("quantidadeCotas",1)))},
                    "dataCadastro":    {"stringValue": str(p.get("dataCadastro",""))},
                }
            }
        })

    documento = {
        "fields": {
            "titulo":      {"stringValue": titulo},
            "loteria":     {"stringValue": loteria},
            "valorPorCota":{"doubleValue": round(float(valor_cota), 2)},
            "dataLimite":  {"stringValue": data_limite},
            "admin":       {"booleanValue": True},
            "participantes": {
                "arrayValue": {"values": valores_part}
            }
        }
    }
    corpo = json.dumps(documento, ensure_ascii=False).encode("utf-8")
    headers = _firebase_headers()

    campos_doc = ["titulo","loteria","valorPorCota","dataLimite","admin","participantes"]
    try:
        doc_id = doc_id_fixo or _bolao_doc_id(titulo)
        # Quando já existe um doc_id_fixo salvo, o bolão foi publicado
        # antes: exige que o documento ainda exista, pra não recriar um
        # bolão que o admin excluiu do site.
        url = _firestore_patch_url(f"{FIREBASE_BASE}/{doc_id}", campos_doc,
                                    exigir_existente=bool(doc_id_fixo))
        print(f"[Firebase] PATCH → {doc_id}")

        req = urllib.request.Request(url, data=corpo, method="PATCH", headers=headers)
        with urllib.request.urlopen(req, timeout=15) as resp:
            status = resp.status
            print(f"[Firebase] Status: {status}")
            if status == 200:
                print("✅ Bolão publicado com sucesso!")
                return {"sucesso": True, "status": status, "erro": None, "doc_id": doc_id}
            else:
                return {"sucesso": False, "status": status,
                        "erro": f"Status inesperado: {status}", "doc_id": doc_id}

    except urllib.error.HTTPError as e:
        corpo_erro = ""
        try: corpo_erro = e.read().decode("utf-8")
        except: pass
        if doc_id_fixo and e.code in (400, 404, 409) and "FAILED_PRECONDITION" in corpo_erro:
            erro = "Bolão foi excluído no site — não será recriado automaticamente."
            print(f"⚠️ {erro}")
            return {"sucesso": False, "status": e.code, "erro": erro,
                    "doc_id": doc_id_fixo, "excluido_no_site": True}
        erro = f"HTTP {e.code}: {e.reason} — {corpo_erro[:300]}"
        print(f"❌ {erro}")
        return {"sucesso": False, "status": e.code, "erro": erro, "doc_id": ""}

    except urllib.error.URLError as e:
        erro = f"Sem conexão: {e.reason}"
        print(f"❌ {erro}")
        return {"sucesso": False, "status": 0, "erro": erro, "doc_id": ""}

    except Exception as e:
        erro = f"Erro inesperado: {str(e)}"
        print(f"❌ {erro}")
        return {"sucesso": False, "status": 0, "erro": erro, "doc_id": ""}


def remover_bolao_do_site(titulo, doc_id_fixo=None):
    """Remove o documento do bolão do Firebase.

    doc_id_fixo: ID salvo localmente (firebase_doc_id) — usado direto
    quando disponível. Sem ele, cai no fallback de buscar pelo título
    atual, que falha se o bolão foi renomeado desde a última publicação
    (o Firestore ainda tem o título antigo)."""
    import urllib.request, urllib.error

    try:
        if doc_id_fixo:
            doc_id = doc_id_fixo
        else:
            doc_name = _firebase_buscar_doc_por_titulo(titulo)
            if not doc_name:
                return {"sucesso": False, "erro": f"Bolão '{titulo}' não encontrado no site."}
            doc_id = doc_name.split("/")[-1]
        # URL completa para DELETE no Firestore REST API
        url = (
            f"https://firestore.googleapis.com/v1/projects/mega-sena-sistema"
            f"/databases/(default)/documents/participantes/{doc_id}"
        )
        req = urllib.request.Request(url, method="DELETE",
              headers=_firebase_headers())
        with urllib.request.urlopen(req, timeout=10) as resp:
            print(f"[Firebase] Documento {doc_id} removido. Status: {resp.status}")
            return {"sucesso": True, "doc_id": doc_id}

    except urllib.error.HTTPError as e:
        erro = f"HTTP {e.code}: {e.reason}"
        return {"sucesso": False, "erro": erro}
    except Exception as e:
        return {"sucesso": False, "erro": str(e)}


# ─── Exemplo de teste (mude False → True para testar isoladamente) ──
if False:
    r = enviar_bolao_para_site(
        titulo       = "Bolão Mega-Sena Junho 2026",
        loteria      = "mega",
        valor_cota   = 500,
        data_limite  = "2026-06-20",
        participantes= [
            {"nome": "João Silva",  "valorPago": 500, "situacao": "quitado"},
            {"nome": "Maria Souza", "valorPago": 0,   "situacao": "em_andamento"},
        ]
    )
    print("Resultado:", r)


# ═══════════════════════════════════════════════════════════════
#  APLICAÇÃO
# ═══════════════════════════════════════════════════════════════
class BolaoApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema de Gestão de Bolões PRO v5.1")
        self.root.geometry("1300x800")
        self.root.minsize(1050, 680)
        self.root.configure(bg=CORES["header_bg"])

        self.db  = Database()
        self.bkp = BackupManager()
        self.bid  = tk.IntVar(value=0)

        self._build_header()
        self._build_status_bar()
        self._build_tabs()
        self._load_boloes_combo()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

        # Pede a senha do Firebase logo na entrada (em vez de só no
        # fechamento): a senha fica em memoria pro resto da sessao e o
        # fechamento sincroniza direto, sem popup de senha no meio do caminho.
        self.root.after(400, self._login_inicial)

    def _status(self, texto, cor="#90caf9"):
        """Atualiza a barra de status no rodapé da janela principal — o
        usuário pediu pra ver o que o sistema está fazendo ao abrir, em vez
        de tudo acontecer em silêncio (erros iam parar num print() que não
        aparece no .exe empacotado sem console)."""
        try:
            self._status_lbl.configure(text=texto, fg=cor)
            self.root.update_idletasks()
        except Exception:
            pass

    def _login_inicial(self):
        """Autentica no Firebase assim que a janela principal abre."""
        self._status("🔄 Conectando ao Firebase...")
        try:
            _firebase_login()
        except Exception as ex:
            self._status("❌ Não entrou no Firebase — sincronização adiada", cor="#ff6b6b")
            messagebox.showwarning("Login Firebase",
                "Não foi possível entrar no Firebase agora:\n\n" + str(ex) +
                "\n\nVocê pode continuar usando o sistema normalmente; a "
                "sincronização com o site vai pedir a senha novamente mais tarde.")
            return
        self._status("🔄 Verificando reservas lançadas no site...")
        self._importar_movimentos_pendentes_web()

    def _importar_movimentos_pendentes_web(self):
        """Importa depósitos/saques de reserva registrados no admin do site
        (fila reservas_movimentos_pendentes) pro SQLite local, e some com o
        item da fila depois de importar com sucesso. Antes disso só existia
        sincronização desktop→site; isso fecha o caminho contrário."""
        import urllib.request, urllib.error, json, re as _re

        url = ("https://firestore.googleapis.com/v1/projects/mega-sena-sistema"
               "/databases/(default)/documents/reservas_movimentos_pendentes")
        try:
            req = urllib.request.Request(url, method="GET", headers=_firebase_headers())
            with urllib.request.urlopen(req, timeout=15) as r:
                dados = json.loads(r.read().decode("utf-8"))
        except Exception as ex:
            self._status("❌ Erro ao verificar reservas do site: " + str(ex)[:90], cor="#ff6b6b")
            return

        docs = dados.get("documents", [])
        if not docs:
            self._status("✅ Reservas em dia — nada novo do site.")
            self.root.after(5000, lambda: self._status(""))
            return

        self._status(f"🔄 Importando {len(docs)} movimento(s) de reserva do site...")

        def campo(fields, nome, tipo, default=None):
            return fields.get(nome, {}).get(tipo, default)

        importados = 0; erros = 0
        total_dep = 0.0; total_saq = 0.0
        erros_detalhe = []

        for doc in docs:
            nome_erro = "?"
            try:
                fields   = doc.get("fields", {})
                doc_id   = doc.get("name", "").split("/")[-1]
                nome     = (campo(fields, "nome", "stringValue", "") or "").strip()
                nome_erro = nome or doc_id
                telefone = campo(fields, "telefone", "stringValue", "") or ""
                chavePix = campo(fields, "chavePix", "stringValue", "") or ""
                tipo_web = campo(fields, "tipo", "stringValue", "deposito")
                valor    = float(campo(fields, "valor", "doubleValue")
                                  or campo(fields, "valor", "integerValue") or 0)
                data_web = campo(fields, "data", "stringValue", "") or ""
                descricao= campo(fields, "descricao", "stringValue", "") or ""

                if valor <= 0 or not nome:
                    raise ValueError(f"dados incompletos (valor={valor!r}, nome={nome!r})")

                tel_digits = _re.sub(r"\D", "", telefone)

                # Acha a pessoa local por telefone (mais confiável) ou nome;
                # se não achar, cria — mesmo que o site achasse que já existia
                # (participanteId pode ter sido gerado antes da pessoa
                # existir localmente, ou o app pode ter sido reinstalado).
                pessoa = None
                if tel_digits:
                    pessoa = self.db.fetchone(
                        "SELECT id FROM reservas_pessoas WHERE telefone=?", (tel_digits,))
                if not pessoa:
                    pessoa = self.db.fetchone(
                        "SELECT id FROM reservas_pessoas WHERE LOWER(nome)=?", (nome.lower(),))

                if pessoa:
                    pessoa_id = pessoa["id"]
                else:
                    self.db.execute(
                        "INSERT INTO reservas_pessoas (nome,telefone,chave_pix,observacoes) "
                        "VALUES (?,?,?,?)",
                        (nome, tel_digits, chavePix, "Criado a partir de movimento do site"))
                    pessoa_id = self.db.fetchone("SELECT last_insert_rowid() as id")["id"]

                tipo_db = "CRÉDITO" if tipo_web == "deposito" else "DÉBITO"

                # "AAAA-MM-DD" (input date do site) -> "DD/MM/AAAA" (padrão desktop)
                data_mov = data_web
                if len(data_web) == 10 and data_web[4:5] == "-":
                    a, m, d = data_web.split("-")
                    data_mov = f"{d}/{m}/{a}"

                desc_final = (descricao + " (via site)").strip() if descricao else "Via site"
                self.db.execute(
                    "INSERT INTO reservas_movimentos "
                    "(pessoa_id, tipo, valor, data_mov, loteria, concurso, descricao) "
                    "VALUES (?,?,?,?,?,?,?)",
                    (pessoa_id, tipo_db, valor, data_mov, "", "", desc_final))

                # Só sai da fila depois de importar com sucesso no SQLite local
                del_req = urllib.request.Request(
                    url + "/" + doc_id, method="DELETE", headers=_firebase_headers())
                urllib.request.urlopen(del_req, timeout=10)

                importados += 1
                if tipo_db == "CRÉDITO": total_dep += valor
                else: total_saq += valor
            except Exception as ex:
                erros += 1
                erros_detalhe.append(f"  • {nome_erro}: {str(ex)[:100]}")

        if importados == 0 and erros == 0:
            return

        partes = []
        if importados > 0:
            partes.append(f"✅ {importados} movimento(s) de reserva importado(s) do site:")
            if total_dep > 0: partes.append(f"  💰 Depósitos: {fmt_brl(total_dep)}")
            if total_saq > 0: partes.append(f"  💸 Saques/uso: {fmt_brl(total_saq)}")
        if erros:
            partes.append(f"⚠ {erros} item(ns) da fila NÃO foram importados:")
            partes.extend(erros_detalhe)
            partes.append("\nEsses itens continuam na fila — corrija no site e feche/abra o app de novo.")

        if erros and importados == 0:
            self._status(f"❌ {erros} movimento(s) de reserva falharam ao importar", cor="#ff6b6b")
            messagebox.showerror("Erro ao sincronizar reservas", "\n".join(partes))
        elif erros:
            self._status(f"⚠️ {importados} importado(s), {erros} com erro", cor="#f39c12")
            messagebox.showwarning("Reservas sincronizadas com pendências", "\n".join(partes))
        else:
            self._status(f"✅ {importados} movimento(s) de reserva importado(s) do site")
            messagebox.showinfo("Reservas sincronizadas", "\n".join(partes))

        if importados > 0:
            try: self._rsv_load()
            except Exception: pass
        self.root.after(8000, lambda: self._status(""))

    # ── HEADER ──────────────────────────────────────────────────
    def _build_header(self):
        hdr = tk.Frame(self.root, bg=CORES["header_bg"], pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text="🎰  SISTEMA DE GESTÃO DE BOLÕES PRO v5.1",
                 bg=CORES["header_bg"], fg="white",
                 font=("Arial",15,"bold")).pack(side="left", padx=18)
        right = tk.Frame(hdr, bg=CORES["header_bg"])
        right.pack(side="right", padx=18)
        tk.Label(right, text="Bolão Ativo:", bg=CORES["header_bg"],
                 fg="#aad4f5", font=("Arial",9,"bold")).pack(side="left")
        self.cb_bolao = ttk.Combobox(right, width=34, state="readonly", font=("Arial",9))
        self.cb_bolao.pack(side="left", padx=6)
        self.cb_bolao.bind("<<ComboboxSelected>>", self._on_bolao_sel)
        btn(right,"⚙ Gerenciar Bolões",CORES["btn_azul"],self._gerenciar_boloes,width=18).pack(side="left",padx=4)
        btn(right,"+ Novo Bolão",CORES["btn_verde"],self._novo_bolao,width=14).pack(side="left",padx=4)

    # ── BARRA DE STATUS (rodapé) ───────────────────────────────────
    def _build_status_bar(self):
        """Faixa discreta no rodapé pra mostrar o que o sistema está
        fazendo ao abrir (login, sincronização) — pedido explícito do
        usuário depois de uma sincronização de reservas ter falhado em
        silêncio, sem nenhum aviso visível."""
        barra = tk.Frame(self.root, bg="#0d1b2a", height=24)
        barra.pack(fill="x", side="bottom")
        barra.pack_propagate(False)
        self._status_lbl = tk.Label(barra, text="", bg="#0d1b2a", fg="#90caf9",
                                     font=("Arial",8), anchor="w")
        self._status_lbl.pack(fill="both", expand=True, padx=10)

    # ── TABS ────────────────────────────────────────────────────
    def _build_tabs(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Inner.TNotebook", background=CORES["bg_frame"])
        style.configure("Inner.TNotebook.Tab",
                        padding=[14,5], font=("Arial",9,"bold"),
                        background="#c8d8e8", foreground="#333")
        style.map("Inner.TNotebook.Tab",
                  background=[("selected","#e67e22")],
                  foreground=[("selected","white")])

        self.nb = ttk.Notebook(self.root)
        self.nb.pack(fill="both", expand=True, padx=6, pady=4)

        # ── 5 guias reorganizadas ──────────────────────────────────
        for attr, label in [
            ("tab_grp_inicio", "🏠 Inicio"),
            ("tab_grp_part",   "👥 Participantes"),
            ("tab_grp_fin",    "💰 Financeiro"),
            ("tab_grp_gestao", "🏆 Gestao"),
            ("tab_grp_sys",    "⚙ Sistema"),
        ]:
            frame = tk.Frame(self.nb, bg=CORES["bg_frame"])
            setattr(self, attr, frame); self.nb.add(frame, text=label)

        # aliases para compatibilidade
        self.tab_cad     = self.tab_grp_part
        self.tab_pag_grp = self.tab_grp_fin

        # ── Inicio: Visao Geral (Dashboard + Administracao fundidos,
        # com rolagem) + Pendencias por Bolao como aba propria ───────
        nb_inicio = ttk.Notebook(self.tab_grp_inicio, style="Inner.TNotebook")
        nb_inicio.pack(fill="both", expand=True, padx=4, pady=4)
        self.tab_dash  = tk.Frame(nb_inicio, bg=CORES["bg_frame"])
        self.tab_bolao = tk.Frame(nb_inicio, bg=CORES["bg_frame"])
        self.tab_pend  = tk.Frame(nb_inicio, bg=CORES["bg_frame"])
        nb_inicio.add(self.tab_dash,  text="🏠 Visão Geral")
        nb_inicio.add(self.tab_bolao, text="🎯 Bolão Selecionado")
        nb_inicio.add(self.tab_pend,  text="📅 Pendências por Bolão")

        # ── Participantes (4 sub-abas) ───────────────────────────────
        nb_part = ttk.Notebook(self.tab_grp_part, style="Inner.TNotebook")
        nb_part.pack(fill="both", expand=True, padx=4, pady=4)
        self.tab_cad_novo  = tk.Frame(nb_part, bg=CORES["bg_frame"])
        self.tab_cad_edit  = tk.Frame(nb_part, bg=CORES["bg_frame"])
        self.tab_cad_lista = tk.Frame(nb_part, bg=CORES["bg_frame"])
        self.tab_pessoas   = tk.Frame(nb_part, bg=CORES["bg_frame"])
        nb_part.add(self.tab_cad_novo,  text="➕ Novo Participante")
        nb_part.add(self.tab_cad_edit,  text="✏ Editar Participante")
        nb_part.add(self.tab_cad_lista, text="📋 Lista / Remover")
        nb_part.add(self.tab_pessoas,   text="🔗 Pessoas / Unificar")

        # ── Financeiro (7 sub-abas — "Editar" foi removida, redundante
        # com "Histórico": busca por nome + duplo-clique já cobria tudo,
        # e "Histórico" ganhou um botão Excluir pra fechar a diferença) ──
        nb_fin = ttk.Notebook(self.tab_grp_fin, style="Inner.TNotebook")
        nb_fin.pack(fill="both", expand=True, padx=4, pady=4)
        self.tab_pag    = tk.Frame(nb_fin, bg=CORES["bg_frame"])
        self.tab_vis    = tk.Frame(nb_fin, bg=CORES["bg_frame"])
        self.tab_dep    = tk.Frame(nb_fin, bg=CORES["bg_frame"])
        self.tab_rel    = tk.Frame(nb_fin, bg=CORES["bg_frame"])
        self.tab_rsv    = tk.Frame(nb_fin, bg=CORES["bg_frame"])
        self.tab_hist   = tk.Frame(nb_fin, bg=CORES["bg_frame"])
        self.tab_import = tk.Frame(nb_fin, bg=CORES["bg_frame"])
        nb_fin.add(self.tab_pag,    text="💳 Registrar")
        nb_fin.add(self.tab_vis,    text="🔍 Visualizar / Recibo")
        nb_fin.add(self.tab_dep,    text="🏦 Depositos")
        nb_fin.add(self.tab_rel,    text="📊 Relatorio")
        nb_fin.add(self.tab_rsv,    text="💰 Reservas Pessoais")
        nb_fin.add(self.tab_hist,   text="📋 Historico")
        nb_fin.add(self.tab_import, text="📥 Importar Extrato")

        self._pag_part_sync = tk.StringVar()

        # ── Gestao: Caixa/Premios + Pendencias por Bolao ─────────────
        nb_gestao = ttk.Notebook(self.tab_grp_gestao, style="Inner.TNotebook")
        nb_gestao.pack(fill="both", expand=True, padx=4, pady=4)
        self.tab_res      = tk.Frame(nb_gestao, bg=CORES["bg_frame"])
        self.tab_prem     = tk.Frame(nb_gestao, bg=CORES["bg_frame"])
        self.tab_caixa_pr = self.tab_grp_gestao  # alias para compat
        nb_gestao.add(self.tab_res,  text="💼 Caixa por Loteria")
        nb_gestao.add(self.tab_prem, text="🏆 Premiacoes")

        # ── Sistema: Backup + Site ────────────────────────────────────
        nb_sys = ttk.Notebook(self.tab_grp_sys, style="Inner.TNotebook")
        nb_sys.pack(fill="both", expand=True, padx=4, pady=4)
        self.tab_bkp = tk.Frame(nb_sys, bg=CORES["bg_frame"])
        self.tab_pub = tk.Frame(nb_sys, bg=CORES["bg_frame"])
        nb_sys.add(self.tab_bkp, text="💾 Backup / Restore")
        nb_sys.add(self.tab_pub, text="🌐 Site / Publicar")

        # Auto-atualiza ao entrar no grupo "Inicio" vindo de outro grupo
        # (o bind mais abaixo, no notebook INTERNO nb_inicio, cobre trocar
        # entre as sub-abas Dashboard/Administracao já dentro dela — os
        # dois binds precisam estar em widgets diferentes: bind() duas
        # vezes no mesmo widget substitui o anterior em silêncio, e foi
        # exatamente isso que deixava esse auto-refresh sempre morto)
        def _on_tab_changed(event):
            try:
                tab = self.nb.tab(self.nb.select(), "text")
                if "Inicio" in tab:
                    self.root.after(50, self._adm_load)
            except Exception:
                pass
        self.nb.bind("<<NotebookTabChanged>>", _on_tab_changed)

        # ── Constrói o conteúdo de todas as abas ────────────────
        self._build_dashboard()
        self._build_bolao_sel()
        self._build_cad()
        self._build_cad_editar()
        self._build_cad_lista()
        self._build_pag()
        self._build_vis()
        self._build_importar()
        self._build_rel()
        self._build_dep()
        self._build_historico()
        self._build_prem()
        self._build_res()
        self._build_adm()
        self._build_reservas()
        self._build_pessoas()
        self._build_publicar()
        self._build_bkp()

        # Auto-atualiza ao trocar entre as 3 sub-abas de Inicio (Visao Geral,
        # Bolao Selecionado, Pendencias por Bolao) — as 3 dependem de
        # _dash_load/_adm_load, entao os dois carregadores rodam sempre
        # que o usuario troca de sub-aba dentro de Inicio.
        def _on_tab_changed_inicio(event):
            self.root.after(50, self._dash_load)
            self.root.after(50, self._adm_load)
        nb_inicio.bind("<<NotebookTabChanged>>", _on_tab_changed_inicio)

    # ════════════════════════════════════════════════════════════
    #  HELPERS CENTRALIZADOS — status e cotas
    # ════════════════════════════════════════════════════════════
    def _get_cotas_ocupadas(self, bid):
        b = self.db.fetchone(
            "SELECT num_participantes, valor_total, adm_nome, adm_paga FROM boloes WHERE id=?", (bid,))
        if not b: return 0, 0
        vt        = float(b["valor_total"] or 0)
        max_cotas = int(b["num_participantes"] or 0)
        if vt <= 0: return 0, max_cotas
        adm_nome = (b["adm_nome"] or "").strip()
        adm_paga = b["adm_paga"] or 0
        adm_low  = adm_nome.lower()
        partic = self.db.fetchall(
            "SELECT nome, valor_esperado, is_adm FROM participantes "
            "WHERE bolao_id=? AND ativo=1", (bid,))
        cotas = 0; adm_cad = False
        for pt in partic:
            ve = float(pt["valor_esperado"] or 0)
            eh_adm = bool(pt["is_adm"]) or (adm_low and adm_low in pt["nome"].lower())
            if eh_adm:
                adm_cad = True
                if not adm_paga:
                    cotas += 1; continue
            cotas += max(1, round(ve/vt)) if ve > 0 else 0
        adm_extra = 1 if (adm_nome and not adm_paga and not adm_cad) else 0
        return cotas + adm_extra, max_cotas

    def _status_part_adm(self, pt_d, pago, ve, parc_esp, parc, adm_paga):
        """
        ADM isento → QUITADO, pago_exib=None (card sintético mostra valor_total do bolão).
        ADM que paga → tratamento normal.
        """
        if pt_d.get("is_adm") and not adm_paga:
            # None = card vai mostrar valor_total do bolão (cota completa)
            return "QUITADO", "quitado", None, 0
        saldo  = max(0, ve - pago)
        st, tag = self._status_part(pago, ve, parc_esp, parc)
        status  = st.replace("✅ ","").replace("⚠ ","").replace("🟦 ","")
        return status, tag, pago, saldo

    # ════════════════════════════════════════════════════════════
    #  ABA CADASTRAR — Novo Participante
    # ════════════════════════════════════════════════════════════
    def _build_cad(self):
        p = self.tab_cad_novo

        outer = tk.Frame(p, bg=CORES["bg_frame"])
        outer.pack(fill="both", expand=True, padx=40, pady=20)

        sec = section(outer, "CADASTRAR NOVO PARTICIPANTE", pady=16)
        sec.pack(fill="x")
        sec.columnconfigure(1, weight=1)

        self._cv = {}

        # Nome
        tk.Label(sec, text="Nome Completo:*", bg=CORES["bg_section"], fg=CORES["fg_label"],
                 font=("Arial",9,"bold")).grid(row=0, column=0, sticky="w", padx=(0,12), pady=8)
        self._cv["nome"] = entry(sec, width=45)
        self._cv["nome"].grid(row=0, column=1, sticky="ew", pady=8)

        # Telefone
        tk.Label(sec, text="Telefone (WhatsApp):", bg=CORES["bg_section"], fg=CORES["fg_label"],
                 font=("Arial",9,"bold")).grid(row=1, column=0, sticky="w", padx=(0,12), pady=8)
        self._cv["tel"] = entry(sec, width=45)
        self._cv["tel"].grid(row=1, column=1, sticky="ew", pady=8)

        # Chave PIX
        tk.Label(sec, text="Chave PIX:", bg=CORES["bg_section"], fg=CORES["fg_label"],
                 font=("Arial",9,"bold")).grid(row=2, column=0, sticky="w", padx=(0,12), pady=8)
        self._cv["pix"] = entry(sec, width=45)
        self._cv["pix"].grid(row=2, column=1, sticky="ew", pady=8)

        # Nº de Cotas + cálculo automático
        tk.Label(sec, text="Nº de Cotas:", bg=CORES["bg_section"], fg=CORES["fg_label"],
                 font=("Arial",9,"bold")).grid(row=3, column=0, sticky="w", padx=(0,12), pady=8)
        cotas_frame = tk.Frame(sec, bg=CORES["bg_section"])
        cotas_frame.grid(row=3, column=1, sticky="w", pady=8)

        self._cv["cotas"] = entry(cotas_frame, width=6)
        self._cv["cotas"].insert(0, "1")
        self._cv["cotas"].pack(side="left")

        tk.Label(cotas_frame, text="  ×  parcela do bolão  =  valor esperado calculado automaticamente",
                 bg=CORES["bg_section"], fg="#888",
                 font=("Arial",8,"italic")).pack(side="left")

        btn(cotas_frame, "🔢 Calcular", CORES["btn_azul"],
            self._calcular_valor_cotas, width=12).pack(side="left", padx=8)

        # Ao sair do campo cotas, recalcula automaticamente sem messagebox
        self._cv["cotas"].bind("<FocusOut>", lambda e: self._preencher_valor_cad())
        self._cv["cotas"].bind("<Return>",   lambda e: self._preencher_valor_cad())

        # Valor Total Esperado
        tk.Label(sec, text="Valor Total Esperado (R$):", bg=CORES["bg_section"], fg=CORES["fg_label"],
                 font=("Arial",9,"bold")).grid(row=4, column=0, sticky="w", padx=(0,12), pady=8)
        val_frame = tk.Frame(sec, bg=CORES["bg_section"])
        val_frame.grid(row=4, column=1, sticky="w", pady=8)
        self._cv["valor"] = entry(val_frame, width=16)
        self._cv["valor"].insert(0, "0,00")
        self._cv["valor"].pack(side="left")
        tk.Label(val_frame,
                 text="  (ou informe manualmente — campo livre)",
                 bg=CORES["bg_section"], fg="#888",
                 font=("Arial",8,"italic")).pack(side="left")

        # Observações
        tk.Label(sec, text="Observações:", bg=CORES["bg_section"], fg=CORES["fg_label"],
                 font=("Arial",9,"bold")).grid(row=5, column=0, sticky="nw", padx=(0,12), pady=8)
        self._cv["obs"] = tk.Text(sec, height=3, relief="solid", bd=1, font=("Arial",9))
        self._cv["obs"].grid(row=5, column=1, sticky="ew", pady=8)

        # Checkbox ADM
        self._cv_is_adm = tk.IntVar(value=0)
        adm_frame = tk.Frame(sec, bg=CORES["bg_section"])
        adm_frame.grid(row=6, column=0, columnspan=2, sticky="w", pady=(4,8))

        chk = tk.Checkbutton(
            adm_frame, text="👑  Sou o Administrador deste bolão",
            variable=self._cv_is_adm,
            bg=CORES["bg_section"], fg=CORES["fg_label"],
            font=("Arial",9,"bold"), activebackground=CORES["bg_section"],
            command=self._cad_adm_toggle)
        chk.pack(side="left")

        self._cad_adm_lbl = tk.Label(
            adm_frame,
            text="",
            bg=CORES["bg_section"], fg="#888", font=("Arial",8,"italic"))
        self._cad_adm_lbl.pack(side="left", padx=8)

        bf = tk.Frame(sec, bg=CORES["bg_section"])
        bf.grid(row=7, column=0, columnspan=2, sticky="w", pady=(6,6))
        btn(bf, "✔ CADASTRAR", CORES["btn_verde"],
            self._cadastrar, width=18).pack(side="left")
        btn(bf, "💳 CADASTRAR + PAGAR", CORES["btn_azul"],
            self._cadastrar_e_pagar, width=22).pack(side="left", padx=8)

        tk.Label(sec, text="* campo obrigatório", bg=CORES["bg_section"],
                 fg="#999", font=("Arial",8,"italic")).grid(
            row=8, column=0, columnspan=2, sticky="w", pady=(0,4))

        # ── Painel pagamento rápido (aparece após cadastro) ───────
        self._cad_pag_pid = None  # ID do participante recém-cadastrado

        # ── Seção: Importar membro de bolão anterior ─────────────
        sec2 = section(outer, "📋 IMPORTAR MEMBRO DE BOLÃO ANTERIOR", pady=12)
        sec2.pack(fill="x", pady=(14,0))

        tk.Label(sec2,
                 text="Selecione um participante já cadastrado em outro bolão para importá-lo automaticamente.",
                 bg=CORES["bg_section"], fg="#555", font=("Arial",8,"italic")).pack(anchor="w", pady=(0,6))

        imp_row = tk.Frame(sec2, bg=CORES["bg_section"]); imp_row.pack(fill="x", pady=4)

        tk.Label(imp_row, text="Buscar membro:", bg=CORES["bg_section"],
                 fg=CORES["fg_label"], font=("Arial",9,"bold")).pack(side="left", padx=(0,6))

        self._imp_entry_var = tk.StringVar()
        self._imp_entry = tk.Entry(imp_row, textvariable=self._imp_entry_var,
                                   width=40, font=("Arial",9), relief="solid", bd=1)
        self._imp_entry.pack(side="left", padx=(0,8))
        self._imp_entry.bind("<Return>", lambda e: self._imp_buscar())

        btn(imp_row, "🔍 Buscar", CORES["btn_azul"],
            self._imp_buscar, width=12).pack(side="left", padx=4)
        btn(imp_row, "⬇ Importar Selecionado", CORES["btn_verde"],
            self._imp_importar, width=22).pack(side="left", padx=4)

        # Tabela de resultados da busca — nome diferente do extrato
        cols_imp = {"Nome":220, "Telefone":140, "PIX":200, "Bolão de origem":200}
        fr_imp, self._imp_busca_tree = make_tree(sec2, cols_imp, height=5)
        fr_imp.pack(fill="x", pady=(6,0))
        # Duplo-clique já importa, sem precisar clicar em "Importar Selecionado" depois
        self._imp_busca_tree.bind("<Double-1>", lambda e: self._imp_importar())

        self._imp_status_lbl = tk.Label(sec2,
            text="Após importar, ajuste o valor de cotas/valor esperado no formulário acima se necessário.",
            bg=CORES["bg_section"], fg="#888", font=("Arial",8,"italic"))
        self._imp_status_lbl.pack(anchor="w", pady=(4,0))

    def _cadastrar_e_pagar(self):
        """Cadastra participante e abre janela de pagamento imediatamente."""
        pid = self._cadastrar(retornar_pid=True)
        if pid:
            self._cad_pag_abrir(pid)

    def _cad_adm_toggle(self):
        """Ao marcar como ADM: verifica se bolão tem isento configurado e zera valor."""
        bid = self.bid.get()
        if not self._cv_is_adm.get():
            # Desmarcou — restaura valor normal
            self._cad_adm_lbl.configure(text="")
            self._preencher_valor_cad()
            return

        if not bid:
            self._cad_adm_lbl.configure(text="Selecione um bolão primeiro.")
            self._cv_is_adm.set(0); return

        b = self.db.fetchone(
            "SELECT adm_paga, adm_nome FROM boloes WHERE id=?", (bid,))
        if not b:
            self._cv_is_adm.set(0); return

        adm_paga = b["adm_paga"] or 0
        adm_nome = (b["adm_nome"] or "").strip()

        if not adm_paga:
            # Bolão configurado como ADM isento → zera valor
            self._cv["valor"].delete(0, "end")
            self._cv["valor"].insert(0, "0,00")
            self._cv["cotas"].delete(0, "end")
            self._cv["cotas"].insert(0, "0")
            self._cad_adm_lbl.configure(
                text="✅ ADM isento neste bolão — valor zerado automaticamente.",
                fg="#1D9E75")
            # Preenche nome se campo vazio e ADM configurado
            if adm_nome and not self._cv["nome"].get().strip():
                self._cv["nome"].insert(0, adm_nome)
        else:
            # ADM paga normalmente
            self._preencher_valor_cad()
            self._cad_adm_lbl.configure(
                text="ℹ ADM participa pagando — valor normal aplicado.",
                fg="#2196F3")

    def _calcular_valor_cotas(self):
        """Lê o valor total do bolão ativo, multiplica pelo nº de cotas e preenche o campo."""
        bid = self.bid.get()
        if not bid:
            messagebox.showwarning("Atenção","Selecione um bolão primeiro!"); return
        b = self.db.fetchone("SELECT valor_total, valor_parcela FROM boloes WHERE id=?",(bid,))
        if not b:
            messagebox.showwarning("Atenção","Bolão não encontrado!"); return
        try:
            n_cotas = int(self._cv["cotas"].get().strip())
            if n_cotas < 1: raise ValueError
        except:
            messagebox.showwarning("Atenção","Informe um número de cotas válido (mínimo 1)!"); return
        val_total_part = b["valor_total"] or 0
        if val_total_part <= 0:
            messagebox.showwarning("Atenção",
                "O bolão não tem valor total definido.\n"
                "Edite o bolão e informe o valor total por participante."); return
        valor_calc = n_cotas * val_total_part
        self._cv["valor"].delete(0,"end")
        self._cv["valor"].insert(0, f"{valor_calc:.2f}".replace(".",","))

    def _preencher_valor_cad(self):
        """Preenche o valor esperado com base no bolão ativo e nº de cotas."""
        try:
            bid = self.bid.get()
            if not bid: return
            b = self.db.fetchone("SELECT valor_total FROM boloes WHERE id=?", (bid,))
            if not b: return
            vt = float(b["valor_total"] or 0)
            if vt <= 0: return
            try:
                n_cotas = int(self._cv["cotas"].get().strip() or "1")
                if n_cotas < 1: n_cotas = 1
            except:
                n_cotas = 1
            valor_calc = n_cotas * vt
            self._cv["valor"].delete(0, "end")
            self._cv["valor"].insert(0, f"{valor_calc:.2f}".replace(".", ","))
        except Exception:
            pass  # silencioso — campo ainda pode não existir na inicialização

    def _cadastrar(self, retornar_pid=False):
        bid = self.bid.get()
        if not bid: messagebox.showwarning("Atenção","Selecione um bolão!"); return
        nome = self._cv["nome"].get().strip()
        if not nome: messagebox.showwarning("Atenção","Informe o nome!"); return
        tel  = self._cv["tel"].get().strip()
        if not tel: messagebox.showwarning("Atenção","Informe o telefone!"); return

        n_cotas_str = self._cv["cotas"].get().strip() or "1"
        try:
            n_cotas = int(n_cotas_str)
            if n_cotas < 1: raise ValueError
        except:
            messagebox.showwarning("Atenção","Número de cotas inválido!"); return

        valor_esp  = to_float(self._cv["valor"].get())
        b = self.db.fetchone(
            "SELECT valor_parcela, valor_total, nome, num_participantes FROM boloes WHERE id=?",(bid,))
        parc       = b["valor_parcela"]     if b else 0
        bolao_nome = b["nome"]              if b else "-"
        max_cotas  = b["num_participantes"] if b else 0

        # ── Verifica se telefone já existe na tabela pessoas ─────
        pessoa_existente = self.db.fetchone(
            "SELECT * FROM pessoas WHERE telefone=?", (tel,))

        if pessoa_existente:
            # Verifica se já está cadastrado NESTE bolão
            ja_neste = self.db.fetchone(
                "SELECT id FROM participantes WHERE pessoa_id=? AND bolao_id=? AND ativo=1",
                (pessoa_existente["id"], bid))
            if ja_neste:
                messagebox.showwarning("Participante já cadastrado",
                    f"'{pessoa_existente['nome']}' (tel: {tel})\n"
                    f"já está cadastrado neste bolão!")
                return
            # Pergunta se quer usar os dados existentes
            resp = messagebox.askyesno("Participante encontrado",
                f"Telefone {tel} já cadastrado:\n\n"
                f"Nome: {pessoa_existente['nome']}\n"
                f"PIX:  {pessoa_existente['chave_pix'] or '—'}\n\n"
                f"Usar esses dados para o cadastro?")
            if resp:
                nome = pessoa_existente["nome"]
                pix  = pessoa_existente["chave_pix"] or self._cv["pix"].get().strip()
                self._cv["nome"].delete(0,"end"); self._cv["nome"].insert(0, nome)
            else:
                pix = self._cv["pix"].get().strip()
            pessoa_id = pessoa_existente["id"]
            # Atualiza PIX se novo
            if pix and not pessoa_existente["chave_pix"]:
                self.db.execute("UPDATE pessoas SET chave_pix=? WHERE id=?",
                                (pix, pessoa_id))
        else:
            # Cria nova pessoa
            pix = self._cv["pix"].get().strip()
            self.db.execute(
                "INSERT INTO pessoas (nome,telefone,chave_pix) VALUES (?,?,?)",
                (nome, tel, pix))
            pessoa_id = self.db.fetchone("SELECT last_insert_rowid() as id")["id"]

        # ── Verificar limite de cotas do bolão ───────────────────
        cotas_ocupadas, max_cotas = self._get_cotas_ocupadas(bid)
        if max_cotas > 0:
            cotas_livres = max_cotas - cotas_ocupadas
            if n_cotas > cotas_livres:
                messagebox.showwarning("Limite Atingido",
                    f"O bolão tem {max_cotas} cotas no total.\n"
                    f"Cotas ocupadas: {cotas_ocupadas}\n"
                    f"Cotas livres:   {cotas_livres}\n\n"
                    f"Você está tentando adicionar {n_cotas} cota(s).\n\n"
                    f"Para adicionar mais, edite o bolão em\n"
                    f"'Gerenciar Bolões → Editar' e aumente o número de cotas.")
                return

        is_adm = getattr(self, "_cv_is_adm", tk.IntVar(value=0)).get()

        self.db.execute(
            "INSERT INTO participantes (bolao_id,pessoa_id,nome,telefone,chave_pix,"
            "valor_esperado,observacoes,is_adm) VALUES (?,?,?,?,?,?,?,?)",
            (bid, pessoa_id, nome, tel, pix,
             valor_esp, self._cv["obs"].get("1.0","end").strip(), is_adm))

        pid_novo = self.db.fetchone("SELECT last_insert_rowid() as id")["id"]

        # Resumo na messagebox de confirmação
        linhas = [
            "✅  Participante cadastrado com sucesso!",
            "",
            f"Nome:           {nome}",
            f"Bolão:          {bolao_nome}",
            f"Nº de Cotas:    {n_cotas}",
            f"Valor Esperado: {fmt_brl(valor_esp)}",
        ]
        if is_adm:
            linhas.append("👑  Cadastrado como Administrador")
        if parc and parc > 0:
            try:
                total_parc = round(valor_esp / parc) if parc > 0 else 0
                if total_parc > 0:
                    linhas.append(f"Parcelas:       {total_parc}x de {fmt_brl(parc)}")
            except: pass
        messagebox.showinfo("Participante Cadastrado", "\n".join(linhas))

        for k,w in self._cv.items():
            (w.delete("1.0","end") if isinstance(w,tk.Text) else w.delete(0,"end"))
        self._cv["cotas"].insert(0,"1")
        # Limpa checkbox ADM e label
        try:
            self._cv_is_adm.set(0)
            self._cad_adm_lbl.configure(text="", fg="#888")
        except: pass
        self._refresh_all()
        self._preencher_valor_cad()
        if retornar_pid:
            return pid_novo

    def _cad_pag_abrir(self, pid):
        """Abre janela de pagamento rápido para o participante recém-cadastrado."""
        pt = self.db.fetchone("SELECT * FROM participantes WHERE id=?", (pid,))
        if not pt: return
        b  = self.db.fetchone("SELECT valor_parcela FROM boloes WHERE id=?", (self.bid.get(),))
        parc = float(b["valor_parcela"] or 0) if b else 0

        win = tk.Toplevel(self.root)
        win.title("Registrar Pagamento")
        win.geometry("520x280")
        win.configure(bg=CORES["bg_section"])
        win.grab_set(); win.lift(); win.focus_force()
        # Centraliza na tela
        win.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width()  - 520) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - 280) // 2
        win.geometry(f"520x280+{x}+{y}")

        tk.Label(win, text="💳 REGISTRAR PAGAMENTO",
                 bg=CORES["bg_section"], fg=CORES["fg_title"],
                 font=("Arial",12,"bold")).pack(pady=12)

        f = tk.Frame(win, bg=CORES["bg_section"], padx=24); f.pack(fill="x")

        # Nome
        r0 = tk.Frame(f, bg=CORES["bg_section"]); r0.pack(fill="x", pady=4)
        tk.Label(r0, text="Participante:", bg=CORES["bg_section"],
                 fg=CORES["fg_label"], font=("Arial",9,"bold"), width=14, anchor="w").pack(side="left")
        tk.Label(r0, text=pt["nome"], bg=CORES["bg_section"],
                 fg=CORES["fg_title"], font=("Arial",10,"bold")).pack(side="left")

        # Valor
        r1 = tk.Frame(f, bg=CORES["bg_section"]); r1.pack(fill="x", pady=6)
        tk.Label(r1, text="Valor (R$):", bg=CORES["bg_section"],
                 fg=CORES["fg_label"], font=("Arial",9,"bold"), width=14, anchor="w").pack(side="left")
        e_val = entry(r1, width=14)
        e_val.insert(0, f"{parc:.2f}".replace(".",",") if parc > 0 else "")
        e_val.pack(side="left"); e_val.focus_set(); e_val.selection_range(0,"end")

        # Data
        r2 = tk.Frame(f, bg=CORES["bg_section"]); r2.pack(fill="x", pady=6)
        tk.Label(r2, text="Data:", bg=CORES["bg_section"],
                 fg=CORES["fg_label"], font=("Arial",9,"bold"), width=14, anchor="w").pack(side="left")
        e_dt = entry(r2, width=14)
        e_dt.insert(0, date.today().strftime("%d/%m/%Y"))
        e_dt.pack(side="left")

        # Mês referência
        r3 = tk.Frame(f, bg=CORES["bg_section"]); r3.pack(fill="x", pady=6)
        tk.Label(r3, text="Mês Referência:", bg=CORES["bg_section"],
                 fg=CORES["fg_label"], font=("Arial",9,"bold"), width=14, anchor="w").pack(side="left")
        cb_mes = ttk.Combobox(r3, values=MESES, width=8, state="readonly")
        cb_mes.set(MESES[date.today().month-1]); cb_mes.pack(side="left", padx=4)
        tk.Label(r3, text="/", bg=CORES["bg_section"], font=("Arial",10,"bold")).pack(side="left")
        cb_ano = ttk.Combobox(r3, values=[str(y) for y in range(2020,2036)],
                               width=6, state="readonly")
        cb_ano.set(str(date.today().year)); cb_ano.pack(side="left", padx=4)

        # Status inline (some no lugar do popup final de confirmação)
        lbl_status = tk.Label(win, text="", bg=CORES["bg_section"],
                              fg="#1D9E75", font=("Arial",9,"bold"))
        lbl_status.pack(pady=(0,2))

        # Botões
        bf = tk.Frame(win, bg=CORES["bg_section"]); bf.pack(pady=14)

        def registrar():
            val = to_float(e_val.get())
            if val <= 0: messagebox.showwarning("Atenção","Informe o valor!"); return
            dt = e_dt.get().strip()
            mes = cb_mes.get(); ano = cb_ano.get()
            mes_ref = f"{MESES.index(mes)+1:02d}/{ano}" if mes in MESES else ""
            self.db.execute("""
                INSERT INTO pagamentos
                (participante_id,bolao_id,mes_referencia,valor,
                 data_pagamento,depositado,observacoes)
                VALUES (?,?,?,?,?,0,'Cadastro + pagamento simultâneo')
            """, (pid, self.bid.get(), mes_ref, val, dt))
            self._refresh_all()
            # Aviso inline em vez de popup bloqueante — fecha sozinho logo em seguida
            lbl_status.configure(text=f"✅ Pagamento de {fmt_brl(val)} registrado para {pt['nome']}!")
            btn_registrar.configure(state="disabled")
            btn_pular.configure(state="disabled")
            win.after(900, win.destroy)

        btn_registrar = btn(bf, "💳 REGISTRAR PAGAMENTO", CORES["btn_verde"], registrar, width=24)
        btn_registrar.pack(side="left", padx=6)
        btn_pular = btn(bf, "✖ Pular", CORES["btn_cinza"], win.destroy, width=10)
        btn_pular.pack(side="left", padx=6)

        # Enter em qualquer campo já registra — os valores já vêm
        # preenchidos com o padrão (valor da cota, data de hoje), então
        # aceitar o padrão vira só apertar Enter.
        e_val.bind("<Return>", lambda e: registrar())
        e_dt.bind("<Return>", lambda e: registrar())

    def _imp_buscar(self):
        """Busca participantes em TODOS os bolões pelo nome digitado."""
        bid_atual = self.bid.get()
        termo = self._imp_entry_var.get().strip().lower()
        if not termo:
            messagebox.showwarning("Atenção","Digite parte do nome para buscar!"); return

        self._imp_status_lbl.configure(
            text="Após importar, ajuste o valor de cotas/valor esperado no formulário acima se necessário.",
            fg="#888")
        self._imp_busca_tree.delete(*self._imp_busca_tree.get_children())

        rows = self.db.fetchall("""
            SELECT DISTINCT p.nome, p.telefone, p.chave_pix, b.nome as bolao_nome, p.id
            FROM participantes p
            JOIN boloes b ON p.bolao_id = b.id
            WHERE LOWER(p.nome) LIKE ?
              AND p.ativo = 1
              AND p.bolao_id != ?
            ORDER BY p.nome, b.nome
        """, (f"%{termo}%", bid_atual or 0))

        if not rows:
            messagebox.showinfo("Busca","Nenhum participante encontrado com esse nome.")
            return

        vistos = set()
        for r in rows:
            chave = (r["nome"].strip().lower(), (r["telefone"] or "").strip())
            if chave not in vistos:
                vistos.add(chave)
                self._imp_busca_tree.insert("","end", iid=str(r["id"]), values=(
                    r["nome"], r["telefone"] or "-",
                    r["chave_pix"] or "-", r["bolao_nome"]))

    def _imp_importar(self):
        """Importa o membro selecionado na tabela de busca para o formulário de cadastro."""
        sel = self._imp_busca_tree.selection()
        if not sel:
            messagebox.showwarning("Atenção","Selecione um participante na lista!"); return

        pid_orig = int(sel[0])
        pt = self.db.fetchone("SELECT * FROM participantes WHERE id=?", (pid_orig,))
        if not pt: return

        # Preenche o formulário principal com os dados do membro selecionado
        for k, w in self._cv.items():
            if isinstance(w, tk.Text):
                w.delete("1.0","end")
            else:
                w.delete(0,"end")

        self._cv["nome"].insert(0, pt["nome"])
        self._cv["tel"].insert(0,  pt["telefone"] or "")
        self._cv["pix"].insert(0,  pt["chave_pix"] or "")
        self._cv["cotas"].insert(0, "1")
        if pt["observacoes"]:
            self._cv["obs"].insert("1.0", pt["observacoes"])

        # Calcula o valor esperado (1 cota x valor do bolão ativo) em vez
        # de deixar "0,00" fixo — mesma lógica usada ao digitar cotas.
        self._preencher_valor_cad()

        # Limpa a tabela de busca — apenas a da aba Cadastrar
        self._imp_busca_tree.delete(*self._imp_busca_tree.get_children())
        self._imp_entry_var.set("")

        self._imp_status_lbl.configure(
            text=f"✅ '{pt['nome']}' importado — confira cotas/valor acima e clique em CADASTRAR.",
            fg="#1D9E75")

    # ════════════════════════════════════════════════════════════
    #  IMPORTAR EXTRATO — métodos funcionais
    # ════════════════════════════════════════════════════════════
    def _imp_escolher_pdf(self):
        path = filedialog.askopenfilename(
            title="Selecionar extrato Nubank",
            filetypes=[("PDF","*.pdf"),("Todos","*.*")])
        if path:
            self._imp_pdf_path.set(path)
            self._imp_status.configure(text="PDF carregado. Clique em ANALISAR.")

    def _imp_analisar(self):
        path = self._imp_pdf_path.get()
        if not path:
            messagebox.showwarning("Atenção","Selecione um arquivo PDF!"); return
        data_ini_str = self._imp_data_ini.get().strip()
        try:
            data_ini = datetime.strptime(data_ini_str, "%d/%m/%Y")
        except:
            messagebox.showwarning("Atenção","Data inválida! Use DD/MM/AAAA"); return

        try:
            import pdfplumber
        except ImportError:
            messagebox.showerror("Biblioteca ausente",
                "Instale o pdfplumber:\n\npython -m pip install pdfplumber\n\n"
                "Depois reinicie o sistema."); return

        self._imp_status.configure(text="Lendo PDF..."); self.root.update_idletasks()

        try:
            texto = ""
            with pdfplumber.open(path) as pdf:
                for pg in pdf.pages:
                    texto += (pg.extract_text() or "") + "\n"
        except Exception as ex:
            messagebox.showerror("Erro lendo PDF", str(ex)); return

        self._imp_status.configure(text="Analisando..."); self.root.update_idletasks()

        import re as _re
        MESES = {"JAN":1,"FEV":2,"MAR":3,"ABR":4,"MAI":5,"JUN":6,
                 "JUL":7,"AGO":8,"SET":9,"OUT":10,"NOV":11,"DEZ":12}
        CREDITOS = [
            "Transferência Recebida","Transferência recebida pelo Pix",
            "Transferencia Recebida","Transferencia recebida pelo Pix",
            "PIX recebido","Pix recebido","Pagamento recebido","Crédito em conta",
        ]

        data_atual = None
        transacoes = []
        for linha in texto.split("\n"):
            linha = linha.strip()
            if not linha: continue
            md = _re.match(
                r"^(\d{1,2})\s+(JAN|FEV|MAR|ABR|MAI|JUN|JUL|AGO|SET|OUT|NOV|DEZ)\s+(\d{4})",
                linha, _re.IGNORECASE)
            if md:
                try:
                    data_atual = datetime(int(md.group(3)),
                                         MESES[md.group(2).upper()],
                                         int(md.group(1)))
                except: pass
                continue
            if not data_atual or data_atual < data_ini: continue
            tipo = None
            for t in CREDITOS:
                if t.lower() in linha.lower():
                    tipo = t; break
            if not tipo: continue
            mv = _re.search(r"(\d{1,3}(?:\.\d{3})*,\d{2})", linha)
            if not mv: continue
            valor = to_float(mv.group(1))
            if valor <= 0: continue
            rem = linha
            rem = _re.sub(_re.escape(tipo), "", rem, flags=_re.IGNORECASE)
            rem = _re.sub(r"\d{1,3}(?:\.\d{3})*,\d{2}", "", rem)
            rem = _re.sub(r"-\s*[\•\*\d][\•\*\d]+.*$", "", rem)
            rem = _re.sub(r"(agência|agencia|conta|banco|S\.A\.|IP\s*\()",
                          "", rem, flags=_re.IGNORECASE)
            rem = _re.sub(r"\s{2,}", " ", rem).strip(" -•·+")
            if len(rem) < 2: rem = "Não identificado"
            transacoes.append({"data": data_atual.strftime("%d/%m/%Y"),
                                "valor": valor, "remetente": rem, "manual": False})

        # Deduplica
        vistos = set(); unicas = []
        for tx in transacoes:
            ch = (tx["data"], tx["valor"], tx["remetente"][:25])
            if ch not in vistos: vistos.add(ch); unicas.append(tx)

        if not unicas:
            messagebox.showinfo("Resultado",
                "Nenhuma transação de crédito encontrada.\n\n"
                "Verifique:\n• Data inicial correta?\n"
                "• PDF é o extrato Nubank?\n"
                "• Período inclui a data informada?")
            self._imp_status.configure(text="Nenhuma transação encontrada."); return

        self._imp_carregar_tabela(unicas)
        self._imp_status.configure(
            text=f"✅ {len(unicas)} crédito(s) encontrado(s) a partir de {data_ini_str}")

    def _imp_carregar_tabela(self, transacoes):
        """Popula a tabela com matching por valor da parcela."""
        import unicodedata
        self._imp_tree.delete(*self._imp_tree.get_children())
        self._imp_linhas  = []
        self._imp_sel_ids = set()

        boloes   = self.db.fetchall("SELECT id, nome, valor_parcela FROM boloes WHERE encerrado=0 ORDER BY nome")
        nomes_bol = ["Outros"] + [f"{b['nome']} (ID: {b['id']})" for b in boloes]
        try: self._imp_cb_bolao["values"] = nomes_bol; self._imp_cb_bolao.set("Outros")
        except: pass
        try: self._imp_man_bolao["values"] = nomes_bol
        except: pass

        todos_part = self.db.fetchall("""
            SELECT p.id, p.nome, p.bolao_id, b.nome as bolao_nome, b.valor_parcela
            FROM participantes p JOIN boloes b ON p.bolao_id=b.id
            WHERE p.ativo=1 ORDER BY p.nome, b.nome
        """)
        nomes_part = sorted([f"{p['nome']} — {p['bolao_nome']} (ID: {p['id']})"
                             for p in todos_part])
        try: self._imp_cb_part["values"] = nomes_part
        except: pass

        # Agrupa por bolão para lookup rápido
        parts_por_bolao = {}
        for pt in todos_part:
            parts_por_bolao.setdefault(pt["bolao_id"], []).append(pt)

        def norm(s):
            return ''.join(c for c in unicodedata.normalize('NFD',s.lower())
                           if unicodedata.category(c)!='Mn')

        def score(rem, nome):
            pr = set(norm(rem).split()); pn = set(norm(nome).split())
            if not pr: return 0
            return len(pr & pn) / max(len(pr), len(pn))

        vinculos = getattr(self, "_imp_vinculos", {})
        linhas_prep = []

        for tx in transacoes:
            valor = tx["valor"]
            boloes_match = [b for b in boloes
                            if float(b["valor_parcela"] or 0) > 0 and
                            abs(valor / float(b["valor_parcela"]) -
                                round(valor / float(b["valor_parcela"]))) < 0.02 and
                            round(valor / float(b["valor_parcela"])) >= 1]

            if len(boloes_match) == 1:
                bm = boloes_match[0]
                bolao_sug = f"{bm['nome']} (ID: {bm['id']})"
                melhor = max(parts_por_bolao.get(bm["id"],[]),
                             key=lambda p: score(tx["remetente"], p["nome"]),
                             default=None)
                if melhor and score(tx["remetente"], melhor["nome"]) >= 0.4:
                    part_sug = f"{melhor['nome']} — {bm['nome']} (ID: {melhor['id']})"
                    tag = "match_ok" if score(tx["remetente"],melhor["nome"])>=0.6 else "match_par"
                else:
                    part_sug = ""; tag = "match_par"
            elif len(boloes_match) > 1:
                bolao_sug = ""; part_sug = ""; tag = "match_par"
            else:
                bolao_sug = "Outros"; part_sug = ""; tag = "match_no"

            chave = f"{tx['data']}|{tx['valor']}|{tx['remetente'][:20]}"
            if chave in vinculos:
                v = vinculos[chave]
                part_sug = v.get("part", part_sug)
                bolao_sug = v.get("bolao", bolao_sug)
                tag = "match_ok"
            if tx.get("manual"): tag = "manual"

            linhas_prep.append({**tx, "part_sug":part_sug,
                                 "bolao_sug":bolao_sug, "tag":tag, "chave":chave})

        # Ordena: data → bolão → remetente
        def sk(l):
            try: dt = datetime.strptime(l["data"],"%d/%m/%Y")
            except: dt = datetime(2000,1,1)
            return (dt, l["bolao_sug"] or "zzz", l["remetente"].lower())
        linhas_prep.sort(key=sk)

        for i, ln in enumerate(linhas_prep):
            iid = str(i)
            self._imp_tree.insert("","end", iid=iid, tags=(ln["tag"],), values=(
                "☐", ln["data"], fmt_brl(ln["valor"]), ln["remetente"],
                ln["part_sug"], ln["bolao_sug"],
                "Manual" if ln.get("manual") else "Extrato"))
            self._imp_linhas.append({**ln, "iid": iid})

    def _imp_toggle_sel(self, e=None):
        row = self._imp_tree.identify_row(e.y)
        col = self._imp_tree.identify_column(e.x)
        if not row: return
        if col == "#1":
            if row in self._imp_sel_ids:
                self._imp_sel_ids.discard(row)
                vals = list(self._imp_tree.item(row,"values")); vals[0]="☐"
                self._imp_tree.item(row, values=vals)
            else:
                self._imp_sel_ids.add(row)
                vals = list(self._imp_tree.item(row,"values")); vals[0]="✅"
                self._imp_tree.item(row, values=vals)

    def _imp_sel_para_edicao(self, e=None):
        sel = self._imp_tree.selection()
        if not sel: return
        vals = self._imp_tree.item(sel[0],"values")
        try: self._imp_cb_part.set(vals[4] if len(vals)>4 else "")
        except: pass
        try: self._imp_cb_bolao.set(vals[5] if len(vals)>5 else "")
        except: pass

    def _imp_aplicar_vinculo(self):
        sel = self._imp_tree.selection()
        if not sel: messagebox.showwarning("Atenção","Selecione uma linha!"); return
        iid  = sel[0]
        part = self._imp_cb_part.get().strip()
        bol  = self._imp_cb_bolao.get().strip()
        vals = list(self._imp_tree.item(iid,"values"))
        vals[4] = part; vals[5] = bol
        self._imp_tree.item(iid, values=vals, tags=("match_ok",))
        idx = int(iid)
        if idx < len(self._imp_linhas):
            self._imp_linhas[idx]["part_sug"]  = part
            self._imp_linhas[idx]["bolao_sug"] = bol
            chave = self._imp_linhas[idx].get("chave","")
            if chave:
                if not hasattr(self,"_imp_vinculos"): self._imp_vinculos = {}
                self._imp_vinculos[chave] = {"part": part, "bolao": bol}

    def _imp_add_manual(self):
        nome = self._imp_man_nome.get().strip()
        val  = to_float(self._imp_man_val.get())
        dt   = self._imp_man_dt.get().strip()
        bol  = self._imp_man_bolao.get() if hasattr(self,"_imp_man_bolao") else "Outros"
        if not nome or val <= 0:
            messagebox.showwarning("Atenção","Informe nome e valor!"); return
        nova = {"data":dt,"valor":val,"remetente":nome,"manual":True,"bolao_sug":bol,"part_sug":""}
        todas = [dict(ln) for ln in self._imp_linhas] + [nova]
        self._imp_carregar_tabela(todas)
        self._imp_man_nome.delete(0,"end"); self._imp_man_val.delete(0,"end")

    def _imp_importar_pagamentos(self):
        import re as _re
        if not self._imp_sel_ids:
            messagebox.showwarning("Atenção","Marque as transações com ✅ que deseja registrar!"); return
        por_bolao = {}
        erros = []
        for iid in sorted(self._imp_sel_ids, key=int):
            vals      = self._imp_tree.item(iid,"values")
            bolao_str = vals[5] if len(vals)>5 else ""
            part_str  = vals[4] if len(vals)>4 else ""
            if bolao_str in ("Outros","","—") or not bolao_str: continue
            if not part_str: erros.append(f"Linha {int(iid)+1}: sem participante"); continue
            m_bid = _re.search(r"\(ID: (\d+)\)", bolao_str)
            m_pid = _re.search(r"\(ID: (\d+)\)", part_str)
            if not m_bid or not m_pid:
                erros.append(f"Linha {int(iid)+1}: vínculo inválido"); continue
            por_bolao.setdefault(bolao_str,[]).append({
                "iid":iid,"bid":int(m_bid.group(1)),"pid":int(m_pid.group(1)),
                "data":vals[1],"valor":to_float(vals[2]),
                "part_nome":_re.sub(r"\(ID:\s*\d+\)","",part_str).split("—")[0].strip(),
            })
        if not por_bolao:
            messagebox.showinfo("Nada a registrar",
                "Nenhuma transação vinculada a bolão gerenciado.\n"
                "Transações 'Outros' só aparecem na mensagem WhatsApp.")
            return
        total_imp = total_dup = 0
        for bolao_str, linhas in por_bolao.items():
            nome_bol = bolao_str.split(" (ID:")[0]
            resumo = "\n".join(f"  • {l['part_nome']} — {fmt_brl(l['valor'])} em {l['data']}"
                               for l in linhas)
            if not messagebox.askyesno("Confirmar",
                f"Registrar no bolão:\n{nome_bol}\n\n{resumo}\n\n"
                f"Total: {len(linhas)} pagamento(s)"): continue
            for l in linhas:
                existe = self.db.fetchone(
                    "SELECT id FROM pagamentos WHERE participante_id=? AND bolao_id=? "
                    "AND data_pagamento=? AND valor=?",
                    (l["pid"],l["bid"],l["data"],l["valor"]))
                if existe: total_dup+=1; continue
                mes_ref = ""
                try: mes_ref = datetime.strptime(l["data"],"%d/%m/%Y").strftime("%m/%Y")
                except: pass
                self.db.execute(
                    "INSERT INTO pagamentos (participante_id,bolao_id,mes_referencia,valor,"
                    "data_pagamento,depositado,data_deposito,observacoes) VALUES (?,?,?,?,?,1,?,"
                    "'Importado do extrato Nubank')",
                    (l["pid"],l["bid"],mes_ref,l["valor"],l["data"],l["data"]))
                total_imp+=1
        msg = f"✅ {total_imp} pagamento(s) registrado(s)."
        if total_dup: msg += f"\n⚠ {total_dup} duplicata(s) ignorada(s)."
        if erros: msg += "\n❌ " + "\n".join(erros[:5])
        if total_imp > 0: messagebox.showinfo("Concluído",msg); self._refresh_all()

    def _imp_gerar_whatsapp(self):
        import re as _re
        ids_usar = self._imp_sel_ids if self._imp_sel_ids else \
                   {str(i) for i in range(len(self._imp_linhas))}
        if not ids_usar:
            messagebox.showwarning("Atenção","Não há transações!"); return
        confirmados_raw = []
        for iid in sorted(ids_usar, key=int):
            vals = self._imp_tree.item(iid,"values")
            if not vals: continue
            try: dc = datetime.strptime(vals[1],"%d/%m/%Y").strftime("%d/%m")
            except: dc = vals[1]
            nome = _re.sub(r"\(ID:\s*\d+\)","",vals[4]).split("—")[0].strip() if vals[4] else vals[3]
            confirmados_raw.append({"nome":nome,"valor":vals[2],"data":dc,"bolao":vals[5] if len(vals)>5 else ""})
        if not confirmados_raw:
            messagebox.showwarning("Atenção","Nenhum participante para a mensagem!"); return
        boloes_pres = sorted(set(r["bolao"] for r in confirmados_raw
                                 if r["bolao"] not in ("","Outros","—")))
        win = tk.Toplevel(self.root); win.title("Gerar Mensagem WhatsApp")
        win.geometry("500x280"); win.configure(bg=CORES["bg_section"]); win.grab_set(); win.lift()
        tk.Label(win,text="📲 CONFIGURAR MENSAGEM",bg=CORES["bg_section"],
                 fg=CORES["fg_title"],font=("Arial",11,"bold")).pack(pady=10)
        form = tk.Frame(win,bg=CORES["bg_section"],padx=24); form.pack(fill="x")
        tk.Label(form,text="Para qual bolão:",bg=CORES["bg_section"],
                 fg=CORES["fg_label"],font=("Arial",9,"bold")).grid(row=0,column=0,sticky="w",pady=6)
        cb_bol = ttk.Combobox(form,width=38,state="readonly",font=("Arial",9),
                               values=["Outros (avulso)"]+boloes_pres)
        cb_bol.set("Outros (avulso)"); cb_bol.grid(row=0,column=1,sticky="w",padx=8,pady=6)
        tk.Label(form,text="Loteria:",bg=CORES["bg_section"],
                 fg=CORES["fg_label"],font=("Arial",9,"bold")).grid(row=1,column=0,sticky="w",pady=6)
        cb_lot = ttk.Combobox(form,width=16,state="readonly",font=("Arial",9),values=LOTERIAS)
        cb_lot.set("Mega-Sena"); cb_lot.grid(row=1,column=1,sticky="w",padx=8,pady=6)
        tk.Label(form,text="Concurso:",bg=CORES["bg_section"],
                 fg=CORES["fg_label"],font=("Arial",9,"bold")).grid(row=2,column=0,sticky="w",pady=6)
        e_conc = entry(form,width=14); e_conc.grid(row=2,column=1,sticky="w",padx=8,pady=6)
        tk.Label(form,text="Gestor:",bg=CORES["bg_section"],
                 fg=CORES["fg_label"],font=("Arial",9,"bold")).grid(row=3,column=0,sticky="w",pady=6)
        e_gest = entry(form,width=20)
        adm = self.db.fetchone("SELECT adm_nome FROM boloes WHERE status='ATIVO' LIMIT 1")
        e_gest.insert(0, adm["adm_nome"] if adm and adm["adm_nome"] else "Elton Luis")
        e_gest.grid(row=3,column=1,sticky="w",padx=8,pady=6)
        def gerar():
            bolao_sel = cb_bol.get(); concurso = e_conc.get().strip()
            gestor = e_gest.get().strip() or "Elton Luis"; loteria = cb_lot.get()
            if bolao_sel == "Outros (avulso)":
                lista = confirmados_raw
                cab = f"*{loteria.upper()}"
                cab += f" — CONCURSO {concurso}*" if concurso else "*"
            else:
                lista = [r for r in confirmados_raw if r["bolao"]==bolao_sel]
                nb = bolao_sel.split(" (ID:")[0]
                cab = f"*{nb.upper()}*" + (f"\n*Concurso: {concurso}*" if concurso else "")
            if not lista:
                messagebox.showwarning("Atenção","Nenhum participante para este filtro!"); return
            msg = "\n".join([f"🎰 {cab}","📋 Pagamentos confirmados:",""] +
                            [f"✅ {r['nome']} — {r['valor']} ({r['data']})" for r in lista] +
                            ["",f"👥 Participantes confirmados: {len(lista)}","",
                             "📌 *Importante*","",
                             "Sozinho: 1 jogo.","Com o grupo: centenas ou milhares de jogos.","",
                             "Prêmio nunca é certo. Mas chance maior é matemática.","",
                             "Seguimos buscando o prêmio. 🍀","",f"_Gestão: {gestor}_"])
            w2 = tk.Toplevel(win); w2.title("Mensagem WhatsApp")
            w2.geometry("520x580"); w2.configure(bg=CORES["bg_section"]); w2.lift()
            tk.Label(w2,text="Copie e cole no WhatsApp:",bg=CORES["bg_section"],
                     fg=CORES["fg_label"],font=("Arial",9,"bold")).pack(pady=8)
            txt = tk.Text(w2,font=("Arial",10),wrap="word",padx=10,pady=8)
            txt.pack(fill="both",expand=True,padx=16,pady=(0,8))
            txt.insert("1.0",msg)
            def cp(): w2.clipboard_clear(); w2.clipboard_append(msg); messagebox.showinfo("Copiado!","Mensagem copiada!")
            bf2 = tk.Frame(w2,bg=CORES["bg_section"]); bf2.pack(pady=8)
            btn(bf2,"📋 COPIAR",CORES["btn_verde"],cp,width=16).pack(side="left",padx=4)
            btn(bf2,"✖ Fechar",CORES["btn_cinza"],w2.destroy,width=10).pack(side="left",padx=4)
            win.destroy()
        bf = tk.Frame(win,bg=CORES["bg_section"]); bf.pack(pady=10)
        btn(bf,"📲 GERAR",CORES["btn_verde"],gerar,width=14).pack(side="left",padx=4)
        btn(bf,"✖ Cancelar",CORES["btn_cinza"],win.destroy,width=12).pack(side="left",padx=4)

    def _flb_conferencia(self):
        ids_usar = self._imp_sel_ids if self._imp_sel_ids else \
                   {str(i) for i in range(len(self._imp_linhas))}
        confirmados = []
        for iid in sorted(ids_usar, key=int):
            vals = self._imp_tree.item(iid,"values")
            if not vals: continue
            import re as _re
            try: dt = datetime.strptime(vals[1],"%d/%m/%Y").strftime("%d/%m")
            except: dt = vals[1]
            nome = _re.sub(r"\(ID:\s*\d+\)","",vals[4]).split("—")[0].strip() if vals[4] else vals[3]
            confirmados.append(f"✅ {nome} — {vals[2]} ({dt})")
        if not confirmados: messagebox.showinfo("Vazio","Nenhum participante."); return
        win = tk.Toplevel(self.root); win.title("Conferência de Participantes")
        win.geometry("480x500"); win.configure(bg=CORES["bg_section"]); win.grab_set(); win.lift()
        tk.Label(win,text=f"PARTICIPANTES — {len(confirmados)}",
                 bg=CORES["bg_section"],fg=CORES["fg_title"],font=("Arial",11,"bold")).pack(pady=10)
        txt = tk.Text(win,font=("Arial",10),wrap="word",padx=10,pady=8)
        txt.pack(fill="both",expand=True,padx=16)
        txt.insert("1.0","\n".join(confirmados)); txt.configure(state="disabled")
        def cp(): win.clipboard_clear(); win.clipboard_append("\n".join(confirmados)); messagebox.showinfo("Copiado","Lista copiada!")
        bf = tk.Frame(win,bg=CORES["bg_section"]); bf.pack(pady=8)
        btn(bf,"📋 Copiar",CORES["btn_verde"],cp,width=14).pack(side="left",padx=4)
        btn(bf,"✖ Fechar",CORES["btn_cinza"],win.destroy,width=10).pack(side="left",padx=4)

    def _imp_limpar(self):
        self._imp_tree.delete(*self._imp_tree.get_children())
        self._imp_linhas  = []
        self._imp_sel_ids = set()
        self._imp_pdf_path.set("")
        self._imp_status.configure(text="")
    def _build_cad_editar(self):
        p = self.tab_cad_edit

        top = tk.Frame(p, bg=CORES["bg_frame"]); top.pack(fill="x", padx=20, pady=12)
        tk.Label(top, text="Participante:", bg=CORES["bg_frame"],
                 fg=CORES["fg_label"], font=("Arial",9,"bold")).pack(side="left")
        self.cad_edit_cb = ttk.Combobox(top, width=50, state="readonly", font=("Arial",10))
        self.cad_edit_cb.pack(side="left", padx=8)
        self.cad_edit_cb.bind("<<ComboboxSelected>>", self._cad_edit_sel)
        btn(top, "🔄", CORES["btn_azul"], self._refresh_all, width=4).pack(side="left")

        self._cad_edit_form_frame = tk.Frame(p, bg=CORES["bg_frame"])
        self._cad_edit_form_frame.pack(fill="x", padx=20)
        # Formulário preenchido ao selecionar
        self._cad_edit_vars = {}
        fields = [("Nome Completo:*","nome"),("Telefone:","telefone"),
                  ("Chave PIX:","chave_pix"),("Valor Esperado (R$):","valor_esperado"),
                  ("Observações:","observacoes")]
        sec_e = section(self._cad_edit_form_frame, "EDITAR DADOS DO PARTICIPANTE")
        sec_e.pack(fill="x", pady=8); sec_e.columnconfigure(1, weight=1)
        for i,(lbl,key) in enumerate(fields):
            tk.Label(sec_e, text=lbl, bg=CORES["bg_section"], fg=CORES["fg_label"],
                     font=("Arial",9,"bold")).grid(row=i,column=0,sticky="w",padx=(0,12),pady=6)
            w = entry(sec_e, width=45); w.grid(row=i, column=1, sticky="ew", pady=6)
            self._cad_edit_vars[key] = w
        bf = tk.Frame(sec_e, bg=CORES["bg_section"]); bf.grid(row=6,column=0,columnspan=2,sticky="w",pady=8)
        btn(bf,"💾 SALVAR ALTERAÇÕES",CORES["btn_verde"],self._cad_edit_salvar,width=22).pack(side="left",padx=4)

    def _cad_edit_sel(self, e=None):
        sel = self.cad_edit_cb.get()
        if not sel: return
        import re as _re
        m = _re.search(r"\(ID: (\d+)\)", sel)
        if not m: return
        pt = self.db.fetchone("SELECT * FROM participantes WHERE id=?", (int(m.group(1)),))
        if not pt: return
        for key, w in self._cad_edit_vars.items():
            w.delete(0,"end")
            val = pt[key] if pt[key] is not None else ""
            if key == "valor_esperado":
                try: val = f"{float(val):.2f}".replace(".",",")
                except: val = "0,00"
            w.insert(0, str(val))

    def _cad_edit_salvar(self):
        sel = self.cad_edit_cb.get()
        if not sel: messagebox.showwarning("Atenção","Selecione um participante!"); return
        import re as _re
        m = _re.search(r"\(ID: (\d+)\)", sel)
        if not m: return
        pid = int(m.group(1))
        nome = self._cad_edit_vars["nome"].get().strip()
        if not nome: messagebox.showwarning("Atenção","Nome é obrigatório!"); return
        self.db.execute(
            "UPDATE participantes SET nome=?,telefone=?,chave_pix=?,valor_esperado=?,observacoes=? WHERE id=?",
            (nome, self._cad_edit_vars["telefone"].get(),
             self._cad_edit_vars["chave_pix"].get(),
             to_float(self._cad_edit_vars["valor_esperado"].get()),
             self._cad_edit_vars["observacoes"].get(), pid))
        messagebox.showinfo("Salvo","Participante atualizado!")
        self._refresh_all()

    # ════════════════════════════════════════════════════════════
    #  ABA CADASTRAR — Lista / Remover
    # ════════════════════════════════════════════════════════════
    def _build_cad_lista(self):
        p = self.tab_cad_lista

        top = tk.Frame(p, bg=CORES["bg_frame"]); top.pack(fill="x", padx=20, pady=10)
        btn(top,"🔄 Atualizar",CORES["btn_azul"],self._cad_lista_load,width=14).pack(side="left",padx=4)
        btn(top,"🗑 Remover Selecionado",CORES["btn_vermelho"],self._cad_remover,width=22).pack(side="left",padx=4)
        tk.Label(top, text="  ⚠ Não é possível remover participante com pagamentos depositados.",
                 bg=CORES["bg_frame"], fg="#e67e22", font=("Arial",8,"italic")).pack(side="left",padx=8)

        sec = section(p,"PARTICIPANTES DO BOLÃO ATIVO")
        sec.pack(fill="both", expand=True, padx=20, pady=(0,10))
        cols = {"ID":50,"Nome":220,"Telefone":140,"Cotas":60,"Valor Esperado":130,"Status":120}
        fr, self._cad_lista_tree = make_tree(sec, cols, height=20)
        fr.pack(fill="both", expand=True)
        self._cad_lista_tree.tag_configure("quitado", background="#d5f5e3")
        self._cad_lista_tree.tag_configure("pendente", background="#fde8d8")

    def _cad_lista_load(self):
        bid = self.bid.get()
        self._cad_lista_tree.delete(*self._cad_lista_tree.get_children())
        if not bid: return
        b   = self.db.fetchone("SELECT * FROM boloes WHERE id=?", (bid,))
        bd  = dict(b) if b else {}
        vt  = float(bd.get("valor_total",0) or 0)
        _,parc_esp,parc = self._calc_parcela_atual(bd)
        parts = self.db.fetchall(
            "SELECT * FROM participantes WHERE bolao_id=? AND ativo=1 ORDER BY nome",(bid,))
        for pt in parts:
            pt_d = dict(pt)
            pago_row = self.db.fetchone(
                "SELECT SUM(valor) as t FROM pagamentos WHERE participante_id=? AND bolao_id=?",
                (pt_d["id"], bid))
            pago  = pago_row["t"] or 0
            ve    = pt_d["valor_esperado"] or 0
            cotas = round(ve/vt) if vt>0 else 1
            saldo = max(0, ve-pago)
            status = "✅ Quitado" if saldo<=0 else "⚠ Pendente"
            tag    = "quitado" if saldo<=0 else "pendente"
            self._cad_lista_tree.insert("","end", iid=str(pt_d["id"]), tags=(tag,), values=(
                pt_d["id"], pt_d["nome"], pt_d["telefone"] or "-",
                cotas, fmt_brl(ve), status))

    def _cad_remover(self):
        sel = self._cad_lista_tree.selection()
        if not sel: messagebox.showwarning("Atenção","Selecione um participante!"); return
        pid = int(sel[0])
        pt  = self.db.fetchone("SELECT * FROM participantes WHERE id=?", (pid,))

        # Proteção: verificar se tem pagamentos depositados
        dep = self.db.fetchone(
            "SELECT COUNT(*) as n FROM pagamentos WHERE participante_id=? AND depositado=1",(pid,))
        if dep["n"] > 0:
            messagebox.showerror("Operação Bloqueada",
                f"Não é possível remover '{pt['nome']}'.\n\n"
                f"Este participante possui {dep['n']} pagamento(s) já depositado(s).\n\n"
                f"Para remover, primeiro corrija os depósitos na aba Depósitos.")
            return

        if messagebox.askyesno("Confirmar",
            f"Remover '{pt['nome']}' do bolão?\n\nEsta ação não pode ser desfeita."):
            self.db.execute("UPDATE participantes SET ativo=0 WHERE id=?", (pid,))
            messagebox.showinfo("Removido","Participante removido.")
            self._refresh_all()

    # ════════════════════════════════════════════════════════════
    #  ABA 2 — REGISTRAR PAGAMENTOS
    # ════════════════════════════════════════════════════════════
    def _build_pag(self):
        p = self.tab_pag

        # ── Seleção do participante — Combobox simples ───────────
        sec1 = section(p, "SELECIONAR PARTICIPANTE")
        sec1.pack(fill="x", padx=20, pady=(16,6))

        row_cb = tk.Frame(sec1, bg=CORES["bg_section"]); row_cb.pack(fill="x", pady=6)
        tk.Label(row_cb, text="Participante:", bg=CORES["bg_section"],
                 font=("Arial",9,"bold"), fg=CORES["fg_label"]).pack(side="left", padx=(0,8))

        self.pag_cb = ttk.Combobox(row_cb, width=50, state="readonly", font=("Arial",10))
        self.pag_cb.pack(side="left", padx=(0,8))
        self.pag_cb.bind("<<ComboboxSelected>>", self._pag_cb_sel)

        btn(row_cb, "🔄 Atualizar lista", CORES["btn_azul"],
            self._refresh_all, width=18).pack(side="left", padx=4)

        # ── Informações do participante — cards visuais ──────────
        self._sec_pag_info = tk.Frame(p, bg=CORES["bg_frame"])
        self._sec_pag_info.pack(fill="x", padx=20, pady=6)
        self._pag_cards_frame = tk.Frame(self._sec_pag_info, bg=CORES["bg_frame"])
        self._pag_cards_frame.pack(fill="x")

        # ── Formulário de pagamento ──────────────────────────────
        sec3 = section(p, "REGISTRAR PAGAMENTO")
        sec3.pack(fill="x", padx=20, pady=6)

        r1 = tk.Frame(sec3, bg=CORES["bg_section"]); r1.pack(fill="x", pady=4)
        tk.Label(r1, text="Mês Ref.:", bg=CORES["bg_section"],
                 font=("Arial",9,"bold"), fg=CORES["fg_label"]).pack(side="left")
        self.pag_mes = ttk.Combobox(r1, values=MESES, width=8, state="readonly")
        self.pag_mes.set(MESES[date.today().month-1]); self.pag_mes.pack(side="left", padx=4)
        tk.Label(r1, text="/", bg=CORES["bg_section"],
                 font=("Arial",10,"bold")).pack(side="left")
        self.pag_ano = ttk.Combobox(r1, values=[str(y) for y in range(2020,2036)],
                                    width=7, state="readonly")
        self.pag_ano.set(str(date.today().year)); self.pag_ano.pack(side="left", padx=4)

        r2 = tk.Frame(sec3, bg=CORES["bg_section"]); r2.pack(fill="x", pady=4)
        tk.Label(r2, text="Valor (R$):", bg=CORES["bg_section"],
                 font=("Arial",9,"bold"), fg=CORES["fg_label"]).pack(side="left")
        self.pag_val = entry(r2, width=14); self.pag_val.pack(side="left", padx=8)
        self.pag_val.bind("<FocusIn>", lambda e: self.pag_val.selection_range(0,"end"))
        tk.Label(r2, text="Data:", bg=CORES["bg_section"],
                 font=("Arial",9,"bold"), fg=CORES["fg_label"]).pack(side="left")
        self.pag_dt = entry(r2, width=14)
        self.pag_dt.insert(0, date.today().strftime("%d/%m/%Y"))
        self.pag_dt.pack(side="left", padx=8)

        r3 = tk.Frame(sec3, bg=CORES["bg_section"]); r3.pack(fill="x", pady=4)
        tk.Label(r3, text="Obs.:", bg=CORES["bg_section"],
                 font=("Arial",9,"bold"), fg=CORES["fg_label"]).pack(side="left")
        self.pag_obs = entry(r3, width=50); self.pag_obs.pack(side="left", padx=8)

        bf = tk.Frame(sec3, bg=CORES["bg_section"]); bf.pack(fill="x", pady=8)
        btn(bf, "💳  REGISTRAR PAGAMENTO", CORES["btn_azul"],
            self._registrar_pag, width=26).pack(side="left")

    def _pag_cb_sel(self, e=None):
        """Sincroniza seleção entre abas de Pagamentos e chama _pag_info."""
        sel = self.pag_cb.get()
        self._pag_part_sync.set(sel)
        # Sincroniza vis_cb
        try:
            if sel in self.vis_cb["values"]:
                self.vis_cb.set(sel)
                self._vis_sel()
        except: pass
        self._pag_info()

    def _pag_info(self, e=None):
        sel = self.pag_cb.get()
        # Limpa cards anteriores
        for w in self._pag_cards_frame.winfo_children():
            w.destroy()
        if not sel: return
        m = re.search(r"\(ID: (\d+)\)", sel)
        if not m: return
        pid = int(m.group(1))
        bid = self.bid.get()
        pt  = self.db.fetchone("SELECT * FROM participantes WHERE id=?",(pid,))
        if not pt: return
        pgs   = self.db.fetchall(
            "SELECT * FROM pagamentos WHERE participante_id=? AND bolao_id=?",(pid,bid))
        pago  = sum(x["valor"] for x in pgs)
        ve    = pt["valor_esperado"] or 0
        saldo = max(0, ve - pago)
        quitado = saldo <= 0

        # ── Linha 1: info do participante ─────────────────────────
        r1 = tk.Frame(self._pag_cards_frame, bg=CORES["bg_frame"])
        r1.pack(fill="x", pady=(0,4))

        # Card nome/contato
        c_nome = tk.Frame(r1, bg="#1e3348", padx=14, pady=10, relief="flat")
        c_nome.pack(side="left", fill="both", expand=True, padx=(0,4))
        tk.Label(c_nome, text="👤  "+pt["nome"], bg="#1e3348", fg="white",
                 font=("Arial",10,"bold")).pack(anchor="w")
        tel = pt["telefone"] or "—"; pix = pt["chave_pix"] or "—"
        tk.Label(c_nome, text=f"📱 {tel}   |   PIX: {pix}", bg="#1e3348",
                 fg="#aad4f5", font=("Arial",8)).pack(anchor="w", pady=(1,0))

        # Card status — mais compacto
        cor_st = "#1D9E75" if quitado else "#e67e22"
        st_txt = "✅ QUITADO" if quitado else "⚠ PENDENTE"
        c_st = tk.Frame(r1, bg=cor_st, padx=10, pady=6, width=110)
        c_st.pack(side="right", fill="y", padx=(4,0))
        c_st.pack_propagate(False)
        tk.Label(c_st, text=st_txt, bg=cor_st, fg="white",
                 font=("Arial",9,"bold")).pack(expand=True)

        # ── Linha 2: cards financeiros compactos ─────────────────
        r2 = tk.Frame(self._pag_cards_frame, bg=CORES["bg_frame"])
        r2.pack(fill="x")
        for titulo, valor, cor in [
            ("💰 Valor Esperado", fmt_brl(ve),      "#8e44ad"),
            ("✅ Total Pago",     fmt_brl(pago),     "#1D9E75"),
            ("📋 Saldo Restante", fmt_brl(saldo),    "#e74c3c" if saldo>0 else "#1D9E75"),
            ("🔢 Pagamentos",     str(len(pgs)),      "#2196F3"),
        ]:
            c = tk.Frame(r2, bg=cor, padx=10, pady=6)
            c.pack(side="left", fill="both", expand=True, padx=2)
            tk.Label(c, text=titulo, bg=cor, fg="white",
                     font=("Arial",7,"bold")).pack(anchor="w")
            tk.Label(c, text=valor, bg=cor, fg="white",
                     font=("Arial",11,"bold")).pack(anchor="w", pady=(1,0))

        self.pag_val.delete(0,"end")
        self.pag_val.focus_set()

    def _registrar_pag(self):
        sel = self.pag_cb.get()
        if not sel: messagebox.showwarning("Atenção","Selecione um participante!"); return
        pid = int(re.search(r"\(ID: (\d+)\)",sel).group(1))
        bid = self.bid.get()
        v   = to_float(self.pag_val.get())
        if v<=0: messagebox.showwarning("Atenção","Informe um valor válido!"); return

        # Avisa antes de registrar pagamento pro ADM isento nesse bolão —
        # é o tipo de lançamento que fica "escondido" nos totais (o ADM
        # isento não entra em Total Esperado/Arrecadado, mas um pagamento
        # real registrado pra ele ENTRA na soma de "pagamentos", inflando
        # a aba Depósitos mesmo assim). Avisa, não bloqueia — pode ser
        # intencional (contribuição voluntária).
        pt_check = self.db.fetchone("SELECT * FROM participantes WHERE id=?", (pid,))
        b_check  = self.db.fetchone("SELECT adm_paga, adm_nome FROM boloes WHERE id=?", (bid,))
        if pt_check and b_check and not (b_check["adm_paga"] or 0):
            adm_nome_low = (b_check["adm_nome"] or "").strip().lower()
            eh_adm_check = bool(pt_check["is_adm"]) or (
                adm_nome_low and adm_nome_low in (pt_check["nome"] or "").lower())
            if eh_adm_check:
                if not messagebox.askyesno("Participante isento",
                    f"'{pt_check['nome']}' está configurado como ADM ISENTO neste "
                    f"bolão (não paga). Registrar um pagamento mesmo assim vai "
                    f"aparecer no 'Total Recebido' da aba Depósitos.\n\n"
                    f"Tem certeza que quer registrar?"):
                    return

        self.db.execute(
            "INSERT INTO pagamentos (participante_id,bolao_id,mes_referencia,valor,data_pagamento,observacoes)"
            " VALUES (?,?,?,?,?,?)",
            (pid,bid,f"{self.pag_mes.get()}/{self.pag_ano.get()}",v,
             self.pag_dt.get().strip(),self.pag_obs.get().strip()))
        messagebox.showinfo("Sucesso",f"Pagamento de {fmt_brl(v)} registrado!")
        self._pag_info()
        self.pag_obs.delete(0,"end")
        # Sem isso, Dashboard/Relatorio/Cards ficavam com os valores de
        # antes do pagamento ate o usuario trocar de bolao manualmente
        # (o auto-refresh ao trocar de aba nao cobre ficar na mesma aba).
        try: self._dash_load()
        except Exception: pass

    # ════════════════════════════════════════════════════════════
    #  ABA 3 — RELATÓRIO
    # ════════════════════════════════════════════════════════════
    def _build_rel(self):
        p = self.tab_rel
        bf = tk.Frame(p,bg=CORES["bg_frame"]); bf.pack(fill="x",padx=20,pady=10)
        btn(bf,"🔄 Atualizar",CORES["btn_azul"],self._gerar_rel,width=18).pack(side="left",padx=4)
        btn(bf,"📊 Exportar Excel",CORES["btn_verde"],self._exportar_excel,width=20).pack(side="left",padx=4)
        btn(bf,"🖼 Cards Visuais",CORES["btn_roxo"],self._cards_visuais,width=20).pack(side="left",padx=4)
        btn(bf,"📄 Relatório Completo",CORES["btn_teal"],self._rel_completo_janela,width=22).pack(side="left",padx=4)
        btn(bf,"📲 WhatsApp",CORES["btn_laranja"],self._rel_whatsapp,width=14).pack(side="left",padx=4)

        self._rel_resumo = section(p,"RESUMO DO BOLÃO")
        self._rel_resumo.pack(fill="x",padx=20,pady=(0,4))
        self._rel_resumo_lbl = tk.Label(self._rel_resumo,text="",bg=CORES["bg_section"],
                                         fg=CORES["fg_label"],font=("Arial",9),justify="left")
        self._rel_resumo_lbl.pack(anchor="w")
        # Legenda fixa para evitar confusão
        tk.Label(self._rel_resumo,
                 text="ℹ  Esperado = soma dos valores esperados de cada participante cadastrado  |  "
                      "Arrecadado = total já recebido em pagamentos  |  "
                      "Pendente = diferença ainda a receber",
                 bg=CORES["bg_section"], fg="#888", font=("Arial",8,"italic"),
                 justify="left").pack(anchor="w", pady=(0,4))

        st = section(p,"PARTICIPANTES")
        st.pack(fill="both",expand=True,padx=20,pady=(0,10))
        cols = {"ID":60,"Nome":220,"Telefone":130,"Valor Esp.":120,
                "Total Pago":120,"Saldo":110,"Status":120,"Parc.":60}
        fr,self.rel_tree = make_tree(st,cols,height=18)
        fr.pack(fill="both",expand=True)
        self.rel_tree.tag_configure("quitado",  background="#d5f5e3")
        self.rel_tree.tag_configure("pendente", background="#fde8d8")
        self.rel_tree.tag_configure("emdia",    background="#d6eaf8")

    def _get_bolao_info(self):
        bid = self.bid.get()
        if not bid: return None,[]
        b = self.db.fetchone("SELECT * FROM boloes WHERE id=?",(bid,))
        p = self.db.fetchall(
            "SELECT * FROM participantes WHERE bolao_id=? AND ativo=1 ORDER BY nome",(bid,))
        return dict(b) if b else None, p

    def _calc_parcela_atual(self, bolao_d):
        """Calcula parcelas esperadas até o fim do mês ANTERIOR (mês atual está em aberto)."""
        try:
            di = datetime.strptime(bolao_d.get("data_inicio","2025-01-01"),"%Y-%m-%d")
        except:
            try:
                di = datetime.strptime(bolao_d.get("data_inicio","01/01/2025"),"%d/%m/%Y")
            except:
                di = datetime.now()
        hoje  = datetime.now()
        # Meses encerrados = do início até o mês ANTERIOR ao atual
        # O mês atual ainda está em aberto → não conta como vencido
        meses_total   = (hoje.year - di.year)*12 + (hoje.month - di.month) + 1
        meses_vencidos = max(0, meses_total - 1)   # exclui o mês corrente
        parc  = bolao_d.get("valor_parcela",0) or 0
        total = bolao_d.get("valor_total",0)  or 0
        n_parcelas_total = round(total/parc) if parc>0 else 0
        parc_esperada = min(meses_vencidos, n_parcelas_total)
        return meses_total, parc_esperada, parc

    def _status_part(self, pago, val_esp, parc_esp, parc):
        saldo = val_esp - pago
        if saldo <= 0:
            return "✅ QUITADO","quitado"
        n_pago = round(pago/parc,1) if parc>0 else 0
        if n_pago >= parc_esp:
            return "🟦 EM DIA","emdia"
        return "⚠ PENDENTE","pendente"

    def _gerar_rel(self):
        bid = self.bid.get()
        if not bid: return
        self.rel_tree.delete(*self.rel_tree.get_children())
        b, partic = self._get_bolao_info()
        if not b: return
        _, parc_esp, parc = self._calc_parcela_atual(b)
        te = tp = ts = 0

        # ADM invisível no relatório
        adm_paga = b.get("adm_paga",0)
        adm_nome = b.get("adm_nome","").strip().lower()

        for pt in partic:
            pt_d = dict(pt)
            # Se ADM não paga, mostra como quitado sem destacar. Checa a
            # flag OU o nome batendo com adm_nome — mesmo critério usado
            # no Dashboard/Cards/publicação pro Firebase; usar só a flag
            # aqui fazia o Relatório cobrar do ADM quando ele foi
            # cadastrado sem marcar "is_adm", divergindo dos outros lugares.
            eh_adm_r = bool(pt_d.get("is_adm")) or (adm_nome and adm_nome in pt_d["nome"].lower())
            if eh_adm_r and not adm_paga:
                self.rel_tree.insert("","end",tags=("quitado",),values=(
                    pt_d["id"],pt_d["nome"],pt_d["telefone"] or "-",
                    fmt_brl(0), fmt_brl(0),
                    fmt_brl(0),"✅ QUITADO (isento)","-"))
                # Não soma em te/tp: ADM isento não deve nunca dinheiro nem
                # entra no "arrecadado" — somar aqui inflava os dois totais
                # igualmente (o "esperado" ficava maior que o que precisa
                # ser cobrado de verdade, e o "arrecadado" incluía dinheiro
                # que nunca entrou de fato).
                continue
            pgs  = self.db.fetchall(
                "SELECT * FROM pagamentos WHERE participante_id=? AND bolao_id=?",
                (pt_d["id"],bid))
            pago = sum(x["valor"] for x in pgs)
            saldo= (pt_d["valor_esperado"] or 0) - pago
            te  += pt_d["valor_esperado"] or 0
            tp  += pago
            ts  += max(0,saldo)
            status,tag = self._status_part(pago,pt_d["valor_esperado"] or 0,parc_esp,parc)
            n_pago = round(pago/parc,1) if parc>0 else 0
            n_tot  = round((pt_d["valor_esperado"] or 0)/parc,0) if parc>0 else 0
            self.rel_tree.insert("","end",tags=(tag,),values=(
                pt_d["id"],pt_d["nome"],pt_d["telefone"] or "-",
                fmt_brl(pt_d["valor_esperado"]),fmt_brl(pago),
                fmt_brl(max(0,saldo)),status,f"{n_pago}/{int(n_tot)}"))

        # Contagem correta:
        # - n_pessoas = todos cadastrados + ADM se não estiver cadastrado
        # - n_pagantes = exclui ADM isento (não paga)
        adm_nome_lower_r = b.get("adm_nome","").strip().lower()
        adm_cadastrado   = any(
            adm_nome_lower_r and adm_nome_lower_r in dict(pt)["nome"].lower()
            for pt in partic
        ) if adm_nome_lower_r else False
        adm_no_bolao     = bool(b.get("adm_nome","").strip())

        n_cadastrados = len(partic)
        n_pessoas     = n_cadastrados + (1 if adm_no_bolao and not adm_cadastrado else 0)
        n_pagantes    = n_pessoas - (1 if not adm_paga and adm_no_bolao else 0)

        cotas_ocup_r, max_cotas_r = self._get_cotas_ocupadas(bid)

        self._rel_resumo_lbl.configure(text=(
            f"Bolão: {b['nome']}  |  Loteria: {b.get('loteria','Mega-Sena')}  |  "
            f"Total no bolão: {n_pessoas}  |  Pagantes: {n_pagantes}  |  "
            f"Cotas: {cotas_ocup_r}/{max_cotas_r}  |  "
            f"💰 Esperado: {fmt_brl(te)}  |  "
            f"✅ Arrecadado: {fmt_brl(tp)}  |  "
            f"⚠ Pendente: {fmt_brl(ts)}"
        ))

    # ─── Relatório completo estilo texto ────────────────────────
    def _rel_whatsapp(self):
        """Gera relatório resumido formatado para WhatsApp."""
        b, partic = self._get_bolao_info()
        if not b: messagebox.showwarning("Atenção","Selecione um bolão!"); return

        bid       = self.bid.get()
        adm_paga  = b.get("adm_paga", 0)
        adm_nome  = (b.get("adm_nome","") or "").strip()
        adm_low   = adm_nome.lower()
        vt        = float(b.get("valor_total",0) or 0)
        pix_chave = "61998507770"  # fixo conforme solicitado

        # Cotas
        cotas_ocup, max_cotas = self._get_cotas_ocupadas(bid)
        vagas_livres = max(0, max_cotas - cotas_ocup)
        fechado = vagas_livres == 0

        # Monta lista de confirmados (quitados + em dia)
        _,parc_esp,parc = self._calc_parcela_atual(b)
        confirmados = []
        for pt in partic:
            pt_d = dict(pt)
            pago_row = self.db.fetchone(
                "SELECT SUM(valor) as t FROM pagamentos WHERE participante_id=? AND bolao_id=?",
                (pt_d["id"], bid))
            pago  = pago_row["t"] or 0
            ve    = pt_d["valor_esperado"] or 0
            eh_adm = bool(pt_d.get("is_adm")) or (adm_low and adm_low in pt_d["nome"].lower())
            if eh_adm and not adm_paga:
                confirmados.append(pt_d["nome"])
            else:
                saldo = max(0, ve - pago)
                if saldo <= 0 or pago >= parc:
                    confirmados.append(pt_d["nome"])

        # ADM não cadastrado mas confirmado
        if adm_nome and not adm_paga:
            adm_ja = any(adm_low in n.lower() for n in confirmados)
            if not adm_ja:
                confirmados.append(adm_nome)

        confirmados.sort()

        # Monta mensagem
        linhas = []
        if fechado:
            linhas.append("🔒 *BOLÃO FECHADO*")
            linhas.append("")
        linhas += [
            f"📋 *{b['nome'].upper()}*",
            f"💰 Valor da Cota: {fmt_brl(vt)}",
            f"📱 PIX: {pix_chave}",
            "",
            f"✅ Confirmados ({len(confirmados)}):",
            "",
        ]
        for i, nome in enumerate(confirmados, 1):
            linhas.append(f"{i:2d}. {nome}")
        linhas += [
            "",
            "Ao realizar o PIX, sua vaga está garantida.",
            "Não é necessário enviar comprovante —",
            "confirmo pelo extrato bancário. 🎯",
            "",
        ]
        if fechado:
            linhas.append("🔒 *BOLÃO FECHADO — VAGAS ESGOTADAS*")
        else:
            linhas.append(f"🎫 Vagas disponíveis: {vagas_livres} de {max_cotas}")
        linhas += [
            "",
            f"_Gestão: {adm_nome or 'Elton Luis'}_",
            f"📅 {datetime.now().strftime('%d/%m/%Y às %H:%M')}",
        ]
        mensagem = "\n".join(linhas)

        # Janela de exibição e cópia
        win = tk.Toplevel(self.root); win.title("Relatório WhatsApp")
        win.geometry("520px" if False else "520x620")
        win.configure(bg=CORES["bg_section"]); win.grab_set(); win.lift()
        tk.Label(win, text="📲 RELATÓRIO PARA WHATSAPP",
                 bg=CORES["bg_section"], fg=CORES["fg_title"],
                 font=("Arial",11,"bold")).pack(pady=10)
        txt = tk.Text(win, font=("Arial",10), wrap="word", padx=10, pady=8)
        txt.pack(fill="both", expand=True, padx=16, pady=(0,8))
        txt.insert("1.0", mensagem)

        def copiar():
            win.clipboard_clear(); win.clipboard_append(mensagem)
            messagebox.showinfo("Copiado!","Mensagem copiada! Cole no WhatsApp.")

        bf2 = tk.Frame(win, bg=CORES["bg_section"]); bf2.pack(pady=8)
        btn(bf2,"📋 COPIAR",CORES["btn_verde"],copiar,width=16).pack(side="left",padx=4)
        btn(bf2,"✖ Fechar",CORES["btn_cinza"],win.destroy,width=10).pack(side="left",padx=4)

    def _rel_completo_janela(self):
        bid = self.bid.get()
        if not bid: messagebox.showwarning("Atenção","Selecione um bolão!"); return
        b, partic = self._get_bolao_info()
        if not b: return
        txt = self._montar_rel_completo(b, partic)
        win = tk.Toplevel(self.root)
        win.title(f"Relatório Completo — {b['nome']}")
        win.geometry("900x700")
        win.configure(bg=CORES["bg_section"])
        tk.Label(win,text=f"RELATÓRIO COMPLETO — {b['nome'].upper()}",
                 bg=CORES["bg_section"],fg=CORES["fg_title"],
                 font=("Arial",12,"bold")).pack(pady=8)
        frm = tk.Frame(win,bg=CORES["bg_section"]); frm.pack(fill="both",expand=True,padx=10,pady=4)
        sb  = tk.Scrollbar(frm)
        sb.pack(side="right",fill="y")
        tx  = tk.Text(frm,yscrollcommand=sb.set,font=("Courier",9),
                      bg="#1e2a35",fg="#e0e0e0",relief="flat")
        tx.pack(fill="both",expand=True)
        sb.configure(command=tx.yview)
        tx.insert("1.0",txt)
        tx.configure(state="disabled")
        bf = tk.Frame(win,bg=CORES["bg_section"]); bf.pack(pady=8)
        def salvar():
            path = filedialog.asksaveasfilename(defaultextension=".txt",
                filetypes=[("Texto","*.txt")],initialfile=f"relatorio_{bid}.txt")
            if path:
                with open(path,"w",encoding="utf-8") as f: f.write(txt)
                messagebox.showinfo("Salvo",f"Arquivo salvo em:\n{path}")
        btn(bf,"💾 Salvar TXT",CORES["btn_verde"],salvar,width=18).pack(side="left",padx=6)
        btn(bf,"❌ Fechar",CORES["btn_cinza"],win.destroy,width=12).pack(side="left")

    def _montar_rel_completo(self, b, partic):
        bid = self.bid.get()
        sep = "="*70
        now = datetime.now().strftime("%d/%m/%Y %H:%M")
        mes_ref = MESES[date.today().month-1]+"/"+str(date.today().year)

        adm_paga      = b.get("adm_paga",0)
        adm_nome_str  = b.get("adm_nome","").strip()
        adm_nome_low  = adm_nome_str.lower()
        adm_cadastrado = any(
            adm_nome_low and adm_nome_low in dict(pt)["nome"].lower()
            for pt in partic
        ) if adm_nome_low else False

        n_cadastrados = len(partic)
        n_pessoas     = n_cadastrados + (1 if adm_nome_str and not adm_cadastrado else 0)
        n_pagantes    = n_pessoas - (1 if not adm_paga and adm_nome_str else 0)
        cotas_ocup_r, n_total = self._get_cotas_ocupadas(bid)

        _,parc_esp,parc = self._calc_parcela_atual(b)

        # Totais depósitos
        tot_rec = self.db.fetchone("SELECT SUM(valor) as t FROM pagamentos WHERE bolao_id=?",(bid,))
        tot_dep = self.db.fetchone("SELECT SUM(valor) as t FROM pagamentos WHERE bolao_id=? AND depositado=1",(bid,))
        tr = tot_rec["t"] or 0
        td = tot_dep["t"] or 0
        pend_dep = tr - td
        n_pend   = self.db.fetchone("SELECT COUNT(*) as n FROM pagamentos WHERE bolao_id=? AND depositado=0",(bid,))["n"]

        ult_deps = self.db.fetchall("""
            SELECT pg.data_deposito, pt.nome, pg.valor
            FROM pagamentos pg JOIN participantes pt ON pg.participante_id=pt.id
            WHERE pg.bolao_id=? AND pg.depositado=1
            ORDER BY pg.data_deposito DESC LIMIT 10
        """,(bid,))

        # Estatísticas pagamentos
        all_pgs = self.db.fetchall("SELECT valor FROM pagamentos WHERE bolao_id=?",(bid,))
        vals    = [x["valor"] for x in all_pgs]
        tot_pags= len(vals)
        media   = sum(vals)/tot_pags if vals else 0
        menor   = min(vals) if vals else 0
        maior   = max(vals) if vals else 0

        # Dados participantes
        dados_pt = []
        quitados = em_dia = pendentes_cnt = 0
        te = tp = ts = 0
        for pt in partic:
            pt_d = dict(pt)
            pgs  = self.db.fetchall(
                "SELECT * FROM pagamentos WHERE participante_id=? AND bolao_id=?",(pt_d["id"],bid))
            pago = sum(x["valor"] for x in pgs)
            ve   = pt_d["valor_esperado"] or 0
            saldo= ve - pago
            te  += ve; tp += pago; ts += max(0,saldo)
            if pt_d.get("is_adm") and not adm_paga:
                status="QUITADO"; quitados+=1; pago_show=ve; saldo_show=0
                n_p=round(ve/parc,1) if parc>0 else 0
            else:
                n_p   = round(pago/parc,1) if parc>0 else 0
                n_tot = round(ve/parc,0)   if parc>0 else 0
                st,_  = self._status_part(pago,ve,parc_esp,parc)
                status= st.replace("✅ ","").replace("⚠ ","").replace("🟦 ","")
                if status=="QUITADO": quitados+=1
                elif status=="EM DIA": em_dia+=1
                else: pendentes_cnt+=1
                pago_show=pago; saldo_show=max(0,saldo); n_tot=round(ve/parc,0) if parc>0 else 0
            dados_pt.append((pt_d["nome"],n_p,n_tot,pago_show,ve,saldo_show,status))

        linhas = []
        def A(s=""): linhas.append(s)
        A(sep)
        A(f"{'BOLÃO '+b['nome'].upper():^70}")
        A(sep)
        A(f"Relatório Administrativo - {mes_ref}")
        A(f"Gerado em: {now}")
        A(f"Loteria: {b.get('loteria','Mega-Sena')}   |   Descrição: {b.get('descricao') or '-'}")
        A(sep)
        A("INFORMAÇÕES DO BOLÃO:")
        A(f"• Cotas máximas do bolão: {n_total}")
        A(f"• Cotas ocupadas:         {cotas_ocup_r}")
        A(f"• Total de pessoas:       {n_pessoas}  (pagantes: {n_pagantes})")
        A(f"• Valor total por cota:   {fmt_brl(b.get('valor_total',0))}")
        A(f"• Valor de cada parcela:  {fmt_brl(parc)}")
        A(sep)
        A("CONTROLE DE DEPÓSITOS NO FUNDO:")
        A(sep)
        A(f"💰 Total Recebido dos Participantes:  {fmt_brl(tr):>14}")
        A(f"✅ Total Já Depositado no Fundo:       {fmt_brl(td):>14}")
        A(f"⚠  Pendente de Depósito:               {fmt_brl(pend_dep):>14}  ({n_pend} pagamentos)")
        A(f"📊 Progresso de Depósitos: {(td/tr*100 if tr>0 else 0):.1f}%")
        if ult_deps:
            A("ÚLTIMOS DEPÓSITOS:")
            for d in ult_deps:
                A(f"  • {d['data_deposito'] or '-'} - {d['nome']}: {fmt_brl(d['valor'])}")
        A(sep)
        A("TOTAIS FINANCEIROS:")
        A(sep)
        A(f"• TOTAL ARRECADADO: {fmt_brl(tp):>14}")
        A(f"• TOTAL ESPERADO:   {fmt_brl(te):>14}")
        A(f"• PERCENTUAL:       {(tp/te*100 if te>0 else 0):>12.1f}%")
        A("ESTATÍSTICAS DE VALORES:")
        A(f"• Valor médio por pagamento: {fmt_brl(media)}")
        A(f"• Menor pagamento:           {fmt_brl(menor)}")
        A(f"• Maior pagamento:           {fmt_brl(maior)}")
        A(f"• Total de pagamentos:       {tot_pags}")
        A("ESTATÍSTICAS DE PAGAMENTO:")
        A(f"• Participantes quitados:    {quitados}")
        A(f"• Participantes em dia:      {em_dia}")
        A(f"• Participantes pendentes:   {pendentes_cnt}")
        A(sep)
        A("DETALHAMENTO COMPLETO DOS PARTICIPANTES:")
        A(sep)
        hdr = f"{'NOME':<26}{'PROGRESSO':<12}{'TOTAL PAGO':<14}{'VALOR ESP.':<14}{'SALDO':<14}{'STATUS'}"
        A(hdr)
        A("-"*70)
        for nome,np,nt,pago,ve,saldo,status in sorted(dados_pt,key=lambda x:x[0]):
            prog = f"{np}/{int(nt)}"
            A(f"{nome:<26}{prog:<12}{fmt_brl(pago):<14}{fmt_brl(ve):<14}{fmt_brl(saldo):<14}{status}")
        A(sep)
        A("OBSERVAÇÕES ADMINISTRATIVAS:")
        A(f"• {n_pend} pagamento(s) aguardando depósito no valor total de {fmt_brl(pend_dep)}")
        A(f"• Valor da parcela deste bolão: {fmt_brl(parc)}")
        A("RECOMENDAÇÕES:")
        A("1. Deposite regularmente os valores recebidos")
        A("2. Mantenha o controle atualizado na aba 'Depósitos'")
        A("3. Faça backup regular do sistema")
        return "\n".join(linhas)

    # ─── Exportar Excel ──────────────────────────────────────────
    def _exportar_excel(self):
        bid = self.bid.get()
        if not bid: messagebox.showwarning("Atenção","Selecione um bolão!"); return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror("Erro","Execute: pip install openpyxl"); return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                filetypes=[("Excel","*.xlsx")],initialfile=f"relatorio_{bid}.xlsx")
        if not path: return
        b, partic = self._get_bolao_info()
        if not b: return
        _,parc_esp,parc = self._calc_parcela_atual(b)
        adm_paga = b.get("adm_paga",0)

        wb = Workbook(); ws = wb.active; ws.title = "Relatório"
        thin = "thin"
        brd = Border(left=Side(style=thin),right=Side(style=thin),
                     top=Side(style=thin),bottom=Side(style=thin))
        ctr = Alignment(horizontal="center",vertical="center")
        esq = Alignment(horizontal="left",vertical="center")

        def hdr_cell(r,c,val,fg="FFFFFF",bg="1A2A3A",bold=True,sz=10):
            cell = ws.cell(r,c,val)
            cell.font=Font(bold=bold,size=sz,color=fg)
            cell.fill=PatternFill("solid",fgColor=bg)
            cell.alignment=ctr; cell.border=brd; return cell

        ws.merge_cells("A1:H1")
        hdr_cell(1,1,f"RELATÓRIO — {b['nome'].upper()}",sz=13)
        ws.row_dimensions[1].height=28
        ws.merge_cells("A2:H2")
        hdr_cell(2,1,f"Loteria: {b.get('loteria','')}  |  Gerado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                 bg="2C3E50",sz=9,bold=False)
        ws.row_dimensions[2].height=16
        ws.append([])
        hdrs=["#","Nome","Telefone","Valor Esperado","Total Pago","Saldo","Status","Parcelas"]
        ws.append(hdrs)
        for c,h in enumerate(hdrs,1):
            hdr_cell(4,c,h,bg="2980B9")
        ws.row_dimensions[4].height=20

        te=tp=ts=0
        for i,pt in enumerate(partic,1):
            pt_d=dict(pt)
            pgs=self.db.fetchall("SELECT * FROM pagamentos WHERE participante_id=? AND bolao_id=?",
                                 (pt_d["id"],bid))
            pago=sum(x["valor"] for x in pgs)
            ve=pt_d["valor_esperado"] or 0
            if pt_d.get("is_adm") and not adm_paga:
                saldo=0;pago=ve;status="✅ Quitado";bg="D5F5E3"
            else:
                saldo=max(0,ve-pago)
                st,_=self._status_part(pago,ve,parc_esp,parc)
                status=st;bg=("D5F5E3" if "QUITADO" in st else ("D6EAF8" if "DIA" in st else "FDE8D8"))
            te+=ve;tp+=pago;ts+=saldo
            n_p=round(pago/parc,1) if parc>0 else 0
            n_t=round(ve/parc,0) if parc>0 else 0
            ws.append([i,pt_d["nome"],pt_d["telefone"] or "-",ve,pago,saldo,status,f"{n_p}/{int(n_t)}"])
            r=ws.max_row
            for c in range(1,9):
                cell=ws.cell(r,c)
                cell.fill=PatternFill("solid",fgColor=bg)
                cell.border=brd
                cell.alignment=esq if c==2 else ctr
                if c in(4,5,6): cell.number_format='R$ #,##0.00'
            ws.row_dimensions[r].height=20

        tr=ws.max_row+1
        for c,v in enumerate(["-","TOTAL","-",te,tp,ts,f"✅{sum(1 for _,__,___,p,ve,s,st,_ in [] for _ in []) }","-"],1):
            pass
        ws.cell(tr,2,"TOTAIS").font=Font(bold=True)
        ws.cell(tr,4,te);ws.cell(tr,5,tp);ws.cell(tr,6,ts)
        for c in range(1,9):
            cell=ws.cell(tr,c)
            cell.fill=PatternFill("solid",fgColor="D6EAF8")
            cell.font=Font(bold=True); cell.border=brd; cell.alignment=ctr
            if c in(4,5,6): cell.number_format='R$ #,##0.00'
        ws.row_dimensions[tr].height=22

        for i,w in enumerate([5,35,18,18,16,16,18,12],1):
            ws.column_dimensions[get_column_letter(i)].width=w
        wb.save(path)
        messagebox.showinfo("Excel Gerado!",f"Salvo em:\n{path}")

    # ─── Cards visuais ───────────────────────────────────────────
    def _cards_visuais(self):
        bid = self.bid.get()
        if not bid: messagebox.showwarning("Atenção","Selecione um bolão!"); return
        b, partic = self._get_bolao_info()
        if not b: return
        _,parc_esp,parc = self._calc_parcela_atual(b)
        adm_paga = b.get("adm_paga",0)

        # Janela maximizada para aproveitar toda a tela
        win = tk.Toplevel(self.root)
        win.title(f"Painel de Situação — {b['nome']}")
        win.state("zoomed")          # maximiza no Windows
        win.configure(bg="#1a2a3a")

        # ── Cabeçalho ──────────────────────────────────────────
        hdr = tk.Frame(win, bg="#1a2a3a", pady=8)
        hdr.pack(fill="x", padx=16)
        tk.Label(hdr, text=f"🎰  {b['nome'].upper()}  —  {b.get('loteria','Mega-Sena')}",
                 bg="#1a2a3a", fg="white",
                 font=("Arial",14,"bold")).pack(side="left")
        tk.Label(hdr, text=datetime.now().strftime("%d/%m/%Y %H:%M"),
                 bg="#1a2a3a", fg="#aad4f5",
                 font=("Arial",10)).pack(side="right")

        # ── Coleta dados ────────────────────────────────────────
        dados = []
        q = p_ = em = 0
        adm_nome_lower = b.get("adm_nome","").strip().lower()
        adm_nome_real  = b.get("adm_nome","").strip()
        adm_ja_listado = False  # controle para não duplicar

        for pt in partic:
            pt_d = dict(pt)
            pgs  = self.db.fetchall(
                "SELECT * FROM pagamentos WHERE participante_id=? AND bolao_id=?",
                (pt_d["id"], bid))
            pago = sum(x["valor"] for x in pgs)
            ve   = pt_d["valor_esperado"] or 0

            # Detecta ADM por flag OU por nome configurado
            eh_adm = bool(pt_d.get("is_adm")) or (
                adm_nome_lower and adm_nome_lower in pt_d["nome"].lower())

            if eh_adm:
                adm_ja_listado = True

            status, tag, pago_f, saldo_f = self._status_part_adm(
                {"is_adm": eh_adm}, pago, ve, parc_esp, parc, adm_paga)

            if   "QUITADO" in status: q  += 1
            elif "EM DIA"  in status: em += 1
            else:                     p_ += 1
            dados.append((pt_d["nome"], ve, pago_f, saldo_f, status))

        # Se ADM tem nome configurado mas não está cadastrado como participante
        # → adiciona card sintético para ele aparecer sempre
        if adm_nome_real and not adm_ja_listado:
            if not adm_paga:
                # ADM isento: card verde quitado sem valor
                dados.append((adm_nome_real, 0, None, 0, "QUITADO"))
                q += 1
            # ADM que paga mas não está cadastrado → não adiciona (situação incomum)

        total = len(dados)  # total de pessoas cadastradas
        cotas_ocup_c, max_cotas_c = self._get_cotas_ocupadas(bid)

        # ── Barra de resumo colorida ────────────────────────────
        bar = tk.Frame(win, bg="#243447", pady=6)
        bar.pack(fill="x", padx=16, pady=(0,6))
        resumo_txt = (
            f"  ✅ Quitados: {q}   |   🟦 Em dia: {em}   |   "
            f"⚠ Pendentes: {p_}   |   👥 Pessoas: {total}   |   "
            f"🎫 Cotas: {cotas_ocup_c}/{max_cotas_c}   |   "
            f"📊 {(q/total*100 if total else 0):.0f}% quitado"
        )
        tk.Label(bar, text=resumo_txt, bg="#243447", fg="white",
                 font=("Arial",10,"bold")).pack(side="left")

        # Legenda de cores
        leg = tk.Frame(bar, bg="#243447"); leg.pack(side="right", padx=10)
        for cor, txt in [("#1e7e44","Quitado"),("#1a5276","Em dia"),("#a93226","Pendente")]:
            tk.Frame(leg, bg=cor, width=14, height=14).pack(side="left", padx=(6,2))
            tk.Label(leg, text=txt, bg="#243447", fg="white",
                     font=("Arial",8)).pack(side="left", padx=(0,8))

        # ── Canvas com scroll ───────────────────────────────────
        canvas = tk.Canvas(win, bg="#1a2a3a", highlightthickness=0)
        sb = ttk.Scrollbar(win, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True, padx=16, pady=(0,10))

        inner = tk.Frame(canvas, bg="#1a2a3a")
        canvas.create_window((0,0), window=inner, anchor="nw")

        # ── 10 cards por linha ──────────────────────────────────
        NCOLS   = 10
        C_W     = 118   # largura do card
        C_H     = 76    # altura do card
        PAD_X   = 4
        PAD_Y   = 4

        dados_ord = sorted(dados, key=lambda x: x[0])  # ordem alfabética

        for idx, (nome, ve, pago, saldo, status) in enumerate(dados_ord):
            row_ = idx // NCOLS
            col_ = idx  % NCOLS

            # Cor por status
            if   "QUITADO" in status: bg_c="#1e7e44"; fg_stat="#afffca"; ic="✅"
            elif "EM DIA"  in status: bg_c="#1a5276"; fg_stat="#aed6f1"; ic="🟦"
            else:                     bg_c="#8b1a1a"; fg_stat="#ffe082"; ic="⚠"

            card = tk.Frame(inner, bg=bg_c, width=C_W, height=C_H,
                            relief="flat", bd=0)
            card.grid(row=row_, column=col_,
                      padx=PAD_X, pady=PAD_Y, sticky="nsew")
            card.grid_propagate(False)

            # Nome — trunca se necessário, fonte menor para caber
            n_max = 15
            nn = nome if len(nome) <= n_max else nome[:n_max-1]+"…"
            tk.Label(card, text=nn, bg=bg_c, fg="white",
                     font=("Arial",8,"bold"), anchor="w",
                     wraplength=C_W-8).pack(fill="x", padx=4, pady=(4,0))

            # Valor pago — ADM isento mostra valor TOTAL da cota (não a parcela, não conta no somatório)
            if pago is not None:
                tk.Label(card, text=f"Pago: {fmt_brl(pago)}", bg=bg_c,
                         fg="#c8f7dc" if "QUITADO" in status else "#cde",
                         font=("Arial",7), anchor="w").pack(fill="x", padx=4)
            else:
                # ADM isento: exibe o valor total de uma cota (não a parcela)
                vt_card = b.get("valor_total", 0) or 0
                tk.Label(card,
                         text=fmt_brl(vt_card),
                         bg=bg_c, fg="#c8f7dc",
                         font=("Arial",7), anchor="w").pack(fill="x", padx=4)

            # Status + saldo se pendente
            if "QUITADO" in status:
                st_txt = f"{ic} QUITADO"
            elif "EM DIA" in status:
                st_txt = f"{ic} EM DIA"
            else:
                st_txt = f"{ic} {fmt_brl(saldo)}"
            tk.Label(card, text=st_txt, bg=bg_c, fg=fg_stat,
                     font=("Arial",8,"bold"), anchor="w").pack(fill="x", padx=4, pady=(1,4))

        # Separadores visuais de linha (linha tracejada entre grupos de 10)
        inner.update_idletasks()
        canvas.configure(scrollregion=canvas.bbox("all"))

        # Scroll com mouse
        def _scroll(e):
            canvas.yview_scroll(int(-1*(e.delta/120)), "units")
        canvas.bind("<MouseWheel>", _scroll)
        win.bind("<MouseWheel>", _scroll)

        # Botão de fechar acessível
        tk.Button(win, text="✕  Fechar", bg="#c0392b", fg="white",
                  font=("Arial",9,"bold"), relief="flat", padx=16, pady=4,
                  command=win.destroy, cursor="hand2").pack(pady=(0,8))

    # ════════════════════════════════════════════════════════════
    #  ABA — HISTÓRICO DE PAGAMENTOS
    # ════════════════════════════════════════════════════════════
    def _build_historico(self):
        p = self.tab_hist

        # ── Barra de controles ──────────────────────────────────
        top = tk.Frame(p, bg=CORES["bg_frame"])
        top.pack(fill="x", padx=20, pady=(10,4))

        btn(top, "🔄 Atualizar", CORES["btn_azul"],
            self._hist_load, width=14).pack(side="left", padx=4)
        btn(top, "📊 Exportar Excel", CORES["btn_verde"],
            self._hist_exportar, width=18).pack(side="left", padx=4)
        btn(top, "🗑 Excluir Selecionado", CORES["btn_vermelho"],
            self._hist_excluir, width=20).pack(side="left", padx=4)
        btn(top, "🧹 Pagamentos Órfãos", CORES["btn_cinza"],
            self._sinc_limpar_orfaos, width=20).pack(side="left", padx=4)

        tk.Label(top, text="  Ordenar por:", bg=CORES["bg_frame"],
                 fg=CORES["fg_label"], font=("Arial",9,"bold")).pack(side="left", padx=(12,4))
        self._hist_ordem = ttk.Combobox(top, width=16, state="readonly", font=("Arial",9),
                                         values=["Data Pagamento ↑","Data Pagamento ↓",
                                                 "Participante A-Z","Valor ↓","Depositado"])
        self._hist_ordem.set("Data Pagamento ↑")
        self._hist_ordem.pack(side="left", padx=4)
        self._hist_ordem.bind("<<ComboboxSelected>>", lambda e: self._hist_load())

        tk.Label(top, text="  Status:", bg=CORES["bg_frame"],
                 fg=CORES["fg_label"], font=("Arial",9,"bold")).pack(side="left", padx=(12,4))
        self._hist_filtro = ttk.Combobox(top, width=14, state="readonly", font=("Arial",9),
                                          values=["Todos","Depositados","Não Depositados"])
        self._hist_filtro.set("Todos")
        self._hist_filtro.pack(side="left", padx=4)
        self._hist_filtro.bind("<<ComboboxSelected>>", lambda e: self._hist_load())

        # Busca por nome
        tk.Label(top, text="  Buscar:", bg=CORES["bg_frame"],
                 fg=CORES["fg_label"], font=("Arial",9,"bold")).pack(side="left", padx=(12,4))
        self._hist_busca_var = tk.StringVar()
        self._hist_busca = tk.Entry(top, textvariable=self._hist_busca_var,
                                     width=18, font=("Arial",9), relief="solid", bd=1)
        self._hist_busca.pack(side="left", padx=4)
        self._hist_busca_var.trace_add("write", lambda *a: self._hist_load())

        btn(top, "✖", CORES["btn_cinza"],
            lambda: (self._hist_busca_var.set(""), self._hist_load()),
            width=3).pack(side="left", padx=2)

        # ── Cards de resumo ──────────────────────────────────────
        krow = tk.Frame(p, bg=CORES["bg_frame"]); krow.pack(fill="x", padx=20, pady=(4,6))
        self._hist_cards = {}
        for attr, titulo, cor in [
            ("h_total",   "Registros",         "#2196F3"),
            ("h_valor",   "Valor Total",        "#8e44ad"),
            ("h_dep",     "✅ Depositados",     "#1D9E75"),
            ("h_ndep",    "⏳ Pendentes dep.",  "#e67e22"),
        ]:
            c = tk.Frame(krow, bg=cor, padx=10, pady=6)
            c.pack(side="left", fill="both", expand=True, padx=3)
            tk.Label(c, text=titulo, bg=cor, fg="white",
                     font=("Arial",7,"bold")).pack(anchor="w")
            lv = tk.Label(c, text="—", bg=cor, fg="white", font=("Arial",11,"bold"))
            lv.pack(anchor="w")
            self._hist_cards[attr] = lv

        # ── Tabela principal ─────────────────────────────────────
        sec = section(p, "HISTÓRICO DE PAGAMENTOS")
        sec.pack(fill="both", expand=True, padx=20, pady=(0,10))

        cols = {
            "#":             40,
            "Data Pagamento":110,
            "Participante":  200,
            "Mês Ref.":       90,
            "Valor (R$)":    110,
            "Depositado":     90,
            "Data Depósito": 110,
            "Observações":   180,
        }
        fr, self.hist_tree = make_tree(sec, cols, height=24)
        fr.pack(fill="both", expand=True)
        self.hist_tree.tag_configure("dep",  background="#d5f5e3")
        self.hist_tree.tag_configure("ndep", background="#fde8d8")
        self.hist_tree.bind("<Double-1>", self._hist_editar_duplo)

    def _hist_load(self):
        bid = self.bid.get()
        self.hist_tree.delete(*self.hist_tree.get_children())
        if not bid:
            for k in self._hist_cards: self._hist_cards[k].configure(text="—")
            return

        # ── Filtro de status ─────────────────────────────────────
        filtro_val = self._hist_filtro.get()
        where_dep = ""
        if filtro_val == "Depositados":       where_dep = "AND pg.depositado = 1"
        elif filtro_val == "Não Depositados": where_dep = "AND pg.depositado = 0"

        # ── Filtro de busca por nome ─────────────────────────────
        busca = self._hist_busca_var.get().strip().lower()
        where_busca = f"AND LOWER(pt.nome) LIKE '%{busca}%'" if busca else ""

        # Busca sem ORDER BY — ordenamos em Python para garantir datas DD/MM/YYYY corretas
        rows = self.db.fetchall(f"""
            SELECT pg.id, pg.data_pagamento, pt.nome, pg.mes_referencia,
                   pg.valor, pg.depositado, pg.data_deposito, pg.observacoes
            FROM pagamentos pg
            JOIN participantes pt ON pg.participante_id = pt.id
            WHERE pg.bolao_id = ? {where_dep} {where_busca}
        """, (bid,))

        # ── Converte data DD/MM/YYYY → tuple (YYYY,MM,DD) para ordenação correta ─
        def data_key(d):
            """Converte DD/MM/YYYY ou YYYY-MM-DD para tuple ordenável. Null vai para o fim."""
            if not d or d == "-":
                return (9999, 99, 99)
            d = str(d).strip()
            try:
                if "/" in d:           # DD/MM/YYYY
                    p = d.split("/")
                    return (int(p[2]), int(p[1]), int(p[0]))
                elif "-" in d:         # YYYY-MM-DD
                    p = d.split("-")
                    return (int(p[0]), int(p[1]), int(p[2]))
            except:
                pass
            return (9999, 99, 99)

        # ── Ordenação escolhida pelo usuário ─────────────────────
        ordem_sel = self._hist_ordem.get()
        rows = list(rows)

        if ordem_sel == "Data Pagamento ↑":
            rows.sort(key=lambda r: (data_key(r["data_pagamento"]), r["id"]))
        elif ordem_sel == "Data Pagamento ↓":
            rows.sort(key=lambda r: (data_key(r["data_pagamento"]), r["id"]), reverse=True)
        elif ordem_sel == "Participante A-Z":
            rows.sort(key=lambda r: (r["nome"].lower(), data_key(r["data_pagamento"])))
        elif ordem_sel == "Valor ↓":
            rows.sort(key=lambda r: (-r["valor"], data_key(r["data_pagamento"])))
        elif ordem_sel == "Depositado":
            rows.sort(key=lambda r: (r["depositado"], data_key(r["data_pagamento"])))
        else:
            # Padrão: data crescente
            rows.sort(key=lambda r: (data_key(r["data_pagamento"]), r["id"]))

        # ── Cards de resumo ──────────────────────────────────────
        total_val  = sum(r["valor"] for r in rows)
        total_dep  = sum(r["valor"] for r in rows if r["depositado"])
        total_ndep = total_val - total_dep
        n_dep      = sum(1 for r in rows if r["depositado"])
        n_ndep     = len(rows) - n_dep

        self._hist_cards["h_total"].configure(text=str(len(rows)))
        self._hist_cards["h_valor"].configure(text=fmt_brl(total_val))
        self._hist_cards["h_dep"].configure(  text=f"{n_dep}  ({fmt_brl(total_dep)})")
        self._hist_cards["h_ndep"].configure( text=f"{n_ndep}  ({fmt_brl(total_ndep)})")

        # ── Linhas da tabela ─────────────────────────────────────
        for i, r in enumerate(rows, 1):
            dep = "✅ Sim" if r["depositado"] else "⏳ Não"
            tag = "dep" if r["depositado"] else "ndep"
            self.hist_tree.insert("", "end", iid=str(r["id"]), tags=(tag,), values=(
                i,
                r["data_pagamento"]  or "-",
                r["nome"],
                r["mes_referencia"]  or "-",
                fmt_brl(r["valor"]),
                dep,
                r["data_deposito"]   or "-",
                r["observacoes"]     or "-",
            ))

    def _hist_editar_duplo(self, e=None):
        """Clique duplo no histórico abre o form de edição do pagamento."""
        sel = self.hist_tree.selection()
        if not sel: return
        pid_pag = int(sel[0])
        pg = self.db.fetchone("SELECT * FROM pagamentos WHERE id=?", (pid_pag,))
        if not pg: return

        win = tk.Toplevel(self.root)
        win.title(f"Editar Pagamento ID {pid_pag}")
        win.geometry("460x380")
        win.configure(bg=CORES["bg_section"])
        win.grab_set()
        tk.Label(win, text=f"EDITAR PAGAMENTO  (ID: {pid_pag})",
                 bg=CORES["bg_section"], fg=CORES["fg_title"],
                 font=("Arial",12,"bold")).pack(pady=10)
        form = tk.Frame(win, bg=CORES["bg_section"], padx=24)
        form.pack(fill="both", expand=True)
        form.columnconfigure(0, weight=1)
        fields = [("Mês Referência:","mes_referencia"),
                  ("Data Pagamento (DD/MM/AAAA):","data_pagamento"),
                  ("Valor (R$):","valor"),
                  ("Observações:","observacoes")]
        vars_ = {}
        for i, (lbl, key) in enumerate(fields):
            tk.Label(form, text=lbl, bg=CORES["bg_section"], fg=CORES["fg_label"],
                     font=("Arial",9,"bold")).grid(row=i*2, column=0, sticky="w", pady=(8,0))
            w = entry(form, width=40)
            val = pg[key] if pg[key] is not None else ""
            if key == "valor":
                val = f"{float(val or 0):.2f}".replace(".", ",")
            w.insert(0, str(val))
            w.grid(row=i*2+1, column=0, sticky="ew")
            vars_[key] = w
        dep_var = tk.IntVar(value=int(pg["depositado"] or 0))
        dep_f = tk.Frame(form, bg=CORES["bg_section"])
        dep_f.grid(row=8, column=0, sticky="w", pady=8)
        tk.Label(dep_f, text="Depositado:", bg=CORES["bg_section"],
                 font=("Arial",9,"bold"), fg=CORES["fg_label"]).pack(side="left")
        tk.Radiobutton(dep_f, text="✅ Sim", variable=dep_var, value=1,
                       bg=CORES["bg_section"]).pack(side="left", padx=8)
        tk.Radiobutton(dep_f, text="⏳ Não", variable=dep_var, value=0,
                       bg=CORES["bg_section"]).pack(side="left")
        def salvar():
            novo_val = to_float(vars_["valor"].get())
            if novo_val <= 0:
                messagebox.showwarning("Atenção", "Valor inválido!"); return
            dep = dep_var.get()
            dd  = pg["data_deposito"]
            if dep and not dd:
                dd = date.today().strftime("%d/%m/%Y")
            elif not dep:
                dd = None
            self.db.execute(
                "UPDATE pagamentos SET mes_referencia=?,data_pagamento=?,valor=?,"
                "observacoes=?,depositado=?,data_deposito=? WHERE id=?",
                (vars_["mes_referencia"].get(), vars_["data_pagamento"].get(),
                 novo_val, vars_["observacoes"].get(),
                 dep, dd, pid_pag))
            messagebox.showinfo("Atualizado", f"Pagamento ID {pid_pag} atualizado!")
            win.destroy()
            self._hist_load()
        btn(form, "💾 SALVAR ALTERAÇÕES", CORES["btn_verde"], salvar, width=22).grid(
            row=9, column=0, pady=14, sticky="e")

    def _hist_excluir(self):
        """Exclui o pagamento selecionado no Histórico (mesma trava de
        segurança que a antiga aba Editar: não deixa excluir se já
        depositado, senão desconta do fundo sem ninguém perceber)."""
        sel = self.hist_tree.selection()
        if not sel:
            messagebox.showwarning("Atenção", "Selecione um pagamento na lista!"); return
        pid_pag = int(sel[0])
        pg = self.db.fetchone("SELECT * FROM pagamentos WHERE id=?", (pid_pag,))
        if not pg: return
        if pg["depositado"]:
            messagebox.showerror("Operação Bloqueada",
                "Este pagamento já foi depositado e não pode ser excluído.\n\n"
                "Para excluir, primeiro dê duplo-clique nele e marque "
                "'Depositado: Não'.")
            return
        if messagebox.askyesno("Confirmar", "Excluir este pagamento?"):
            self.db.execute("DELETE FROM pagamentos WHERE id=?", (pid_pag,))
            messagebox.showinfo("Excluído", "Pagamento excluído.")
            self._hist_load()

    def _hist_exportar(self):
        bid = self.bid.get()
        if not bid: messagebox.showwarning("Atenção","Selecione um bolão!"); return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror("Erro","Execute: pip install openpyxl"); return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx")],
            initialfile=f"historico_pagamentos_{bid}.xlsx")
        if not path: return

        rows = self.db.fetchall("""
            SELECT pg.id, pg.data_pagamento, pt.nome, pg.mes_referencia,
                   pg.valor, pg.depositado, pg.data_deposito, pg.observacoes
            FROM pagamentos pg
            JOIN participantes pt ON pg.participante_id = pt.id
            WHERE pg.bolao_id = ?
            ORDER BY pg.data_pagamento, pg.id
        """, (bid,))

        wb = Workbook(); ws = wb.active; ws.title = "Histórico"
        thin = "thin"
        brd = Border(left=Side(style=thin), right=Side(style=thin),
                     top=Side(style=thin),  bottom=Side(style=thin))
        ctr = Alignment(horizontal="center", vertical="center")

        # Cabeçalho
        headers = ["#","Data Pagamento","Participante","Mês Ref.",
                   "Valor (R$)","Depositado","Data Depósito","Observações"]
        ws.append(headers)
        for c in range(1, len(headers)+1):
            cell = ws.cell(1, c)
            cell.font  = Font(bold=True, color="FFFFFF")
            cell.fill  = PatternFill("solid", fgColor="1A2A3A")
            cell.alignment = ctr
            cell.border = brd
        ws.row_dimensions[1].height = 20

        for i, r in enumerate(rows, 1):
            dep = "Sim" if r["depositado"] else "Não"
            bg  = "D5F5E3" if r["depositado"] else "FDE8D8"
            ws.append([i, r["data_pagamento"] or "", r["nome"],
                       r["mes_referencia"] or "", r["valor"],
                       dep, r["data_deposito"] or "", r["observacoes"] or ""])
            row_idx = ws.max_row
            for c in range(1, len(headers)+1):
                cell = ws.cell(row_idx, c)
                cell.fill   = PatternFill("solid", fgColor=bg)
                cell.border = brd
                cell.alignment = ctr
            ws.cell(row_idx, 5).number_format = 'R$ #,##0.00'

        # Linha de total
        tot_row = ws.max_row + 1
        ws.cell(tot_row, 2, "TOTAL").font = Font(bold=True)
        ws.cell(tot_row, 5, sum(r["valor"] for r in rows))
        ws.cell(tot_row, 5).number_format = 'R$ #,##0.00'
        ws.cell(tot_row, 5).font = Font(bold=True)

        for i, w in enumerate([5,16,28,12,14,12,14,30], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        wb.save(path)
        messagebox.showinfo("Exportado", f"Histórico salvo em:\n{path}")

    # ════════════════════════════════════════════════════════════
    #  ABA DASHBOARD
    # ════════════════════════════════════════════════════════════
    def _build_dashboard(self):
        p = self.tab_dash
        p.configure(bg="#1a2a3a")

        # Reestruturado a pedido do usuário: visão geral (todos os bolões)
        # primeiro, depois um seletor de bolão fácil de enxergar, depois o
        # resumo+detalhe do bolão escolhido. Tela única com rolagem.
        canvas = tk.Canvas(p, bg="#1a2a3a", highlightthickness=0)
        sb = ttk.Scrollbar(p, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        p = tk.Frame(canvas, bg="#1a2a3a")
        canvas_window = canvas.create_window((0, 0), window=p, anchor="nw")
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(canvas_window, width=e.width))

        # ══════════════ PARTE 1 — VISÃO GERAL (TODOS OS BOLÕES) ═════════
        hdr = tk.Frame(p, bg="#1a2a3a", pady=10)
        hdr.pack(fill="x", padx=20)
        tk.Label(hdr, text="📊  Visão Geral",
                 bg="#1a2a3a", fg="white", font=("Arial",16,"bold")).pack(side="left")
        self._dash_hora_lbl = tk.Label(hdr, text="",
                 bg="#1a2a3a", fg="#556677", font=("Arial",9))
        self._dash_hora_lbl.pack(side="right", padx=4)
        btn(hdr, "🔄", CORES["btn_azul"], self._recarregar_visao_geral, width=4).pack(side="right", padx=4)

        # ── KPIs gerais — 7 cards, paleta própria (não repete nenhuma cor
        # dos KPIs do bolão selecionado, lá embaixo) ─────────────────────
        kpi_bar = tk.Frame(p, bg="#1a2a3a")
        kpi_bar.pack(fill="x", padx=20, pady=(0,8))
        self._adm_kpis = {}
        kpi_defs = [
            ("geral_boloes",   "BOLÕES ATIVOS",          "#78716c", "0"),
            ("geral_arrec",    "ARRECADADO (GERAL)",     "#65a30d", "R$ 0,00"),
            ("geral_pend_dep", "PENDENTE DEPÓSITO",      "#9333ea", "R$ 0,00"),
            ("geral_atrasados","PARTICIPANTES ATRASADOS","#ca8a04", "0"),
            ("adm_total",      "TOTAL GANHO (ADM)",      "#4f46e5", "R$ 0,00"),
            ("adm_sacado",     "TOTAL SACADO",           "#db2777", "R$ 0,00"),
            ("adm_saldo",      "SALDO DISPONÍVEL",       "#0891b2", "R$ 0,00"),
        ]
        for attr, titulo, cor, default in kpi_defs:
            card = tk.Frame(kpi_bar, bg=cor, padx=10, pady=8)
            card.pack(side="left", fill="both", expand=True, padx=3)
            tk.Label(card, text=titulo, bg=cor, fg="#f0f0f0",
                     font=("Arial",7,"bold")).pack(anchor="w")
            val_lbl = tk.Label(card, text=default, bg=cor, fg="white",
                               font=("Arial",13,"bold"))
            val_lbl.pack(anchor="w")
            self._adm_kpis[attr] = val_lbl

        # ── Ações rápidas ─────────────────────────────────────────
        acoes_fr = tk.Frame(p, bg="#1a2a3a")
        acoes_fr.pack(fill="x", padx=20, pady=(0,10))
        btn(acoes_fr, "📊 Ver Ganhos por Loteria", CORES["btn_roxo"],
            self._abrir_ganhos_por_loteria, width=24).pack(side="left")
        btn(acoes_fr, "➕ Registrar Lançamento", CORES["btn_verde"],
            self._abrir_registrar_lancamento, width=22).pack(side="left", padx=8)

        # ── Depósitos Pendentes | Participantes Atrasados ────────
        mid1 = tk.Frame(p, bg="#1a2a3a")
        mid1.pack(fill="both", padx=20, pady=(0,8))
        mid1.columnconfigure(0, weight=3); mid1.columnconfigure(1, weight=2)

        sec_dep_pend = tk.LabelFrame(mid1, text="  DEPÓSITOS PENDENTES  ",
            bg="#243447", fg="#ffcc88", font=("Arial",9,"bold"), bd=1, padx=6, pady=4)
        sec_dep_pend.grid(row=0, column=0, sticky="nsew", padx=(0,4))
        fr_dp, self.adm_tree_dep = make_tree(sec_dep_pend,
            {"ID":42,"Participante":150,"Data Pag.":86,"Valor":86,"Bolao":120}, height=8)
        fr_dp.pack(fill="both", expand=True)
        self.adm_tree_dep.tag_configure("row1", background="#fff8e8")
        self.adm_tree_dep.tag_configure("row2", background="#ffffff")
        self._adm_dep_total_lbl = tk.Label(sec_dep_pend, text="", bg="#243447",
                                            fg="#ffcc88", font=("Arial",8,"bold"))
        self._adm_dep_total_lbl.pack(anchor="e", pady=(2,0))

        sec_atr = tk.LabelFrame(mid1, text="  PARTICIPANTES ATRASADOS  ",
            bg="#243447", fg="#ffcc88", font=("Arial",9,"bold"), bd=1, padx=6, pady=4)
        sec_atr.grid(row=0, column=1, sticky="nsew")
        # "Faltam" (não "Devidas"): antes mostrava o total de parcelas
        # esperadas desde o início do bolão (um número cumulativo, ex.: 8),
        # não quanto realmente falta (ex.: pagou 7 de 8 esperadas → falta
        # só 1) — confundia porque a coluna "Saldo" ao lado já mostrava o
        # valor certo, e os dois números pareciam contraditórios.
        fr_a, self.adm_tree_atr = make_tree(sec_atr,
            {"Participante":120,"Bolao":110,"Pagas":46,"Faltam":52,"Saldo":86}, height=8)
        fr_a.pack(fill="both", expand=True)
        self.adm_tree_atr.tag_configure("atr1", background="#fde8d8")
        self.adm_tree_atr.tag_configure("atr2", background="#f9c0b0")
        self._adm_atr_lbl = tk.Label(sec_atr, text="", bg="#243447",
                                      fg="#ffcc88", font=("Arial",8,"bold"))
        self._adm_atr_lbl.pack(anchor="e", pady=(2,0))

        # ── Últimos Pagamentos (geral) | Histórico de Lançamentos ─
        mid2 = tk.Frame(p, bg="#1a2a3a")
        mid2.pack(fill="both", padx=20, pady=(0,14))
        mid2.columnconfigure(0, weight=2); mid2.columnconfigure(1, weight=3)

        sec_ult_geral = tk.LabelFrame(mid2, text="  ÚLTIMOS PAGAMENTOS (GERAL)  ",
            bg="#243447", fg="white", font=("Arial",9,"bold"), bd=1, padx=6, pady=4)
        sec_ult_geral.grid(row=0, column=0, sticky="nsew", padx=(0,4))
        cols_ug = {"Participante":150,"Bolão":140,"Data":86,"Valor":86}
        fr_ug, self.geral_tree_ult = make_tree(sec_ult_geral, cols_ug, height=8)
        fr_ug.pack(fill="both", expand=True)
        self.geral_tree_ult.tag_configure("linha", background="#f8f8f8")

        sec_hist = tk.LabelFrame(mid2, text="  HISTÓRICO DE LANÇAMENTOS  ",
            bg="#243447", fg="white", font=("Arial",9,"bold"), bd=1, padx=6, pady=4)
        sec_hist.grid(row=0, column=1, sticky="nsew")
        cols_h = {"ID":40,"Bolao":120,"Loteria":80,"Concurso":65,
                  "Valor":90,"Tipo":70,"Descricao":140,"Data":90}
        fr_h, self.adm_tree_hist = make_tree(sec_hist, cols_h, height=8)
        fr_h.pack(fill="both", expand=True)
        self.adm_tree_hist.tag_configure("ganho", background="#d5f5e3")
        self.adm_tree_hist.tag_configure("saque", background="#fde8d8")
        bh = tk.Frame(sec_hist, bg="#243447"); bh.pack(fill="x", pady=2)
        btn(bh, "Editar",  CORES["btn_azul"],    self._adm_editar,  width=10).pack(side="left", padx=3)
        btn(bh, "Excluir", CORES["btn_vermelho"], self._adm_excluir, width=10).pack(side="left", padx=3)

        # Região de rolagem — precisa ser calculada depois que TODO o
        # conteúdo já foi montado
        p.update_idletasks()
        canvas.configure(scrollregion=canvas.bbox("all"))

        # bind() direto no canvas só dispara quando o mouse está sobre a
        # área vazia dele — como a tela é preenchida por labels/treeviews,
        # isso quase nunca acontecia na prática. O jeito certo é ligar o
        # scroll globalmente (bind_all) só enquanto o mouse estiver sobre
        # a área do canvas (Enter/Leave), cobrindo o mouse em cima de
        # qualquer widget filho também.
        def _scroll(e):
            canvas.yview_scroll(int(-1*(e.delta/120)), "units")
        def _ligar_scroll(e):
            canvas.bind_all("<MouseWheel>", _scroll)
        def _desligar_scroll(e):
            canvas.unbind_all("<MouseWheel>")
        canvas.bind("<Enter>", _ligar_scroll)
        canvas.bind("<Leave>", _desligar_scroll)

    def _build_bolao_sel(self):
        """Aba própria "🎯 Bolão Selecionado" — separada da Visão Geral a
        pedido do usuário: numa aba só, a rolagem ficava longa e incômoda.
        Aqui: seletor de cartões, depois resumo, depois detalhe de UM
        bolão por vez — a Visão Geral (todos os bolões) fica na aba anterior."""
        p = self.tab_bolao
        p.configure(bg="#1a2a3a")

        canvas = tk.Canvas(p, bg="#1a2a3a", highlightthickness=0)
        sb = ttk.Scrollbar(p, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        p = tk.Frame(canvas, bg="#1a2a3a")
        canvas_window = canvas.create_window((0, 0), window=p, anchor="nw")
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(canvas_window, width=e.width))

        # ── Cabeçalho + seletor de bolão — cartões clicáveis (mais fácil
        # de achar e trocar do que o combo pequeno do cabeçalho) ─────────
        hdr2 = tk.Frame(p, bg="#1a2a3a", pady=10)
        hdr2.pack(fill="x", padx=20)
        tk.Label(hdr2, text="🎯  Selecione um bolão",
                 bg="#1a2a3a", fg="white", font=("Arial",14,"bold")).pack(side="left")
        btn(hdr2, "🔄", CORES["btn_azul"], self._recarregar_visao_geral, width=4).pack(side="right")
        tk.Label(p, text="Clique num bolão abaixo para ver o resumo e os detalhes dele.",
                 bg="#1a2a3a", fg="#8899aa", font=("Arial",8)).pack(padx=20, anchor="w", pady=(0,6))

        self._sel_bolao_frame = tk.Frame(p, bg="#1a2a3a")
        self._sel_bolao_frame.pack(fill="x", padx=20, pady=(0,12))

        # ── Divisor fino antes do resumo/detalhe ──────────────────
        tk.Frame(p, bg="#334455", height=2).pack(fill="x", padx=20, pady=(0,10))

        hdr3 = tk.Frame(p, bg="#1a2a3a", pady=4)
        hdr3.pack(fill="x", padx=20)
        self._dash_bolao_lbl = tk.Label(hdr3, text="",
                 bg="#1a2a3a", fg="white", font=("Arial",13,"bold"))
        self._dash_bolao_lbl.pack(side="left")

        # ── Resumo — 6 cards compactos deste bolão ────────────────
        kpi_row = tk.Frame(p, bg="#1a2a3a")
        kpi_row.pack(fill="x", padx=20, pady=(6,6))
        kpi_defs2 = [
            ("dash_participantes", "Cotas Ocupadas", "#2196F3"),
            ("dash_quitados",      "Quitados",       "#1D9E75"),
            ("dash_pendentes",     "Pendentes",      "#e67e22"),
            ("dash_arrecadado",    "Arrecadado",     "#8e44ad"),
            ("dash_pendente_val",  "A Receber",      "#e74c3c"),
            ("dash_depositos",     "Depositado",     "#16a085"),
        ]
        self._dash_kpis = {}
        for attr, titulo, cor in kpi_defs2:
            card = tk.Frame(kpi_row, bg=cor, padx=12, pady=8, relief="flat")
            card.pack(side="left", fill="both", expand=True, padx=3)
            tk.Label(card, text=titulo, bg=cor, fg="white",
                     font=("Arial",7,"bold")).pack(anchor="w")
            lbl_v = tk.Label(card, text="—", bg=cor, fg="white",
                             font=("Arial",15,"bold"))
            lbl_v.pack(anchor="w", pady=(2,0))
            self._dash_kpis[attr] = lbl_v

        # ── Barra de progresso ───────────────────────────────────
        prog_fr = tk.Frame(p, bg="#1e3348", padx=16, pady=8)
        prog_fr.pack(fill="x", padx=20, pady=(0,6))
        top_pr = tk.Frame(prog_fr, bg="#1e3348"); top_pr.pack(fill="x")
        tk.Label(top_pr, text="Progresso de Arrecadação",
                 bg="#1e3348", fg="white", font=("Arial",9,"bold")).pack(side="left")
        self._dash_pct_lbl = tk.Label(top_pr, text="",
                 bg="#1e3348", fg="#aad4f5", font=("Arial",9,"bold"))
        self._dash_pct_lbl.pack(side="right")
        self._dash_prog = ttk.Progressbar(prog_fr, mode="determinate", maximum=100)
        self._dash_prog.pack(fill="x", pady=(4,0))

        # ── Detalhe: 3 colunas ────────────────────────────────────
        main = tk.Frame(p, bg="#1a2a3a")
        main.pack(fill="both", expand=True, padx=20, pady=(0,6))
        main.columnconfigure(0, weight=2)
        main.columnconfigure(1, weight=2)
        main.columnconfigure(2, weight=1)
        main.rowconfigure(0, weight=1)

        # ── Coluna esquerda: situação dos participantes ──────────
        sec_sit = tk.LabelFrame(main, text="  Situação dos Participantes  ",
                                bg="#243447", fg="white",
                                font=("Arial",9,"bold"), bd=1, padx=6, pady=6)
        sec_sit.grid(row=0, column=0, sticky="nsew", padx=(0,4))
        sec_sit.rowconfigure(0, weight=1); sec_sit.columnconfigure(0, weight=1)
        cols_s = {"Nome":190, "Pago":100, "Saldo":100, "Status":90}
        fr_s, self._dash_tree_sit = make_tree(sec_sit, cols_s, height=10)
        fr_s.grid(row=0, column=0, sticky="nsew")
        self._dash_tree_sit.tag_configure("quitado",  background="#d5f5e3")
        self._dash_tree_sit.tag_configure("pendente", background="#fde8d8")

        # ── Coluna central: últimos pagamentos ───────────────────
        sec_ult = tk.LabelFrame(main, text="  Últimos Pagamentos Recebidos  ",
                                bg="#243447", fg="white",
                                font=("Arial",9,"bold"), bd=1, padx=6, pady=6)
        sec_ult.grid(row=0, column=1, sticky="nsew", padx=4)
        sec_ult.rowconfigure(0, weight=1); sec_ult.columnconfigure(0, weight=1)
        cols_u = {"Participante":190, "Data":100, "Valor":100}
        fr_u, self._dash_tree_ult = make_tree(sec_ult, cols_u, height=10)
        fr_u.grid(row=0, column=0, sticky="nsew")
        self._dash_tree_ult.tag_configure("linha", background="#f8f8f8")

        # ── Coluna direita: resumo financeiro + premiações ───────
        col3 = tk.Frame(main, bg="#1a2a3a")
        col3.grid(row=0, column=2, sticky="nsew", padx=(4,0))
        col3.columnconfigure(0, weight=1)
        col3.rowconfigure(0, weight=1)
        col3.rowconfigure(1, weight=1)
        col3.rowconfigure(2, weight=1)

        # Bloco financeiro
        sec_fin = tk.LabelFrame(col3, text="  Financeiro  ",
                                bg="#243447", fg="white",
                                font=("Arial",9,"bold"), bd=1, padx=10, pady=8)
        sec_fin.grid(row=0, column=0, sticky="nsew", pady=(0,4))

        # Só os 2 números que NÃO aparecem no resumo acima (Arrecadado,
        # Depositado e A Receber já estão lá — repetir aqui era a mesma
        # informação duas vezes, em dois estilos visuais diferentes).
        self._dash_fin_labels = {}
        fin_items = [
            ("fin_esp",  "Total Esperado", "#aad4f5"),
            ("fin_pct",  "% Arrecadado",   "#aad4f5"),
        ]
        for attr, lbl_txt, cor in fin_items:
            row_f = tk.Frame(sec_fin, bg="#243447"); row_f.pack(fill="x", pady=2)
            tk.Label(row_f, text=lbl_txt, bg="#243447", fg="#8899aa",
                     font=("Arial",8)).pack(side="left")
            lv = tk.Label(row_f, text="—", bg="#243447", fg=cor,
                          font=("Arial",9,"bold"))
            lv.pack(side="right")
            self._dash_fin_labels[attr] = lv

        # Bloco Situação do Fundo
        sec_fundo = tk.LabelFrame(col3, text="  🏦 Situação do Fundo  ",
                                  bg="#243447", fg="white",
                                  font=("Arial",9,"bold"), bd=1, padx=10, pady=8)
        sec_fundo.grid(row=1, column=0, sticky="nsew", pady=4)

        self._dash_fundo_labels = {}
        fundo_items = [
            ("fundo_dep",   "Total Depositado",  "#afffca"),
            ("fundo_saques","Saques Emerg.",     "#ffaaaa"),
            ("fundo_saldo", "Saldo na Conta",    "#ffd700"),
            ("fundo_pend",  "Pendente Depósito", "#ffcc88"),
        ]
        for attr, lbl_txt, cor in fundo_items:
            row_f = tk.Frame(sec_fundo, bg="#243447"); row_f.pack(fill="x", pady=2)
            tk.Label(row_f, text=lbl_txt, bg="#243447", fg="#8899aa",
                     font=("Arial",8)).pack(side="left")
            lv = tk.Label(row_f, text="—", bg="#243447", fg=cor,
                          font=("Arial",9,"bold"))
            lv.pack(side="right")
            self._dash_fundo_labels[attr] = lv

        # Bloco Informações do Bolão
        sec_info = tk.LabelFrame(col3, text="  📋 Informações do Bolão  ",
                                  bg="#243447", fg="white",
                                  font=("Arial",9,"bold"), bd=1, padx=10, pady=8)
        sec_info.grid(row=2, column=0, sticky="nsew", pady=(4,0))

        self._dash_info_labels = {}
        info_items = [
            ("info_lot",   "Loteria",           "#aad4f5"),
            ("info_inicio","Início",            "#aad4f5"),
            ("info_parc",  "Valor da Parcela",  "#afffca"),
            ("info_nparc", "Total de Parcelas", "#aad4f5"),
            ("info_atual", "Parcela Atual",     "#ffd700"),
            ("info_cotas", "Cotas",             "#aad4f5"),
        ]
        for attr, lbl_txt, cor in info_items:
            row_i = tk.Frame(sec_info, bg="#243447"); row_i.pack(fill="x", pady=2)
            tk.Label(row_i, text=lbl_txt, bg="#243447", fg="#8899aa",
                     font=("Arial",8)).pack(side="left")
            lv = tk.Label(row_i, text="—", bg="#243447", fg=cor,
                          font=("Arial",9,"bold"))
            lv.pack(side="right")
            self._dash_info_labels[attr] = lv

        # ── Rodapé ───────────────────────────────────────────────
        rod = tk.Frame(p, bg="#151f2b", pady=5)
        rod.pack(fill="x", padx=20, pady=(0,8))
        self._dash_footer = tk.Label(rod, text="",
                 bg="#151f2b", fg="#445566", font=("Arial",8))
        self._dash_footer.pack(anchor="w", padx=8)

        # Região de rolagem — precisa ser calculada depois que TODO o
        # conteúdo já foi montado
        p.update_idletasks()
        canvas.configure(scrollregion=canvas.bbox("all"))

        # bind() direto no canvas só dispara quando o mouse está sobre a
        # área vazia dele — como a tela é preenchida por labels/treeviews,
        # isso quase nunca acontecia na prática. O jeito certo é ligar o
        # scroll globalmente (bind_all) só enquanto o mouse estiver sobre
        # a área do canvas (Enter/Leave), cobrindo o mouse em cima de
        # qualquer widget filho também.
        def _scroll(e):
            canvas.yview_scroll(int(-1*(e.delta/120)), "units")
        def _ligar_scroll(e):
            canvas.bind_all("<MouseWheel>", _scroll)
        def _desligar_scroll(e):
            canvas.unbind_all("<MouseWheel>")
        canvas.bind("<Enter>", _ligar_scroll)
        canvas.bind("<Leave>", _desligar_scroll)

        self._atualizar_cartoes_bolao()

    def _recarregar_visao_geral(self):
        """Botão de atualizar manual — recarrega os dois blocos da tela."""
        self._adm_load()
        self._dash_load()

    def _atualizar_cartoes_bolao(self):
        """Redesenha os cartões clicáveis de seleção de bolão. Cada cartão
        mostra um badge de status (✅ em dia / ⚠ N atrasado(s)) calculado
        por _adm_load, pra dar uma pista do bolão sem precisar abri-lo."""
        frame = self._sel_bolao_frame
        for w in frame.winfo_children():
            w.destroy()
        rows_b = self.db.fetchall("SELECT * FROM boloes WHERE encerrado=0 ORDER BY nome")
        bid_atual = self.bid.get()
        contagem = getattr(self, "_atrasados_count_por_bolao", {})
        NCOLS = 5
        for idx, b in enumerate(rows_b):
            bd = dict(b)
            selecionado = bd["id"] == bid_atual
            cor_bg = "#4f46e5" if selecionado else "#243447"
            cor_fg = "white" if selecionado else "#aad4f5"
            marca = "✅ " if selecionado else ""

            n_atr_bolao = contagem.get(bd["nome"])
            if n_atr_bolao is None:
                badge_txt, badge_fg = "", cor_fg
            elif n_atr_bolao == 0:
                badge_txt = "✅ em dia"
                badge_fg  = "#c8ffe0" if selecionado else "#7ee2a8"
            else:
                badge_txt = f"⚠ {n_atr_bolao} atrasado(s)"
                badge_fg  = "#ffe0c0" if selecionado else "#ffb066"

            card = tk.Frame(frame, bg=cor_bg, padx=14, pady=8, cursor="hand2",
                             highlightthickness=1, highlightbackground="#334455")
            card.grid(row=idx // NCOLS, column=idx % NCOLS, padx=4, pady=4, sticky="w")
            lbl_nome = tk.Label(card, text=marca + bd["nome"], bg=cor_bg, fg=cor_fg,
                                 font=("Arial",9,"bold"), wraplength=170, justify="left")
            lbl_nome.pack(anchor="w")
            lbl_badge = tk.Label(card, text=badge_txt, bg=cor_bg, fg=badge_fg,
                                  font=("Arial",7,"bold"))
            lbl_badge.pack(anchor="w")
            for w2 in (card, lbl_nome, lbl_badge):
                w2.bind("<Button-1>", lambda e, i=bd["id"]: self._selecionar_bolao_via_cartao(i))
        if not rows_b:
            tk.Label(frame, text="Nenhum bolão ativo — cadastre um em '+ Novo Bolão'.",
                     bg="#1a2a3a", fg="#8899aa", font=("Arial",9,"italic")).grid(row=0, column=0)

    def _selecionar_bolao_via_cartao(self, bid):
        self._selecionar_bolao_por_id(bid)
        self._atualizar_cartoes_bolao()

    def _abrir_ganhos_por_loteria(self):
        """Janela com o resumo de ganhos por loteria (taxa_adm) — tirado
        da tela principal a pedido do usuário: informação de consulta
        ocasional, não precisa competir por espaço todo dia."""
        win = tk.Toplevel(self.root)
        win.title("Ganhos por Loteria")
        win.geometry("420x360")
        win.configure(bg=CORES["bg_section"])
        win.grab_set(); win.lift(); win.focus_force()

        tk.Label(win, text="📊 GANHOS POR LOTERIA", bg=CORES["bg_section"],
                 fg=CORES["fg_title"], font=("Arial",12,"bold")).pack(pady=(14,8))

        fr = tk.Frame(win, bg=CORES["bg_section"], padx=16); fr.pack(fill="both", expand=True)
        cols = {"Loteria":140,"Lançamentos":90,"Total Ganho":120}
        fr_t, tree = make_tree(fr, cols, height=10)
        fr_t.pack(fill="both", expand=True)
        tree.tag_configure("pos", background="#d5f5e3")

        from collections import defaultdict
        todos = self.db.fetchall("SELECT * FROM taxa_adm WHERE tipo='GANHO'")
        por_lot = defaultdict(lambda: {"g": 0.0, "n": 0})
        for r in todos:
            lot = r["loteria"] or "Mega-Sena"
            por_lot[lot]["g"] += r["valor_ganho"]
            por_lot[lot]["n"] += 1
        for lot in sorted(por_lot.keys()):
            d = por_lot[lot]
            tree.insert("","end", tags=("pos",), values=(lot, d["n"], fmt_brl(d["g"])))
        if not por_lot:
            tk.Label(win, text="Nenhum ganho registrado ainda.", bg=CORES["bg_section"],
                     fg="#888", font=("Arial",9,"italic")).pack(pady=8)

        btn(win, "Fechar", CORES["btn_cinza"], win.destroy, width=12).pack(pady=12)

    def _abrir_registrar_lancamento(self):
        """Janela pra registrar um GANHO/SAQUE do organizador — tirada da
        tela principal (ação, não informação) a pedido do usuário."""
        win = tk.Toplevel(self.root)
        win.title("Registrar Lançamento")
        win.geometry("380x420")
        win.configure(bg=CORES["bg_section"])
        win.grab_set(); win.lift(); win.focus_force()

        tk.Label(win, text="➕ REGISTRAR LANÇAMENTO", bg=CORES["bg_section"],
                 fg=CORES["fg_title"], font=("Arial",12,"bold")).pack(pady=(14,8))

        fr = tk.Frame(win, bg=CORES["bg_section"], padx=20); fr.pack(fill="both", expand=True)
        def lbl_f(text):
            tk.Label(fr, text=text, bg=CORES["bg_section"], fg=CORES["fg_label"],
                     font=("Arial",9,"bold"), anchor="w").pack(fill="x", pady=(8,0))

        lbl_f("Tipo:")
        self.adm_tipo = ttk.Combobox(fr,
            values=["GANHO (taxa organizacao)","SAQUE (retirada)"],
            state="readonly", font=("Arial",9))
        self.adm_tipo.set("GANHO (taxa organizacao)"); self.adm_tipo.pack(fill="x")

        self._adm_bolao_frame = tk.Frame(fr, bg=CORES["bg_section"])
        self._adm_bolao_frame.pack(fill="x")
        tk.Label(self._adm_bolao_frame, text="Bolão:", bg=CORES["bg_section"],
                 fg=CORES["fg_label"], font=("Arial",9,"bold"), anchor="w").pack(fill="x", pady=(8,0))
        self.adm_cb_bolao = ttk.Combobox(self._adm_bolao_frame, state="readonly", font=("Arial",9))
        self.adm_cb_bolao.pack(fill="x")
        rows_b = self.db.fetchall("SELECT * FROM boloes WHERE encerrado=0 ORDER BY id")
        items_b = [f"{dict(b)['nome']} (ID: {dict(b)['id']})" for b in rows_b]
        items_b.append("-- Outros (bolão independente) --")
        self.adm_cb_bolao["values"] = items_b
        bid_atual = self.bid.get()
        selecionado = False
        for item in items_b:
            m = re.search(r"\(ID: (\d+)\)", item)
            if m and int(m.group(1)) == bid_atual:
                self.adm_cb_bolao.set(item); selecionado = True; break
        if not selecionado and items_b:
            self.adm_cb_bolao.set(items_b[0])

        self._adm_lot_frame = tk.Frame(fr, bg=CORES["bg_section"])
        self._adm_lot_frame.pack(fill="x")
        tk.Label(self._adm_lot_frame, text="Loteria:", bg=CORES["bg_section"],
                 fg=CORES["fg_label"], font=("Arial",9,"bold"), anchor="w").pack(fill="x", pady=(8,0))
        self.adm_lot = ttk.Combobox(self._adm_lot_frame, values=LOTERIAS, state="readonly", font=("Arial",9))
        self.adm_lot.set("Mega-Sena"); self.adm_lot.pack(fill="x")
        tk.Label(self._adm_lot_frame, text="Concurso:", bg=CORES["bg_section"],
                 fg=CORES["fg_label"], font=("Arial",9,"bold"), anchor="w").pack(fill="x", pady=(8,0))
        self.adm_conc = entry(self._adm_lot_frame, width=40); self.adm_conc.pack(fill="x")

        lbl_f("Valor (R$):")
        self.adm_val = entry(fr, width=40); self.adm_val.pack(fill="x")
        lbl_f("Data:")
        self.adm_dt = entry(fr, width=40)
        self.adm_dt.insert(0, date.today().strftime("%d/%m/%Y")); self.adm_dt.pack(fill="x")
        lbl_f("Descrição:")
        self.adm_desc = entry(fr, width=40); self.adm_desc.pack(fill="x")

        def _on_tipo_change(e=None):
            if "SAQUE" in self.adm_tipo.get():
                self._adm_bolao_frame.pack_forget(); self._adm_lot_frame.pack_forget()
            else:
                self._adm_bolao_frame.pack(fill="x", after=self.adm_tipo)
                self._adm_lot_frame.pack(fill="x", after=self._adm_bolao_frame)
        self.adm_tipo.bind("<<ComboboxSelected>>", _on_tipo_change)

        btn(win, "💾 SALVAR", CORES["btn_verde"], self._adm_registrar, width=20).pack(pady=14)

    def _dash_load(self):
        bid = self.bid.get()
        self._dash_hora_lbl.configure(
            text=f"Atualizado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        try: self._atualizar_cartoes_bolao()
        except Exception: pass

        if not bid:
            self._dash_bolao_lbl.configure(text="Nenhum bolão selecionado")
            return

        b = self.db.fetchone("SELECT * FROM boloes WHERE id=?", (bid,))
        if not b: return
        bd = dict(b)
        self._dash_bolao_lbl.configure(
            text=f"🎯  {bd['nome']}  —  {bd.get('loteria','Mega-Sena')}")

        partic   = self.db.fetchall(
            "SELECT * FROM participantes WHERE bolao_id=? AND ativo=1 ORDER BY nome", (bid,))
        adm_paga = bd.get("adm_paga", 0)
        adm_nome_lower = bd.get("adm_nome","").strip().lower()
        adm_nome_real  = bd.get("adm_nome","").strip()
        adm_ja_listado = False
        _,parc_esp_d,parc_d = self._calc_parcela_atual(bd)

        total_esp = total_pago = total_saldo = 0
        quitados = pendentes_n = 0
        dados_sit = []

        for pt in partic:
            pt_d = dict(pt)
            row  = self.db.fetchone(
                "SELECT SUM(valor) as t FROM pagamentos WHERE participante_id=? AND bolao_id=?",
                (pt_d["id"], bid))
            pago = row["t"] or 0
            ve   = pt_d["valor_esperado"] or 0

            # Detecta ADM por flag OU por nome
            eh_adm = bool(pt_d.get("is_adm")) or (
                adm_nome_lower and adm_nome_lower in pt_d["nome"].lower())
            if eh_adm:
                adm_ja_listado = True

            status_txt, tag, pago_f, saldo_f = self._status_part_adm(
                {"is_adm": eh_adm}, pago, ve, parc_esp_d, parc_d, adm_paga)

            # ADM isento não entra em NENHUM total financeiro (nem esperado,
            # nem arrecadado, nem pendente) — ele não paga, então o
            # "esperado" real do bolão é só a soma de quem paga de verdade.
            # Antes "Total Esperado" vinha de um SUM(valor_esperado) cru que
            # somava até o ADM isento (o campo só é zerado automaticamente
            # se ele foi cadastrado DEPOIS do bolão já estar marcado como
            # isento; se o bolão virou isento depois, o valor antigo ficava
            # salvo e inflava o total).
            if not (eh_adm and not adm_paga):
                total_esp += ve
            total_pago  += (pago_f if pago_f is not None else 0)
            total_saldo += saldo_f
            if saldo_f <= 0: quitados    += 1
            else:            pendentes_n += 1
            st_show = "✅ Quitado" if "QUITADO" in status_txt else "⚠ Pendente"
            pago_show = pago_f if pago_f is not None else 0
            dados_sit.append((pt_d["nome"], pago_show, saldo_f, st_show, tag))

        # ADM configurado mas não cadastrado como participante → adiciona
        # linha sintética só pra exibição (não entra em nenhum total)
        if adm_nome_real and not adm_ja_listado and not adm_paga:
            quitados += 1
            dados_sit.append((adm_nome_real, 0, 0, "✅ Quitado", "quitado"))

        # ── KPIs ────────────────────────────────────────────────
        # Cotas totais = soma das cotas individuais (quem tem 2 cotas conta 2)
        cotas_ocup, max_cotas = self._get_cotas_ocupadas(bid)
        self._dash_kpis["dash_participantes"].configure(
            text=f"{cotas_ocup}/{max_cotas}")
        self._dash_kpis["dash_quitados"].configure(text=str(quitados))
        self._dash_kpis["dash_pendentes"].configure(text=str(pendentes_n))
        self._dash_kpis["dash_arrecadado"].configure(text=fmt_brl(total_pago))
        self._dash_kpis["dash_pendente_val"].configure(text=fmt_brl(total_saldo))

        tot_dep_row = self.db.fetchone(
            "SELECT SUM(valor) as t FROM pagamentos WHERE bolao_id=? AND depositado=1", (bid,))
        tot_dep = tot_dep_row["t"] or 0
        self._dash_kpis["dash_depositos"].configure(text=fmt_brl(tot_dep))

        # ── Barra progresso ──────────────────────────────────────
        pct = (total_pago / total_esp * 100) if total_esp > 0 else 0
        self._dash_prog["value"] = pct
        self._dash_pct_lbl.configure(
            text=f"{pct:.1f}%  —  {fmt_brl(total_pago)} de {fmt_brl(total_esp)}")

        # ── Situação participantes ───────────────────────────────
        self._dash_tree_sit.delete(*self._dash_tree_sit.get_children())
        for nome, pago, saldo, status, tag in sorted(dados_sit, key=lambda x: x[0]):
            self._dash_tree_sit.insert("","end", tags=(tag,), values=(
                nome, fmt_brl(pago), fmt_brl(saldo), status))

        # ── Últimos 20 pagamentos ────────────────────────────────
        self._dash_tree_ult.delete(*self._dash_tree_ult.get_children())
        ults = self.db.fetchall("""
            SELECT pt.nome, pg.data_pagamento, pg.valor
            FROM pagamentos pg
            JOIN participantes pt ON pg.participante_id = pt.id
            WHERE pg.bolao_id = ?
            ORDER BY pg.id DESC LIMIT 20
        """, (bid,))
        for i, r in enumerate(ults):
            tag = "linha" if i % 2 == 0 else ""
            self._dash_tree_ult.insert("","end", tags=(tag,), values=(
                r["nome"], r["data_pagamento"] or "-", fmt_brl(r["valor"])))

        # ── Bloco financeiro ─────────────────────────────────────
        pend_dep = total_pago - tot_dep
        self._dash_fin_labels["fin_esp"].configure(text=fmt_brl(total_esp))
        self._dash_fin_labels["fin_pct"].configure(text=f"{pct:.1f}%")

        # ── Bloco Situação do Fundo ──────────────────────────────
        td_row = self.db.fetchone(
            "SELECT SUM(valor) as t FROM pagamentos WHERE bolao_id=? AND depositado=1", (bid,))
        tot_dep_fundo = td_row["t"] or 0
        sq_row = self.db.fetchone(
            "SELECT SUM(valor) as t FROM saques_emergenciais WHERE bolao_id=?", (bid,))
        tot_saques = sq_row["t"] or 0
        saldo_conta = tot_dep_fundo - tot_saques
        pend_dep_val = total_pago - tot_dep_fundo

        self._dash_fundo_labels["fundo_dep"].configure(text=fmt_brl(tot_dep_fundo))
        self._dash_fundo_labels["fundo_saques"].configure(
            text=fmt_brl(tot_saques) if tot_saques > 0 else "—")
        self._dash_fundo_labels["fundo_saldo"].configure(
            text=fmt_brl(max(0, saldo_conta)))
        self._dash_fundo_labels["fundo_pend"].configure(
            text=fmt_brl(pend_dep_val) if pend_dep_val > 0 else "✅ Em dia")

        # ── Bloco Informações do Bolão ───────────────────────────
        try:
            di = datetime.strptime(bd.get("data_inicio",""), "%Y-%m-%d")
            inicio_fmt = di.strftime("%d/%m/%Y")
        except:
            inicio_fmt = bd.get("data_inicio","—") or "—"

        parc_val = bd.get("valor_parcela", 0) or 0
        tot_val  = bd.get("valor_total", 0) or 0
        n_parc_total = round(tot_val / parc_val) if parc_val > 0 else 0
        _, parc_atual_n, _ = self._calc_parcela_atual(bd)
        cotas_ocup_d, max_cotas_d = self._get_cotas_ocupadas(bid)

        self._dash_info_labels["info_lot"].configure(
            text=bd.get("loteria","Mega-Sena") or "—")
        self._dash_info_labels["info_inicio"].configure(text=inicio_fmt)
        self._dash_info_labels["info_parc"].configure(text=fmt_brl(parc_val))
        self._dash_info_labels["info_nparc"].configure(text=str(n_parc_total))
        self._dash_info_labels["info_atual"].configure(
            text=f"{parc_atual_n}/{n_parc_total}")
        self._dash_info_labels["info_cotas"].configure(
            text=f"{cotas_ocup_d}/{max_cotas_d}")

        # ── Rodapé ───────────────────────────────────────────────
        self._dash_footer.configure(text=(
            f"Bolão: {bd['nome']}  |  Loteria: {bd.get('loteria','—')}  |  "
            f"Valor da parcela: {fmt_brl(parc_val)}  |  "
            f"Cotas: {cotas_ocup_d}/{max_cotas_d}  |  "
            f"Atualizado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ))

    # ════════════════════════════════════════════════════════════
    #  RECIBO DE QUITAÇÃO (imagem PNG)
    # ════════════════════════════════════════════════════════════
    def _emitir_recibo(self):
        try:
            sel = self.vis_cb.get()
            if not sel:
                messagebox.showwarning("Atencao","Selecione um participante!"); return
            pid = int(re.search(r"\(ID: (\d+)\)", sel).group(1))
            bid = self.bid.get()
            pt  = dict(self.db.fetchone("SELECT * FROM participantes WHERE id=?", (pid,)) or {})
            pgs = self.db.fetchall(
                "SELECT * FROM pagamentos WHERE participante_id=? AND bolao_id=? ORDER BY id",
                (pid, bid))
            b     = dict(self.db.fetchone("SELECT * FROM boloes WHERE id=?", (bid,)) or {})
            pago  = sum(float(x["valor"] or 0) for x in pgs)
            ve    = float(pt["valor_esperado"] or 0)
            saldo = max(0, ve - pago)
            quitado = saldo <= 0
            bd = dict(b)
            import tempfile, webbrowser
    
            rows_pag = ""
            for i, pg in enumerate([dict(x) for x in pgs]):
                bg = "#f0fff8" if i%2==0 else "#ffffff"
                dep_ic = "✓" if pg["depositado"] else "⏳"
                dep_cor = "#1D9E75" if pg["depositado"] else "#e67e22"
                dep_txt = "Depositado" if pg["depositado"] else "Pendente"
                obs = pg.get("observacoes","") or ""
                rows_pag += f"""<tr style="background:{bg}">
                  <td class="tc">{i+1}</td>
                  <td>{pg["data_pagamento"] or "-"}</td>
                  <td style="color:#555">{pg["mes_referencia"] or "-"}</td>
                  <td class="money" style="color:#1D9E75">{fmt_brl(float(pg["valor"] or 0))}</td>
                  <td><span style="color:{dep_cor};font-weight:600">{dep_ic} {dep_txt}</span></td>
                  <td style="color:#777;font-size:11px">{obs[:40] if obs else "-"}</td>
                </tr>"""
    
            cor_st = "#1D9E75" if quitado else "#e74c3c"
            ic_st  = "✅ QUITADO" if quitado else "⚠️ EM ANDAMENTO"
            n_rec  = "REC-"+datetime.now().strftime("%Y")+f"-{pid:04d}"
            ini    = "".join(w[0].upper() for w in (pt["nome"] or "?").split()[:2])
            loteria = bd.get("loteria") or "—"
            concurso = bd.get("concurso") or bd.get("num_concurso") or "—"
    
            html = f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
    <title>Recibo {n_rec}</title>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&display=swap');
    *{{margin:0;padding:0;box-sizing:border-box}}
    body{{font-family:'Inter',Arial,sans-serif;background:#0f2027;background:linear-gradient(135deg,#0f2027,#203a43,#2c5364);min-height:100vh;display:flex;align-items:flex-start;justify-content:center;padding:30px 16px}}
    .wrap{{width:100%;max-width:780px}}
    .card{{background:white;border-radius:20px;overflow:hidden;box-shadow:0 20px 60px rgba(0,0,0,.4)}}
    /* HEADER */
    .hdr{{background:linear-gradient(135deg,#1b5e20,#2e7d32,#388e3c);padding:0;position:relative;overflow:hidden}}
    .hdr-bg{{position:absolute;inset:0;opacity:.1;background:repeating-linear-gradient(45deg,transparent,transparent 20px,rgba(255,255,255,.3) 20px,rgba(255,255,255,.3) 21px)}}
    .hdr-inner{{position:relative;z-index:1;padding:28px 36px;display:flex;justify-content:space-between;align-items:center}}
    .hdr-left .badge{{background:rgba(255,255,255,.15);backdrop-filter:blur(4px);border:1px solid rgba(255,255,255,.25);color:#c8f7dc;font-size:10px;font-weight:700;letter-spacing:2px;text-transform:uppercase;padding:4px 12px;border-radius:20px;display:inline-block;margin-bottom:10px}}
    .hdr-left h1{{color:white;font-size:24px;font-weight:800;line-height:1.2}}
    .hdr-left .sub{{color:#a5d6a7;font-size:12px;margin-top:6px}}
    .hdr-right{{text-align:right}}
    .hdr-right .rec-num{{background:rgba(255,255,255,.15);backdrop-filter:blur(4px);border:1px solid rgba(255,255,255,.3);border-radius:12px;padding:12px 18px}}
    .hdr-right .rec-label{{color:#a5d6a7;font-size:9px;text-transform:uppercase;letter-spacing:1.5px}}
    .hdr-right .rec-val{{color:white;font-size:16px;font-weight:800;margin-top:2px}}
    /* PARTICIPANTE */
    .part{{background:linear-gradient(135deg,#f9fffe,#e8f5e9);padding:20px 36px;display:flex;align-items:center;gap:16px;border-bottom:2px solid #e8f5e9}}
    .avatar{{width:56px;height:56px;border-radius:50%;background:linear-gradient(135deg,#1b5e20,#43a047);color:white;font-size:18px;font-weight:800;display:flex;align-items:center;justify-content:center;flex-shrink:0;box-shadow:0 4px 14px rgba(27,94,32,.35)}}
    .part-info h2{{font-size:17px;font-weight:700;color:#1a1a1a}}
    .part-info .tags{{display:flex;gap:8px;margin-top:6px;flex-wrap:wrap}}
    .tag{{background:#e8f5e9;color:#2e7d32;font-size:11px;font-weight:600;padding:3px 10px;border-radius:20px;border:1px solid #c8e6c9}}
    .tag.blue{{background:#e3f2fd;color:#1565c0;border-color:#bbdefb}}
    /* KPIs */
    .kpis{{display:grid;grid-template-columns:repeat(3,1fr);gap:0;border-bottom:1px solid #f0f0f0}}
    .kpi{{padding:18px 24px;border-right:1px solid #f0f0f0;position:relative;overflow:hidden}}
    .kpi:last-child{{border-right:none}}
    .kpi::before{{content:'';position:absolute;bottom:-20px;right:-20px;width:80px;height:80px;border-radius:50%;opacity:.06}}
    .kpi.k1::before{{background:#1565c0}}
    .kpi.k2::before{{background:#1D9E75}}
    .kpi.k3::before{{background:#e74c3c}}
    .kpi label{{font-size:10px;color:#999;text-transform:uppercase;letter-spacing:1px;font-weight:600}}
    .kpi .kval{{font-size:20px;font-weight:800;margin-top:6px}}
    .kpi.k1 .kval{{color:#1565c0}}
    .kpi.k2 .kval{{color:#1D9E75}}
    .kpi.k3 .kval{{color:{"#1D9E75" if quitado else "#e74c3c"}}}
    /* STATUS */
    .status-bar{{margin:0 36px 20px;border-radius:12px;background:{"linear-gradient(135deg,#e8f5e9,#c8e6c9)" if quitado else "linear-gradient(135deg,#fde8d8,#ffccbc)"};border:2px solid {cor_st};padding:14px 20px;display:flex;justify-content:space-between;align-items:center}}
    .status-bar .st-txt{{font-size:18px;font-weight:800;color:{cor_st}}}
    .status-bar .st-date{{font-size:11px;color:#888}}
    /* TABELA */
    .sec{{padding:0 36px 24px}}
    .sec-title{{font-size:10px;color:#999;font-weight:700;text-transform:uppercase;letter-spacing:1.5px;padding:16px 0 10px}}
    table{{width:100%;border-collapse:collapse;font-size:13px}}
    thead tr{{background:linear-gradient(135deg,#1b5e20,#388e3c)}}
    thead th{{color:white;padding:10px 12px;text-align:left;font-size:11px;font-weight:600;letter-spacing:.5px}}
    tbody tr:hover{{background:#f0fff4!important}}
    td{{padding:9px 12px;border-bottom:1px solid #f5f5f5;vertical-align:middle}}
    .tc{{text-align:center;color:#999;font-weight:600}}
    .money{{font-weight:700}}
    /* FOOTER */
    .footer{{background:linear-gradient(135deg,#1b5e20,#2e7d32);padding:14px 36px;display:flex;justify-content:space-between;align-items:center}}
    .footer span{{color:#a5d6a7;font-size:11px}}
    .footer .brand{{color:white;font-weight:700;font-size:12px}}
    @media print{{
      body{{background:white;padding:0}}
      .card{{box-shadow:none;border-radius:0;max-width:100%}}
      .hdr,.hdr-bg,.status-bar,.kpi::before,thead tr,.footer{{-webkit-print-color-adjust:exact!important;print-color-adjust:exact!important}}
    }}
    </style></head><body>
    <div class="wrap"><div class="card">
      <div class="hdr">
        <div class="hdr-bg"></div>
        <div class="hdr-inner">
          <div class="hdr-left">
            <span class="badge">Recibo de Pagamento</span>
            <h1>Sistema de Gestão<br>de Bolões</h1>
            <p class="sub">📅 {datetime.now().strftime("%d/%m/%Y  às  %H:%M")}</p>
          </div>
          <div class="hdr-right">
            <div class="rec-num">
              <div class="rec-label">Número do Recibo</div>
              <div class="rec-val">{n_rec}</div>
            </div>
          </div>
        </div>
      </div>
      <div class="part">
        <div class="avatar">{ini}</div>
        <div class="part-info">
          <h2>{pt['nome']}</h2>
          <div class="tags">
            <span class="tag">🎯 {bd.get('nome','')}</span>
            <span class="tag blue">🎰 {loteria}</span>
            <span class="tag blue">🔢 Concurso {concurso}</span>
          </div>
        </div>
      </div>
      <div class="kpis">
        <div class="kpi k1"><label>Valor da Cota</label><div class="kval">{fmt_brl(ve)}</div></div>
        <div class="kpi k2"><label>Total Pago</label><div class="kval">{fmt_brl(pago)}</div></div>
        <div class="kpi k3"><label>Saldo Devedor</label><div class="kval">{fmt_brl(saldo)}</div></div>
      </div>
      <div style="height:16px"></div>
      <div class="status-bar">
        <span class="st-txt">{ic_st}</span>
        <span class="st-date">Emitido em {datetime.now().strftime("%d/%m/%Y")}</span>
      </div>
      <div class="sec">
        <div class="sec-title">Histórico de Pagamentos</div>
        <table>
          <thead><tr><th>#</th><th>Data</th><th>Mês Ref.</th><th>Valor</th><th>Depósito</th><th>Observação</th></tr></thead>
          <tbody>{rows_pag}</tbody>
        </table>
      </div>
      <div class="footer">
        <span>Sistema de Gestão de Bolões v5.1</span>
        <span class="brand">✨ Desenvolvido por Elton Luis</span>
      </div>
    </div></div>
    
    </body></html>"""
    
            import os as _os2
            nome_sugerido = (pt.get("nome","recibo") or "recibo").replace(" ","_")
            path_html = filedialog.asksaveasfilename(
                defaultextension=".html",
                filetypes=[("HTML","*.html"),("Todos","*.*")],
                initialfile=nome_sugerido+"_"+datetime.now().strftime("%Y%m%d_%H%M")+".html")
            if not path_html: return
            with open(path_html,"w",encoding="utf-8") as _f2: _f2.write(html)
            webbrowser.open("file:///" + _os2.path.abspath(path_html).replace("\\", "/"))
    
        except Exception as _ex_rec:
            import traceback
            messagebox.showerror("Erro no Recibo",
                "Erro ao gerar recibo:\n\n" + str(_ex_rec) +
                "\n\n" + traceback.format_exc()[-400:])

    # ════════════════════════════════════════════════════════════
    #  ABA 4 — VISUALIZAR / EDITAR DADOS
    # ════════════════════════════════════════════════════════════
    #  ABA IMPORTAR EXTRATO NUBANK
    # ════════════════════════════════════════════════════════════

    def _build_importar(self):
        p = self.tab_import

        # ── Passo 1: carregar PDF ────────────────────────────────
        sec1 = section(p, "📄 PASSO 1 — CARREGAR EXTRATO NUBANK (PDF)")
        sec1.pack(fill="x", padx=20, pady=(10,4))

        r1 = tk.Frame(sec1, bg=CORES["bg_section"]); r1.pack(fill="x", pady=3)
        tk.Label(r1, text="Arquivo PDF:", bg=CORES["bg_section"],
                 fg=CORES["fg_label"], font=("Arial",9,"bold")).pack(side="left")
        self._imp_pdf_path = tk.StringVar()
        tk.Entry(r1, textvariable=self._imp_pdf_path, width=50,
                 state="readonly", font=("Arial",9),
                 relief="solid", bd=1).pack(side="left", padx=6)
        btn(r1, "📂 Selecionar PDF", CORES["btn_azul"],
            self._imp_escolher_pdf, width=16).pack(side="left", padx=4)

        r2 = tk.Frame(sec1, bg=CORES["bg_section"]); r2.pack(fill="x", pady=3)
        tk.Label(r2, text="A partir de:", bg=CORES["bg_section"],
                 fg=CORES["fg_label"], font=("Arial",9,"bold")).pack(side="left")
        self._imp_data_ini = entry(r2, width=12)
        self._imp_data_ini.insert(0, date.today().replace(day=1).strftime("%d/%m/%Y"))
        self._imp_data_ini.pack(side="left", padx=6)
        btn(r2, "🔍 ANALISAR", CORES["btn_verde"],
            self._imp_analisar, width=14).pack(side="left", padx=8)
        self._imp_status = tk.Label(r2, text="", bg=CORES["bg_section"],
                                     fg="#888", font=("Arial",8,"italic"))
        self._imp_status.pack(side="left", padx=8)

        # ── Passo 2: tabela de transações ───────────────────────
        sec2 = section(p, "📋 PASSO 2 — VINCULE E SELECIONE AS TRANSAÇÕES")
        sec2.pack(fill="both", expand=True, padx=20, pady=(4,4))

        leg = tk.Frame(sec2, bg=CORES["bg_section"]); leg.pack(fill="x", pady=(0,4))
        for cor, txt in [("#d5f5e3","🟢 Bolão identificado pelo valor"),
                         ("#fffde7","🟡 Valor ambíguo — revisar"),
                         ("#ffffff","⚪ Não identificado — preencher"),
                         ("#e8f4fd","🔵 Manual")]:
            f = tk.Frame(leg, bg=cor, padx=6, pady=2, relief="solid", bd=1)
            f.pack(side="left", padx=3)
            tk.Label(f, text=txt, bg=cor, font=("Arial",8)).pack()

        cols_imp = {"✓":30,"Data":90,"Valor":90,"Remetente":190,
                    "Participante":170,"Bolão":160,"Tipo":70}
        fr_imp, self._imp_tree = make_tree(sec2, cols_imp, height=10)
        fr_imp.pack(fill="both", expand=True)
        self._imp_tree.tag_configure("match_ok",  background="#d5f5e3")
        self._imp_tree.tag_configure("match_par", background="#fffde7")
        self._imp_tree.tag_configure("match_no",  background="#ffffff")
        self._imp_tree.tag_configure("manual",    background="#e8f4fd")
        self._imp_tree.bind("<ButtonRelease-1>", self._imp_toggle_sel)
        self._imp_tree.bind("<<TreeviewSelect>>", self._imp_sel_para_edicao)

        # Edição inline
        ed = tk.Frame(sec2, bg=CORES["bg_section"]); ed.pack(fill="x", pady=(4,0))
        tk.Label(ed, text="✏ Editar linha:", bg=CORES["bg_section"],
                 fg=CORES["fg_label"], font=("Arial",9,"bold")).pack(side="left")
        tk.Label(ed, text="Participante:", bg=CORES["bg_section"],
                 fg=CORES["fg_label"], font=("Arial",8)).pack(side="left", padx=(8,2))
        self._imp_cb_part = ttk.Combobox(ed, width=28, font=("Arial",9))
        self._imp_cb_part.pack(side="left", padx=4)
        tk.Label(ed, text="Bolão:", bg=CORES["bg_section"],
                 fg=CORES["fg_label"], font=("Arial",8)).pack(side="left", padx=(6,2))
        self._imp_cb_bolao = ttk.Combobox(ed, width=24, font=("Arial",9))
        self._imp_cb_bolao.pack(side="left", padx=4)
        btn(ed, "✔ Salvar", CORES["btn_azul"],
            self._imp_aplicar_vinculo, width=9).pack(side="left", padx=6)

        # ── Passo 3: manual + ações ──────────────────────────────
        sec3 = section(p, "➕ PASSO 3 — MANUAL + AÇÕES")
        sec3.pack(fill="x", padx=20, pady=(0,6))

        tk.Label(sec3, text="Pagamentos manuais entram apenas na mensagem WhatsApp.",
                 bg=CORES["bg_section"], fg="#888", font=("Arial",8,"italic")).pack(anchor="w")

        r3 = tk.Frame(sec3, bg=CORES["bg_section"]); r3.pack(fill="x", pady=4)
        for lbl, attr, w in [("Nome:","_imp_man_nome",20),("Valor:","_imp_man_val",10),
                              ("Data:","_imp_man_dt",12)]:
            tk.Label(r3, text=lbl, bg=CORES["bg_section"], fg=CORES["fg_label"],
                     font=("Arial",9,"bold")).pack(side="left")
            e = entry(r3, width=w); e.pack(side="left", padx=6)
            setattr(self, attr, e)
        self._imp_man_dt.insert(0, date.today().strftime("%d/%m/%Y"))
        tk.Label(r3, text="Bolão:", bg=CORES["bg_section"], fg=CORES["fg_label"],
                 font=("Arial",9,"bold")).pack(side="left")
        self._imp_man_bolao = ttk.Combobox(r3, width=22, font=("Arial",9))
        self._imp_man_bolao.pack(side="left", padx=6)
        btn(r3, "➕ Adicionar", CORES["btn_laranja"],
            self._imp_add_manual, width=12).pack(side="left", padx=8)

        r4 = tk.Frame(sec3, bg=CORES["bg_section"]); r4.pack(fill="x", pady=6)
        btn(r4, "✅ REGISTRAR NOS BOLÕES", CORES["btn_verde"],
            self._imp_importar_pagamentos, width=24).pack(side="left", padx=4)
        btn(r4, "📲 GERAR MENSAGEM", CORES["btn_roxo"],
            self._imp_gerar_whatsapp, width=20).pack(side="left", padx=4)
        btn(r4, "📋 CONFERÊNCIA", CORES["btn_azul"],
            self._flb_conferencia, width=16).pack(side="left", padx=4)
        btn(r4, "🗑 Limpar", CORES["btn_cinza"],
            self._imp_limpar, width=10).pack(side="left", padx=4)

        self._imp_linhas   = []
        self._imp_sel_ids  = set()
        self._imp_vinculos = {}


    # ════════════════════════════════════════════════════════════
    def _build_vis(self):
        p = self.tab_vis
        top=tk.Frame(p,bg=CORES["bg_frame"]); top.pack(fill="x",padx=20,pady=10)
        tk.Label(top,text="Participante:",bg=CORES["bg_frame"],font=("Arial",9,"bold"),
                 fg=CORES["fg_label"]).pack(side="left")
        self.vis_cb=ttk.Combobox(top,width=40,state="readonly",font=("Arial",9))
        self.vis_cb.pack(side="left",padx=8)
        self.vis_cb.bind("<<ComboboxSelected>>",self._vis_sel)
        btn(top,"🔄 Atualizar",CORES["btn_azul"],self._vis_sel,width=14).pack(side="left",padx=4)
        btn(top,"🧾 Emitir Recibo",CORES["btn_verde"],self._emitir_recibo,width=16).pack(side="left",padx=4)

        si=section(p,"DADOS DO PARTICIPANTE"); si.pack(fill="x",padx=20,pady=(0,8))
        self.vis_info=tk.Text(si,height=5,state="disabled",relief="flat",
                              bg=CORES["bg_section"],font=("Arial",9),fg=CORES["fg_label"])
        self.vis_info.pack(fill="x")

        sp=section(p,"HISTÓRICO DE PAGAMENTOS"); sp.pack(fill="both",expand=True,padx=20,pady=(0,10))
        cols={"ID Pag.":70,"Mês Ref.":100,"Data Pag.":130,"Valor (R$)":120,"Depositado":110,"Obs.":200}
        fr,self.vis_tree=make_tree(sp,cols,height=10); fr.pack(fill="both",expand=True)

    def _vis_limpar(self):
        """Limpa o painel de visualização ao trocar de bolão."""
        try:
            self.vis_info.configure(state="normal")
            self.vis_info.delete("1.0","end")
            self.vis_info.configure(state="disabled")
            self.vis_tree.delete(*self.vis_tree.get_children())
        except: pass

    def _vis_sel(self, e=None):
        sel=self.vis_cb.get()
        if not sel: return
        pid=int(re.search(r"\(ID: (\d+)\)",sel).group(1))
        bid=self.bid.get()
        pt=self.db.fetchone("SELECT * FROM participantes WHERE id=?",(pid,))
        pgs=self.db.fetchall("SELECT * FROM pagamentos WHERE participante_id=? AND bolao_id=? ORDER BY id",(pid,bid))
        pago=sum(x["valor"] for x in pgs)
        saldo=(pt["valor_esperado"] or 0)-pago
        status="✅ QUITADO" if saldo<=0 else "⚠ PENDENTE"
        info=(f"Nome: {pt['nome']}  |  Tel: {pt['telefone'] or '-'}  |  PIX: {pt['chave_pix'] or '-'}\n"
              f"Valor Esperado: {fmt_brl(pt['valor_esperado'])}  |  Total Pago: {fmt_brl(pago)}  |  "
              f"Saldo: {fmt_brl(max(0,saldo))}  |  Status: {status}\nObs.: {pt['observacoes'] or '-'}")
        self.vis_info.configure(state="normal")
        self.vis_info.delete("1.0","end"); self.vis_info.insert("1.0",info)
        self.vis_info.configure(state="disabled")
        self.vis_tree.delete(*self.vis_tree.get_children())
        for pg in pgs:
            dep="✅ Sim" if pg["depositado"] else "❌ Não"
            self.vis_tree.insert("","end",iid=str(pg["id"]),values=(
                pg["id"],pg["mes_referencia"],pg["data_pagamento"],
                fmt_brl(pg["valor"]),dep,pg["observacoes"] or "-"))

    def _editar_part(self):
        sel=self.vis_cb.get()
        if not sel: messagebox.showwarning("Atenção","Selecione um participante!"); return
        pid=int(re.search(r"\(ID: (\d+)\)",sel).group(1))
        pt=self.db.fetchone("SELECT * FROM participantes WHERE id=?",(pid,))
        self._form_part(pt)

    def _form_part(self, pt):
        win=tk.Toplevel(self.root); win.title("Editar Participante")
        win.geometry("500x420"); win.configure(bg=CORES["bg_section"]); win.grab_set()
        tk.Label(win,text="EDITAR PARTICIPANTE",bg=CORES["bg_section"],
                 fg=CORES["fg_title"],font=("Arial",12,"bold")).pack(pady=10)
        form=tk.Frame(win,bg=CORES["bg_section"],padx=25); form.pack(fill="both",expand=True)
        fields=[("Nome Completo:","nome"),("Telefone (WhatsApp):","telefone"),
                ("Chave PIX:","chave_pix"),("Valor Total Esperado (R$):","valor_esperado"),
                ("Observações:","observacoes")]
        vars_={}
        for i,(lbl,key) in enumerate(fields):
            tk.Label(form,text=lbl,bg=CORES["bg_section"],fg=CORES["fg_label"],
                     font=("Arial",9,"bold")).grid(row=i*2,column=0,sticky="w",pady=(6,0))
            w=entry(form,width=46)
            # Formata valor_esperado com vírgula para evitar confusão com ponto decimal
            if key == "valor_esperado":
                try:
                    val_fmt = f"{float(pt[key] or 0):.2f}".replace(".", ",")
                except:
                    val_fmt = "0,00"
                w.insert(0, val_fmt)
            else:
                val = pt[key] if pt[key] is not None else ""
                w.insert(0, str(val))
            w.grid(row=i*2+1,column=0,sticky="ew"); vars_[key]=w
        form.columnconfigure(0,weight=1)
        def salvar():
            self.db.execute(
                "UPDATE participantes SET nome=?,telefone=?,chave_pix=?,valor_esperado=?,observacoes=? WHERE id=?",
                (vars_["nome"].get(), vars_["telefone"].get(), vars_["chave_pix"].get(),
                 to_float(vars_["valor_esperado"].get()), vars_["observacoes"].get(), pt["id"]))
            messagebox.showinfo("Sucesso","Participante atualizado!"); win.destroy()
            self._refresh_all(); self._vis_sel()
        btn(form,"💾 SALVAR",CORES["btn_verde"],salvar,width=18).grid(row=20,column=0,pady=14,sticky="e")

    def _remover_part(self):
        sel=self.vis_cb.get()
        if not sel: messagebox.showwarning("Atenção","Selecione um participante!"); return
        pid=int(re.search(r"\(ID: (\d+)\)",sel).group(1))
        pt=self.db.fetchone("SELECT * FROM participantes WHERE id=?",(pid,))
        if messagebox.askyesno("Confirmar",f"Remover '{pt['nome']}'?"):
            self.db.execute("UPDATE participantes SET ativo=0 WHERE id=?",(pid,))
            messagebox.showinfo("Removido","Participante removido."); self._refresh_all()

    # ════════════════════════════════════════════════════════════
    #  [ANTIGA ABA 5 — EDITAR PAGAMENTOS — REMOVIDA]
    #  Era redundante com "📋 Histórico": busca por nome já cobria
    #  "selecionar participante", e duplo-clique já editava. Histórico
    #  ganhou um botão "🗑 Excluir Selecionado" (ver _hist_excluir) pra
    #  fechar a única diferença real que havia entre as duas abas.
    # ════════════════════════════════════════════════════════════

    # ════════════════════════════════════════════════════════════
    #  ABA 6 — CONTROLE DE DEPÓSITOS
    # ════════════════════════════════════════════════════════════
    def _build_dep(self):
        p = self.tab_dep

        # ── Resumo financeiro do fundo ───────────────────────────
        sr = section(p,"RESUMO DO FUNDO"); sr.pack(fill="x",padx=20,pady=(16,6))

        krow = tk.Frame(sr, bg=CORES["bg_section"]); krow.pack(fill="x", pady=(4,8))
        self._dep_cards = {}
        card_defs = [
            ("dep_recebido",  "💰 Total Recebido",       "#2196F3"),
            ("dep_depositado","✅ Total Depositado",      "#1D9E75"),
            ("dep_pendente",  "⏳ Pendente de Depósito", "#e67e22"),
            ("dep_sacado",    "💸 Saques Emergenciais",  "#e74c3c"),
            ("dep_saldo",     "🏦 Saldo Real na Conta",  "#8e44ad"),
        ]
        for attr, titulo, cor in card_defs:
            c = tk.Frame(krow, bg=cor, padx=12, pady=8)
            c.pack(side="left", fill="both", expand=True, padx=4)
            tk.Label(c, text=titulo, bg=cor, fg="white",
                     font=("Arial",8,"bold")).pack(anchor="w")
            lv = tk.Label(c, text="R$ 0,00", bg=cor, fg="white",
                          font=("Arial",13,"bold"))
            lv.pack(anchor="w", pady=(2,0))
            self._dep_cards[attr] = lv

        self._dep_bar = ttk.Progressbar(sr, mode="determinate", maximum=100)
        self._dep_bar.pack(fill="x", pady=(0,4))
        self._dep_bar_lbl = tk.Label(sr, text="", bg=CORES["bg_section"],
                                      font=("Arial",8), fg="#666")
        self._dep_bar_lbl.pack(anchor="e")

        # ── Botões de controle ───────────────────────────────────
        sc = section(p,"CONTROLE DE DEPÓSITOS"); sc.pack(fill="x",padx=20,pady=(0,4))
        row = tk.Frame(sc, bg=CORES["bg_section"]); row.pack(pady=4)
        btn(row,"🔄 Atualizar",CORES["btn_azul"],self._dep_refresh,width=14).pack(side="left",padx=4)
        btn(row,"✅ Depositar Selecionados",CORES["btn_verde"],self._dep_sel,width=22).pack(side="left",padx=4)
        btn(row,"💰 Depositar Todos Pendentes",CORES["btn_teal"],self._dep_todos,width=24).pack(side="left",padx=4)
        btn(row,"📋 Ver Histórico",CORES["btn_roxo"],self._dep_historico,width=16).pack(side="left",padx=4)

        # ── Tabela de pendentes — expande para ocupar todo o espaço disponível ──
        st = section(p,"PAGAMENTOS PENDENTES DE DEPÓSITO")
        st.pack(fill="both", expand=True, padx=20, pady=(0,4))
        cols = {"ID":70,"Participante":210,"Mês Ref.":100,"Data Pag.":130,"Valor (R$)":130}
        fr, self.dep_tree = make_tree(st, cols, height=18)
        fr.pack(fill="both", expand=True)
        self.dep_tree.configure(selectmode="extended")

        # ── Saques Emergenciais — compacto, colapsável, no rodapé ──
        self._saq_expandido = tk.BooleanVar(value=False)

        saq_hdr = tk.Frame(p, bg="#2c3e50", pady=4)
        saq_hdr.pack(fill="x", padx=20, pady=(2,0))

        self._saq_toggle_btn = tk.Button(
            saq_hdr, text="▶  💸 SAQUES EMERGENCIAIS  (clique para expandir)",
            bg="#2c3e50", fg="#aad4f5", font=("Arial",9,"bold"),
            relief="flat", cursor="hand2", anchor="w",
            command=self._saq_toggle)
        self._saq_toggle_btn.pack(fill="x", padx=8)

        # Frame colapsável — começa oculto
        self._saq_body = tk.Frame(p, bg=CORES["bg_section"])
        # Não faz pack ainda — aparece só ao expandir

        # Formulário interno
        form_row = tk.Frame(self._saq_body, bg=CORES["bg_section"])
        form_row.pack(fill="x", padx=10, pady=6)

        tk.Label(form_row, text="Valor (R$):", bg=CORES["bg_section"],
                 font=("Arial",9,"bold"), fg=CORES["fg_label"]).pack(side="left", padx=(0,4))
        self._saq_val = entry(form_row, width=12); self._saq_val.pack(side="left", padx=(0,10))

        tk.Label(form_row, text="Data:", bg=CORES["bg_section"],
                 font=("Arial",9,"bold"), fg=CORES["fg_label"]).pack(side="left", padx=(0,4))
        self._saq_dt = entry(form_row, width=12)
        self._saq_dt.insert(0, date.today().strftime("%d/%m/%Y"))
        self._saq_dt.pack(side="left", padx=(0,10))

        tk.Label(form_row, text="Motivo:", bg=CORES["bg_section"],
                 font=("Arial",9,"bold"), fg=CORES["fg_label"]).pack(side="left", padx=(0,4))
        self._saq_motivo = entry(form_row, width=28); self._saq_motivo.pack(side="left", padx=(0,8))

        btn(form_row, "💸 Registrar Saque", CORES["btn_vermelho"],
            self._saq_registrar, width=18).pack(side="left", padx=4)

        # Tabela de saques — compacta, 4 linhas
        cols_s = {"ID":45,"Data":100,"Valor":110,"Motivo":230,"Reposto":80,"Data Reposição":120}
        fr_s, self.saq_tree = make_tree(self._saq_body, cols_s, height=4)
        fr_s.pack(fill="x", padx=10, pady=(4,0))
        self.saq_tree.tag_configure("reposto",  background="#d5f5e3")
        self.saq_tree.tag_configure("pendente", background="#fde8d8")

        bf_s = tk.Frame(self._saq_body, bg=CORES["bg_section"]); bf_s.pack(fill="x", padx=10, pady=4)
        btn(bf_s,"✅ Marcar Reposto", CORES["btn_verde"],
            self._saq_repor, width=18).pack(side="left", padx=4)
        btn(bf_s,"🗑 Excluir", CORES["btn_vermelho"],
            self._saq_excluir, width=12).pack(side="left", padx=4)

    def _saq_toggle(self):
        """Expande ou colapsa a seção de saques emergenciais."""
        if self._saq_expandido.get():
            self._saq_body.pack_forget()
            self._saq_expandido.set(False)
            self._saq_toggle_btn.configure(
                text="▶  💸 SAQUES EMERGENCIAIS  (clique para expandir)")
        else:
            self._saq_body.pack(fill="x", padx=20, pady=(0,8))
            self._saq_expandido.set(True)
            self._saq_toggle_btn.configure(
                text="▼  💸 SAQUES EMERGENCIAIS  (clique para recolher)")

    def _dep_refresh(self):
        bid = self.bid.get()
        if not bid: return

        # Totais de pagamentos
        tr_row = self.db.fetchone("SELECT SUM(valor) as t FROM pagamentos WHERE bolao_id=?",(bid,))
        td_row = self.db.fetchone("SELECT SUM(valor) as t FROM pagamentos WHERE bolao_id=? AND depositado=1",(bid,))
        tr = tr_row["t"] or 0
        td = td_row["t"] or 0
        pend = tr - td
        pct  = (td/tr*100) if tr>0 else 0

        # Saques emergenciais
        sq_row = self.db.fetchone(
            "SELECT SUM(valor) as t FROM saques_emergenciais WHERE bolao_id=?",(bid,))
        total_sacado = sq_row["t"] or 0
        saldo_real   = td - total_sacado

        # Atualiza cards
        self._dep_cards["dep_recebido"].configure(text=fmt_brl(tr))
        self._dep_cards["dep_depositado"].configure(text=fmt_brl(td))
        self._dep_cards["dep_pendente"].configure(text=fmt_brl(pend))
        self._dep_cards["dep_sacado"].configure(text=fmt_brl(total_sacado))
        self._dep_cards["dep_saldo"].configure(
            text=fmt_brl(saldo_real),
            fg="#ffe082" if saldo_real < td * 0.9 else "white")

        self._dep_bar["value"] = pct
        self._dep_bar_lbl.configure(
            text=f"{pct:.1f}% depositado  |  Pendente de depósito: {fmt_brl(pend)}")

        # Pendentes de depósito
        rows = self.db.fetchall("""
            SELECT pg.id,pt.nome,pg.mes_referencia,pg.data_pagamento,pg.valor
            FROM pagamentos pg JOIN participantes pt ON pg.participante_id=pt.id
            WHERE pg.bolao_id=? AND pg.depositado=0 ORDER BY pg.id
        """,(bid,))
        self.dep_tree.delete(*self.dep_tree.get_children())
        for r in rows:
            self.dep_tree.insert("","end",iid=str(r["id"]),values=(
                r["id"],r["nome"],r["mes_referencia"],r["data_pagamento"],fmt_brl(r["valor"])))

        # Saques emergenciais
        saques = self.db.fetchall(
            "SELECT * FROM saques_emergenciais WHERE bolao_id=? ORDER BY id DESC",(bid,))
        self.saq_tree.delete(*self.saq_tree.get_children())
        for s in saques:
            rep_txt  = "✅ Sim" if s["reposto"] else "⏳ Não"
            rep_data = s["data_reposicao"] or "-"
            tag = "reposto" if s["reposto"] else "pendente"
            self.saq_tree.insert("","end", iid=str(s["id"]), tags=(tag,), values=(
                s["id"], s["data_saque"], fmt_brl(s["valor"]),
                s["motivo"] or "-", rep_txt, rep_data))

    def _saq_registrar(self):
        bid = self.bid.get()
        if not bid: messagebox.showwarning("Atenção","Selecione um bolão!"); return
        v = to_float(self._saq_val.get())
        if v <= 0: messagebox.showwarning("Atenção","Informe o valor do saque!"); return
        motivo = self._saq_motivo.get().strip()
        if not motivo: messagebox.showwarning("Atenção","Informe o motivo do saque!"); return
        dt = self._saq_dt.get().strip() or date.today().strftime("%d/%m/%Y")

        # Verifica saldo disponível
        td_row = self.db.fetchone(
            "SELECT SUM(valor) as t FROM pagamentos WHERE bolao_id=? AND depositado=1",(bid,))
        sq_row = self.db.fetchone(
            "SELECT SUM(valor) as t FROM saques_emergenciais WHERE bolao_id=?",(bid,))
        saldo = (td_row["t"] or 0) - (sq_row["t"] or 0)

        if v > saldo:
            messagebox.showwarning("Saldo Insuficiente",
                f"Saldo disponível na conta: {fmt_brl(saldo)}\n"
                f"Valor do saque: {fmt_brl(v)}\n\n"
                f"Não é possível sacar mais do que o saldo disponível.")
            return

        if not messagebox.askyesno("Confirmar Saque",
            f"⚠ REGISTRAR SAQUE EMERGENCIAL\n\n"
            f"Valor:  {fmt_brl(v)}\n"
            f"Data:   {dt}\n"
            f"Motivo: {motivo}\n\n"
            f"Saldo após o saque: {fmt_brl(saldo - v)}"):
            return

        self.db.execute(
            "INSERT INTO saques_emergenciais (bolao_id,valor,data_saque,motivo,reposto)"
            " VALUES (?,?,?,?,0)",
            (bid, v, dt, motivo))

        messagebox.showinfo("Saque Registrado",
            f"Saque de {fmt_brl(v)} registrado.\n"
            f"Saldo real na conta: {fmt_brl(saldo - v)}")
        self._saq_val.delete(0,"end")
        self._saq_motivo.delete(0,"end")
        self._saq_dt.delete(0,"end")
        self._saq_dt.insert(0, date.today().strftime("%d/%m/%Y"))
        self._dep_refresh()

    def _saq_repor(self):
        sel = self.saq_tree.selection()
        if not sel: messagebox.showwarning("Atenção","Selecione um saque!"); return
        sid = int(sel[0])
        s   = self.db.fetchone("SELECT * FROM saques_emergenciais WHERE id=?",(sid,))
        if s["reposto"]:
            messagebox.showinfo("Info","Este saque já foi marcado como reposto."); return
        if messagebox.askyesno("Confirmar Reposição",
            f"Marcar o saque de {fmt_brl(s['valor'])} ({s['data_saque']}) como REPOSTO?\n\n"
            f"Isso indica que o valor foi devolvido à conta do bolão."):
            self.db.execute(
                "UPDATE saques_emergenciais SET reposto=1, data_reposicao=? WHERE id=?",
                (date.today().strftime("%d/%m/%Y"), sid))
            messagebox.showinfo("Reposto","Saque marcado como reposto!")
            self._dep_refresh()

    def _saq_excluir(self):
        sel = self.saq_tree.selection()
        if not sel: messagebox.showwarning("Atenção","Selecione um saque!"); return
        sid = int(sel[0])
        s   = self.db.fetchone("SELECT * FROM saques_emergenciais WHERE id=?",(sid,))
        if messagebox.askyesno("Confirmar",
            f"Excluir registro do saque de {fmt_brl(s['valor'])} em {s['data_saque']}?"):
            self.db.execute("DELETE FROM saques_emergenciais WHERE id=?",(sid,))
            self._dep_refresh()

    def _dep_sel(self):
        sel=self.dep_tree.selection()
        if not sel: messagebox.showwarning("Atenção","Selecione pagamentos!"); return
        # Calcula valor total dos selecionados
        valor_total=0.0
        nomes_vals=[]
        for iid in sel:
            vals=self.dep_tree.item(iid,"values")
            # vals: (ID, Participante, MesRef, DataPag, Valor)
            v_str=vals[4]  # "R$ 1.234,56"
            v=float(v_str.replace("R$","").replace(".","").replace(",",".").strip())
            valor_total+=v
            nomes_vals.append(f"  • {vals[1]} — {vals[3]} — {v_str}")
        det="\n".join(nomes_vals[:10])+("\n  ..." if len(nomes_vals)>10 else "")
        if not messagebox.askyesno("Confirmar Depósito",
            f"Depositar {len(sel)} pagamento(s)?\n\n"
            f"VALOR TOTAL A DEPOSITAR: {fmt_brl(valor_total)}\n\n{det}"):
            return
        dd=date.today().strftime("%d/%m/%Y")
        for iid in sel:
            self.db.execute("UPDATE pagamentos SET depositado=1,data_deposito=? WHERE id=?",(dd,int(iid)))
        messagebox.showinfo("Sucesso",f"{len(sel)} pagamento(s) depositados!\nTotal: {fmt_brl(valor_total)}")
        self._dep_refresh()

    def _dep_todos(self):
        bid=self.bid.get()
        if not bid: return
        row=self.db.fetchone("SELECT SUM(valor) as t,COUNT(*) as n FROM pagamentos WHERE bolao_id=? AND depositado=0",(bid,))
        vt=row["t"] or 0; n=row["n"] or 0
        if n==0: messagebox.showinfo("Info","Não há pagamentos pendentes de depósito."); return
        if not messagebox.askyesno("Confirmar",
            f"Depositar TODOS os {n} pagamentos pendentes?\n\nVALOR TOTAL: {fmt_brl(vt)}"):
            return
        dd=date.today().strftime("%d/%m/%Y")
        self.db.execute("UPDATE pagamentos SET depositado=1,data_deposito=? WHERE bolao_id=? AND depositado=0",
                        (dd,bid))
        messagebox.showinfo("Sucesso",f"Todos depositados! Total: {fmt_brl(vt)}")
        self._dep_refresh()

    def _dep_historico(self):
        bid=self.bid.get()
        if not bid: return
        win=tk.Toplevel(self.root); win.title("Histórico de Depósitos")
        win.geometry("920x520"); win.configure(bg=CORES["bg_section"])
        tk.Label(win,text="HISTÓRICO COMPLETO DE DEPÓSITOS",bg=CORES["bg_section"],
                 fg=CORES["fg_title"],font=("Arial",12,"bold")).pack(pady=10)
        cols={"ID":60,"Participante":200,"Mês Ref.":100,"Data Pag.":130,
              "Valor":120,"Depositado":100,"Data Depósito":130}
        fr,tree=make_tree(win,cols,height=18); fr.pack(fill="both",expand=True,padx=10,pady=8)
        todos=self.db.fetchall("""
            SELECT pg.id,pt.nome,pg.mes_referencia,pg.data_pagamento,pg.valor,pg.depositado,pg.data_deposito
            FROM pagamentos pg JOIN participantes pt ON pg.participante_id=pt.id
            WHERE pg.bolao_id=? ORDER BY pg.id
        """,(bid,))
        for r in todos:
            dep="✅ Sim" if r["depositado"] else "❌ Não"
            tree.insert("","end",values=(r["id"],r["nome"],r["mes_referencia"],
                r["data_pagamento"],fmt_brl(r["valor"]),dep,r["data_deposito"] or "-"))

    # ════════════════════════════════════════════════════════════
    #  ABA 7 — PREMIAÇÕES  (independente da reserva)
    # ════════════════════════════════════════════════════════════
    def _build_prem(self):
        p = self.tab_prem
        tk.Label(p,text="🏆 REGISTRO DE PREMIAÇÕES",bg=CORES["bg_frame"],
                 fg=CORES["fg_title"],font=("Arial",12,"bold")).pack(pady=(14,4))
        tk.Label(p,text="Registre cada prêmio ganho pelo bolão. Isso não afeta automaticamente a reserva.",
                 bg=CORES["bg_frame"],fg="#555",font=("Arial",9)).pack()

        sec=section(p,"REGISTRAR NOVA PREMIAÇÃO"); sec.pack(fill="x",padx=20,pady=10)

        r1=tk.Frame(sec,bg=CORES["bg_section"]); r1.pack(fill="x",pady=4)
        # Loteria
        tk.Label(r1,text="Loteria:",bg=CORES["bg_section"],font=("Arial",9,"bold"),
                 fg=CORES["fg_label"]).pack(side="left",padx=4)
        self.prem_lot=ttk.Combobox(r1,values=LOTERIAS,width=14,state="readonly",font=("Arial",9))
        self.prem_lot.set("Mega-Sena"); self.prem_lot.pack(side="left",padx=4)
        # Concurso / Data / Valor
        for lbl,attr,w_ in [("Nº Concurso:","prem_conc",10),("Data Sorteio:","prem_dt",12),
                              ("Valor Ganho (R$):","prem_val",14)]:
            tk.Label(r1,text=lbl,bg=CORES["bg_section"],font=("Arial",9,"bold"),
                     fg=CORES["fg_label"]).pack(side="left",padx=4)
            w=entry(r1,width=w_)
            if attr=="prem_dt": w.insert(0,date.today().strftime("%d/%m/%Y"))
            w.pack(side="left",padx=4); setattr(self,attr,w)

        r2=tk.Frame(sec,bg=CORES["bg_section"]); r2.pack(fill="x",pady=4)
        tk.Label(r2,text="Descrição:",bg=CORES["bg_section"],font=("Arial",9,"bold"),
                 fg=CORES["fg_label"]).pack(side="left",padx=4)
        self.prem_desc=entry(r2,width=55); self.prem_desc.pack(side="left",padx=4)

        bf=tk.Frame(sec,bg=CORES["bg_section"]); bf.pack(fill="x",pady=8)
        btn(bf,"🏆 REGISTRAR PREMIAÇÃO",CORES["btn_dourado"],self._reg_prem,width=26).pack(side="right",padx=4)

        sh=section(p,"RESUMO POR LOTERIA"); sh.pack(fill="x",padx=20,pady=(0,6))

        # Cabeçalho de totais geral
        self._prem_tot_lbl=tk.Label(sh,text="",bg=CORES["bg_section"],
                                     font=("Arial",11,"bold"),fg="#1a2a3a")
        self._prem_tot_lbl.pack(anchor="w",pady=(4,2))

        # Tabela resumo por loteria
        cols_r={"Loteria":160,"Qtd Prêmios":100,"Total Ganho":150,"Último Concurso":130,"Última Data":120}
        fr_r,self.prem_tree_sum=make_tree(sh,cols_r,height=5)
        fr_r.pack(fill="x")
        self.prem_tree_sum.tag_configure("lot_row",background="#fef9e7")

        sh2=section(p,"HISTÓRICO DETALHADO DE PREMIAÇÕES — cada prêmio registrado")
        sh2.pack(fill="both",expand=True,padx=20,pady=(4,4))
        cols={"ID":45,"Loteria":110,"Concurso":90,"Data Sorteio":110,
              "Valor Prêmio":130,"Descrição":240,"Data Registro":140}
        fr,self.prem_tree=make_tree(sh2,cols,height=10)
        fr.pack(fill="both",expand=True)
        self.prem_tree.tag_configure("prem_mega",background="#fde8d8")
        self.prem_tree.tag_configure("prem_loto",background="#d5f5e3")
        self.prem_tree.tag_configure("prem_quin",background="#d6eaf8")
        self.prem_tree.tag_configure("prem_outx",background="#fef9e7")

        bf2=tk.Frame(p,bg=CORES["bg_frame"]); bf2.pack(fill="x",padx=20,pady=(0,4))
        btn(bf2,"🗑 Excluir Premiação",CORES["btn_vermelho"],self._del_prem,width=22).pack(side="left",padx=4)

        # ── Calculadora de Rateio ────────────────────────────────
        sc=section(p,"🧮 CALCULADORA DE RATEIO"); sc.pack(fill="x",padx=20,pady=(0,12))
        sc.columnconfigure(1,weight=1)

        tk.Label(sc,text="Valor do Prêmio (R$):",bg=CORES["bg_section"],
                 fg=CORES["fg_label"],font=("Arial",9,"bold")).grid(row=0,column=0,sticky="w",padx=(0,10),pady=6)
        self._rat_val = entry(sc,width=18); self._rat_val.grid(row=0,column=1,sticky="w",pady=6)

        tk.Label(sc,text="Total de Cotas:",bg=CORES["bg_section"],
                 fg=CORES["fg_label"],font=("Arial",9,"bold")).grid(row=1,column=0,sticky="w",padx=(0,10),pady=6)
        rat_cotas_fr = tk.Frame(sc,bg=CORES["bg_section"]); rat_cotas_fr.grid(row=1,column=1,sticky="w",pady=6)
        self._rat_cotas = entry(rat_cotas_fr,width=10); self._rat_cotas.pack(side="left")
        btn(rat_cotas_fr,"↺ Auto",CORES["btn_azul"],self._rat_auto_cotas,width=10).pack(side="left",padx=6)
        tk.Label(rat_cotas_fr,text="(busca cotas reais do bolão)",bg=CORES["bg_section"],
                 fg="#888",font=("Arial",8,"italic")).pack(side="left")

        bf_r=tk.Frame(sc,bg=CORES["bg_section"]); bf_r.grid(row=2,column=0,columnspan=2,sticky="w",pady=8)
        btn(bf_r,"🧮 CALCULAR RATEIO",CORES["btn_verde"],self._calcular_rateio,width=22).pack(side="left")

        self._rat_resultado=tk.Label(sc,text="",bg=CORES["bg_section"],
                                      fg=CORES["fg_title"],font=("Arial",11,"bold"))
        self._rat_resultado.grid(row=3,column=0,columnspan=2,sticky="w",pady=4)

        self._rat_detalhe=tk.Label(sc,text="",bg=CORES["bg_section"],
                                    fg=CORES["fg_label"],font=("Arial",9),justify="left")
        self._rat_detalhe.grid(row=4,column=0,columnspan=2,sticky="w",pady=(0,6))

    def _reg_prem(self):
        bid=self.bid.get()
        if not bid: messagebox.showwarning("Atenção","Selecione um bolão!"); return
        v=to_float(self.prem_val.get())
        if v<=0: messagebox.showwarning("Atenção","Informe o valor do prêmio!"); return
        lot=self.prem_lot.get()
        self.db.execute(
            "INSERT INTO premiacoes (bolao_id,loteria,concurso,data_sorteio,valor_premio,descricao,data_registro)"
            " VALUES (?,?,?,?,?,?,?)",
            (bid,lot,self.prem_conc.get().strip(),self.prem_dt.get().strip(),v,
             self.prem_desc.get().strip(),datetime.now().strftime("%d/%m/%Y %H:%M")))
        messagebox.showinfo("Sucesso",f"Premiação de {fmt_brl(v)} ({lot}) registrada!")
        for w in [self.prem_conc,self.prem_val,self.prem_desc]:
            w.delete(0,"end")
        self.prem_dt.delete(0,"end"); self.prem_dt.insert(0,date.today().strftime("%d/%m/%Y"))
        self._load_prem()

    def _load_prem(self):
        self.prem_tree.delete(*self.prem_tree.get_children())
        self.prem_tree_sum.delete(*self.prem_tree_sum.get_children())

        # Carrega TODAS as premiações (todos os bolões) — aba é geral
        rows = self.db.fetchall("""
            SELECT p.*, b.nome as bolao_nome
            FROM premiacoes p
            LEFT JOIN boloes b ON p.bolao_id = b.id
            ORDER BY p.data_sorteio DESC, p.id DESC
        """)

        # ── Totais gerais ──────────────────────────────────────
        total_geral = sum(r["valor_premio"] for r in rows)
        n_total     = len(rows)
        self._prem_tot_lbl.configure(
            text=f"🏆 Total de prêmios ganhos: {fmt_brl(total_geral)}   |   "
                 f"Premiações registradas: {n_total}   |   "
                 f"Loterias diferentes: {len(set((r['loteria'] or 'Mega-Sena') for r in rows))}"
        )

        # ── Resumo por loteria ─────────────────────────────────
        from collections import defaultdict
        por_lot = defaultdict(lambda: {"total":0.0,"qtd":0,"ultimo_conc":"","ultima_data":""})
        for r in rows:  # já em ordem decrescente de data
            lot = r["loteria"] or "Mega-Sena"
            por_lot[lot]["total"] += r["valor_premio"]
            por_lot[lot]["qtd"]   += 1
            if not por_lot[lot]["ultimo_conc"]:  # pega o mais recente (primeiro do loop desc)
                por_lot[lot]["ultimo_conc"]  = r["concurso"] or "-"
                por_lot[lot]["ultima_data"]  = r["data_sorteio"] or "-"

        for lot in sorted(por_lot.keys()):
            d=por_lot[lot]
            self.prem_tree_sum.insert("","end",tags=("lot_row",),values=(
                lot, d["qtd"], fmt_brl(d["total"]),
                d["ultimo_conc"], d["ultima_data"]))

        # ── Histórico detalhado — SEM iid para evitar conflitos ──
        tag_map = {
            "Mega-Sena":"prem_mega","Lotofácil":"prem_loto",
            "Quina":"prem_quin",
        }
        for r in rows:
            lot = r["loteria"] or "Mega-Sena"
            tag = tag_map.get(lot,"prem_outx")
            # Não usa iid — deixa o Treeview gerar automaticamente
            self.prem_tree.insert("","end",tags=(tag,),values=(
                r["id"],
                lot,
                r["concurso"]    or "-",
                r["data_sorteio"]or "-",
                fmt_brl(r["valor_premio"]),
                r["descricao"]   or "-",
                r["data_registro"]or "-"))

    def _del_prem(self):
        sel=self.prem_tree.selection()
        if not sel: messagebox.showwarning("Atenção","Selecione uma premiação na lista detalhada!"); return
        item_vals = self.prem_tree.item(sel[0],"values")
        if not item_vals: return
        prem_id = int(item_vals[0])
        loteria  = item_vals[1]
        valor    = item_vals[4]
        if messagebox.askyesno("Confirmar",
            f"Excluir premiação ID {prem_id}?\nLoteria: {loteria}  |  Valor: {valor}"):
            self.db.execute("DELETE FROM premiacoes WHERE id=?",(prem_id,))
            self._load_prem()

    def _rat_auto_cotas(self):
        """Preenche automaticamente o campo de cotas com o total real do bolão."""
        bid = self.bid.get()
        if not bid: messagebox.showwarning("Atenção","Selecione um bolão!"); return
        cotas_ocup, max_cotas = self._get_cotas_ocupadas(bid)
        self._rat_cotas.delete(0,"end")
        self._rat_cotas.insert(0, str(cotas_ocup))
        messagebox.showinfo("Cotas carregadas",
            f"Total de cotas ocupadas: {cotas_ocup} de {max_cotas}\n"
            f"Use este valor para o rateio proporcional.")

    def _calcular_rateio(self):
        """Calcula quanto cada cota recebe e lista participantes com seus valores."""
        bid = self.bid.get()
        if not bid: messagebox.showwarning("Atenção","Selecione um bolão!"); return
        try:
            premio = to_float(self._rat_val.get())
            n_cotas = int(self._rat_cotas.get().strip() or 0)
        except:
            messagebox.showwarning("Atenção","Preencha valor e número de cotas!"); return
        if premio <= 0: messagebox.showwarning("Atenção","Informe o valor do prêmio!"); return
        if n_cotas <= 0: messagebox.showwarning("Atenção","Informe o número de cotas!"); return

        por_cota = premio / n_cotas
        b = self.db.fetchone("SELECT valor_total FROM boloes WHERE id=?", (bid,))
        vt = float(b["valor_total"] or 0) if b else 0

        # Resultado principal
        self._rat_resultado.configure(
            text=f"Valor por cota: {fmt_brl(por_cota)}   |   "
                 f"Total: {fmt_brl(premio)}   |   {n_cotas} cotas")

        # Detalhe por participante
        partic = self.db.fetchall(
            "SELECT nome, valor_esperado FROM participantes "
            "WHERE bolao_id=? AND ativo=1 ORDER BY nome", (bid,))
        linhas = []
        for pt in partic:
            ve = float(pt["valor_esperado"] or 0)
            n = round(ve/vt) if vt>0 and ve>0 else 1
            n = max(1,n)
            val_part = por_cota * n
            linhas.append(f"{pt['nome']}: {n} cota(s) → {fmt_brl(val_part)}")
        # Trunca se muitos participantes
        if len(linhas) > 15:
            self._rat_detalhe.configure(
                text="\n".join(linhas[:15]) + f"\n... e mais {len(linhas)-15} participantes")
        else:
            self._rat_detalhe.configure(text="\n".join(linhas))

    # ════════════════════════════════════════════════════════════
    #  ABA 8 — RESERVA / CAIXA  (por loteria, independente)
    # ════════════════════════════════════════════════════════════
    def _build_res(self):
        p = self.tab_res
        tk.Label(p,text="RESERVA / CAIXA POR LOTERIA",bg=CORES["bg_frame"],
                 fg=CORES["fg_title"],font=("Arial",12,"bold")).pack(pady=(14,4))
        tk.Label(p,text="Controle independente de reserva por loteria. Adicione incrementos ou saques conforme desejar.",
                 bg=CORES["bg_frame"],fg="#555",font=("Arial",9)).pack()

        # Card destaque — saldo total
        card_tot = tk.Frame(p, bg="#27ae60", padx=24, pady=14, relief="flat")
        card_tot.pack(fill="x", padx=20, pady=(10,4))
        tk.Label(card_tot, text="SALDO TOTAL DA RESERVA",
                 bg="#27ae60", fg="white", font=("Arial",9,"bold")).pack(anchor="w")
        self._res_saldo_destaque = tk.Label(card_tot, text="R$ 0,00",
                 bg="#27ae60", fg="white", font=("Arial",28,"bold"))
        self._res_saldo_destaque.pack(anchor="w")

        # Resumo geral
        sr=section(p,"SALDO POR LOTERIA"); sr.pack(fill="x",padx=20,pady=(4,10))
        cols_r={"Loteria":160,"Entradas":140,"Saídas":140,"Saldo Atual":150,"Movimentos":100}
        fr,self.res_tree_sum=make_tree(sr,cols_r,height=6); fr.pack(fill="both",expand=True)
        self.res_tree_sum.tag_configure("pos",background="#d5f5e3")
        self.res_tree_sum.tag_configure("neg",background="#fde8d8")
        self._res_total_lbl=tk.Label(sr,text="",bg=CORES["bg_section"],
                                      font=("Arial",10,"bold"),fg=CORES["fg_title"])
        self._res_total_lbl.pack(anchor="w",pady=4)

        # Lançar movimento
        sm=section(p,"LANÇAR MOVIMENTO DE RESERVA"); sm.pack(fill="x",padx=20,pady=8)
        r1=tk.Frame(sm,bg=CORES["bg_section"]); r1.pack(fill="x",pady=4)
        tk.Label(r1,text="Loteria:",bg=CORES["bg_section"],font=("Arial",9,"bold"),
                 fg=CORES["fg_label"]).pack(side="left",padx=4)
        self.res_lot=ttk.Combobox(r1,values=LOTERIAS,width=16,state="readonly",font=("Arial",9))
        self.res_lot.set("Mega-Sena"); self.res_lot.pack(side="left",padx=4)

        tk.Label(r1,text="Tipo:",bg=CORES["bg_section"],font=("Arial",9,"bold"),
                 fg=CORES["fg_label"]).pack(side="left",padx=8)
        self.res_tipo=ttk.Combobox(r1,values=["ENTRADA (incremento)","SAÍDA (uso da reserva)"],
                                    width=24,state="readonly",font=("Arial",9))
        self.res_tipo.set("ENTRADA (incremento)"); self.res_tipo.pack(side="left",padx=4)

        r2=tk.Frame(sm,bg=CORES["bg_section"]); r2.pack(fill="x",pady=4)
        tk.Label(r2,text="Valor (R$):",bg=CORES["bg_section"],font=("Arial",9,"bold"),
                 fg=CORES["fg_label"]).pack(side="left",padx=4)
        self.res_val=entry(r2,width=16); self.res_val.pack(side="left",padx=4)
        tk.Label(r2,text="Data:",bg=CORES["bg_section"],font=("Arial",9,"bold"),
                 fg=CORES["fg_label"]).pack(side="left",padx=8)
        self.res_dt=entry(r2,width=14)
        self.res_dt.insert(0,date.today().strftime("%d/%m/%Y"))
        self.res_dt.pack(side="left",padx=4)

        r3=tk.Frame(sm,bg=CORES["bg_section"]); r3.pack(fill="x",pady=4)
        tk.Label(r3,text="Descrição:",bg=CORES["bg_section"],font=("Arial",9,"bold"),
                 fg=CORES["fg_label"]).pack(side="left",padx=4)
        self.res_desc=entry(r3,width=52); self.res_desc.pack(side="left",padx=4)

        bf=tk.Frame(sm,bg=CORES["bg_section"]); bf.pack(fill="x",pady=8)
        btn(bf,"💾 REGISTRAR MOVIMENTO",CORES["btn_verde"],self._reg_mov_res,width=26).pack(side="right",padx=4)

        # Histórico de movimentos
        sh=section(p,"HISTÓRICO DE MOVIMENTOS DA RESERVA")
        sh.pack(fill="both",expand=True,padx=20,pady=(0,10))
        cols_h={"ID":50,"Loteria":120,"Tipo":100,"Valor":110,"Descrição":220,"Data":120,"Bolão":160}
        fr2,self.res_tree_hist=make_tree(sh,cols_h,height=8); fr2.pack(fill="both",expand=True)
        self.res_tree_hist.tag_configure("entrada",background="#d5f5e3")
        self.res_tree_hist.tag_configure("saida",  background="#fde8d8")

        bf2=tk.Frame(p,bg=CORES["bg_frame"]); bf2.pack(fill="x",padx=20,pady=(0,10))
        btn(bf2,"🗑 Excluir Movimento",CORES["btn_vermelho"],self._del_mov_res,width=22).pack(side="left",padx=4)
        btn(bf2,"🔄 Atualizar",CORES["btn_azul"],self._load_res,width=16).pack(side="left",padx=4)

    def _reg_mov_res(self):
        bid=self.bid.get()
        if not bid: messagebox.showwarning("Atenção","Selecione um bolão!"); return
        v=to_float(self.res_val.get())
        if v<=0: messagebox.showwarning("Atenção","Informe um valor!"); return
        lot=self.res_lot.get()
        tipo_full=self.res_tipo.get()
        tipo="ENTRADA" if "ENTRADA" in tipo_full else "SAÍDA"
        desc=self.res_desc.get().strip()
        dt=self.res_dt.get().strip() or date.today().strftime("%d/%m/%Y")
        self.db.execute(
            "INSERT INTO reserva_caixa (bolao_id,loteria,tipo,valor,descricao,data_movimento)"
            " VALUES (?,?,?,?,?,?)",
            (bid,lot,tipo,v,desc,dt))
        messagebox.showinfo("Sucesso",
            f"{'Entrada' if tipo=='ENTRADA' else 'Saída'} de {fmt_brl(v)} registrada na reserva de {lot}!")
        self.res_val.delete(0,"end")
        self.res_desc.delete(0,"end")
        self._load_res()

    def _saldo_loteria(self, bid, lot):
        """Saldo de uma loteria dentro de um bolão específico."""
        ent = self.db.fetchone(
            "SELECT SUM(valor) as t FROM reserva_caixa WHERE bolao_id=? AND loteria=? AND tipo='ENTRADA'",
            (bid, lot))
        sai = self.db.fetchone(
            "SELECT SUM(valor) as t FROM reserva_caixa WHERE bolao_id=? AND loteria=? AND tipo='SAÍDA'",
            (bid, lot))
        return (ent["t"] or 0) - (sai["t"] or 0)

    def _load_res(self):
        self.res_tree_sum.delete(*self.res_tree_sum.get_children())
        self.res_tree_hist.delete(*self.res_tree_hist.get_children())

        # Resumo por loteria — TODOS os bolões
        lots = self.db.fetchall(
            "SELECT DISTINCT loteria FROM reserva_caixa ORDER BY loteria")
        total_geral = 0
        for row in lots:
            lot = row["loteria"]
            ent = self.db.fetchone(
                "SELECT SUM(valor) as t FROM reserva_caixa WHERE loteria=? AND tipo='ENTRADA'", (lot,))
            sai = self.db.fetchone(
                "SELECT SUM(valor) as t FROM reserva_caixa WHERE loteria=? AND tipo='SAÍDA'", (lot,))
            n   = self.db.fetchone(
                "SELECT COUNT(*) as n FROM reserva_caixa WHERE loteria=?", (lot,))
            e = ent["t"] or 0; s = sai["t"] or 0; saldo = e - s
            total_geral += saldo
            tag = "pos" if saldo >= 0 else "neg"
            self.res_tree_sum.insert("","end", tags=(tag,), values=(
                lot, fmt_brl(e), fmt_brl(s), fmt_brl(saldo), n["n"]))

        self._res_total_lbl.configure(
            text="SALDO TOTAL (todas as loterias): " + fmt_brl(total_geral))
        cor = "#27ae60" if total_geral >= 0 else "#e74c3c"
        try:
            self._res_saldo_destaque.configure(text=fmt_brl(total_geral), bg=cor)
            for w in self._res_saldo_destaque.master.winfo_children():
                try: w.configure(bg=cor)
                except: pass
            self._res_saldo_destaque.master.configure(bg=cor)
        except Exception:
            pass

        # Histórico completo — TODOS os bolões com nome do bolão
        hist = self.db.fetchall("""
            SELECT r.*, b.nome as bolao_nome
            FROM reserva_caixa r
            LEFT JOIN boloes b ON r.bolao_id = b.id
            ORDER BY r.id DESC
        """)
        for r in hist:
            tag   = "entrada" if r["tipo"] == "ENTRADA" else "saida"
            sinal = "➕" if r["tipo"] == "ENTRADA" else "➖"
            bolao_txt = r["bolao_nome"] or "-"
            self.res_tree_hist.insert("","end", tags=(tag,), iid=str(r["id"]), values=(
                r["id"], r["loteria"], sinal+" "+r["tipo"],
                fmt_brl(r["valor"]), r["descricao"] or "-",
                r["data_movimento"], bolao_txt))

    def _del_mov_res(self):
        sel=self.res_tree_hist.selection()
        if not sel: messagebox.showwarning("Atenção","Selecione um movimento!"); return
        if messagebox.askyesno("Confirmar","Excluir este movimento da reserva?"):
            self.db.execute("DELETE FROM reserva_caixa WHERE id=?",(int(sel[0]),))
            self._load_res()

    # ════════════════════════════════════════════════════════════
    #  ABA — ADMINISTRAÇÃO (taxa de organização)
    # ════════════════════════════════════════════════════════════
    def _build_adm(self):
        # A "Visão Geral" (KPIs de ganhos, registrar lançamento, ganhos por
        # loteria, histórico) foi fundida na tela do Dashboard
        # (_build_dashboard) — essa função agora só monta "Pendências por
        # Bolão", que virou aba própria dentro de "Início" em vez de ficar
        # 2 níveis de abas para dentro.
        p = self.tab_pend
        p.configure(bg="#1a2a3a")
        tab_adm_pend = p

        ctrl = tk.Frame(tab_adm_pend, bg="#1a2a3a"); ctrl.pack(fill="x", padx=16, pady=(12,4))
        self._pend_mes_lbl = tk.Label(ctrl, text="", bg="#1a2a3a",
                                       fg="#aad4f5", font=("Arial",10,"bold"))
        self._pend_mes_lbl.pack(side="left")
        btn(ctrl, "Atualizar",        CORES["btn_azul"],    self._adm_load,           width=12).pack(side="right", padx=4)
        btn(ctrl, "Reativar Bolao",   CORES["btn_cinza"],   self._adm_reativar_bolao, width=16).pack(side="right", padx=4)
        btn(ctrl, "Encerrar Bolao",   CORES["btn_vermelho"],self._adm_encerrar_bolao, width=16).pack(side="right", padx=4)
        self._pend_nb = ttk.Notebook(tab_adm_pend, style="Inner.TNotebook")
        self._pend_nb.pack(fill="both", expand=True, padx=8, pady=(0,8))

    def _adm_registrar(self):
        tipo_full = self.adm_tipo.get()
        tipo = "GANHO" if "GANHO" in tipo_full else "SAQUE"

        # SAQUE: bolão não é obrigatório — desconta do saldo geral
        if tipo == "GANHO":
            sel_b = self.adm_cb_bolao.get()
            if not sel_b:
                messagebox.showwarning("Atenção", "Selecione o bolão para o GANHO!"); return
            m = re.search(r"\(ID: (\d+)\)", sel_b)
            if m:
                bid_adm = int(m.group(1))
            elif "Outros" in sel_b or "independente" in sel_b:
                bid_adm = 0
            else:
                messagebox.showwarning("Atenção", "Selecione um bolão válido!"); return
        else:
            bid_adm = 0  # saque sem vínculo com bolão específico

        v = to_float(self.adm_val.get())
        if v <= 0:
            messagebox.showwarning("Atenção", "Informe um valor!"); return

        # Verificar saldo total disponível para saque
        if tipo == "SAQUE":
            g_all = self.db.fetchone("SELECT SUM(valor_ganho) as t FROM taxa_adm WHERE tipo='GANHO'")
            s_all = self.db.fetchone("SELECT SUM(valor_sacado) as t FROM taxa_adm WHERE tipo='SAQUE'")
            saldo_disp = (g_all["t"] or 0) - (s_all["t"] or 0)
            if v > saldo_disp:
                messagebox.showwarning("Atenção",
                    f"Saldo insuficiente!\nDisponível: {fmt_brl(saldo_disp)}"); return

        lot   = self.adm_lot.get()
        conc  = self.adm_conc.get().strip()
        desc  = self.adm_desc.get().strip()
        dt    = self.adm_dt.get().strip() or date.today().strftime("%d/%m/%Y")
        vg    = v if tipo == "GANHO" else 0
        vs    = v if tipo == "SAQUE" else 0

        self.db.execute(
            "INSERT INTO taxa_adm (bolao_id,loteria,concurso,valor_ganho,valor_sacado,"
            "descricao,data_registro,tipo) VALUES (?,?,?,?,?,?,?,?)",
            (bid_adm, lot, conc, vg, vs, desc, dt, tipo))

        emoji = "💵" if tipo == "GANHO" else "💸"
        messagebox.showinfo("Registrado",
            f"{emoji} {tipo} de {fmt_brl(v)} registrado!\nLoteria: {lot}  |  Concurso: {conc or '-'}")
        for w in [self.adm_val, self.adm_conc, self.adm_desc]:
            w.delete(0, "end")
        self.adm_dt.delete(0, "end")
        self.adm_dt.insert(0, date.today().strftime("%d/%m/%Y"))
        self._adm_load()


    def _adm_load(self):
        # O combo de bolões do popup "Registrar Lançamento" e a tabela de
        # ganhos por loteria do popup próprio são montados só quando os
        # respectivos popups abrem (ver _abrir_registrar_lancamento /
        # _abrir_ganhos_por_loteria) — não existem mais nessa tela.

        # Todos os lançamentos — COALESCE exibe "Outros" quando bolao_id=0
        todos = self.db.fetchall("""
            SELECT t.*,
                   COALESCE(b.nome, '📦 Outros (independente)') as bolao_nome
            FROM taxa_adm t
            LEFT JOIN boloes b ON t.bolao_id = b.id AND t.bolao_id > 0
            ORDER BY t.id DESC
        """)

        total_ganho  = sum(r["valor_ganho"]  for r in todos)
        total_sacado = sum(r["valor_sacado"] for r in todos)
        saldo_geral  = total_ganho - total_sacado

        self._adm_kpis["adm_total"].configure(text=fmt_brl(total_ganho))
        self._adm_kpis["adm_sacado"].configure(text=fmt_brl(total_sacado))
        self._adm_kpis["adm_saldo"].configure(text=fmt_brl(saldo_geral))

        # ── KPIs gerais (todos os bolões) ─────────────────────────
        n_boloes_ativos = self.db.fetchone(
            "SELECT COUNT(*) as t FROM boloes WHERE encerrado=0")["t"] or 0
        self._adm_kpis["geral_boloes"].configure(text=str(n_boloes_ativos))

        arrec_geral_row = self.db.fetchone("SELECT SUM(valor) as t FROM pagamentos")
        arrec_geral = arrec_geral_row["t"] or 0
        self._adm_kpis["geral_arrec"].configure(text=fmt_brl(arrec_geral))

        # ── Participantes atrasados — TODOS os bolões ────────────
        self.adm_tree_atr.delete(*self.adm_tree_atr.get_children())
        atrasados = []

        todos_boloes = self.db.fetchall("SELECT * FROM boloes")
        for bol in todos_boloes:
            bd_loop = dict(bol)
            _, parc_devidas, parc_val = self._calc_parcela_atual(bd_loop)
            if parc_devidas <= 0: continue
            if parc_val <= 0: continue

            partic_bol = self.db.fetchall(
                "SELECT * FROM participantes WHERE bolao_id=? AND ativo=1",
                (bol["id"],))
            adm_nome_low = (bd_loop.get("adm_nome","") or "").strip().lower()
            adm_paga     = bd_loop.get("adm_paga", 0)

            for pt in partic_bol:
                pt_d = dict(pt)
                eh_adm = bool(pt_d.get("is_adm")) or (
                    adm_nome_low and adm_nome_low in pt_d["nome"].lower())
                if eh_adm and not adm_paga: continue

                pago_row = self.db.fetchone(
                    "SELECT SUM(valor) as t FROM pagamentos "
                    "WHERE participante_id=? AND bolao_id=?",
                    (pt_d["id"], bol["id"]))
                pago = pago_row["t"] or 0

                n_pagas = round(pago / parc_val) if parc_val > 0 else 0
                if n_pagas < parc_devidas:
                    saldo_dev = (parc_devidas - n_pagas) * parc_val
                    atrasados.append({
                        "nome":      pt_d["nome"],
                        "bolao":     bd_loop["nome"],
                        "pagas":     n_pagas,
                        "devidas":   parc_devidas,
                        "saldo":     saldo_dev,
                        "gravidade": parc_devidas - n_pagas,
                    })

        # Ordena: mais atrasados primeiro
        atrasados.sort(key=lambda x: (-x["gravidade"], x["bolao"], x["nome"]))
        for a in atrasados:
            tag = "atr2" if a["gravidade"] >= 2 else "atr1"
            # Mostra "gravidade" (quanto REALMENTE falta), não "devidas"
            # (o acumulado esperado até hoje) — ver comentário na criação
            # da coluna "Faltam" em _build_dashboard.
            self.adm_tree_atr.insert("","end", tags=(tag,), values=(
                a["nome"], a["bolao"],
                a["pagas"], a["gravidade"], fmt_brl(a["saldo"])))

        n_atr = len(atrasados)
        self._adm_atr_lbl.configure(
            text=f"{'⚠' if n_atr else '✅'} "
                 f"{n_atr} participante(s) com parcelas em atraso")
        self._adm_kpis["geral_atrasados"].configure(
            text=f"{'✅' if n_atr == 0 else '⚠'} {n_atr}")

        # Contagem de atrasados por bolão — alimenta o badge dos cartões
        # de seleção (_atualizar_cartoes_bolao), pra dar uma pista do
        # status de cada bolão sem precisar clicar nele.
        self._atrasados_count_por_bolao = {bol["nome"]: 0 for bol in todos_boloes}
        for a in atrasados:
            self._atrasados_count_por_bolao[a["bolao"]] = \
                self._atrasados_count_por_bolao.get(a["bolao"], 0) + 1

        # ── Depósitos pendentes — todos os bolões ativos ─────────
        self.adm_tree_dep.delete(*self.adm_tree_dep.get_children())
        pend_rows = self.db.fetchall("""
            SELECT pg.id, pt.nome, pg.data_pagamento, pg.valor, b.nome as bolao_nome
            FROM pagamentos pg
            JOIN participantes pt ON pg.participante_id = pt.id
            JOIN boloes b ON pg.bolao_id = b.id
            WHERE pg.depositado = 0
            ORDER BY pg.id DESC
        """)
        total_pend = 0
        for i, r in enumerate(pend_rows):
            tag = "row1" if i % 2 == 0 else "row2"
            total_pend += r["valor"]
            self.adm_tree_dep.insert("","end", tags=(tag,), values=(
                r["id"], r["nome"],
                r["data_pagamento"] or "-",
                fmt_brl(r["valor"]),
                r["bolao_nome"] or "-"))
        self._adm_dep_total_lbl.configure(
            text=f"Total pendente: {fmt_brl(total_pend)}  ({len(pend_rows)} pagamentos)")
        self._adm_kpis["geral_pend_dep"].configure(
            text=f"{'✅' if total_pend == 0 else '⚠'} {fmt_brl(total_pend)}")

        # Atualiza os cartões de seleção de bolão com os badges recém-
        # calculados acima (contagem de atrasados por bolão).
        try: self._atualizar_cartoes_bolao()
        except Exception: pass

        # ── Últimos pagamentos — TODOS os bolões ─────────────────
        self.geral_tree_ult.delete(*self.geral_tree_ult.get_children())
        ults_geral = self.db.fetchall("""
            SELECT pt.nome, pg.data_pagamento, pg.valor, b.nome as bolao_nome
            FROM pagamentos pg
            JOIN participantes pt ON pg.participante_id = pt.id
            JOIN boloes b ON pg.bolao_id = b.id
            ORDER BY pg.id DESC LIMIT 20
        """)
        for i, r in enumerate(ults_geral):
            tag = "linha" if i % 2 == 0 else ""
            self.geral_tree_ult.insert("","end", tags=(tag,), values=(
                r["nome"], r["bolao_nome"] or "-",
                r["data_pagamento"] or "-", fmt_brl(r["valor"])))

        # Painel de pendencias mensais — usa _calc_parcela_atual
        from datetime import date as _date
        mes_atual = _date.today().strftime("%m/%Y")
        self._pend_mes_lbl.configure(
            text="Parcela de " + mes_atual + "  |  Vermelho=pendente  |  Verde=em dia/quitado")

        for tab in self._pend_nb.tabs():
            self._pend_nb.forget(tab)

        boloes_ativos = self.db.fetchall(
            "SELECT * FROM boloes WHERE encerrado=0 ORDER BY nome")

        for bol in boloes_ativos:
            bd = dict(bol)
            bid_p = bd["id"]
            vt = float(bd.get("valor_total") or 0)
            if vt <= 0: continue

            _, parc_esp, parc_val = self._calc_parcela_atual(bd)

            partic_bol = self.db.fetchall(
                "SELECT * FROM participantes WHERE bolao_id=? AND ativo=1 ORDER BY nome",
                (bid_p,))
            adm_nome_low = (bd.get("adm_nome","") or "").strip().lower()
            adm_paga = bd.get("adm_paga", 0)

            pag_rows = self.db.fetchall(
                "SELECT participante_id, SUM(valor) as t FROM pagamentos "
                "WHERE bolao_id=? GROUP BY participante_id", (bid_p,))
            pag_map = {r["participante_id"]: float(r["t"] or 0) for r in pag_rows}

            pendentes = []; em_dia = []
            for pt in partic_bol:
                pt_d = dict(pt)
                eh_adm = bool(pt_d.get("is_adm")) or (
                    adm_nome_low and adm_nome_low in pt_d["nome"].lower())
                if eh_adm and not adm_paga: continue
                ve    = float(pt_d.get("valor_esperado") or 0)
                pago  = pag_map.get(pt_d["id"], 0)
                n_cotas = max(1, round(ve / vt)) if vt > 0 and ve > 0 else 1
                val_esp_agora = parc_esp * parc_val * n_cotas
                status, _ = self._status_part(pago, ve, parc_esp, parc_val * n_cotas)
                if "PENDENTE" in status:
                    pendentes.append((pt_d["nome"], pago, max(0, val_esp_agora - pago), n_cotas))
                else:
                    em_dia.append((pt_d["nome"], pago, n_cotas, status))

            n_pend = len(pendentes)
            tab_label = bd["nome"][:18] + " (" + str(n_pend) + " pend)" if n_pend > 0 else bd["nome"][:22] + " OK"

            tab_frame = tk.Frame(self._pend_nb, bg="#1a2a3a")
            self._pend_nb.add(tab_frame, text=tab_label)
            cols_p = {"Participante":200,"Total Pago":110,"Falta":110,"Cotas":55,"Situacao":130}
            fr_p, tv_p = make_tree(tab_frame, cols_p, height=10)
            fr_p.pack(fill="both", expand=True, padx=4, pady=4)
            tv_p.tag_configure("pend",  background="#fde8d8", foreground="#8b1a1a")
            tv_p.tag_configure("emdia", background="#d5f5e3", foreground="#1a5c2a")
            tv_p.tag_configure("quit",  background="#afffca", foreground="#0a3a1a")

            for nome, pago, falta, nc in sorted(pendentes, key=lambda x: x[0]):
                tv_p.insert("","end", iid=bd["nome"]+"|"+nome, tags=("pend",), values=(
                    nome, fmt_brl(pago), fmt_brl(falta), str(nc)+"x", "PENDENTE"))

            for nome, pago, nc, st in sorted(em_dia, key=lambda x: x[0]):
                tag = "quit" if "QUITADO" in st else "emdia"
                tv_p.insert("","end", tags=(tag,), values=(
                    nome, fmt_brl(pago), "—", str(nc)+"x", st))

            rodape = tk.Frame(tab_frame, bg="#1a2a3a"); rodape.pack(fill="x", padx=4, pady=(0,4))
            tk.Label(rodape,
                text=str(n_pend)+" pendente(s)  |  "+str(len(em_dia))+" em dia/quitado(s)  |  "
                     "Parcelas esperadas: "+str(parc_esp)+"  |  Duplo clique para registrar pagamento",
                bg="#1a2a3a", fg="#aad4f5", font=("Arial",8)).pack(side="left")

            # Duplo clique para registrar pagamento
            _bid_pend = bid_p; _vt_pend = vt
            def _reg_pend(event, tv=tv_p, bid_=_bid_pend, vt_=_vt_pend):
                sel = tv.selection()
                if not sel: return
                iid = sel[0]
                vals = tv.item(iid, "values")
                if not vals or "PENDENTE" not in str(vals[4]): return
                nome_p = vals[0]
                # Busca participante
                pt = self.db.fetchone(
                    "SELECT id FROM participantes WHERE bolao_id=? AND ativo=1 AND nome=?",
                    (bid_, nome_p))
                if not pt: return
                pid_p = pt["id"]
                falta_str = vals[2].replace("R$","").replace(".","").replace(",",".").strip()
                try: falta_v = float(falta_str)
                except: falta_v = vt_

                # Mini popup de registro
                from datetime import date as _d2
                win = tk.Toplevel(self.root); win.title("Registrar Pagamento")
                win.geometry("400x220"); win.configure(bg=CORES["bg_section"])
                win.grab_set(); win.lift()
                tk.Label(win, text="Registrar pagamento — " + nome_p,
                         bg=CORES["bg_section"], fg="white",
                         font=("Arial",10,"bold")).pack(pady=(14,8), padx=16, anchor="w")
                r_ = tk.Frame(win, bg=CORES["bg_section"]); r_.pack(fill="x", padx=16, pady=4)
                tk.Label(r_, text="Valor (R$):", bg=CORES["bg_section"],
                         fg=CORES["fg_label"], font=("Arial",9,"bold")).pack(side="left")
                e_val = entry(r_, width=14); e_val.pack(side="left", padx=8)
                e_val.insert(0, "{:.2f}".format(falta_v).replace(".",","))
                tk.Label(r_, text="Data:", bg=CORES["bg_section"],
                         fg=CORES["fg_label"], font=("Arial",9,"bold")).pack(side="left")
                e_dt = entry(r_, width=12); e_dt.pack(side="left", padx=4)
                e_dt.insert(0, _d2.today().strftime("%d/%m/%Y"))
                r2_ = tk.Frame(win, bg=CORES["bg_section"]); r2_.pack(fill="x", padx=16, pady=4)
                tk.Label(r2_, text="Obs.:", bg=CORES["bg_section"],
                         fg=CORES["fg_label"], font=("Arial",9,"bold")).pack(side="left")
                e_obs = entry(r2_, width=32); e_obs.pack(side="left", padx=8)
                def _confirmar():
                    from datetime import datetime as _dt3
                    try: v_pg = float(e_val.get().replace(",","."))
                    except: messagebox.showerror("Erro","Valor inválido"); return
                    dt_raw = e_dt.get().strip()
                    try: mes_ref = _dt3.strptime(dt_raw,"%d/%m/%Y").strftime("%m/%Y")
                    except: mes_ref = ""
                    self.db.execute(
                        "INSERT INTO pagamentos (participante_id,bolao_id,mes_referencia,"
                        "valor,data_pagamento,depositado,observacoes) VALUES (?,?,?,?,?,0,?)",
                        (pid_p, bid_, mes_ref, v_pg, dt_raw, e_obs.get() or "Registrado via painel"))
                    win.destroy(); self._adm_load()
                    messagebox.showinfo("OK", "Pagamento de " + nome_p + " registrado!")
                bf_ = tk.Frame(win, bg=CORES["bg_section"]); bf_.pack(pady=12)
                btn(bf_, "Registrar", CORES["btn_verde"], _confirmar, width=14).pack(side="left", padx=6)
                btn(bf_, "Cancelar",  CORES["btn_cinza"], win.destroy, width=10).pack(side="left", padx=6)

            tv_p.bind("<Double-1>", _reg_pend)

        encerrados = self.db.fetchall("SELECT * FROM boloes WHERE encerrado=1 ORDER BY nome")
        if encerrados:
            tab_enc = tk.Frame(self._pend_nb, bg="#1a2a3a")
            self._pend_nb.add(tab_enc, text="Encerrados ("+str(len(encerrados))+")")
            fr_e, tv_e = make_tree(tab_enc, {"Bolao":300,"Status":150}, height=8)
            fr_e.pack(fill="both", expand=True, padx=4, pady=4)
            tv_e.tag_configure("enc", background="#f0f0f0", foreground="#666")
            for bol in encerrados:
                tv_e.insert("","end", tags=("enc",), iid=str(bol["id"]),
                             values=(dict(bol)["nome"], "Encerrado"))
            self._pend_enc_tree = tv_e

        # Histórico completo
        self.adm_tree_hist.delete(*self.adm_tree_hist.get_children())
        for r in todos:
            tag = "ganho" if r["tipo"] == "GANHO" else "saque"
            val = r["valor_ganho"] if r["tipo"] == "GANHO" else r["valor_sacado"]
            self.adm_tree_hist.insert("","end", values=(
                r["id"], r["bolao_nome"] or "-", r["loteria"] or "-",
                r["concurso"] or "-", fmt_brl(val),
                r["tipo"], r["descricao"] or "-", r["data_registro"] or "-"),
                tags=(tag,))

    def _adm_encerrar_bolao(self):
        """Encerra o bolão da aba ativa no painel de pendências."""
        bid = self._pend_nb_bolao_id()
        if not bid:
            messagebox.showwarning("Atenção",
                "Nenhum bolão ativo selecionado.\n"
                "Clique na aba do bolão que deseja encerrar e tente novamente."); return
        b = self.db.fetchone("SELECT nome FROM boloes WHERE id=?", (bid,))
        if not b: return
        if messagebox.askyesno("Encerrar Bolão",
            "Encerrar o bolão '" + b["nome"] + "'?\n\n"
            "Ele desaparecerá de todas as listas de seleção.\n"
            "Você poderá reativá-lo a qualquer momento."):
            self.db.execute("UPDATE boloes SET encerrado=1 WHERE id=?", (bid,))
            messagebox.showinfo("Encerrado", "Bolão '" + b["nome"] + "' encerrado.")
            self._refresh_all()

    def _pend_nb_bolao_id(self):
        """Retorna o ID do bolão da aba ativa no notebook de pendências."""
        try:
            tab_idx = self._pend_nb.index(self._pend_nb.select())
            tab_text = self._pend_nb.tab(tab_idx, "text")
            # Aba de encerrados não tem bolão ativo
            if "Encerrado" in tab_text or "ncerrad" in tab_text:
                return None
            # Busca bolão pelo nome (remove badge de pendentes)
            import re as _re2
            nome_limpo = _re2.sub(r"\s*\(.*\)$|\s*OK$", "", tab_text).strip()
            b = self.db.fetchone(
                "SELECT id FROM boloes WHERE encerrado=0 AND nome LIKE ?",
                ("%" + nome_limpo[:15] + "%",))
            return b["id"] if b else None
        except Exception:
            return None

    def _adm_reativar_bolao(self):
        """Reativa um bolão encerrado — volta a aparecer nas listas."""
        # Verifica se tem a tree de encerrados visível
        try:
            sel = self._pend_enc_tree.selection()
            if not sel:
                messagebox.showwarning("Atenção","Selecione um bolão na lista de Encerrados."); return
            bid = int(sel[0])
            b = self.db.fetchone("SELECT nome FROM boloes WHERE id=?", (bid,))
            if not b: return
            if messagebox.askyesno("Reativar Bolão",
                "Reativar o bolão '" + b["nome"] + "'?\n\n"
                "Ele voltará a aparecer em todas as listas."):
                self.db.execute("UPDATE boloes SET encerrado=0 WHERE id=?", (bid,))
                messagebox.showinfo("Reativado", "Bolão '" + b["nome"] + "' reativado.")
                self._refresh_all()
        except AttributeError:
            messagebox.showinfo("Sem encerrados",
                "Nenhum bolão encerrado encontrado.\n"
                "Para encerrar, selecione um bolão no campo Bolão e clique em Encerrar.")

    def _adm_editar(self):
        sel = self.adm_tree_hist.selection()
        if not sel:
            messagebox.showwarning("Atenção","Selecione um lançamento no histórico!"); return
        vals = self.adm_tree_hist.item(sel[0],"values")
        if not vals: return
        rid = int(vals[0])
        reg = self.db.fetchone("SELECT * FROM taxa_adm WHERE id=?",(rid,))
        if not reg: return

        win = tk.Toplevel(self.root)
        win.title(f"Editar Lançamento ADM — ID {rid}")
        win.geometry("520x400")
        win.configure(bg="#243447")
        win.grab_set()

        tk.Label(win, text=f"✏  EDITAR LANÇAMENTO  (ID: {rid})",
                 bg="#243447", fg="white",
                 font=("Arial",12,"bold")).pack(pady=14)

        form = tk.Frame(win, bg="#243447", padx=28); form.pack(fill="both", expand=True)
        form.columnconfigure(1, weight=1)

        def lbl_e(r, texto):
            tk.Label(form, text=texto, bg="#243447", fg="#aad4f5",
                     font=("Arial",9,"bold")).grid(row=r, column=0, sticky="w", pady=7, padx=(0,12))

        def ent_e(r, val="", w=32):
            e = entry(form, width=w)
            if val: e.insert(0,str(val))
            e.grid(row=r, column=1, sticky="ew", pady=7)
            return e

        lbl_e(0,"Loteria:")
        lot_var = tk.StringVar(value=reg["loteria"] or "Mega-Sena")
        lot_cb  = ttk.Combobox(form, textvariable=lot_var, values=LOTERIAS,
                                state="readonly", width=18, font=("Arial",9))
        lot_cb.grid(row=0, column=1, sticky="w", pady=7)

        lbl_e(1,"Nº Concurso:")
        conc_e = ent_e(1, reg["concurso"] or "")

        # Valor — exibe ganho ou sacado conforme tipo
        val_atual = reg["valor_ganho"] if reg["tipo"]=="GANHO" else reg["valor_sacado"]
        lbl_e(2, f"Valor (R$)  [{reg['tipo']}]:")
        val_e  = ent_e(2, f"{val_atual:.2f}".replace(".",","))

        lbl_e(3,"Data:")
        dt_e   = ent_e(3, reg["data_registro"] or "")

        lbl_e(4,"Descrição:")
        desc_e = ent_e(4, reg["descricao"] or "")

        def salvar():
            v    = to_float(val_e.get())
            lot  = lot_var.get()
            conc = conc_e.get().strip()
            dt   = dt_e.get().strip()
            desc = desc_e.get().strip()
            vg   = v if reg["tipo"]=="GANHO" else 0
            vs   = v if reg["tipo"]=="SAQUE" else 0
            self.db.execute(
                "UPDATE taxa_adm SET loteria=?,concurso=?,valor_ganho=?,valor_sacado=?,"
                "data_registro=?,descricao=? WHERE id=?",
                (lot, conc, vg, vs, dt, desc, rid))
            messagebox.showinfo("Atualizado", f"Lançamento ID {rid} atualizado!")
            win.destroy()
            self._adm_load()

        bf = tk.Frame(form, bg="#243447")
        bf.grid(row=5, column=0, columnspan=2, sticky="e", pady=16)
        btn(bf,"💾 SALVAR ALTERAÇÕES", CORES["btn_verde"], salvar, width=22).pack(side="left",padx=6)
        btn(bf,"❌ Cancelar", CORES["btn_cinza"], win.destroy, width=12).pack(side="left")

    def _adm_excluir(self):
        sel = self.adm_tree_hist.selection()
        if not sel:
            messagebox.showwarning("Atenção", "Selecione um lançamento no histórico!"); return
        vals = self.adm_tree_hist.item(sel[0], "values")
        if not vals: return
        rid  = int(vals[0])
        tipo = vals[5]; val = vals[4]
        if messagebox.askyesno("Confirmar",
            f"Excluir lançamento ID {rid}?\nTipo: {tipo}  |  Valor: {val}"):
            self.db.execute("DELETE FROM taxa_adm WHERE id=?", (rid,))
            self._adm_load()

    # ════════════════════════════════════════════════════════════
    #  ABA — BACKUP / RESTORE
    # ════════════════════════════════════════════════════════════
    #  ABA RESERVAS PESSOAIS
    # ════════════════════════════════════════════════════════════
    def _pub_sincronizar_reservas(self):
        """Botão manual — sincroniza reservas com Firebase."""
        self._pub_status_rsv.configure(text="⏳ Sincronizando reservas...", fg="#aad4f5")
        self.root.update_idletasks()
        import threading
        def _run():
            try:
                env, err = enviar_reservas_para_site(DB_FILE)
                def _ui():
                    if err:
                        self._pub_status_rsv.configure(
                            text="⚠ "+str(env)+" ok, "+str(len(err))+" erro(s): "+err[0],
                            fg="#e67e22")
                    else:
                        self._pub_status_rsv.configure(
                            text="✅ "+str(env)+" reserva(s) sincronizadas com sucesso!",
                            fg="#1D9E75")
                self.root.after(0, _ui)
            except Exception as ex:
                def _ui_err():
                    self._pub_status_rsv.configure(text="❌ Erro: "+str(ex)[:80], fg="#e74c3c")
                self.root.after(0, _ui_err)
        threading.Thread(target=_run, daemon=True).start()

    def _build_reservas(self):
        p = self.tab_rsv
        # Barra de sincronização Firebase no topo
        bsf = tk.Frame(p, bg="#1a3a5a", padx=12, pady=6)
        bsf.pack(fill="x")
        from tkinter import ttk as _ttk2
        btn(bsf, "📤 Sincronizar Reservas com Site",
            CORES["btn_verde"], self._pub_sincronizar_reservas,
            width=30).pack(side="left", padx=4)
        self._pub_status_rsv = tk.Label(bsf, text="", bg="#1a3a5a",
            fg="#1D9E75", font=("Arial",9,"bold"))
        self._pub_status_rsv.pack(side="left", padx=12)

        # ── Cards de totais ──────────────────────────────────────
        krow = tk.Frame(p, bg=CORES["bg_frame"]); krow.pack(fill="x", padx=20, pady=(14,6))
        self._rsv_cards = {}
        for attr, titulo, cor in [
            ("rsv_pessoas", "👥 Participantes",    "#2196F3"),
            ("rsv_total",   "💰 Total em Reserva", "#1D9E75"),
            ("rsv_credito", "⬆ Total Créditos",   "#8e44ad"),
            ("rsv_debito",  "⬇ Total Débitos",    "#e74c3c"),
        ]:
            c = tk.Frame(krow, bg=cor, padx=14, pady=10)
            c.pack(side="left", fill="both", expand=True, padx=4)
            tk.Label(c, text=titulo, bg=cor, fg="white",
                     font=("Arial",8,"bold")).pack(anchor="w")
            lv = tk.Label(c, text="—", bg=cor, fg="white", font=("Arial",14,"bold"))
            lv.pack(anchor="w", pady=(2,0))
            self._rsv_cards[attr] = lv

        # ── Área principal: 2 colunas ────────────────────────────
        main = tk.Frame(p, bg=CORES["bg_frame"])
        main.pack(fill="both", expand=True, padx=20, pady=(0,8))
        main.columnconfigure(0, weight=1); main.columnconfigure(1, weight=2)
        main.rowconfigure(0, weight=1)

        # ── Coluna esquerda: saldos + cadastro de pessoa ─────────
        col1 = tk.Frame(main, bg=CORES["bg_frame"])
        col1.grid(row=0, column=0, sticky="nsew", padx=(0,6))
        col1.rowconfigure(0, weight=1)

        sec_sal = tk.LabelFrame(col1, text="  💳 SALDO POR PESSOA  ",
                                bg="#243447", fg="white",
                                font=("Arial",9,"bold"), bd=1, padx=8, pady=6)
        sec_sal.pack(fill="both", expand=True)

        cols_s = {"Nome":180, "Saldo":110, "Últ. Mov.":100}
        fr_s, self._rsv_tree_sal = make_tree(sec_sal, cols_s, height=14)
        fr_s.pack(fill="both", expand=True)
        self._rsv_tree_sal.tag_configure("positivo", background="#d5f5e3")
        self._rsv_tree_sal.tag_configure("zerado",   background="#fde8d8")
        self._rsv_tree_sal.bind("<<TreeviewSelect>>", self._rsv_sel_pessoa)
        # Ordenacao por clique no cabecalho
        self._rsv_sort_col = "Saldo"
        self._rsv_sort_asc = False  # default: maior saldo primeiro
        for col in ("Nome", "Saldo", "Últ. Mov."):
            self._rsv_tree_sal.heading(col, text=col,
                command=lambda c=col: self._rsv_sort(c))

        # Botões de gestão de pessoas
        bf_p = tk.Frame(sec_sal, bg="#243447"); bf_p.pack(fill="x", pady=4)
        btn(bf_p, "➕ Nova Pessoa", CORES["btn_verde"],
            self._rsv_nova_pessoa, width=16).pack(side="left", padx=4)
        btn(bf_p, "✏ Editar",      CORES["btn_laranja"],
            self._rsv_editar_pessoa, width=10).pack(side="left", padx=4)
        btn(bf_p, "🧾 Recibo",     CORES["btn_roxo"],
            self._rsv_emitir_recibo, width=10).pack(side="left", padx=4)
        btn(bf_p, "🔄 Atualizar",  CORES["btn_azul"],
            self._rsv_load, width=10).pack(side="left", padx=4)

        # ── Coluna direita: movimentações ────────────────────────
        col2 = tk.Frame(main, bg=CORES["bg_frame"])
        col2.grid(row=0, column=1, sticky="nsew", padx=(6,0))
        col2.rowconfigure(1, weight=1)

        # Formulário de lançamento
        sec_form = tk.LabelFrame(col2, text="  ➕ REGISTRAR MOVIMENTAÇÃO  ",
                                 bg="#243447", fg="white",
                                 font=("Arial",9,"bold"), bd=1, padx=10, pady=8)
        sec_form.pack(fill="x", pady=(0,6))

        r0 = tk.Frame(sec_form, bg="#243447"); r0.pack(fill="x", pady=3)
        tk.Label(r0, text="Pessoa:", bg="#243447", fg="#aad4f5",
                 font=("Arial",9,"bold")).pack(side="left")
        self._rsv_cb_pessoa = ttk.Combobox(r0, width=28, state="readonly", font=("Arial",9))
        self._rsv_cb_pessoa.pack(side="left", padx=8)

        tk.Label(r0, text="Tipo:", bg="#243447", fg="#aad4f5",
                 font=("Arial",9,"bold")).pack(side="left", padx=(8,4))
        self._rsv_tipo = ttk.Combobox(r0, width=14, state="readonly", font=("Arial",9),
                                       values=["CRÉDITO (entrada)", "DÉBITO (uso)"])
        self._rsv_tipo.set("CRÉDITO (entrada)")
        self._rsv_tipo.pack(side="left")

        r1 = tk.Frame(sec_form, bg="#243447"); r1.pack(fill="x", pady=3)
        tk.Label(r1, text="Valor (R$):", bg="#243447", fg="#aad4f5",
                 font=("Arial",9,"bold")).pack(side="left")
        self._rsv_val = entry(r1, width=12); self._rsv_val.pack(side="left", padx=8)

        tk.Label(r1, text="Data:", bg="#243447", fg="#aad4f5",
                 font=("Arial",9,"bold")).pack(side="left")
        self._rsv_dt = entry(r1, width=12)
        self._rsv_dt.insert(0, date.today().strftime("%d/%m/%Y"))
        self._rsv_dt.pack(side="left", padx=8)

        # Campos extras — visíveis só para DÉBITO
        self._rsv_debito_frame = tk.Frame(sec_form, bg="#243447")
        self._rsv_debito_frame.pack(fill="x", pady=3)

        tk.Label(self._rsv_debito_frame, text="Loteria:", bg="#243447", fg="#aad4f5",
                 font=("Arial",9,"bold")).pack(side="left")
        self._rsv_lot = ttk.Combobox(self._rsv_debito_frame, values=LOTERIAS, width=14,
                                      state="readonly", font=("Arial",9))
        self._rsv_lot.set("Mega-Sena")
        self._rsv_lot.pack(side="left", padx=8)

        tk.Label(self._rsv_debito_frame, text="Concurso:", bg="#243447", fg="#aad4f5",
                 font=("Arial",9,"bold")).pack(side="left")
        self._rsv_conc = entry(self._rsv_debito_frame, width=10)
        self._rsv_conc.pack(side="left", padx=8)

        tk.Label(self._rsv_debito_frame, text="Descrição:", bg="#243447", fg="#aad4f5",
                 font=("Arial",9,"bold")).pack(side="left")
        self._rsv_desc = entry(self._rsv_debito_frame, width=22)
        self._rsv_desc.pack(side="left", padx=8)

        # Oculta inicialmente (padrão é CRÉDITO)
        self._rsv_debito_frame.pack_forget()

        def _on_rsv_tipo(e=None):
            if "DÉBITO" in self._rsv_tipo.get():
                self._rsv_debito_frame.pack(fill="x", pady=3)
            else:
                self._rsv_debito_frame.pack_forget()
        self._rsv_tipo.bind("<<ComboboxSelected>>", _on_rsv_tipo)

        bf_f = tk.Frame(sec_form, bg="#243447"); bf_f.pack(fill="x", pady=6)
        btn(bf_f, "💳 REGISTRAR", CORES["btn_verde"],
            self._rsv_registrar, width=18).pack(side="left", padx=4)
        btn(bf_f, "🗑 Excluir Selecionado", CORES["btn_vermelho"],
            self._rsv_excluir_mov, width=22).pack(side="left", padx=4)

        # Tabela de histórico da pessoa selecionada
        sec_hist = tk.LabelFrame(col2, text="  📋 HISTÓRICO DA PESSOA SELECIONADA  ",
                                  bg="#243447", fg="white",
                                  font=("Arial",9,"bold"), bd=1, padx=8, pady=6)
        sec_hist.pack(fill="both", expand=True)

        self._rsv_nome_lbl = tk.Label(sec_hist, text="← Selecione uma pessoa na lista",
                                       bg="#243447", fg="#aad4f5",
                                       font=("Arial",9,"italic"))
        self._rsv_nome_lbl.pack(anchor="w", pady=(0,4))

        cols_h = {"ID":45, "Tipo":100, "Valor":110,
                  "Data":100, "Loteria":110, "Concurso":90, "Descrição":180}
        fr_h, self._rsv_tree_hist = make_tree(sec_hist, cols_h, height=10)
        fr_h.pack(fill="both", expand=True)
        self._rsv_tree_hist.tag_configure("credito", background="#d5f5e3")
        self._rsv_tree_hist.tag_configure("debito",  background="#fde8d8")

        # Saldo da pessoa selecionada
        self._rsv_saldo_lbl = tk.Label(sec_hist, text="",
                                        bg="#243447", fg="#ffd700",
                                        font=("Arial",10,"bold"))
        self._rsv_saldo_lbl.pack(anchor="e", pady=(4,0))

    def _rsv_load(self):
        """Recarrega saldos e popula combo de pessoas."""
        pessoas = self.db.fetchall(
            "SELECT * FROM reservas_pessoas WHERE ativo=1 ORDER BY nome")

        nomes_combo = []
        total_cred = total_deb = 0
        self._rsv_dados = []  # guarda dados para reordenar sem re-consultar BD
        for ps in pessoas:
            pid = ps["id"]
            cred = self.db.fetchone(
                "SELECT COALESCE(SUM(valor),0) as t FROM reservas_movimentos "
                "WHERE pessoa_id=? AND UPPER(tipo) IN "
                "('CRÉDITO','CREDITO','ENTRADA','DEPOSITO','DEPÓSITO')", (pid,))["t"] or 0
            deb  = self.db.fetchone(
                "SELECT COALESCE(SUM(valor),0) as t FROM reservas_movimentos "
                "WHERE pessoa_id=? AND UPPER(tipo) IN "
                "('DÉBITO','DEBITO','SAQUE','USO')", (pid,))["t"] or 0
            saldo = float(cred) - float(deb)
            total_cred += float(cred); total_deb += float(deb)
            ult = self.db.fetchone(
                "SELECT data_mov FROM reservas_movimentos "
                "WHERE pessoa_id=? ORDER BY id DESC LIMIT 1", (pid,))
            ult_dt = ult["data_mov"] if ult else "-"
            self._rsv_dados.append({
                "pid": pid, "nome": ps["nome"], "saldo": saldo, "ult": ult_dt
            })
            nomes_combo.append(f"{ps['nome']} (ID: {pid})")

        self._rsv_cb_pessoa["values"] = nomes_combo
        total_rsv = total_cred - total_deb
        self._rsv_cards["rsv_pessoas"].configure(text=str(len(pessoas)))
        self._rsv_cards["rsv_total"].configure(text=fmt_brl(total_rsv))
        self._rsv_cards["rsv_credito"].configure(text=fmt_brl(total_cred))
        self._rsv_cards["rsv_debito"].configure(text=fmt_brl(total_deb))
        self._rsv_renderizar()

    def _rsv_sort(self, col):
        """Ordena a lista ao clicar no cabecalho."""
        if self._rsv_sort_col == col:
            self._rsv_sort_asc = not self._rsv_sort_asc
        else:
            self._rsv_sort_col = col
            self._rsv_sort_asc = col == "Nome"  # Nome: A→Z; Saldo: maior primeiro
        self._rsv_renderizar()

    def _rsv_renderizar(self):
        """Renderiza a tree com a ordenacao atual."""
        col  = getattr(self, "_rsv_sort_col", "Saldo")
        asc  = getattr(self, "_rsv_sort_asc", False)
        dados = getattr(self, "_rsv_dados", [])

        if col == "Nome":
            dados_ord = sorted(dados, key=lambda x: x["nome"].lower(), reverse=not asc)
        elif col == "Saldo":
            dados_ord = sorted(dados, key=lambda x: (x["saldo"], x["nome"].lower()), reverse=not asc)
        else:  # Últ. Mov.
            dados_ord = sorted(dados, key=lambda x: (x["ult"], x["nome"].lower()), reverse=not asc)

        # Atualiza icone no cabecalho
        for c_ in ("Nome", "Saldo", "Últ. Mov."):
            ic = (" ▲" if asc else " ▼") if c_ == col else ""
            self._rsv_tree_sal.heading(c_, text=c_+ic,
                command=lambda c2=c_: self._rsv_sort(c2))

        sel_atual = self._rsv_tree_sal.selection()
        self._rsv_tree_sal.delete(*self._rsv_tree_sal.get_children())
        for row in dados_ord:
            tag = "positivo" if row["saldo"] > 0 else "zerado"
            self._rsv_tree_sal.insert("","end", iid=str(row["pid"]), tags=(tag,),
                values=(row["nome"], fmt_brl(row["saldo"]), row["ult"]))
        # Restaura selecao
        if sel_atual:
            try: self._rsv_tree_sal.selection_set(sel_atual)
            except: pass

    def _rsv_sel_pessoa(self, e=None):
        """Ao clicar numa pessoa, carrega histórico dela."""
        sel = self._rsv_tree_sal.selection()
        if not sel: return
        pid = int(sel[0])
        ps  = self.db.fetchone("SELECT * FROM reservas_pessoas WHERE id=?", (pid,))
        if not ps: return

        # Atualiza combo para pessoa selecionada
        self._rsv_cb_pessoa.set(f"{ps['nome']} (ID: {pid})")

        # Carrega histórico
        self._rsv_tree_hist.delete(*self._rsv_tree_hist.get_children())
        movs = self.db.fetchall(
            "SELECT * FROM reservas_movimentos WHERE pessoa_id=? ORDER BY id DESC", (pid,))
        cred = deb = 0
        for m in movs:
            eh_cred = m["tipo"] == "CRÉDITO"
            tag = "credito" if eh_cred else "debito"
            sinal = "⬆ CRÉDITO" if eh_cred else "⬇ DÉBITO"
            if eh_cred: cred += m["valor"]
            else:       deb  += m["valor"]
            self._rsv_tree_hist.insert("","end", iid=str(m["id"]), tags=(tag,), values=(
                m["id"], sinal, fmt_brl(m["valor"]),
                m["data_mov"] or "-", m["loteria"] or "-",
                m["concurso"] or "-", m["descricao"] or "-"))

        saldo = cred - deb
        cor_s = "#afffca" if saldo > 0 else "#ffaaaa"
        self._rsv_nome_lbl.configure(text=f"📋 {ps['nome']}")
        self._rsv_saldo_lbl.configure(
            text=f"Saldo: {fmt_brl(saldo)}  |  "
                 f"Créditos: {fmt_brl(cred)}  |  Débitos: {fmt_brl(deb)}",
            fg=cor_s)

    def _rsv_registrar(self):
        sel_p = self._rsv_cb_pessoa.get()
        if not sel_p:
            messagebox.showwarning("Atenção","Selecione uma pessoa!"); return
        m = re.search(r"\(ID: (\d+)\)", sel_p)
        if not m:
            messagebox.showwarning("Atenção","Selecione uma pessoa válida!"); return
        pid = int(m.group(1))

        v = to_float(self._rsv_val.get())
        if v <= 0:
            messagebox.showwarning("Atenção","Informe um valor válido!"); return

        tipo_full = self._rsv_tipo.get()
        tipo = "CRÉDITO" if "CRÉDITO" in tipo_full else "DÉBITO"

        # Valida saldo para débito
        if tipo == "DÉBITO":
            cred = self.db.fetchone(
                "SELECT SUM(valor) as t FROM reservas_movimentos "
                "WHERE pessoa_id=? AND tipo='CRÉDITO'", (pid,))["t"] or 0
            deb  = self.db.fetchone(
                "SELECT SUM(valor) as t FROM reservas_movimentos "
                "WHERE pessoa_id=? AND tipo='DÉBITO'", (pid,))["t"] or 0
            saldo = cred - deb
            if v > saldo:
                if not messagebox.askyesno("Saldo Insuficiente",
                    f"Saldo disponível: {fmt_brl(saldo)}\n"
                    f"Débito: {fmt_brl(v)}\n\n"
                    f"Saldo ficará negativo. Deseja continuar?"):
                    return

        dt   = self._rsv_dt.get().strip() or date.today().strftime("%d/%m/%Y")
        # Loteria/concurso/desc só se aplica a DÉBITO
        lot  = self._rsv_lot.get()  if tipo == "DÉBITO" else ""
        conc = self._rsv_conc.get().strip() if tipo == "DÉBITO" else ""
        desc = self._rsv_desc.get().strip() if tipo == "DÉBITO" else ""

        self.db.execute(
            "INSERT INTO reservas_movimentos "
            "(pessoa_id, tipo, valor, data_mov, loteria, concurso, descricao) "
            "VALUES (?,?,?,?,?,?,?)",
            (pid, tipo, v, dt, lot, conc, desc))

        # ── Sincroniza automaticamente com Firebase ──────────────
        ps = self.db.fetchone("SELECT nome FROM reservas_pessoas WHERE id=?", (pid,))
        nome_rsv  = ps["nome"] if ps else "Desconhecido"
        tipo_fire = "deposito" if tipo == "CRÉDITO" else \
                    ("uso" if (lot or conc) else "saque")
        try:
            sincronizar_reserva(
                pessoa_id = pid,
                nome      = nome_rsv,
                tipo      = tipo_fire,
                valor     = v,
                descricao = desc or f"{tipo_fire.capitalize()} via sistema",
                loteria   = lot  or None,
                concurso  = conc or None,
            )
        except Exception as e:
            print(f"[Sync reserva] {e}")

        # Limpa campos após registrar
        self._rsv_val.delete(0,"end")
        self._rsv_dt.delete(0,"end")
        self._rsv_dt.insert(0, date.today().strftime("%d/%m/%Y"))
        # Limpa campos de débito sempre após registrar
        self._rsv_conc.delete(0,"end")
        self._rsv_desc.delete(0,"end")

        self._rsv_load()
        # Reseleciona a pessoa para atualizar histórico
        self._rsv_tree_sal.selection_set(str(pid))
        self._rsv_sel_pessoa()

    def _rsv_excluir_mov(self):
        sel = self._rsv_tree_hist.selection()
        if not sel:
            messagebox.showwarning("Atenção","Selecione um lançamento!"); return
        mid = int(sel[0])
        m   = self.db.fetchone("SELECT * FROM reservas_movimentos WHERE id=?", (mid,))
        if messagebox.askyesno("Confirmar",
            f"Excluir lançamento {m['tipo']} de {fmt_brl(m['valor'])} em {m['data_mov']}?"):
            self.db.execute("DELETE FROM reservas_movimentos WHERE id=?", (mid,))
            self._rsv_load()
            # Reatualiza histórico da pessoa
            sel_p = self._rsv_cb_pessoa.get()
            mp = re.search(r"\(ID: (\d+)\)", sel_p)
            if mp:
                pid = int(mp.group(1))
                self._rsv_tree_sal.selection_set(str(pid))
                self._rsv_sel_pessoa()

    def _rsv_nova_pessoa(self):
        win = tk.Toplevel(self.root); win.title("Nova Pessoa — Reservas")
        win.geometry("420x300"); win.configure(bg=CORES["bg_section"]); win.grab_set()
        tk.Label(win, text="CADASTRAR PESSOA PARA RESERVA",
                 bg=CORES["bg_section"], fg=CORES["fg_title"],
                 font=("Arial",11,"bold")).pack(pady=12)
        form = tk.Frame(win, bg=CORES["bg_section"], padx=24); form.pack(fill="x")
        fields = [("Nome*:", "nome"), ("Telefone:", "tel"),
                  ("Chave PIX:", "pix"), ("Observações:", "obs")]
        vars_ = {}
        for i, (lbl, key) in enumerate(fields):
            tk.Label(form, text=lbl, bg=CORES["bg_section"], fg=CORES["fg_label"],
                     font=("Arial",9,"bold")).grid(row=i*2, column=0, sticky="w", pady=(6,0))
            w = entry(form, width=44); w.grid(row=i*2+1, column=0, sticky="ew")
            vars_[key] = w
        form.columnconfigure(0, weight=1)
        def salvar():
            nome = vars_["nome"].get().strip()
            if not nome:
                messagebox.showwarning("Atenção","Nome é obrigatório!"); return
            self.db.execute(
                "INSERT INTO reservas_pessoas (nome,telefone,chave_pix,observacoes) "
                "VALUES (?,?,?,?)",
                (nome, vars_["tel"].get(), vars_["pix"].get(), vars_["obs"].get()))
            messagebox.showinfo("Cadastrado",f"'{nome}' cadastrado!")
            win.destroy(); self._rsv_load()
        btn(form, "💾 SALVAR", CORES["btn_verde"], salvar, width=18).grid(
            row=10, column=0, pady=14, sticky="e")

    def _rsv_editar_pessoa(self):
        sel = self._rsv_tree_sal.selection()
        if not sel:
            messagebox.showwarning("Atenção","Selecione uma pessoa!"); return
        pid = int(sel[0])
        ps  = self.db.fetchone("SELECT * FROM reservas_pessoas WHERE id=?", (pid,))
        win = tk.Toplevel(self.root); win.title("Editar Pessoa")
        win.geometry("420x300"); win.configure(bg=CORES["bg_section"]); win.grab_set()
        tk.Label(win, text="EDITAR PESSOA",
                 bg=CORES["bg_section"], fg=CORES["fg_title"],
                 font=("Arial",11,"bold")).pack(pady=12)
        form = tk.Frame(win, bg=CORES["bg_section"], padx=24); form.pack(fill="x")
        fields = [("Nome*:", "nome"), ("Telefone:", "telefone"),
                  ("Chave PIX:", "chave_pix"), ("Observações:", "observacoes")]
        vars_ = {}
        for i, (lbl, key) in enumerate(fields):
            tk.Label(form, text=lbl, bg=CORES["bg_section"], fg=CORES["fg_label"],
                     font=("Arial",9,"bold")).grid(row=i*2, column=0, sticky="w", pady=(6,0))
            w = entry(form, width=44)
            w.insert(0, ps[key] or "")
            w.grid(row=i*2+1, column=0, sticky="ew")
            vars_[key] = w
        form.columnconfigure(0, weight=1)
        def salvar():
            nome = vars_["nome"].get().strip()
            if not nome:
                messagebox.showwarning("Atenção","Nome é obrigatório!"); return
            self.db.execute(
                "UPDATE reservas_pessoas SET nome=?,telefone=?,chave_pix=?,observacoes=? WHERE id=?",
                (nome, vars_["telefone"].get(), vars_["chave_pix"].get(),
                 vars_["observacoes"].get(), pid))
            win.destroy(); self._rsv_load()
        btn(form, "💾 SALVAR", CORES["btn_verde"], salvar, width=18).grid(
            row=10, column=0, pady=14, sticky="e")
    def _rsv_emitir_recibo(self):
        try:
            sel = self._rsv_tree_sal.selection()
            if not sel:
                messagebox.showwarning("Atencao","Selecione uma pessoa!"); return
            pid = int(sel[0])
            ps_d = dict(self.db.fetchone("SELECT * FROM reservas_pessoas WHERE id=?", (pid,)) or {})
            if not ps_d: return
            movs = self.db.fetchall(
                "SELECT * FROM reservas_movimentos WHERE pessoa_id=? ORDER BY id ASC", (pid,))
            cred = sum(float(m["valor"] or 0) for m in movs
                       if (m["tipo"] or "").upper() in
                       ("CREDITO","CRÉDITO","ENTRADA","DEPOSITO","DEPÓSITO"))
            deb  = sum(float(m["valor"] or 0) for m in movs
                       if (m["tipo"] or "").upper() in ("DEBITO","DÉBITO","SAQUE","USO"))
            saldo = cred - deb
            import tempfile, webbrowser
    
            rows_mov = ""
            saldo_acum = 0.0
            for i, m in enumerate([dict(x) for x in movs]):
                tipo_up = (m["tipo"] or "").upper()
                eh_c = tipo_up in ("CREDITO","CRÉDITO","ENTRADA","DEPOSITO","DEPÓSITO")
                val  = float(m["valor"] or 0)
                saldo_ant = saldo_acum
                saldo_acum = saldo_acum + val if eh_c else saldo_acum - val
                cor_t = "#1D9E75" if eh_c else "#e74c3c"
                bg    = "#f0faf5" if i%2==0 else "#ffffff"
                ic    = "⬆" if eh_c else "⬇"
                sinal = "CRÉDITO" if eh_c else "DÉBITO"
                lot   = (m["loteria"] or "")+((" #"+str(m["concurso"])) if m["concurso"] else "")
                desc  = m.get("descricao") or m.get("observacoes") or ""
                rows_mov += f"""<tr style="background:{bg}">
                  <td><span style="color:{cor_t};font-weight:700">{ic} {sinal}</span></td>
                  <td>{m["data_mov"] or "-"}</td>
                  <td style="color:#555">{lot or "-"}</td>
                  <td style="color:#666;font-size:11px">{desc[:35] if desc else "-"}</td>
                  <td class="money" style="color:{cor_t}">{fmt_brl(val)}</td>
                  <td class="money" style="color:{'#1D9E75' if saldo_acum>=0 else '#e74c3c'}">{fmt_brl(saldo_acum)}</td>
                </tr>"""
    
            cor_s = "#1D9E75" if saldo>=0 else "#e74c3c"
            n_ext = "RSV-"+datetime.now().strftime("%Y")+f"-{pid:04d}"
            ini   = "".join(w[0].upper() for w in (ps_d["nome"] or "?").split()[:2])
    
            html = f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
    <title>Extrato {n_ext}</title>
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&display=swap');
    *{{margin:0;padding:0;box-sizing:border-box}}
    body{{font-family:'Inter',Arial,sans-serif;background:linear-gradient(135deg,#0d1b2a,#1a2f4a,#0d2137);min-height:100vh;display:flex;align-items:flex-start;justify-content:center;padding:30px 16px}}
    .wrap{{width:100%;max-width:820px}}
    .card{{background:white;border-radius:20px;overflow:hidden;box-shadow:0 20px 60px rgba(0,0,0,.4)}}
    .hdr{{background:linear-gradient(135deg,#0d2137,#1a3a6e,#1565c0);padding:0;position:relative;overflow:hidden}}
    .hdr-bg{{position:absolute;inset:0;opacity:.1;background:repeating-linear-gradient(45deg,transparent,transparent 20px,rgba(255,255,255,.3) 20px,rgba(255,255,255,.3) 21px)}}
    .hdr-inner{{position:relative;z-index:1;padding:28px 36px;display:flex;justify-content:space-between;align-items:center}}
    .hdr-left .badge{{background:rgba(255,255,255,.15);border:1px solid rgba(255,255,255,.25);color:#90caf9;font-size:10px;font-weight:700;letter-spacing:2px;text-transform:uppercase;padding:4px 12px;border-radius:20px;display:inline-block;margin-bottom:10px}}
    .hdr-left h1{{color:white;font-size:24px;font-weight:800}}
    .hdr-left .sub{{color:#90caf9;font-size:12px;margin-top:6px}}
    .hdr-right .rec-num{{background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.25);border-radius:12px;padding:12px 18px;text-align:right}}
    .hdr-right .rec-label{{color:#90caf9;font-size:9px;text-transform:uppercase;letter-spacing:1.5px}}
    .hdr-right .rec-val{{color:white;font-size:16px;font-weight:800;margin-top:2px}}
    .part{{background:linear-gradient(135deg,#f8fbff,#e3f2fd);padding:20px 36px;display:flex;align-items:center;gap:16px;border-bottom:2px solid #e3f2fd}}
    .avatar{{width:56px;height:56px;border-radius:50%;background:linear-gradient(135deg,#0d2137,#1976d2);color:white;font-size:18px;font-weight:800;display:flex;align-items:center;justify-content:center;flex-shrink:0;box-shadow:0 4px 14px rgba(13,33,55,.35)}}
    .part-info h2{{font-size:17px;font-weight:700;color:#1a1a1a}}
    .part-info .tags{{display:flex;gap:8px;margin-top:6px;flex-wrap:wrap}}
    .tag{{background:#e3f2fd;color:#1565c0;font-size:11px;font-weight:600;padding:3px 10px;border-radius:20px;border:1px solid #bbdefb}}
    .saldo-hero{{background:linear-gradient(135deg,{"#e8f5e9,#c8e6c9" if saldo>=0 else "#fde8d8,#ffccbc"});border:2px solid {cor_s};border-radius:0;padding:20px 36px;display:flex;justify-content:space-between;align-items:center}}
    .saldo-hero .sh-label{{font-size:12px;color:#555;font-weight:600;text-transform:uppercase;letter-spacing:1px}}
    .saldo-hero .sh-val{{font-size:32px;font-weight:800;color:{cor_s}}}
    .kpis{{display:grid;grid-template-columns:repeat(2,1fr);gap:0;border-bottom:1px solid #f0f0f0}}
    .kpi{{padding:16px 24px;border-right:1px solid #f0f0f0}}
    .kpi:last-child{{border-right:none}}
    .kpi label{{font-size:10px;color:#999;text-transform:uppercase;letter-spacing:1px;font-weight:600}}
    .kpi .kval{{font-size:20px;font-weight:800;margin-top:4px}}
    .kpi.k1 .kval{{color:#1D9E75}}
    .kpi.k2 .kval{{color:#e74c3c}}
    .sec{{padding:0 36px 24px}}
    .sec-title{{font-size:10px;color:#999;font-weight:700;text-transform:uppercase;letter-spacing:1.5px;padding:16px 0 10px}}
    table{{width:100%;border-collapse:collapse;font-size:13px}}
    thead tr{{background:linear-gradient(135deg,#0d2137,#1565c0)}}
    thead th{{color:white;padding:10px 12px;text-align:left;font-size:11px;font-weight:600;letter-spacing:.5px}}
    tbody tr:hover{{background:#f0f4ff!important}}
    td{{padding:9px 12px;border-bottom:1px solid #f5f5f5;vertical-align:middle}}
    .money{{font-weight:700}}
    .footer{{background:linear-gradient(135deg,#0d2137,#1a3a6e);padding:14px 36px;display:flex;justify-content:space-between;align-items:center}}
    .footer span{{color:#90caf9;font-size:11px}}
    .footer .brand{{color:white;font-weight:700;font-size:12px}}
    @media print{{
      body{{background:white;padding:0}}
      .card{{box-shadow:none;border-radius:0;max-width:100%}}
      .hdr,.saldo-hero,.kpi,thead tr,.footer{{-webkit-print-color-adjust:exact!important;print-color-adjust:exact!important}}
    }}
    </style></head><body>
    <div class="wrap"><div class="card">
      <div class="hdr">
        <div class="hdr-bg"></div>
        <div class="hdr-inner">
          <div class="hdr-left">
            <span class="badge">Extrato de Reserva Pessoal</span>
            <h1>Sistema de Gestão<br>de Bolões</h1>
            <p class="sub">📅 {datetime.now().strftime("%d/%m/%Y  às  %H:%M")}</p>
          </div>
          <div class="hdr-right">
            <div class="rec-num">
              <div class="rec-label">Número do Extrato</div>
              <div class="rec-val">{n_ext}</div>
            </div>
          </div>
        </div>
      </div>
      <div class="part">
        <div class="avatar">{ini}</div>
        <div class="part-info">
          <h2>{ps_d['nome']}</h2>
          <div class="tags">
            <span class="tag">📱 {ps_d['telefone'] or 'Sem telefone'}</span>
            <span class="tag">🔑 PIX: {ps_d['chave_pix'] or 'Não informado'}</span>
          </div>
        </div>
      </div>
      <div class="saldo-hero">
        <div><div class="sh-label">💰 Saldo Disponível</div></div>
        <div class="sh-val">{fmt_brl(saldo)}</div>
      </div>
      <div class="kpis">
        <div class="kpi k1"><label>Total de Créditos</label><div class="kval">{fmt_brl(cred)}</div></div>
        <div class="kpi k2"><label>Total de Débitos</label><div class="kval">{fmt_brl(deb)}</div></div>
      </div>
      <div class="sec">
        <div class="sec-title">Histórico de Movimentações</div>
        <table>
          <thead><tr><th>Tipo</th><th>Data</th><th>Loteria/Concurso</th><th>Descrição</th><th>Valor</th><th>Saldo Acumulado</th></tr></thead>
          <tbody>{rows_mov}</tbody>
        </table>
      </div>
      <div class="footer">
        <span>Sistema de Gestão de Bolões v5.1</span>
        <span class="brand">✨ Desenvolvido por Elton Luis</span>
      </div>
    </div></div>
    
    </body></html>"""
    
            import os as _os3
            nome_sugerido2 = (ps_d.get("nome","reserva") or "reserva").replace(" ","_")
            path_rsv = filedialog.asksaveasfilename(
                defaultextension=".html",
                filetypes=[("HTML","*.html"),("Todos","*.*")],
                initialfile=nome_sugerido2+"_"+datetime.now().strftime("%Y%m%d_%H%M")+".html")
            if not path_rsv: return
            with open(path_rsv,"w",encoding="utf-8") as _f3: _f3.write(html)
            webbrowser.open("file:///" + _os3.path.abspath(path_rsv).replace("\\", "/"))
    
        except Exception as _ex_rsv:
            import traceback
            messagebox.showerror("Erro no Extrato",
                "Erro ao gerar extrato:\n\n" + str(_ex_rsv) +
                "\n\n" + traceback.format_exc()[-400:])

    # ════════════════════════════════════════════════════════════
    # ════════════════════════════════════════════════════════════
    #  PUBLICAR NO SITE — Firebase Firestore
    # ════════════════════════════════════════════════════════════

    # ════════════════════════════════════════════════════════════
    #  ABA PESSOAS — Cadastro Global
    # ════════════════════════════════════════════════════════════
    def _build_pessoas(self):
        p = self.tab_pessoas

        # ── Barra de ações ───────────────────────────────────────
        top = tk.Frame(p, bg=CORES["bg_frame"]); top.pack(fill="x", padx=20, pady=10)
        btn(top, "🔄 Atualizar", CORES["btn_azul"],
            self._pessoas_load, width=14).pack(side="left", padx=4)
        btn(top, "🔗 Unificar Selecionados", CORES["btn_laranja"],
            self._pessoas_unificar, width=22).pack(side="left", padx=4)
        btn(top, "✏ Editar Pessoa", CORES["btn_verde"],
            self._pessoas_editar, width=16).pack(side="left", padx=4)

        tk.Label(top, text="  Buscar:", bg=CORES["bg_frame"],
                 fg=CORES["fg_label"], font=("Arial",9,"bold")).pack(side="left", padx=(12,4))
        self._pessoas_busca = tk.StringVar()
        tk.Entry(top, textvariable=self._pessoas_busca, width=22,
                 font=("Arial",9), relief="solid", bd=1).pack(side="left")
        self._pessoas_busca.trace_add("write", lambda *a: self._pessoas_load())

        # ── Cards de resumo ──────────────────────────────────────
        krow = tk.Frame(p, bg=CORES["bg_frame"]); krow.pack(fill="x", padx=20, pady=(0,6))
        self._pessoas_cards = {}
        for attr, titulo, cor in [
            ("p_total",  "Total de Pessoas",    "#2196F3"),
            ("p_boloes", "Participações",        "#1D9E75"),
            ("p_dup",    "Possíveis Duplicatas", "#e67e22"),
        ]:
            c = tk.Frame(krow, bg=cor, padx=14, pady=8)
            c.pack(side="left", fill="both", expand=True, padx=4)
            tk.Label(c, text=titulo, bg=cor, fg="white",
                     font=("Arial",8,"bold")).pack(anchor="w")
            lv = tk.Label(c, text="—", bg=cor, fg="white", font=("Arial",14,"bold"))
            lv.pack(anchor="w"); self._pessoas_cards[attr] = lv

        # ── Área principal: lista + detalhes ─────────────────────
        mid = tk.Frame(p, bg=CORES["bg_frame"])
        mid.pack(fill="both", expand=True, padx=20, pady=(0,8))
        mid.columnconfigure(0, weight=2); mid.columnconfigure(1, weight=1)
        mid.rowconfigure(0, weight=1)

        # Lista de pessoas
        sec_l = tk.LabelFrame(mid, text="  👥 CADASTRO GLOBAL DE PESSOAS  ",
                               bg="#243447", fg="white",
                               font=("Arial",9,"bold"), bd=1, padx=6, pady=6)
        sec_l.grid(row=0, column=0, sticky="nsew", padx=(0,4))
        cols_p = {"ID":45, "Nome":220, "Telefone":130, "PIX":160, "Bolões":60}
        fr_p, self._pessoas_tree = make_tree(sec_l, cols_p, height=20)
        fr_p.pack(fill="both", expand=True)
        self._pessoas_tree.tag_configure("dup", background="#fff3cd")
        self._pessoas_tree.bind("<<TreeviewSelect>>", self._pessoas_sel)

        # Detalhes da pessoa selecionada
        sec_d = tk.LabelFrame(mid, text="  📋 PARTICIPAÇÕES  ",
                               bg="#243447", fg="white",
                               font=("Arial",9,"bold"), bd=1, padx=6, pady=6)
        sec_d.grid(row=0, column=1, sticky="nsew", padx=(4,0))
        self._pessoas_det_lbl = tk.Label(sec_d, text="← Selecione uma pessoa",
                                          bg="#243447", fg="#aad4f5",
                                          font=("Arial",9,"italic"))
        self._pessoas_det_lbl.pack(anchor="w", pady=(0,4))
        cols_d = {"Bolão":160, "Status":90, "Pago":90}
        fr_d, self._pessoas_det_tree = make_tree(sec_d, cols_d, height=10)
        fr_d.pack(fill="both", expand=True)

        # Painel de possíveis duplicatas
        sec_dup = tk.LabelFrame(mid, text="  ⚠ POSSÍVEIS DUPLICATAS  ",
                                 bg="#243447", fg="#ffcc88",
                                 font=("Arial",9,"bold"), bd=1, padx=6, pady=6)
        sec_dup.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(4,0))
        mid.rowconfigure(1, weight=0)
        cols_dup = {"ID A":45, "Nome A":200, "Tel A":120,
                    "ID B":45, "Nome B":200, "Tel B":120, "Similaridade":100}
        fr_dup, self._dup_tree = make_tree(sec_dup, cols_dup, height=4)
        fr_dup.pack(fill="both", expand=True)
        self._dup_tree.tag_configure("dup", background="#fff3cd")
        tk.Label(sec_dup,
                 text="Selecione uma linha e clique em '🔗 Unificar Selecionados' para fundir.",
                 bg="#243447", fg="#888", font=("Arial",8,"italic")).pack(anchor="w", pady=(4,0))

    def _pessoas_load(self):
        self._pessoas_tree.delete(*self._pessoas_tree.get_children())
        busca = self._pessoas_busca.get().strip().lower()

        rows = self.db.fetchall("""
            SELECT p.id, p.nome, p.telefone, p.chave_pix,
                   COUNT(pt.id) as n_boloes
            FROM pessoas p
            LEFT JOIN participantes pt ON pt.pessoa_id=p.id AND pt.ativo=1
            GROUP BY p.id ORDER BY p.nome
        """)

        # Filtra em Python (evita SQL injection e simplifica)
        if busca:
            rows = [r for r in rows if busca in r["nome"].lower() or
                    busca in (r["telefone"] or "").lower()]

        total      = len(rows)
        total_part = sum(r["n_boloes"] for r in rows)

        for r in rows:
            tag = "dup" if r["n_boloes"] == 0 else ""
            self._pessoas_tree.insert("","end", iid=str(r["id"]), tags=(tag,), values=(
                r["id"], r["nome"], r["telefone"] or "—",
                r["chave_pix"] or "—", r["n_boloes"]))

        # Detecta possíveis duplicatas por nome similar
        self._dup_tree.delete(*self._dup_tree.get_children())
        import unicodedata
        def norm(s):
            return "".join(c for c in unicodedata.normalize("NFD",s.lower())
                           if unicodedata.category(c)!="Mn")
        dups = []
        lista = list(rows)
        for i in range(len(lista)):
            for j in range(i+1, len(lista)):
                a = lista[i]; b = lista[j]
                na = set(norm(a["nome"]).split())
                nb = set(norm(b["nome"]).split())
                if not na or not nb: continue
                sim = len(na & nb) / max(len(na), len(nb))
                if sim >= 0.7 and a["telefone"] != b["telefone"]:
                    dups.append((a, b, sim))
        dups.sort(key=lambda x: -x[2])
        for a, b, sim in dups[:20]:
            self._dup_tree.insert("","end", tags=("dup",), values=(
                a["id"], a["nome"], a["telefone"] or "—",
                b["id"], b["nome"], b["telefone"] or "—",
                f"{sim*100:.0f}%"))

        self._pessoas_cards["p_total"].configure(text=str(total))
        self._pessoas_cards["p_boloes"].configure(text=str(total_part))
        self._pessoas_cards["p_dup"].configure(text=str(len(dups)))

    def _pessoas_sel(self, e=None):
        sel = self._pessoas_tree.selection()
        if not sel: return
        pid = int(sel[0])
        ps  = self.db.fetchone("SELECT * FROM pessoas WHERE id=?", (pid,))
        if not ps: return
        self._pessoas_det_lbl.configure(
            text=f"👤 {ps['nome']}  |  📱 {ps['telefone'] or '—'}  |  PIX: {ps['chave_pix'] or '—'}")
        self._pessoas_det_tree.delete(*self._pessoas_det_tree.get_children())
        parts = self.db.fetchall("""
            SELECT b.nome as bolao, pt.valor_esperado,
                   COALESCE((SELECT SUM(pg.valor) FROM pagamentos pg
                              WHERE pg.participante_id=pt.id), 0) as pago
            FROM participantes pt
            JOIN boloes b ON pt.bolao_id=b.id
            WHERE pt.pessoa_id=? AND pt.ativo=1
            ORDER BY b.nome
        """, (pid,))
        for r in parts:
            saldo = max(0, (r["valor_esperado"] or 0) - r["pago"])
            status = "✅ Quitado" if saldo<=0 else "⚠ Pendente"
            self._pessoas_det_tree.insert("","end", values=(
                r["bolao"], status, fmt_brl(r["pago"])))

    def _pessoas_unificar(self):
        """Unifica dois registros — mantém o de menor ID, migra dados do outro."""
        import re as _re
        # Verifica seleção na tabela de duplicatas primeiro
        sel_dup = self._dup_tree.selection()
        if sel_dup:
            vals = self._dup_tree.item(sel_dup[0], "values")
            id_a, id_b = int(vals[0]), int(vals[3])
        else:
            # Tenta na lista principal (precisa de 2 selecionados)
            sel = self._pessoas_tree.selection()
            if len(sel) < 2:
                messagebox.showwarning("Atenção",
                    "Selecione 2 pessoas na lista OU uma linha na tabela de duplicatas!"); return
            id_a, id_b = int(sel[0]), int(sel[1])

        pa = self.db.fetchone("SELECT * FROM pessoas WHERE id=?", (id_a,))
        pb = self.db.fetchone("SELECT * FROM pessoas WHERE id=?", (id_b,))
        if not pa or not pb: return

        if not messagebox.askyesno("Confirmar Unificação",
            f"Unificar em UMA pessoa?\n\n"
            f"A: {pa['nome']} ({pa['telefone']})\n"
            f"B: {pb['nome']} ({pb['telefone']})\n\n"
            f"Registro A mantido. Dados de B migrados para A.\n"
            f"Esta ação não pode ser desfeita!"): return

        # Migra participantes de B para A
        self.db.execute(
            "UPDATE participantes SET pessoa_id=? WHERE pessoa_id=?", (id_a, id_b))
        # Remove B
        self.db.execute("DELETE FROM pessoas WHERE id=?", (id_b,))
        messagebox.showinfo("Unificado",
            f"Registros unificados!\n'{pb['nome']}' fundido em '{pa['nome']}'.")
        self._pessoas_load()
        self._refresh_all()

    def _pessoas_editar(self):
        sel = self._pessoas_tree.selection()
        if not sel: messagebox.showwarning("Atenção","Selecione uma pessoa!"); return
        pid = int(sel[0])
        ps  = self.db.fetchone("SELECT * FROM pessoas WHERE id=?", (pid,))

        win = tk.Toplevel(self.root); win.title("Editar Pessoa")
        win.geometry("440x260"); win.configure(bg=CORES["bg_section"]); win.grab_set(); win.lift()
        tk.Label(win, text="EDITAR DADOS DA PESSOA",
                 bg=CORES["bg_section"], fg=CORES["fg_title"],
                 font=("Arial",11,"bold")).pack(pady=10)
        form = tk.Frame(win, bg=CORES["bg_section"], padx=24); form.pack(fill="x")
        fields = [("Nome:","nome"),("Telefone:","telefone"),("PIX:","chave_pix")]
        vars_ = {}
        for i,(lbl,key) in enumerate(fields):
            tk.Label(form, text=lbl, bg=CORES["bg_section"], fg=CORES["fg_label"],
                     font=("Arial",9,"bold")).grid(row=i*2, column=0, sticky="w", pady=(6,0))
            w = entry(form, width=40); w.insert(0, ps[key] or "")
            w.grid(row=i*2+1, column=0, sticky="ew"); vars_[key] = w
        form.columnconfigure(0, weight=1)
        def salvar():
            self.db.execute(
                "UPDATE pessoas SET nome=?,telefone=?,chave_pix=? WHERE id=?",
                (vars_["nome"].get(), vars_["telefone"].get(),
                 vars_["chave_pix"].get(), pid))
            # Atualiza também participantes vinculados
            self.db.execute(
                "UPDATE participantes SET nome=?,telefone=?,chave_pix=? WHERE pessoa_id=?",
                (vars_["nome"].get(), vars_["telefone"].get(),
                 vars_["chave_pix"].get(), pid))
            win.destroy()
            self._pessoas_load(); self._refresh_all()
        btn(form,"💾 SALVAR",CORES["btn_verde"],salvar,width=18).grid(
            row=8, column=0, pady=12, sticky="e")


    def _build_publicar(self):
        # Era um notebook com 2 sub-abas ("Publicar no Site" e a extinta
        # "Sincronizar Participantes" — nunca funcionava, nada escrevia na
        # fila que ela lia). Só sobrou "Publicar", então empacota direto
        # sem o notebook interno — não faz sentido uma aba só dentro de
        # outro nível de abas.
        p = self.tab_pub
        outer = tk.Frame(p, bg=CORES["bg_frame"])
        outer.pack(fill="both", expand=True, padx=20, pady=(16,6))
        sec = section(outer, "🌐 PUBLICAR BOLÃO NO SITE")
        sec.pack(fill="x")
        tk.Label(sec, text="Site: https://eltonluisoc.github.io/mega-sena-sistema/",
                 bg=CORES["bg_section"], fg="#aad4f5",
                 font=("Arial",9,"italic")).pack(anchor="w", pady=(0,8))
        r1 = tk.Frame(sec, bg=CORES["bg_section"]); r1.pack(fill="x", pady=4)
        tk.Label(r1, text="Bolão:", bg=CORES["bg_section"],
                 fg=CORES["fg_label"], font=("Arial",9,"bold")).pack(side="left")
        self._pub_cb = ttk.Combobox(r1, width=40, state="readonly", font=("Arial",9))
        self._pub_cb.pack(side="left", padx=8)
        # Loteria inferida automaticamente do nome do bolão — sem campo manual
        bf = tk.Frame(sec, bg=CORES["bg_section"]); bf.pack(fill="x", pady=10)
        btn(bf, "👁 Pré-visualizar",  CORES["btn_azul"],    self._pub_preview, width=18).pack(side="left", padx=4)
        btn(bf, "🌐 PUBLICAR NO SITE",CORES["btn_verde"],   self._pub_enviar,  width=22).pack(side="left", padx=4)
        btn(bf, "🗑 REMOVER DO SITE", CORES["btn_vermelho"],self._pub_remover, width=20).pack(side="left", padx=4)
        self._pub_status = tk.Label(sec, text="", bg=CORES["bg_section"],
                                     fg="#1D9E75", font=("Arial",10,"bold"))
        self._pub_status.pack(anchor="w", pady=4)
        sec2 = section(outer, "PRÉ-VISUALIZAÇÃO")
        sec2.pack(fill="both", expand=True, pady=(12,0))
        self._pub_txt = tk.Text(sec2, font=("Courier",9), wrap="word",
                                 bg="#1a2a3a", fg="#aad4f5", relief="flat", padx=8, pady=8)
        vsb = ttk.Scrollbar(sec2, orient="vertical", command=self._pub_txt.yview)
        self._pub_txt.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y"); self._pub_txt.pack(fill="both", expand=True)

    def _pub_montar_dados(self):
        try:
            return self._pub_montar_dados_impl()
        except Exception as ex:
            import traceback; traceback.print_exc()
            messagebox.showerror("Erro interno", "Erro ao montar dados do bolao:\n" + str(ex))
            return None

    def _pub_montar_dados_impl(self):
        import re as _re
        sel = self._pub_cb.get()
        if not sel:
            messagebox.showwarning("Atenção","Selecione um bolão!"); return None
        m2 = _re.search(r"\(ID:\s*(\d+)\)", sel)
        if not m2:
            messagebox.showwarning("Atenção","Selecione um bolão válido!"); return None
        bid = int(m2.group(1))
        b   = self.db.fetchone("SELECT * FROM boloes WHERE id=?", (bid,))
        if not b: return None
        bd  = dict(b)
        vt  = float(bd.get("valor_total", 0) or 0)
        if vt <= 0:
            # Mesma trava que a sincronizacao automatica do fechamento ja
            # tem (pula bolao sem valor de cota) - faltava aqui, entao
            # dava pra publicar manualmente um bolao com valorPorCota=0.
            messagebox.showwarning("Atenção",
                "Este bolão não tem valor de cota configurado. Edite o bolão e "
                "informe o valor antes de publicar.")
            return None
        partic = self.db.fetchall(
            "SELECT * FROM participantes WHERE bolao_id=? AND ativo=1 ORDER BY nome", (bid,))
        adm_nome     = bd.get("adm_nome","").strip()
        adm_paga     = bd.get("adm_paga", 0)
        adm_nome_low = adm_nome.lower()
        # Bulk query pagamentos
        pag_rows = self.db.fetchall(
            "SELECT participante_id, SUM(valor) as total_pago, MIN(data_pagamento) as dt_cad "
            "FROM pagamentos WHERE bolao_id=? GROUP BY participante_id", (bid,))
        pag_map = {r["participante_id"]: r for r in pag_rows}
        adm_cadastrado = False; lista_part = []
        for pt in partic:
            pt_d = dict(pt); pid = pt_d["id"]
            pr   = pag_map.get(pid)
            pago = float(pr["total_pago"] or 0) if pr else 0.0
            dt_cad = pr["dt_cad"] if pr and pr["dt_cad"] else ""
            ve   = float(pt_d["valor_esperado"] or 0)
            eh_adm = bool(pt_d.get("is_adm")) or (adm_nome_low and adm_nome_low in pt_d["nome"].lower())
            if eh_adm: adm_cadastrado = True
            if eh_adm and not adm_paga:
                n_cotas=1; situacao="quitado"; valorPago=0
            else:
                n_cotas  = max(1, round(ve/vt)) if vt>0 and ve>0 else 1
                saldo    = max(0, ve-pago)
                situacao = "quitado" if saldo<=0 else "em_andamento"
                valorPago= int(round(pago))
            import re as _re_tel
            tel_raw = str(pt_d.get("telefone") or "")
            tel_digits = _re_tel.sub(r"\D", "", tel_raw)
            lista_part.append({"nome":pt_d["nome"],"telefone":tel_digits,
                "valorPago":valorPago,"situacao":situacao,
                "quantidadeCotas":n_cotas,"dataCadastro":dt_cad})
        if adm_nome and not adm_paga and not adm_cadastrado:
            lista_part.append({"nome":adm_nome,"telefone":"","valorPago":0,
                "situacao":"quitado","quantidadeCotas":1,"dataCadastro":""})
        # Loteria inferida do nome do bolão
        nome_lower = bd["nome"].lower()
        if "lotofacil" in nome_lower or "lotofácil" in nome_lower:
            loteria = "lotofacil"
        elif "mega" in nome_lower:
            loteria = "mega"
        elif "quina" in nome_lower:
            loteria = "quina"
        elif "lotomania" in nome_lower:
            loteria = "lotomania"
        else:
            loteria = "mega"
        # Data limite: usa data do concurso do bolão ou vazia
        data_limite = bd.get("data_concurso") or ""
        if data_limite:
            try:
                from datetime import datetime as _dt2
                if "/" in data_limite:
                    data_limite = _dt2.strptime(data_limite, "%d/%m/%Y").strftime("%Y-%m-%d")
            except: pass
        return {"titulo":bd["nome"],"loteria":loteria,"valorPorCota":round(vt, 2),
                "dataLimite":data_limite,"participantes":lista_part,
                "_bolao_id":bid,"_firebase_doc_id":bd.get("firebase_doc_id") or None}

    def _pub_preview(self):
        dados = self._pub_montar_dados()
        if not dados: return
        import json
        # Mostra com data formatada para leitura humana
        exibir = dict(dados)
        exibir["dataLimite"] = dados.get("dataLimiteFmt", dados["dataLimite"])
        exibir.pop("dataLimiteFmt", None)
        doc_id = _bolao_doc_id(dados["titulo"])
        preview_header = (
            "ID Firebase: " + doc_id + "\n"
            "URL: " + FIREBASE_BASE + "/" + doc_id + "\n"
            + "=" * 60 + "\n"
        )
        self._pub_txt.configure(state="normal")
        self._pub_txt.delete("1.0","end")
        self._pub_txt.insert("1.0", preview_header + json.dumps(exibir, ensure_ascii=False, indent=2))
        self._pub_txt.configure(state="disabled")

    def _pub_enviar(self):
        dados = self._pub_montar_dados()
        if not dados: return
        self._pub_status.configure(text="⏳ Publicando no Firebase...", fg="#aad4f5")
        self.root.update_idletasks()
        def _executar():
            try:
                resultado = enviar_bolao_para_site(
                    titulo=dados["titulo"], loteria=dados["loteria"],
                    valor_cota=dados["valorPorCota"], data_limite=dados["dataLimite"],
                    participantes=dados["participantes"],
                    doc_id_fixo=dados.get("_firebase_doc_id"))
                # Primeira publicação deste bolão: guarda o ID gerado pra
                # sempre reaproveitar dele em diante, mesmo se o nome mudar
                if resultado.get("sucesso") and not dados.get("_firebase_doc_id") and resultado.get("doc_id"):
                    self.db.execute(
                        "UPDATE boloes SET firebase_doc_id=? WHERE id=?",
                        (resultado["doc_id"], dados["_bolao_id"]))
            except Exception as ex:
                resultado = {"sucesso":False,"erro":str(ex),"status":0,"doc_id":""}
            if resultado.get("excluido_no_site"):
                # Bolão foi excluído no site: em vez de recriar o documento,
                # reflete isso localmente encerrando o bolão.
                try:
                    self.db.execute(
                        "UPDATE boloes SET encerrado=1 WHERE id=?", (dados["_bolao_id"],))
                except Exception: pass
            def _ui():
                if resultado.get("sucesso"):
                    self._pub_status.configure(
                        text="✅ Publicado com sucesso!", fg="#1D9E75")
                elif resultado.get("excluido_no_site"):
                    self._pub_status.configure(
                        text="⚠️ Bolão foi excluído no site — marcado como encerrado aqui.",
                        fg="#f39c12")
                    messagebox.showwarning("Bolão excluído no site",
                        "Este bolão foi excluído no site e NÃO será recriado "
                        "automaticamente.\n\nEle foi marcado como encerrado neste "
                        "programa para não tentar publicá-lo de novo.")
                else:
                    erro = resultado.get("erro") or "Erro desconhecido"
                    http_s = resultado.get("status",0)
                    if http_s == 403:
                        dica = "\n\nFirebase 403: verifique regras Firestore."
                    elif http_s == 0:
                        dica = "\n\nVerifique sua conexao."
                    else:
                        dica = ""
                    self._pub_status.configure(
                        text="\u274c Erro HTTP "+str(http_s)+": "+erro[:80], fg="#e74c3c")
                    messagebox.showerror("Erro ao publicar",
                        "Nao foi possivel publicar.\n\n"+erro+dica)
            self.root.after(0, _ui)
        import threading; threading.Thread(target=_executar, daemon=True).start()

    def _pub_remover(self):
        dados = self._pub_montar_dados()
        if not dados: return
        if not messagebox.askyesno("Confirmar",
            f"Remover o bolão '{dados['titulo']}' do site?\n\n"
            f"O documento será excluído do Firebase."):
            return
        self._pub_status.configure(text="⏳ Removendo...", fg="#aad4f5")
        self.root.update()
        resultado = remover_bolao_do_site(dados["titulo"], doc_id_fixo=dados.get("_firebase_doc_id"))
        if resultado["sucesso"]:
            try:
                self.db.execute(
                    "UPDATE boloes SET encerrado=1 WHERE id=?", (dados["_bolao_id"],))
            except Exception: pass
            self._pub_status.configure(
                text="✅ Bolão removido do site com sucesso!", fg="#1D9E75")
        else:
            self._pub_status.configure(
                text=f"❌ Erro: {resultado['erro']}", fg="#e74c3c")
            messagebox.showerror("Erro ao remover", resultado["erro"])


    # ════════════════════════════════════════════════════════════
    #  [SINCRONIZAÇÃO — Firebase participantes_pendentes — REMOVIDA]
    #  Nada mais escreve na coleção participantes_pendentes desde que o
    #  site parou de usar esse fluxo — a fila nunca recebia nada, então
    #  "Sincronizar Participantes" nunca tinha o que mostrar. A limpeza
    #  de pagamentos órfãos (_sinc_limpar_orfaos, logo abaixo) é
    #  independente disso e continua — só mudou de lugar na tela
    #  (agora é um botão em "📋 Histórico").
    # ════════════════════════════════════════════════════════════
    def _sinc_limpar_orfaos(self):
        orfaos = self.db.fetchall("""
            SELECT pg.id as pag_id, pg.participante_id, pg.bolao_id,
                   pg.valor, pg.data_pagamento, pg.mes_referencia, pg.depositado,
                   COALESCE(pt.nome,'(removido)') as nome_part,
                   pt.ativo as part_ativo, b.nome as bolao_nome
            FROM pagamentos pg
            LEFT JOIN participantes pt ON pg.participante_id=pt.id
            LEFT JOIN boloes b ON pg.bolao_id=b.id
            WHERE pt.id IS NULL OR pt.ativo=0
            ORDER BY pg.data_pagamento DESC""")
        if not orfaos:
            messagebox.showinfo("Pagamentos Orfaos","Nenhum pagamento orfao encontrado."); return
        win = tk.Toplevel(self.root)
        win.title("Diagnostico de Pagamentos Orfaos")
        win.geometry("920x540"); win.configure(bg=CORES["bg_frame"])
        win.grab_set(); win.lift()
        hdr = tk.Frame(win, bg="#1a3a5a", padx=12, pady=8); hdr.pack(fill="x")
        tk.Label(hdr, text=str(len(orfaos))+" pagamento(s) vinculados a participantes inativos/removidos",
                 bg="#1a3a5a", fg="white", font=("Arial",10,"bold")).pack(anchor="w")
        tk.Label(hdr, text="Amarelo=inativo (regularizavel)  |  Laranja=removido do banco",
                 bg="#1a3a5a", fg="#aad4f5", font=("Arial",8)).pack(anchor="w")
        fr_t = tk.Frame(win, bg=CORES["bg_frame"]); fr_t.pack(fill="both",expand=True,padx=12,pady=8)
        cols_o = {"Participante":210,"Bolao":160,"Data":90,"Mes":80,"Valor":90,"Dep.":70,"Situacao":160}
        tv = ttk.Treeview(fr_t, columns=list(cols_o.keys()), show="headings",
                          selectmode="extended", height=13)
        for col,w in cols_o.items():
            tv.heading(col,text=col); tv.column(col,width=w,minwidth=w)
        vsb = ttk.Scrollbar(fr_t,orient="vertical",command=tv.yview)
        tv.configure(yscrollcommand=vsb.set); vsb.pack(side="right",fill="y"); tv.pack(fill="both",expand=True)
        tv.tag_configure("removido",background="#fde8d8")
        tv.tag_configure("inativo", background="#fff9c4")
        for o in orfaos:
            tag = "removido" if o["part_ativo"] is None else "inativo"
            sit = "REMOVIDO" if o["part_ativo"] is None else "Inativo"
            dep = "Sim" if o["depositado"] else "Nao"
            vf  = "R$ "+"{:.2f}".format(float(o["valor"] or 0)).replace(".",",")
            tv.insert("","end",iid=str(o["pag_id"]),tags=(tag,),values=(
                o["nome_part"],o["bolao_nome"] or "—",
                o["data_pagamento"] or "—",o["mes_referencia"] or "—",vf,dep,sit))
        bf = tk.Frame(win,bg=CORES["bg_frame"]); bf.pack(fill="x",padx=12,pady=6)
        def sel_todos(): tv.selection_set(tv.get_children())
        def des_todos(): tv.selection_remove(tv.get_children())
        def regularizar():
            sel = tv.selection()
            if len(sel)!=1:
                messagebox.showwarning("Atencao","Selecione exatamente 1 pagamento."); return
            pag_id = int(sel[0])
            pag = self.db.fetchone("SELECT * FROM pagamentos WHERE id=?", (pag_id,))
            if not pag: return
            pt_in = self.db.fetchone("SELECT nome FROM participantes WHERE id=?", (pag["participante_id"],))
            nome_in = (pt_in["nome"] if pt_in else "").strip()
            ativos = self.db.fetchall(
                "SELECT id,nome FROM participantes WHERE bolao_id=? AND ativo=1 ORDER BY nome",
                (pag["bolao_id"],))
            if not ativos:
                messagebox.showwarning("Sem ativos","Nenhum participante ativo neste bolao."); return
            w2 = tk.Toplevel(win); w2.title("Regularizar"); w2.geometry("480x220")
            w2.configure(bg=CORES["bg_section"]); w2.grab_set(); w2.lift()
            tk.Label(w2,text="Vincular pagamento ao participante ATIVO:",
                     bg=CORES["bg_section"],fg=CORES["fg_label"],
                     font=("Arial",9,"bold")).pack(pady=(14,4),padx=16,anchor="w")
            vf2 = "R$ "+"{:.2f}".format(float(pag["valor"] or 0)).replace(".",",")
            tk.Label(w2,text="Pagamento: "+(pag["data_pagamento"] or "")+" — "+vf2+
                     " | Atual: "+(nome_in or "removido"),
                     bg=CORES["bg_section"],fg="#888",font=("Arial",8),wraplength=440).pack(padx=16,anchor="w")
            tk.Label(w2,text="Vincular a:",bg=CORES["bg_section"],
                     fg=CORES["fg_label"],font=("Arial",9,"bold")).pack(pady=(8,2),padx=16,anchor="w")
            import unicodedata as _ud, re as _re3
            def norm2(s):
                s=_ud.normalize("NFD",s)
                return "".join(ch for ch in s if _ud.category(ch)!="Mn").lower()
            nomes_at = [r["nome"]+" (ID: "+str(r["id"])+")" for r in ativos]
            cb2 = ttk.Combobox(w2,values=nomes_at,width=46,font=("Arial",9),state="readonly")
            cb2.pack(padx=16,pady=4)
            if nome_in:
                melhor = max(ativos,key=lambda r:len(set(norm2(nome_in).split())&set(norm2(r["nome"]).split())),default=None)
                if melhor: cb2.set(melhor["nome"]+" (ID: "+str(melhor["id"])+")")
            def confirmar():
                m3 = _re3.search(r"\(ID: (\d+)\)",cb2.get())
                if not m3: messagebox.showwarning("Atencao","Selecione um participante!"); return
                self.db.execute("UPDATE pagamentos SET participante_id=? WHERE id=?",(int(m3.group(1)),pag_id))
                messagebox.showinfo("OK","Pagamento regularizado!")
                w2.destroy(); win.destroy(); self._refresh_all(); self._sinc_limpar_orfaos()
            bf2 = tk.Frame(w2,bg=CORES["bg_section"]); bf2.pack(pady=10)
            btn(bf2,"Vincular",CORES["btn_verde"],confirmar,width=14).pack(side="left",padx=6)
            btn(bf2,"Cancelar",CORES["btn_cinza"],w2.destroy,width=10).pack(side="left",padx=6)
        def excluir_sel():
            sel = tv.selection()
            if not sel: messagebox.showwarning("Atencao","Selecione ao menos um!"); return
            total = sum(float(tv.item(s,"values")[4].replace("R$ ","").replace(",",".")) for s in sel)
            if messagebox.askyesno("Confirmar",
                str(len(sel))+" pagamento(s) serao excluidos.\nTotal: R$ "+
                "{:.2f}".format(total)+"\n\nEsta acao nao pode ser desfeita."):
                for pid2 in sel:
                    self.db.execute("DELETE FROM pagamentos WHERE id=?",(int(pid2),))
                messagebox.showinfo("Concluido",str(len(sel))+" pagamento(s) excluido(s).")
                win.destroy(); self._refresh_all()
        btn(bf,"Selecionar Todos",CORES["btn_cinza"],sel_todos,width=16).pack(side="left",padx=4)
        btn(bf,"Desmarcar Todos",CORES["btn_cinza"],des_todos,width=16).pack(side="left",padx=4)
        btn(bf,"Regularizar (1 sel.)",CORES["btn_verde"],regularizar,width=20).pack(side="left",padx=4)
        btn(bf,"Excluir Selecionados",CORES["btn_vermelho"],excluir_sel,width=20).pack(side="left",padx=4)
        btn(bf,"Fechar",CORES["btn_azul"],win.destroy,width=10).pack(side="right",padx=4)

    def _build_bkp(self):
        p = self.tab_bkp
        sa=section(p,"AÇÕES PRINCIPAIS"); sa.pack(fill="x",padx=20,pady=(16,8))
        row=tk.Frame(sa,bg=CORES["bg_section"]); row.pack(pady=6)
        btn(row,"💾 Backup Manual",CORES["btn_verde"],self._bkp_manual,width=20).pack(side="left",padx=6)
        btn(row,"📋 Listar Backups",CORES["btn_azul"],self._bkp_listar,width=18).pack(side="left",padx=6)
        btn(row,"📂 Importar Backup",CORES["btn_roxo"],self._bkp_importar,width=18).pack(side="left",padx=6)

        sl=section(p,"BACKUPS DISPONÍVEIS"); sl.pack(fill="both",expand=True,padx=20,pady=8)
        cols={"Nome":300,"Tipo":100,"Data/Hora":160,"Tamanho":100}
        fr,self.bkp_tree=make_tree(sl,cols,height=12); fr.pack(fill="both",expand=True)

        bf=tk.Frame(p,bg=CORES["bg_frame"]); bf.pack(fill="x",padx=20,pady=8)
        btn(bf,"♻ Restaurar Backup",CORES["btn_laranja"],self._bkp_restaurar,width=22).pack(side="left",padx=4)
        btn(bf,"🗑 Excluir Backup",CORES["btn_vermelho"],self._bkp_excluir,width=20).pack(side="left",padx=4)

        si=section(p,"INFO DO SISTEMA"); si.pack(fill="x",padx=20,pady=(0,16))
        self._bkp_info=tk.Label(si,text="",bg=CORES["bg_section"],fg=CORES["fg_label"],
                                 font=("Arial",9),justify="left")
        self._bkp_info.pack(anchor="w")
        self._bkp_listar()

    def _bkp_atualizar_info(self):
        bkps=self.bkp.listar()
        db_size=os.path.getsize(DB_FILE)/1024/1024 if os.path.exists(DB_FILE) else 0
        db_mtime=""
        if os.path.exists(DB_FILE):
            db_mtime=datetime.fromtimestamp(os.path.getmtime(DB_FILE)).strftime("%d/%m/%Y %H:%M")
        tot_kb=sum(os.path.getsize(b["path"]) for b in bkps)/1024
        self._bkp_info.configure(text=(
            f"Backups manuais: {sum(1 for b in bkps if b['tipo']=='Manual')}   |   "
            f"Automáticos: {sum(1 for b in bkps if b['tipo']=='Automático')}   |   "
            f"Total: {len(bkps)}   |   Espaço: {tot_kb:.2f} KB\n"
            f"BD atual: {db_size:.2f} MB   |   Última modificação: {db_mtime}\n"
            f"⚠ Faça backups regulares!"))

    def _bkp_listar(self):
        self.bkp_tree.delete(*self.bkp_tree.get_children())
        for b in self.bkp.listar():
            self.bkp_tree.insert("","end",iid=b["path"],values=(b["nome"],b["tipo"],b["data"],b["tamanho"]))
        self._bkp_atualizar_info()

    def _bkp_manual(self):
        dest=self.bkp.fazer_backup(auto=False)
        messagebox.showinfo("Backup Feito",f"Salvo em:\n{dest}"); self._bkp_listar()

    def _bkp_importar(self):
        path=filedialog.askopenfilename(title="Selecionar backup",
             filetypes=[("SQLite DB","*.db"),("Todos","*.*")])
        if not path: return
        dest=os.path.join(BACKUP_DIR,f"importado_{os.path.basename(path)}")
        shutil.copy2(path,dest)
        messagebox.showinfo("Importado","Backup importado!"); self._bkp_listar()

    def _bkp_restaurar(self):
        sel=self.bkp_tree.selection()
        if not sel: messagebox.showwarning("Atenção","Selecione um backup!"); return
        if messagebox.askyesno("Confirmar",
            "Restaurar este backup?\nOS DADOS ATUAIS SERÃO SUBSTITUÍDOS.\nO sistema será encerrado."):
            self.bkp.fazer_backup(auto=False)
            self.bkp.restaurar(sel[0])
            messagebox.showinfo("Restaurado","Reinicie o sistema."); self.root.destroy()

    def _bkp_excluir(self):
        sel=self.bkp_tree.selection()
        if not sel: messagebox.showwarning("Atenção","Selecione um backup!"); return
        if messagebox.askyesno("Confirmar","Excluir este backup?"):
            os.remove(sel[0]); self._bkp_listar()

    # ════════════════════════════════════════════════════════════
    #  GERENCIAR BOLÕES  ← CORRIGIDO
    # ════════════════════════════════════════════════════════════
    def _gerenciar_boloes(self):
        win=tk.Toplevel(self.root)
        win.title("Gerenciar Bolões")
        win.geometry("980x520")
        win.configure(bg=CORES["bg_section"])
        win.grab_set()

        tk.Label(win,text="GERENCIAMENTO DE BOLÕES",bg=CORES["bg_section"],
                 fg=CORES["fg_title"],font=("Arial",13,"bold")).pack(pady=10)

        cols={"ID":50,"Nome":220,"Loteria":120,"Início":100,
              "Part.":70,"Valor Total":120,"Parcela":110,"Status":80,"ADM":120}
        fr,tree=make_tree(win,cols,height=14)
        fr.pack(fill="both",expand=True,padx=10,pady=4)

        def carregar():
            tree.delete(*tree.get_children())
            rows=self.db.fetchall("SELECT * FROM boloes ORDER BY id")
            for b in rows:
                bd=dict(b)
                adm_info=(f"{bd.get('adm_nome','')} ({'Paga' if bd.get('adm_paga') else 'Não paga'})"
                          if bd.get("adm_nome") else "Não config.")
                tree.insert("","end",iid=str(bd["id"]),values=(
                    bd["id"],bd["nome"],bd.get("loteria","Mega-Sena"),bd["data_inicio"],
                    bd["num_participantes"],fmt_brl(bd["valor_total"]),
                    fmt_brl(bd["valor_parcela"]),bd["status"],adm_info))
        carregar()

        bf=tk.Frame(win,bg=CORES["bg_section"]); bf.pack(pady=10)

        def ativar():
            sel=tree.selection()
            if not sel: messagebox.showwarning("Atenção","Selecione um bolão!"); return
            bid=int(sel[0])
            self.db.execute("UPDATE boloes SET status='INATIVO'")
            self.db.execute("UPDATE boloes SET status='ATIVO' WHERE id=?",(bid,))
            self.bid.set(bid)
            self._load_boloes_combo(); self._refresh_all(); carregar()
            messagebox.showinfo("Ativo","Bolão ativado!")

        def editar():
            sel=tree.selection()
            if not sel: messagebox.showwarning("Atenção","Selecione um bolão!"); return
            bid=int(sel[0])
            b=self.db.fetchone("SELECT * FROM boloes WHERE id=?",(bid,))
            win.destroy()
            self._form_bolao(dict(b))

        def excluir():
            sel=tree.selection()
            if not sel: messagebox.showwarning("Atenção","Selecione um bolão!"); return
            bid=int(sel[0])
            b=self.db.fetchone("SELECT nome FROM boloes WHERE id=?",(bid,))
            if messagebox.askyesno("Confirmar",f"Excluir '{b['nome']}'? TUDO será apagado!"):
                for tbl in ["pagamentos","participantes","premiacoes","reserva_caixa",
                            "saques_emergenciais","taxa_adm","boloes"]:
                    if tbl=="boloes":
                        self.db.execute("DELETE FROM boloes WHERE id=?",(bid,))
                    else:
                        self.db.execute(f"DELETE FROM {tbl} WHERE bolao_id=?",(bid,))
                carregar(); self._load_boloes_combo()

        def config_adm():
            sel=tree.selection()
            if not sel: messagebox.showwarning("Atenção","Selecione um bolão!"); return
            bid=int(sel[0])
            b=self.db.fetchone("SELECT * FROM boloes WHERE id=?",(bid,))
            bd=dict(b)
            self._config_adm_win(bid,bd,lambda: carregar())

        btn(bf,"✅ Ativar",CORES["btn_verde"],ativar,width=14).pack(side="left",padx=4)
        btn(bf,"✏ Editar",CORES["btn_laranja"],editar,width=14).pack(side="left",padx=4)
        btn(bf,"🗑 Excluir",CORES["btn_vermelho"],excluir,width=14).pack(side="left",padx=4)
        btn(bf,"👤 Config. ADM",CORES["btn_roxo"],config_adm,width=16).pack(side="left",padx=4)
        btn(bf,"❌ Fechar",CORES["btn_cinza"],win.destroy,width=12).pack(side="left",padx=4)

    def _config_adm_win(self, bid, bd, on_save=None):
        """Configura o ADM do bolão: nome e se paga ou não."""
        win=tk.Toplevel(self.root)
        win.title("Configurar ADM do Bolão")
        win.geometry("460x280")
        win.configure(bg=CORES["bg_section"])
        win.grab_set()
        tk.Label(win,text="CONFIGURAÇÃO DO ADMINISTRADOR",bg=CORES["bg_section"],
                 fg=CORES["fg_title"],font=("Arial",12,"bold")).pack(pady=12)
        tk.Label(win,text="Esta configuração é interna. Não aparece explicitamente nos relatórios.",
                 bg=CORES["bg_section"],fg="#888",font=("Arial",8,"italic")).pack()

        form=tk.Frame(win,bg=CORES["bg_section"],padx=30); form.pack(fill="both",expand=True,pady=10)
        form.columnconfigure(0,weight=1)

        tk.Label(form,text="Nome do ADM (seu nome no bolão):",bg=CORES["bg_section"],
                 font=("Arial",9,"bold"),fg=CORES["fg_label"]).grid(row=0,column=0,sticky="w",pady=(8,0))
        adm_nome_e=entry(form,width=38)
        adm_nome_e.insert(0,bd.get("adm_nome",""))
        adm_nome_e.grid(row=1,column=0,sticky="ew")

        paga_var=tk.IntVar(value=int(bd.get("adm_paga",0)))
        tk.Label(form,text="O ADM participa pagando as parcelas?",bg=CORES["bg_section"],
                 font=("Arial",9,"bold"),fg=CORES["fg_label"]).grid(row=2,column=0,sticky="w",pady=(16,4))
        pf=tk.Frame(form,bg=CORES["bg_section"]); pf.grid(row=3,column=0,sticky="w")
        tk.Radiobutton(pf,text="✅ Sim, pago normalmente",variable=paga_var,value=1,
                       bg=CORES["bg_section"],font=("Arial",9)).pack(side="left",padx=8)
        tk.Radiobutton(pf,text="❌ Não, sou isento",variable=paga_var,value=0,
                       bg=CORES["bg_section"],font=("Arial",9)).pack(side="left")

        tk.Label(form,text="Obs: se 'Não pago', você aparece como Quitado nos relatórios\n"
                           "e o nº de pagantes é calculado sem você.",
                 bg=CORES["bg_section"],fg="#888",font=("Arial",8,"italic")).grid(
            row=4,column=0,sticky="w",pady=4)

        def salvar():
            nome=adm_nome_e.get().strip()
            paga=paga_var.get()
            self.db.execute("UPDATE boloes SET adm_nome=?,adm_paga=? WHERE id=?",
                            (nome,paga,bid))
            # Marcar participante com esse nome como is_adm
            if nome:
                self.db.execute("UPDATE participantes SET is_adm=0 WHERE bolao_id=?",(bid,))
                self.db.execute(
                    "UPDATE participantes SET is_adm=1 WHERE bolao_id=? AND LOWER(nome) LIKE LOWER(?)",
                    (bid,f"%{nome}%"))
            messagebox.showinfo("Salvo","Configuração ADM salva!")
            if on_save: on_save()
            win.destroy()

        btn(form,"💾 SALVAR",CORES["btn_verde"],salvar,width=16).grid(row=5,column=0,pady=14,sticky="e")

    # ════════════════════════════════════════════════════════════
    #  CRIAR / EDITAR BOLÃO
    # ════════════════════════════════════════════════════════════
    def _novo_bolao(self):
        self._form_bolao(None)

    def _form_bolao(self, bd):
        win=tk.Toplevel(self.root)
        win.title("Criar Novo Bolão" if not bd else "Editar Bolão")
        win.geometry("540x560")
        win.configure(bg=CORES["bg_section"])
        win.grab_set()

        tk.Label(win,text="CRIAR NOVO BOLÃO" if not bd else "EDITAR BOLÃO",
                 bg=CORES["bg_section"],fg=CORES["fg_title"],font=("Arial",13,"bold")).pack(pady=12)
        form=tk.Frame(win,bg=CORES["bg_section"],padx=30); form.pack(fill="both",expand=True)
        form.columnconfigure(0,weight=1)

        def rf(r,label,default=""):
            tk.Label(form,text=label,bg=CORES["bg_section"],fg=CORES["fg_label"],
                     font=("Arial",9,"bold")).grid(row=r,column=0,sticky="w",pady=(8,0))
            w=entry(form,width=52)
            if default: w.insert(0,str(default))
            w.grid(row=r+1,column=0,sticky="ew"); return w

        nome_e  = rf(0,"Nome do Bolão:",         bd["nome"]        if bd else "")
        tk.Label(form,text="Loteria:",bg=CORES["bg_section"],fg=CORES["fg_label"],
                 font=("Arial",9,"bold")).grid(row=2,column=0,sticky="w",pady=(8,0))
        lot_var=tk.StringVar(value=bd.get("loteria","Mega-Sena") if bd else "Mega-Sena")
        ttk.Combobox(form,textvariable=lot_var,values=LOTERIAS,state="readonly",
                     width=28,font=("Arial",9)).grid(row=3,column=0,sticky="w")
        data_e  = rf(4,"Data de Início (DD/MM/AAAA):", bd["data_inicio"] if bd else date.today().strftime("%d/%m/%Y"))
        npart_e = rf(6,"Número de Participantes (total de cotas):", bd["num_participantes"] if bd else "")
        vtot_e  = rf(8,"Valor Total por Participante (R$):",
                     f"{bd['valor_total']:.2f}".replace(".",",") if bd else "")
        vpar_e  = rf(10,"Valor de Cada Parcela (R$):",
                     f"{bd['valor_parcela']:.2f}".replace(".",",") if bd else "")
        tk.Label(form,text="Descrição/Observações:",bg=CORES["bg_section"],fg=CORES["fg_label"],
                 font=("Arial",9,"bold")).grid(row=12,column=0,sticky="w",pady=(8,0))
        desc_t=tk.Text(form,height=3,width=52,relief="solid",bd=1,font=("Arial",9))
        if bd and bd.get("descricao"): desc_t.insert("1.0",bd["descricao"])
        desc_t.grid(row=13,column=0,sticky="ew")

        def salvar():
            nome=nome_e.get().strip()
            if not nome: messagebox.showwarning("Atenção","Informe o nome!"); return
            try:
                npart=int(npart_e.get())
                vtot=to_float(vtot_e.get())
                vpar=to_float(vpar_e.get())
            except:
                messagebox.showwarning("Atenção","Valores numéricos inválidos!"); return
            desc=desc_t.get("1.0","end").strip()
            lot=lot_var.get()
            di=data_e.get().strip()
            if bd:
                self.db.execute(
                    "UPDATE boloes SET nome=?,loteria=?,data_inicio=?,num_participantes=?,"
                    "valor_total=?,valor_parcela=?,descricao=? WHERE id=?",
                    (nome,lot,di,npart,vtot,vpar,desc,bd["id"]))
                messagebox.showinfo("Sucesso","Bolão atualizado!")
            else:
                c=self.db.execute(
                    "INSERT INTO boloes (nome,loteria,data_inicio,num_participantes,"
                    "valor_total,valor_parcela,descricao,status) VALUES (?,?,?,?,?,?,?,'ATIVO')",
                    (nome,lot,di,npart,vtot,vpar,desc))
                new_id=c.lastrowid
                self.db.execute("UPDATE boloes SET status='INATIVO' WHERE id!=?",(new_id,))
                self.db.execute("UPDATE boloes SET status='ATIVO' WHERE id=?",(new_id,))
                self.bid.set(new_id)
                messagebox.showinfo("Sucesso","Novo bolão criado e ativado!")
            win.destroy()
            self._load_boloes_combo()
            self._refresh_all()

        bf=tk.Frame(form,bg=CORES["bg_section"])
        bf.grid(row=20,column=0,sticky="e",pady=16)
        btn(bf,"💾 SALVAR BOLÃO",CORES["btn_verde"],salvar,width=18).pack(side="left",padx=6)
        btn(bf,"❌ CANCELAR",CORES["btn_cinza"],win.destroy,width=14).pack(side="left")

    # ════════════════════════════════════════════════════════════
    #  HELPERS GLOBAIS
    # ════════════════════════════════════════════════════════════
    def _load_boloes_combo(self):
        rows  = self.db.fetchall("SELECT * FROM boloes WHERE encerrado=0 ORDER BY id")
        items = [f"{dict(b)['nome']} (ID: {dict(b)['id']})" for b in rows]
        self.cb_bolao["values"] = items
        bid = self.bid.get()
        if not bid:
            ativo = self.db.fetchone("SELECT id FROM boloes WHERE status='ATIVO' ORDER BY id DESC")
            if ativo:
                bid = ativo["id"]
                self.bid.set(bid)
        if bid:
            for item in items:
                m = re.search(r"\(ID: (\d+)\)", item)
                if m and int(m.group(1)) == bid:
                    self.cb_bolao.set(item)
                    break
        # ← CORREÇÃO: popula combos de participantes ao iniciar
        # _load_prem e _load_res são gerais — sempre carregam
        self._load_prem()
        self._load_res()
        if bid:
            self._refresh_all()

    def _on_bolao_sel(self, e=None):
        sel=self.cb_bolao.get()
        if not sel: return
        m=re.search(r"\(ID: (\d+)\)",sel)
        if m:
            self._selecionar_bolao_por_id(int(m.group(1)))

    def _selecionar_bolao_por_id(self, bid):
        """Seleciona um bolão — usado tanto pelo combo do cabeçalho quanto
        pelos cartões clicáveis da Visão Geral. Atualiza bid, status no
        banco, e recarrega tudo que depende do bolão selecionado."""
        self.bid.set(bid)
        self.db.execute("UPDATE boloes SET status='INATIVO'")
        self.db.execute("UPDATE boloes SET status='ATIVO' WHERE id=?",(bid,))
        # mantém a combobox do cabeçalho sincronizada com a seleção
        try:
            for item in self.cb_bolao["values"]:
                m2 = re.search(r"\(ID: (\d+)\)", item)
                if m2 and int(m2.group(1)) == bid:
                    self.cb_bolao.set(item)
                    break
        except Exception:
            pass
        self._refresh_all()
        self._preencher_valor_cad()

    def _pub_carregar_boloes(self):
        boloes = self.db.fetchall(
            "SELECT id, nome FROM boloes WHERE encerrado=0 ORDER BY nome")
        vals = [b["nome"] + " (ID: " + str(b["id"]) + ")" for b in boloes]
        self._pub_cb["values"] = vals
        if vals and not self._pub_cb.get():
            self._pub_cb.current(0)

    def _refresh_all(self):
        bid = self.bid.get()
        # Todos os participantes ativos (para editar/visualizar)
        todos = self.db.fetchall(
            "SELECT * FROM participantes WHERE bolao_id=? AND ativo=1 ORDER BY nome",(bid,))
        items_todos = [f"{dict(p)['nome']} (ID: {dict(p)['id']})" for p in todos]

        self.pag_cb["values"]  = items_todos
        self.vis_cb["values"]  = items_todos
        try: self.cad_edit_cb["values"] = items_todos
        except: pass
        try: self.pag_cb.set("")
        except: pass
        try:
            self.vis_cb.set("")
            self._vis_limpar()
        except: pass
        try: self.cad_edit_cb.set("")
        except: pass
        try: self._cad_lista_load()
        except: pass

        self._gerar_rel()
        self._dep_refresh()
        self._hist_load()
        self._load_prem()   # geral — mostra todos os bolões
        self._load_res()    # geral — mostra todos os bolões
        self._adm_load()
        self._dash_load()
        self._preencher_valor_cad()  # preenche valor esperado no cadastro
        self._rsv_load()
        try: self._pub_carregar_boloes()
        except: pass
    def _on_close(self):
        """Fecha com tela de sincronizacao redesenhada."""
        import threading, sqlite3 as _sq_c, re as _re_c, time as _time, json
        import unicodedata as _uc
        import urllib.request, urllib.error

        try: self.bkp.fazer_backup(auto=True)
        except Exception as ex: print("Backup erro:", ex)

        # ═══════════════════════════════════════════════════════════
        # JANELA DE SAÍDA — design premium
        # ═══════════════════════════════════════════════════════════
        win = tk.Toplevel(self.root)
        win.title("Encerrando o sistema")
        win.geometry("680x560")
        win.configure(bg="#0d1b2a")
        win.resizable(False, False)
        win.protocol("WM_DELETE_WINDOW", lambda: None)

        # Header com gradiente simulado
        hdr = tk.Frame(win, bg="#1a3a6e")
        hdr.pack(fill="x")
        tk.Label(hdr, text="  Sincronizando e encerrando o sistema",
                 bg="#1a3a6e", fg="white",
                 font=("Arial",13,"bold")).pack(anchor="w", padx=20, pady=(14,2))
        tk.Label(hdr, text="  Aguarde enquanto os dados sao enviados ao Firebase.",
                 bg="#1a3a6e", fg="#90caf9",
                 font=("Arial",9)).pack(anchor="w", padx=20, pady=(0,14))

        # Barra de progresso estilizada
        fr_prog = tk.Frame(win, bg="#0d1b2a")
        fr_prog.pack(fill="x", padx=24, pady=(14,4))
        tk.Label(fr_prog, text="Progresso geral", bg="#0d1b2a",
                 fg="#90caf9", font=("Arial",8,"bold")).pack(anchor="w")
        prog = ttk.Progressbar(fr_prog, length=630, mode="determinate",
                               style="Accent.Horizontal.TProgressbar")
        prog.pack(fill="x", pady=4)
        pct_lbl = tk.Label(fr_prog, text="0%", bg="#0d1b2a",
                           fg="#1D9E75", font=("Arial",9,"bold"))
        pct_lbl.pack(anchor="e")

        # Status atual
        status_lbl = tk.Label(win, text="Iniciando...",
                              bg="#0d1b2a", fg="#f39c12",
                              font=("Arial",10,"bold"))
        status_lbl.pack(pady=(0,6))

        # Log area com visual terminal
        fr_log = tk.Frame(win, bg="#050d15", relief="flat",
                          highlightbackground="#1a3a6e", highlightthickness=1)
        fr_log.pack(fill="both", expand=True, padx=24, pady=(0,10))
        vsb = tk.Scrollbar(fr_log, bg="#0d1b2a"); vsb.pack(side="right", fill="y")
        log_box = tk.Text(fr_log, bg="#050d15", fg="#c8e6ff",
                          font=("Consolas",9), relief="flat",
                          yscrollcommand=vsb.set, state="disabled",
                          padx=10, pady=8, spacing1=2)
        log_box.pack(fill="both", expand=True)
        vsb.configure(command=log_box.yview)
        log_box.tag_configure("ok",    foreground="#1D9E75")
        log_box.tag_configure("err",   foreground="#ff5252")
        log_box.tag_configure("warn",  foreground="#ffb74d")
        log_box.tag_configure("info",  foreground="#81d4fa")
        log_box.tag_configure("head",  foreground="#ffd54f", font=("Consolas",9,"bold"))
        log_box.tag_configure("muted", foreground="#37474f")
        log_box.tag_configure("name",  foreground="#e1f5fe", font=("Consolas",9,"bold"))
        log_box.tag_configure("dim",   foreground="#546e7a")

        # Footer
        footer = tk.Frame(win, bg="#1a3a6e", height=32)
        footer.pack(fill="x", side="bottom"); footer.pack_propagate(False)
        footer_lbl = tk.Label(footer, text="",
                              bg="#1a3a6e", fg="#90caf9", font=("Arial",9))
        footer_lbl.pack(side="left", padx=16, pady=6)
        btn_fechar_mesmo_assim = tk.Button(footer, text="Fechar mesmo assim",
            bg="#c0392b", fg="white", relief="flat", font=("Arial",9,"bold"),
            padx=10, activebackground="#a93226", activeforeground="white")
        # só aparece se a sincronização terminar com erro (ver _liberar_fechamento_manual)

        def log(msg, tag="info"):
            log_box.configure(state="normal")
            log_box.insert("end", msg+"\n", tag)
            log_box.see("end")
            log_box.configure(state="disabled")
            status_lbl.configure(text=msg[:90],
                fg={"ok":"#1D9E75","err":"#ff5252","warn":"#ffb74d",
                    "head":"#ffd54f"}.get(tag,"#81d4fa"))
            win.update_idletasks()

        def sp(v):
            prog["value"] = v
            pct_lbl.configure(text=str(int(v))+"%")
            win.update_idletasks()

        def _fechar_tudo():
            try: win.destroy()
            except: pass
            try: self.db.close()
            except: pass
            try: self.root.destroy()
            except: pass

        btn_fechar_mesmo_assim.configure(command=_fechar_tudo)

        def _liberar_fechamento_manual():
            """Chamada quando a sincronização termina com erro: não fecha
            sozinho (o usuário ficava sem saber que algo falhou), libera o
            X da janela e mostra um botão explícito de fechar."""
            win.protocol("WM_DELETE_WINDOW", _fechar_tudo)
            footer_lbl.configure(
                text="Sincronização com erro(s) — revise o log acima.")
            btn_fechar_mesmo_assim.pack(side="right", padx=16, pady=4)

        FIREBASE_RSV  = ("https://firestore.googleapis.com/v1/projects/mega-sena-sistema"
                         "/databases/(default)/documents/reservas_participantes/")
        FIREBASE_BASE = ("https://firestore.googleapis.com/v1/projects/mega-sena-sistema"
                         "/databases/(default)/documents/participantes/")
        FIREBASE_IDS  = {
            "mega da virada 2026":"bolao_mega_da_virada_2026",
            "mega da virada amigos 2026":"bolao_mega_da_virada_amigos_2026",
            "lotofacil da independencia":"bolao_lotofacil_da_independencia",
            "quina de sao joao 2026":"bolao_quina_de_sao_joao_2026",
            "quina de sao joao 2026 ii":"bolao_quina_de_sao_joao_2026_ii",
            "mega-sena 30 anos":"bolao_mega_sena_30_anos",
        }

        def norm_id(nome, tel):
            def n(s):
                s = _uc.normalize("NFD", s)
                return "".join(c for c in s if _uc.category(c)!="Mn")
            partes = [_re_c.sub(r"[^a-zA-Z0-9]","",n(p)) for p in nome.strip().split() if p]
            id_n = (partes[0]+"_"+partes[-1]) if len(partes)>1 else (partes[0] if partes else "x")
            tel_d = _re_c.sub(r"\D","",tel or "")
            return (id_n+"_"+tel_d) if tel_d else id_n

        def slug(s):
            s = _uc.normalize("NFD", s)
            s = "".join(c for c in s if _uc.category(c)!="Mn")
            return _re_c.sub(r"[^a-zA-Z0-9]+","_",s).strip("_").lower()

        def _run():
            try:
                _conn = _sq_c.connect(DB_FILE)
                _conn.row_factory = _sq_c.Row
                class _DB2:
                    def __init__(self,c): self.conn=c
                    def fetchall(self,s,p=()): return self.conn.execute(s,p).fetchall()
                    def fetchone(self,s,p=()): return self.conn.execute(s,p).fetchone()
                db2 = _DB2(_conn)
                from datetime import datetime as _dtnow
                ts_now = _dtnow.utcnow().strftime("%Y-%m-%dT%H:%M:%S.000Z")

                # ═══ RESERVAS ═══════════════════════════════════════
                win.after(0, lambda: log("", "dim"))
                win.after(0, lambda: log("  RESERVAS PESSOAIS", "head"))
                win.after(0, lambda: log("  "+"-"*50, "dim"))
                win.after(0, lambda: sp(2))
                win.after(0, lambda: footer_lbl.configure(text="Etapa 1/2 — Reservas pessoais"))

                pessoas = db2.fetchall(
                    "SELECT * FROM reservas_pessoas WHERE ativo=1 ORDER BY nome")
                total_p = len(pessoas)
                win.after(0, lambda n=total_p: log(
                    "  Total: "+str(n)+" participante(s)", "info"))

                ok_r = 0; err_r = []
                for idx_p, ps in enumerate(pessoas):
                    ps = dict(ps)
                    pid_r = ps["id"]; nome_r = ps["nome"] or "?"
                    tel_r  = ps.get("telefone") or ""
                    win.after(0, lambda nm=nome_r, i=idx_p, t=total_p: log(
                        "  ["+str(i+1).zfill(2)+"/"+str(t).zfill(2)+"] "+nm, "name"))

                    movs = db2.fetchall(
                        "SELECT * FROM reservas_movimentos WHERE pessoa_id=? ORDER BY id", (pid_r,))
                    cred = sum(float(m["valor"] or 0) for m in movs
                               if (m["tipo"] or "").upper() in
                               ("CREDITO","CRÉDITO","ENTRADA","DEPOSITO","DEPÓSITO"))
                    deb  = sum(float(m["valor"] or 0) for m in movs
                               if (m["tipo"] or "").upper() in
                               ("DEBITO","DÉBITO","SAQUE","USO"))
                    saldo_r = cred - deb
                    win.after(0, lambda s=saldo_r, n=len(movs): log(
                        "       saldo R$"+str(round(s,2))+" | "+str(n)+" movimentos", "dim"))

                    hist = []; sa = 0.0
                    for m in movs:
                        v = float(m["valor"] or 0)
                        tp_up = (m["tipo"] or "").upper()
                        eh_c = tp_up in ("CREDITO","CRÉDITO","ENTRADA","DEPOSITO","DEPÓSITO")
                        ant = sa; sa = sa+v if eh_c else sa-v
                        dm = m["data_mov"] or ts_now
                        try:
                            if len(dm)==10 and "/" in dm:
                                dm = _dtnow.strptime(dm,"%d/%m/%Y").strftime("%Y-%m-%dT00:00:00.000Z")
                        except: pass
                        tp_fb = "deposito" if eh_c else "saque"
                        desc_m = str(dict(m).get("descricao") or dict(m).get("loteria") or tp_fb)
                        hist.append({"mapValue":{"fields":{
                            "data":{"stringValue":dm},"tipo":{"stringValue":tp_fb},
                            "valor":{"doubleValue":v},"saldoAnterior":{"doubleValue":round(ant,2)},
                            "saldoNovo":{"doubleValue":round(sa,2)},
                            "descricao":{"stringValue":desc_m},
                        }}})

                    doc_id = norm_id(nome_r, tel_r)
                    doc = {"fields":{
                        "participanteId":{"stringValue":doc_id},
                        "nome":{"stringValue":nome_r},
                        "telefone":{"stringValue":_re_c.sub(r"\D","",tel_r)},
                        "saldoReserva":{"doubleValue":round(saldo_r,2)},
                        "dataAtualizacao":{"stringValue":ts_now},
                        "admin":{"booleanValue":True},
                        "historico":{"arrayValue":{"values":hist}},
                    }}
                    corpo = json.dumps(doc, ensure_ascii=False).encode("utf-8")
                    url_r = _firestore_patch_url(FIREBASE_RSV+doc_id,
                        ["participanteId","nome","telefone","saldoReserva","dataAtualizacao","admin","historico"])
                    try:
                        req = urllib.request.Request(url_r, data=corpo,
                              method="PATCH", headers=_firebase_headers())
                        with urllib.request.urlopen(req, timeout=15) as resp:
                            if resp.status == 200:
                                ok_r += 1
                                win.after(0, lambda: log("       OK enviado", "ok"))
                            else:
                                err_r.append(nome_r)
                                win.after(0, lambda s=resp.status: log(
                                    "       ERRO HTTP "+str(s), "err"))
                    except Exception as ex:
                        err_r.append(nome_r)
                        win.after(0, lambda e=str(ex)[:60]: log("       ERRO: "+e, "err"))

                    perc_r = int(2 + 38*(idx_p+1)/max(total_p,1))
                    win.after(0, lambda p=perc_r: sp(p))

                win.after(0, lambda o=ok_r, e=len(err_r): log(
                    "  Resultado: "+str(o)+" ok"+(", "+str(e)+" erro(s)" if e else ""),
                    "ok" if not e else "warn"))

                # ═══ BOLÕES ══════════════════════════════════════════
                win.after(0, lambda: log("", "dim"))
                win.after(0, lambda: log("  BOLOES NO SITE", "head"))
                win.after(0, lambda: log("  "+"-"*50, "dim"))
                win.after(0, lambda: sp(42))
                win.after(0, lambda: footer_lbl.configure(text="Etapa 2/2 — Publicando boloes"))

                boloes = db2.fetchall(
                    "SELECT * FROM boloes WHERE encerrado=0 ORDER BY id")
                total_b = len(boloes)
                win.after(0, lambda n=total_b: log(
                    "  Total: "+str(n)+" bolao(oes) ativo(s)", "info"))

                ok_b = 0; err_b = []
                for idx_b, bol in enumerate(boloes):
                    bd = dict(bol); bid = bd["id"]
                    vt = float(bd.get("valor_total") or 0)
                    nome_b = bd["nome"]

                    win.after(0, lambda nm=nome_b, i=idx_b, t=total_b: log(
                        "  ["+str(i+1).zfill(2)+"/"+str(t).zfill(2)+"] "+nm, "name"))

                    if vt <= 0:
                        win.after(0, lambda: log("       sem valor de cota — pulado", "warn"))
                        continue

                    partic   = db2.fetchall(
                        "SELECT * FROM participantes WHERE bolao_id=? AND ativo=1 ORDER BY nome", (bid,))
                    adm_low  = (bd.get("adm_nome") or "").strip().lower()
                    adm_pg   = bd.get("adm_paga", 0)
                    pag_rows = db2.fetchall(
                        "SELECT participante_id, SUM(valor) as t, MIN(data_pagamento) as dt "
                        "FROM pagamentos WHERE bolao_id=? GROUP BY participante_id", (bid,))
                    pag_map  = {r["participante_id"]: r for r in pag_rows}

                    adm_cad = False; lista = []
                    for pt in partic:
                        pt_d = dict(pt); pid3 = pt_d["id"]
                        pr   = pag_map.get(pid3)
                        pago = float(pr["t"] or 0) if pr else 0.0
                        dt_c = pr["dt"] if pr and pr["dt"] else ""
                        ve   = float(pt_d.get("valor_esperado") or 0)
                        eh_adm = bool(pt_d.get("is_adm")) or (
                            adm_low and adm_low in pt_d["nome"].lower())
                        if eh_adm: adm_cad = True
                        if eh_adm and not adm_pg:
                            n_c=1; sit="quitado"; vp=0
                        else:
                            n_c = max(1,round(ve/vt)) if vt>0 and ve>0 else 1
                            sit = "quitado" if pago>=ve else "em_andamento"
                            vp  = int(round(pago))
                        tel3 = _re_c.sub(r"\D","",str(pt_d.get("telefone") or ""))
                        sit_ic = "Q" if sit=="quitado" else "A"
                        lista.append({"nome":pt_d["nome"],"telefone":tel3,
                            "valorPago":vp,"situacao":sit,
                            "quantidadeCotas":n_c,"dataCadastro":dt_c})
                        win.after(0, lambda nm=pt_d["nome"], s=sit_ic, v=vp: log(
                            "       ["+s+"] "+nm+" R$"+str(v), "dim"))

                    if adm_low and not adm_pg and not adm_cad:
                        lista.append({"nome":bd.get("adm_nome",""),"telefone":"",
                            "valorPago":0,"situacao":"quitado","quantidadeCotas":1,"dataCadastro":""})

                    n_nome = nome_b.lower()
                    if "lotofacil" in n_nome or "lotofácil" in n_nome: lot="lotofacil"
                    elif "quina" in n_nome: lot="quina"
                    elif "lotomania" in n_nome: lot="lotomania"
                    else: lot="mega"

                    # Reaproveita o ID salvo na primeira publicacao (se
                    # houver) em vez de recalcular do nome atual - renomear
                    # o bolao nao pode gerar um documento novo no Firebase
                    doc_id_salvo = bd.get("firebase_doc_id")
                    if doc_id_salvo:
                        doc_id_b = doc_id_salvo
                    else:
                        chave = slug(nome_b).replace("_"," ")
                        doc_id_b = FIREBASE_IDS.get(chave, "bolao_"+slug(nome_b))

                    valores_part = []
                    for p in lista:
                        tel_fb = _re_c.sub(r"\D","",str(p.get("telefone") or ""))
                        valores_part.append({"mapValue":{"fields":{
                            "nome":{"stringValue":p["nome"]},
                            "telefone":{"stringValue":tel_fb},
                            "valorPago":{"doubleValue":float(p["valorPago"])},
                            "situacao":{"stringValue":p["situacao"]},
                            "quantidadeCotas":{"integerValue":str(int(p["quantidadeCotas"]))},
                            "dataCadastro":{"stringValue":p["dataCadastro"]},
                        }}})

                    doc_b = {"fields":{
                        "titulo":{"stringValue":nome_b},"loteria":{"stringValue":lot},
                        "valorPorCota":{"doubleValue":round(vt, 2)},"dataLimite":{"stringValue":""},
                        "admin":{"booleanValue":True},
                        "participantes":{"arrayValue":{"values":valores_part}},
                    }}

                    corpo_b = json.dumps(doc_b, ensure_ascii=False).encode("utf-8")
                    win.after(0, lambda did=doc_id_b: log(
                        "       enviando "+did+"...", "info"))
                    # Se já existe firebase_doc_id salvo, o bolão foi publicado
                    # antes: exige que o documento ainda exista no Firestore,
                    # pra não recriar um bolão que o admin excluiu no site.
                    url_b = _firestore_patch_url(FIREBASE_BASE+doc_id_b,
                        ["titulo","loteria","valorPorCota","dataLimite","admin","participantes"],
                        exigir_existente=bool(doc_id_salvo))
                    try:
                        req = urllib.request.Request(url_b, data=corpo_b,
                              method="PATCH", headers=_firebase_headers())
                        with urllib.request.urlopen(req, timeout=20) as resp:
                            if resp.status == 200:
                                ok_b += 1
                                if not doc_id_salvo:
                                    _conn.execute(
                                        "UPDATE boloes SET firebase_doc_id=? WHERE id=?",
                                        (doc_id_b, bid))
                                    _conn.commit()
                                n_q   = sum(1 for x in lista if x["situacao"]=="quitado")
                                n_and = len(lista) - n_q
                                win.after(0, lambda nm=nome_b, q=n_q, a=n_and: log(
                                    "       publicado! "+str(q)+" quitados, "+str(a)+" em andamento", "ok"))
                            else:
                                err_b.append(nome_b)
                                win.after(0, lambda s=resp.status: log("       ERRO HTTP "+str(s), "err"))
                    except urllib.error.HTTPError as e:
                        corpo_erro_b = ""
                        try: corpo_erro_b = e.read().decode("utf-8")
                        except: pass
                        if doc_id_salvo and e.code in (400,404,409) and "FAILED_PRECONDITION" in corpo_erro_b:
                            _conn.execute("UPDATE boloes SET encerrado=1 WHERE id=?", (bid,))
                            _conn.commit()
                            win.after(0, lambda: log(
                                "       bolão foi excluído no site — marcado como encerrado aqui", "warn"))
                        else:
                            err_b.append(nome_b)
                            win.after(0, lambda e2=str(e)[:80]: log("       ERRO: "+e2, "err"))
                    except Exception as ex:
                        err_b.append(nome_b)
                        win.after(0, lambda e=str(ex)[:80]: log("       ERRO: "+e, "err"))

                    perc_b = int(42 + 55*(idx_b+1)/max(total_b,1))
                    win.after(0, lambda p=perc_b: sp(p))

                    if idx_b < total_b-1:
                        win.after(0, lambda: log("       aguardando 2s...", "muted"))
                        _time.sleep(2)

                _conn.close()
                win.after(0, lambda: sp(100))
                win.after(0, lambda: log("", "dim"))
                houve_erro = bool(err_r or err_b)
                win.after(0, lambda o1=ok_r, o2=ok_b, e1=len(err_r), e2=len(err_b): log(
                    "  CONCLUIDO — Reservas: "+str(o1)+" ok | Boloes: "+str(o2)+" ok"
                    +(" | "+str(e1+e2)+" erro(s)" if e1+e2 else ""),
                    "ok" if not houve_erro else "warn"))

                if not houve_erro:
                    win.after(0, lambda: footer_lbl.configure(
                        text="Sincronizacao concluida! Fechando em 3 segundos..."))
                    win.after(0, lambda: status_lbl.configure(
                        text="Tudo pronto! Sistema encerrando...", fg="#1D9E75"))
                    # Contador regressivo visual
                    for i in range(3, 0, -1):
                        win.after(0, lambda n=i: footer_lbl.configure(
                            text="Fechando em "+str(n)+"s..."))
                        _time.sleep(1)
                    win.after(0, _fechar_tudo)
                else:
                    # Alguns itens não sincronizaram: não fecha sozinho, senão
                    # o usuário nunca fica sabendo que dados ficaram de fora.
                    win.after(0, lambda: log(
                        "  Alguns itens NAO foram sincronizados. Feche e tente "
                        "novamente mais tarde (confira sua internet/login).", "err"))
                    win.after(0, lambda: status_lbl.configure(
                        text="Sincronizacao incompleta — revise os erros acima", fg="#ff5252"))
                    win.after(0, _liberar_fechamento_manual)

            except Exception as ex:
                win.after(0, lambda e=str(ex): log("ERRO GERAL: "+e, "err"))
                win.after(0, lambda: status_lbl.configure(
                    text="Falha na sincronizacao — revise o erro acima", fg="#ff5252"))
                win.after(0, _liberar_fechamento_manual)

        # Centraliza
        win.update_idletasks()
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        win.geometry("720x640+"+str((sw-720)//2)+"+"+str((sh-640)//2))
        win.lift(); win.focus_force()
        threading.Thread(target=_run, daemon=True).start()

if __name__ == "__main__":
    root=tk.Tk()
    BolaoApp(root)
    root.mainloop()
